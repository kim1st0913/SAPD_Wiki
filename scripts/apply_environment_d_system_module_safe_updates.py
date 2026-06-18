#!/usr/bin/env python3
"""Apply confirmed D-class system/module safe updates to the source workbook.

Only updates format-safe H-column source cells from the generated candidate.
Merged ranges and styles are preserved by writing to the top-left cell only.
"""

from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
SHEET_NAME = "作用域-安全技术服务-安全技术模块映射"
CANDIDATE_PATH = (
    ROOT
    / "data/exports/worker-verify/environment-d-system-module-candidate/environment-d-system-module-candidate.json"
)
OUTPUT_DIR = ROOT / "data/exports/worker-verify/environment-d-system-module-candidate/formal-source-apply"


def merged_ranges_snapshot(ws) -> list[str]:
    return sorted(str(item) for item in ws.merged_cells.ranges)


def main() -> None:
    timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    apply_dir = OUTPUT_DIR / timestamp
    backup_dir = apply_dir / "source-excel-backup"
    backup_dir.mkdir(parents=True, exist_ok=True)

    candidate = json.loads(CANDIDATE_PATH.read_text(encoding="utf-8"))
    targets = candidate.get("formatSafeWriteTargets") or []
    if not targets:
        raise SystemExit("No formatSafeWriteTargets found in candidate JSON.")

    backup_path = backup_dir / f"wiki sample.before-environment-d-system-module-{timestamp}.xlsx"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    before_merges = merged_ranges_snapshot(ws)

    applied: list[dict] = []
    skipped: list[dict] = []
    for target in targets:
        cell_ref = target.get("sourceCell")
        new_value = target.get("candidateSecuritySystem")
        old_value = target.get("currentRawSecuritySystemValue")
        if not cell_ref or not new_value:
            skipped.append({**target, "reason": "missing sourceCell or candidateSecuritySystem"})
            continue
        cell = ws[cell_ref]
        actual_before = cell.value
        if actual_before != old_value:
            skipped.append(
                {
                    **target,
                    "reason": "current value mismatch",
                    "expectedCurrentValue": old_value,
                    "actualCurrentValue": actual_before,
                }
            )
            continue
        cell.value = new_value
        applied.append(
            {
                "sourceCell": cell_ref,
                "mergedRange": target.get("mergedRange"),
                "from": old_value,
                "to": new_value,
                "normalizedIssueRows": target.get("normalizedIssueRows"),
                "excelRows": target.get("excelRows"),
                "modules": target.get("modules"),
                "objectContextCount": target.get("objectContextCount"),
            }
        )

    after_merges = merged_ranges_snapshot(ws)
    if before_merges != after_merges:
        raise SystemExit("Merged ranges changed before save; aborting.")

    if applied:
        wb.save(WORKBOOK_PATH)

    wb_check = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    ws_check = wb_check[SHEET_NAME]
    check_merges = merged_ranges_snapshot(ws_check)
    verification = []
    for item in applied:
        verification.append(
            {
                "sourceCell": item["sourceCell"],
                "expected": item["to"],
                "actual": ws_check[item["sourceCell"]].value,
                "ok": ws_check[item["sourceCell"]].value == item["to"],
            }
        )

    report = {
        "status": "pass"
        if applied and not skipped and check_merges == before_merges and all(item["ok"] for item in verification)
        else "warnings",
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sheet": SHEET_NAME,
        "candidate": str(CANDIDATE_PATH.relative_to(ROOT)),
        "backup": str(backup_path.relative_to(ROOT)),
        "applyDir": str(apply_dir.relative_to(ROOT)),
        "targetCount": len(targets),
        "appliedCount": len(applied),
        "skippedCount": len(skipped),
        "mergedRangesBeforeCount": len(before_merges),
        "mergedRangesAfterCount": len(check_merges),
        "mergedRangesUnchanged": before_merges == check_merges,
        "applied": applied,
        "skipped": skipped,
        "verification": verification,
    }

    report_path = apply_dir / "environment-d-system-module-source-apply-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    md_lines = [
        "# D 类系统-模块原始表写回报告",
        "",
        f"- 状态：`{report['status']}`",
        f"- 原始表：`{report['workbook']}`",
        f"- Sheet：`{SHEET_NAME}`",
        f"- 备份：`{report['backup']}`",
        f"- 候选目标：`{report['targetCount']}`",
        f"- 已写回：`{report['appliedCount']}`",
        f"- 跳过：`{report['skippedCount']}`",
        f"- 合并区域保持不变：`{report['mergedRangesUnchanged']}`",
        "",
        "## 写回目标",
        "",
        "| 单元格 | 合并区域 | 原值 | 新值 | 规范化问题行 | Excel 行 | 模块 |",
        "|---|---|---|---|---:|---|---|",
    ]
    for item in applied:
        md_lines.append(
            "| "
            + " | ".join(
                [
                    f"`{item['sourceCell']}`",
                    f"`{item.get('mergedRange') or ''}`",
                    str(item.get("from") or ""),
                    str(item.get("to") or ""),
                    str(item.get("normalizedIssueRows") or ""),
                    ", ".join(str(row) for row in item.get("excelRows") or []),
                    " / ".join(str(module) for module in item.get("modules") or []),
                ]
            )
            + " |"
        )
    (apply_dir / "environment-d-system-module-source-apply-report.md").write_text(
        "\n".join(md_lines) + "\n", encoding="utf-8"
    )

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
