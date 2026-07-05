#!/usr/bin/env python3
"""Apply confirmed environment object-service supplements to the source workbook."""

from __future__ import annotations

import copy
import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
SHEET_NAME = "作用域-安全技术服务-安全技术模块映射"
AUDIT_PATH = ROOT / "data/exports/worker-verify/environment-object-service-overexpansion-audit/audit.json"
OUTPUT_ROOT = ROOT / "data/exports/worker-verify/environment-object-service-source-supplement-apply"
MODULE_TITLE_ALIASES = {
    "安全工作区": "终端安全工作区",
}


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.replace("\xa0", " ").replace("\u3000", " ")).strip()


def split_lines(value: Any) -> list[str]:
    text = "" if value is None else str(value)
    return [norm(line) for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if norm(line)]


def service_scope_code(service_code: str) -> str:
    return service_code.split("&", 1)[0] if "&" in service_code else ""


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def merged_value_lookup(ws) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, col)] = anchor_value
    return values


def cell_text(ws, merged_values: dict[tuple[int, int], Any], row: int, col: int) -> str:
    return norm(merged_values.get((row, col), ws.cell(row, col).value))


def row_context(ws, merged_values: dict[tuple[int, int], Any], row: int) -> tuple[str, str, str]:
    return (
        cell_text(ws, merged_values, row, 2),
        cell_text(ws, merged_values, row, 3),
        cell_text(ws, merged_values, row, 4),
    )


def source_rows_for(
    ws,
    merged_values: dict[tuple[int, int], Any],
    context: tuple[str, str, str],
    service_code: str,
) -> list[int]:
    rows: list[int] = []
    for row in range(3, ws.max_row + 1):
        if row_context(ws, merged_values, row) == context and service_code in cell_text(ws, merged_values, row, 6):
            rows.append(row)
    return rows


def target_rows_for(ws, merged_values: dict[tuple[int, int], Any], context: tuple[str, str, str]) -> list[int]:
    return [row for row in range(3, ws.max_row + 1) if row_context(ws, merged_values, row) == context]


def authoritative_module_titles(wb) -> set[str]:
    sheet_name = "安全技术模块清单"
    if sheet_name not in wb.sheetnames:
        return set()
    ws = wb[sheet_name]
    merged_values = merged_value_lookup(ws)
    titles: set[str] = set()
    for row in range(3, ws.max_row + 1):
        title = cell_text(ws, merged_values, row, 4)
        if title and not title.replace(".", "", 1).isdigit():
            titles.add(title)
    return titles


def authoritative_scope_text_by_code(wb) -> dict[str, str]:
    sheet_name = "安全能力作用域目录"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    result: dict[str, str] = {}
    for row in range(3, ws.max_row + 1):
        raw = norm(ws.cell(row, 3).value)
        if not raw:
            continue
        if " " in raw:
            code, title = raw.split(" ", 1)
            result[code] = f"{code} {norm(title)}"
    return result


def canonical_module_title(title: str, module_titles: set[str]) -> str:
    text = norm(title)
    if text in module_titles:
        return text
    alias = MODULE_TITLE_ALIASES.get(text)
    if alias in module_titles:
        return alias
    return text


def choose_source_row(ws, merged_values: dict[tuple[int, int], Any], rows: list[int], module_titles: set[str]) -> int:
    if not rows:
        raise ValueError("empty source rows")
    for row in rows:
        if canonical_module_title(cell_text(ws, merged_values, row, 7), module_titles) in module_titles:
            return row
    return rows[0]


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy.copy(source._style)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        copy_cell_style(source, target)


@dataclass
class Addition:
    service_code: str
    service_display: str
    source_context_key: str
    source_row: int
    source_scope_text: str
    module_text: str
    system_text: str
    inserted_row: int | None = None


@dataclass
class TargetGroup:
    context: tuple[str, str, str]
    context_key: str
    first_row: int
    last_row: int
    insert_at: int
    additions: list[Addition] = field(default_factory=list)


def source_context_from_hit(hit: dict[str, Any]) -> tuple[str, str, str]:
    return (norm(hit.get("environment")), norm(hit.get("environmentSegment")), norm(hit.get("informationObject")))


def build_target_groups(wb) -> dict[tuple[str, str, str], TargetGroup]:
    ws = wb[SHEET_NAME]
    merged_values = merged_value_lookup(ws)
    module_titles = authoritative_module_titles(wb)
    audit = load_json(AUDIT_PATH)
    items = audit.get("sameObjectTitleLeakSuspects") or []
    if not items:
        raise SystemExit("No supplement candidates found in audit JSON.")

    groups: dict[tuple[str, str, str], TargetGroup] = {}
    seen_pairs: set[tuple[tuple[str, str, str], str]] = set()
    for item in items:
        target_context = (
            norm(item.get("environment")),
            norm(item.get("environmentSegment")),
            norm(item.get("informationObject")),
        )
        service = item.get("extraService") or {}
        service_code = norm(service.get("code"))
        service_display = norm(service.get("display") or f"{service_code} {service.get('title') or ''}")
        if not service_code:
            continue
        pair_key = (target_context, service_code)
        if pair_key in seen_pairs:
            continue
        seen_pairs.add(pair_key)

        target_rows = target_rows_for(ws, merged_values, target_context)
        if not target_rows:
            raise RuntimeError(f"Target context not found: {target_context}")
        group = groups.setdefault(
            target_context,
            TargetGroup(
                context=target_context,
                context_key=" / ".join(target_context),
                first_row=min(target_rows),
                last_row=max(target_rows),
                insert_at=max(target_rows) + 1,
            ),
        )

        source_candidates: list[tuple[int, str]] = []
        for hit in item.get("sourceContextsWithSameObjectAndService") or []:
            source_context = source_context_from_hit(hit)
            rows = source_rows_for(ws, merged_values, source_context, service_code)
            for row in rows:
                source_candidates.append((row, hit.get("contextKey") or " / ".join(source_context)))
        if not source_candidates:
            raise RuntimeError(f"Source row not found for {group.context_key} -> {service_code}")
        canonical_candidates = [
            (row, context_key)
            for row, context_key in source_candidates
            if canonical_module_title(cell_text(ws, merged_values, row, 7), module_titles) in module_titles
        ]
        source_row, source_context_key = (canonical_candidates or source_candidates)[0]
        group.additions.append(
            Addition(
                service_code=service_code,
                service_display=service_display,
                source_context_key=norm(source_context_key),
                source_row=source_row,
                source_scope_text=cell_text(ws, merged_values, source_row, 5),
                module_text=canonical_module_title(cell_text(ws, merged_values, source_row, 7), module_titles),
                system_text=cell_text(ws, merged_values, source_row, 8) or "N/A",
            )
        )

    for group in groups.values():
        group.additions.sort(key=lambda item: (service_scope_code(item.service_code), item.service_code))
    return groups


def merge_snapshot(ws) -> list[CellRange]:
    return [CellRange(str(item)) for item in ws.merged_cells.ranges]


def adjusted_ranges_after_insertions(ranges: list[CellRange], groups_desc: list[TargetGroup]) -> list[CellRange]:
    adjusted = [CellRange(str(item)) for item in ranges]
    for group in groups_desc:
        amount = len(group.additions)
        index = group.insert_at
        for merged_range in adjusted:
            if merged_range.min_row >= index:
                merged_range.shift(row_shift=amount)
            elif merged_range.min_row < index <= merged_range.max_row:
                merged_range.max_row += amount
            elif (
                merged_range.max_row + 1 == index
                and merged_range.min_row <= group.first_row <= merged_range.max_row
                and merged_range.max_row == group.last_row
                and merged_range.min_col in {2, 3, 4, 5}
                and merged_range.max_col == merged_range.min_col
            ):
                merged_range.max_row += amount
    return adjusted


def row_after_insertions(original_row: int, groups: list[TargetGroup]) -> int:
    return original_row + sum(len(group.additions) for group in groups if group.insert_at <= original_row)


def inserted_row_after_all_insertions(group: TargetGroup, offset: int, groups: list[TargetGroup]) -> int:
    return group.insert_at + offset + sum(len(other.additions) for other in groups if other.insert_at < group.insert_at)


def copy_semantic_source_cell_styles(ws, groups: list[TargetGroup]) -> None:
    for group in groups:
        for addition in group.additions:
            if addition.inserted_row is None:
                continue
            source_row = row_after_insertions(addition.source_row, groups)
            for col in (6, 7, 8):
                copy_cell_style(ws.cell(source_row, col), ws.cell(addition.inserted_row, col))


def expand_scope_cell_if_needed(ws, groups_desc: list[TargetGroup], scope_text_by_code: dict[str, str]) -> list[dict[str, Any]]:
    updates: list[dict[str, Any]] = []
    groups_all = list(groups_desc)
    for group in groups_desc:
        anchor_row = row_after_insertions(group.first_row, groups_all)
        anchor_cell = ws.cell(anchor_row, 5)
        existing_lines = split_lines(anchor_cell.value)
        existing_codes = {line.split(" ", 1)[0] for line in existing_lines if " " in line}
        changed = False
        for addition in group.additions:
            code = service_scope_code(addition.service_code)
            if code and code not in existing_codes:
                scope_text = scope_text_by_code.get(code) or code
                if code == "I-NT" and any(line.startswith("I-US ") for line in existing_lines):
                    insert_at = next(index for index, line in enumerate(existing_lines) if line.startswith("I-US "))
                    existing_lines.insert(insert_at, scope_text)
                else:
                    existing_lines.append(scope_text)
                existing_codes.add(code)
                changed = True
        if changed:
            before = anchor_cell.value
            after = "\n".join(existing_lines)
            anchor_cell.value = after
            updates.append(
                {
                    "contextKey": group.context_key,
                    "cell": anchor_cell.coordinate,
                    "before": before,
                    "after": after,
                }
            )
    return updates


def main() -> None:
    timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    apply_dir = OUTPUT_ROOT / timestamp
    backup_dir = apply_dir / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"wiki-sample.before-environment-object-service-supplement-{timestamp}.xlsx"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]
    groups = build_target_groups(wb)
    groups_desc = sorted(groups.values(), key=lambda group: group.insert_at, reverse=True)
    original_ranges = merge_snapshot(ws)
    scope_text_by_code = authoritative_scope_text_by_code(wb)

    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    for group in groups_desc:
        amount = len(group.additions)
        ws.insert_rows(group.insert_at, amount)
        for offset, addition in enumerate(group.additions):
            row = group.insert_at + offset
            copy_row_style(ws, group.last_row, row)
            ws.cell(row, 6).value = addition.service_display
            ws.cell(row, 7).value = addition.module_text
            ws.cell(row, 8).value = addition.system_text
            addition.inserted_row = inserted_row_after_all_insertions(group, offset, groups_desc)

    copy_semantic_source_cell_styles(ws, groups_desc)

    adjusted_ranges = adjusted_ranges_after_insertions(original_ranges, groups_desc)
    for merged_range in adjusted_ranges:
        ws.merge_cells(str(merged_range))

    scope_updates = expand_scope_cell_if_needed(ws, groups_desc, scope_text_by_code)
    wb.save(WORKBOOK_PATH)
    wb.close()

    check_wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    check_ws = check_wb[SHEET_NAME]
    check_merged_values = merged_value_lookup(check_ws)

    verification: list[dict[str, Any]] = []
    for group in sorted(groups.values(), key=lambda item: item.context_key):
        for addition in group.additions:
            found_rows = source_rows_for(check_ws, check_merged_values, group.context, addition.service_code)
            verification.append(
                {
                    "contextKey": group.context_key,
                    "service": addition.service_display,
                    "insertedRow": addition.inserted_row,
                    "foundRowsAfterApply": found_rows,
                    "ok": bool(found_rows),
                }
            )
    check_wb.close()

    report = {
        "status": "pass" if all(item["ok"] for item in verification) else "failed",
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sheet": SHEET_NAME,
        "audit": str(AUDIT_PATH.relative_to(ROOT)),
        "backup": str(backup_path.relative_to(ROOT)),
        "targetContextCount": len(groups),
        "addedRowCount": sum(len(group.additions) for group in groups.values()),
        "scopeUpdateCount": len(scope_updates),
        "mergedRangeCountBefore": len(original_ranges),
        "mergedRangeCountAfter": len(adjusted_ranges),
        "targets": [
            {
                "contextKey": group.context_key,
                "originalRows": [group.first_row, group.last_row],
                "insertAtOriginal": group.insert_at,
                "added": [
                    {
                        "insertedRow": addition.inserted_row,
                        "service": addition.service_display,
                        "moduleOrMeasure": addition.module_text,
                        "securitySystem": addition.system_text,
                        "sourceContextKey": addition.source_context_key,
                        "sourceRow": addition.source_row,
                    }
                    for addition in group.additions
                ],
            }
            for group in sorted(groups.values(), key=lambda item: item.context_key)
        ],
        "scopeUpdates": scope_updates,
        "verification": verification,
    }

    report_path = apply_dir / "source-supplement-apply-report.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md_lines = [
        "# 信息化对象安全技术服务源表补充写回报告",
        "",
        f"- 状态：`{report['status']}`",
        f"- 原始表：`{report['workbook']}`",
        f"- 备份：`{report['backup']}`",
        f"- 目标上下文数：`{report['targetContextCount']}`",
        f"- 新增行数：`{report['addedRowCount']}`",
        f"- 作用域补充数：`{report['scopeUpdateCount']}`",
        "",
        "| 目标上下文 | 插入行 | 安全技术服务 | 模块/措施 | 安全系统 | 来源上下文 | 来源行 |",
        "|---|---:|---|---|---|---|---:|",
    ]
    for target in report["targets"]:
        for addition in target["added"]:
            md_lines.append(
                "| "
                + " | ".join(
                    [
                        target["contextKey"],
                        str(addition["insertedRow"]),
                        addition["service"],
                        addition["moduleOrMeasure"],
                        addition["securitySystem"],
                        addition["sourceContextKey"],
                        str(addition["sourceRow"]),
                    ]
                )
                + " |"
            )
    (apply_dir / "source-supplement-apply-report.md").write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "status": report["status"],
                "backup": report["backup"],
                "targetContextCount": report["targetContextCount"],
                "addedRowCount": report["addedRowCount"],
                "scopeUpdateCount": report["scopeUpdateCount"],
                "report": str(report_path.relative_to(ROOT)),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
