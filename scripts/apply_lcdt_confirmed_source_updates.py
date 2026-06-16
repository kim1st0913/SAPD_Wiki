#!/usr/bin/env python3
from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
OUT_DIR = ROOT / "data/exports/worker-verify/lcdt-source-update"
BACKUP_DIR = OUT_DIR / "source-excel-backup"
LIFECYCLE_SHEET = "LC-DT 数据生命周期"
MAPPING_SHEET = "LC-DT 安全技术服务、模块、策略映射表"

SERVICE_REPLACEMENTS = {
    "I-DI&T-AS.AD-03 数据备份": "I-DI&T-AS.DG-03 数据备份",
    "I-DI&T-PD.DP-01 应用页面水印": "I-DI&T-PD.DP-01 数据内容水印",
    "I-DI&T-PD.DP-02 应用动态数据脱敏": "I-DI&T-PD.DP-02 静态数据脱敏",
}

MODULE_REPLACEMENTS = {
    "数据加密": "数据加密和令牌化",
    "数据脱敏": "数据脱敏(去标识化)",
    "零信任访问代理（独立部署，接替/协同应用系统相关技术模块）": "零信任访问代理",
}

LIFECYCLE_MODULE_SUPPLEMENT = "数据流转监测和泄漏防护"


def now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def normalize_lines(value: Any) -> str:
    lines = []
    seen = set()
    for raw in str(value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        item = raw.strip()
        if not item or item in seen:
            continue
        lines.append(item)
        seen.add(item)
    return "\n".join(lines)


def replace_value(value: Any) -> str:
    lines = []
    for raw in str(value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        item = raw.strip()
        if not item:
            continue
        lines.append(SERVICE_REPLACEMENTS.get(item, MODULE_REPLACEMENTS.get(item, item)))
    return normalize_lines("\n".join(lines))


def append_line(value: Any, line: str) -> str:
    lines = normalize_lines(value).split("\n") if normalize_lines(value) else []
    if line not in lines:
        lines.append(line)
    return "\n".join(lines)


def cell_ref(sheet_name: str, coord: str, before: Any, after: Any) -> dict[str, Any]:
    return {
        "sheet": sheet_name,
        "cell": coord,
        "before": before,
        "after": after,
    }


def update_target_cells(ws, coords: list[str], changes: list[dict[str, Any]]) -> None:
    for coord in coords:
        cell = ws[coord]
        before = cell.value
        after = replace_value(before)
        if before != after:
            cell.value = after
            changes.append(cell_ref(ws.title, coord, before, after))


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"workbook not found: {WORKBOOK_PATH}")
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f"wiki sample.before-lcdt-confirmed-source-updates.{now_stamp()}.xlsx"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    changes: list[dict[str, Any]] = []

    lifecycle = wb[LIFECYCLE_SHEET]
    mapping = wb[MAPPING_SHEET]

    update_target_cells(lifecycle, ["H12", "H22", "H26", "I12", "I17", "I22", "I26", "I28"], changes)
    update_target_cells(
        mapping,
        [
            "M27",
            "M29",
            "M39",
            "M40",
            "M53",
            "M55",
            "N14",
            "N15",
            "N16",
            "N17",
            "N18",
            "N19",
            "N20",
            "N26",
            "N39",
            "N40",
            "N41",
            "N43",
            "N52",
            "N53",
            "N54",
            "N55",
            "N56",
            "N57",
            "N61",
            "N63",
        ],
        changes,
    )

    before = lifecycle["I22"].value
    after = append_line(before, LIFECYCLE_MODULE_SUPPLEMENT)
    if before != after:
        lifecycle["I22"].value = after
        changes.append(cell_ref(lifecycle.title, "I22", before, after))

    wb.save(WORKBOOK_PATH)
    report = {
        "status": "applied",
        "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "backup": str(backup_path.relative_to(ROOT)),
        "changeCount": len(changes),
        "changes": changes,
    }
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "lcdt-confirmed-source-updates-apply-result.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = [
        "# LC-DT confirmed source updates apply result",
        "",
        f"- status: `{report['status']}`",
        f"- workbook: `{report['workbook']}`",
        f"- backup: `{report['backup']}`",
        f"- changeCount: `{report['changeCount']}`",
        "",
        "## Changes",
        "",
    ]
    for change in changes:
        lines.append(f"- `{change['sheet']}!{change['cell']}`")
    (OUT_DIR / "lcdt-confirmed-source-updates-apply-result.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
