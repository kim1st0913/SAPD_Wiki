#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
OUTPUT_ROOT = ROOT / "data/exports/worker-verify/lcdt-module-alias-normalization"

CONFIRM_FLAG = "--confirm-lcdt-module-alias-normalization"

TARGET_COLUMNS = (
    ("LC-DT 数据生命周期", "I"),
    ("LC-DT 安全技术服务、模块、策略映射表", "N"),
)

ALIASES = {
    "数据加密": "数据加密和令牌化",
    "数据脱敏": "数据脱敏(去标识化)",
    "零信任访问代理（独立部署，接替/协同应用系统相关技术模块）": "零信任访问代理",
}


def now_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def relative(path: Path) -> str:
    return str(path.relative_to(ROOT))


def split_lines(value: Any) -> list[str]:
    return [line.strip() for line in str(value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")]


def normalize_cell_value(value: Any) -> tuple[str, list[dict[str, Any]]]:
    lines = split_lines(value)
    replacements: list[dict[str, Any]] = []
    normalized_lines: list[str] = []
    for index, line in enumerate(lines, start=1):
        replacement = ALIASES.get(line)
        if replacement:
            normalized_lines.append(replacement)
            replacements.append({"line": index, "before": line, "after": replacement})
        else:
            normalized_lines.append(line)
    return "\n".join(normalized_lines), replacements


def build_report(mode: str, out_dir: Path, backup_path: Path | None, changes: list[dict[str, Any]]) -> dict[str, Any]:
    replacement_count = sum(len(change["replacements"]) for change in changes)
    return {
        "status": "applied" if mode == "apply" else "dry_run",
        "mode": mode,
        "workbook": relative(WORKBOOK_PATH),
        "backup": relative(backup_path) if backup_path else None,
        "targetColumns": [{"sheet": sheet, "column": column} for sheet, column in TARGET_COLUMNS],
        "aliasMap": ALIASES,
        "changedCellCount": len(changes),
        "replacementCount": replacement_count,
        "changes": changes,
        "outputDir": relative(out_dir),
    }


def write_report(out_dir: Path, report: dict[str, Any]) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "lcdt-module-alias-normalization-report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    lines = [
        "# LC-DT module alias normalization report",
        "",
        f"- status: `{report['status']}`",
        f"- workbook: `{report['workbook']}`",
        f"- backup: `{report['backup'] or ''}`",
        f"- changedCellCount: `{report['changedCellCount']}`",
        f"- replacementCount: `{report['replacementCount']}`",
        "",
        "## Changes",
        "",
    ]
    if not report["changes"]:
        lines.append("- 无")
    for change in report["changes"]:
        lines.append(f"- `{change['sheet']}!{change['cell']}`")
        for replacement in change["replacements"]:
            lines.append(f"  - `{replacement['before']}` -> `{replacement['after']}`")
    (out_dir / "lcdt-module-alias-normalization-report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize confirmed LC-DT module aliases in the source Excel workbook.")
    parser.add_argument("--apply", action="store_true", help="Write changes to the source workbook. Default is dry-run.")
    parser.add_argument(
        CONFIRM_FLAG,
        dest="confirm",
        action="store_true",
        help="Required with --apply to avoid accidental source workbook edits.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.apply and not args.confirm:
        raise SystemExit(
            json.dumps(
                {
                    "status": "blocked",
                    "reason": "apply_requires_confirmation",
                    "requiredFlag": CONFIRM_FLAG,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"workbook not found: {WORKBOOK_PATH}")

    stamp = now_stamp()
    out_dir = OUTPUT_ROOT / stamp
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    changes: list[dict[str, Any]] = []

    for sheet_name, column in TARGET_COLUMNS:
        ws = wb[sheet_name]
        for row in range(1, ws.max_row + 1):
            cell = ws[f"{column}{row}"]
            if isinstance(cell, MergedCell) or cell.value in (None, ""):
                continue
            after, replacements = normalize_cell_value(cell.value)
            if replacements and after != cell.value:
                changes.append(
                    {
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "before": cell.value,
                        "after": after,
                        "replacements": replacements,
                    }
                )
                if args.apply:
                    cell.value = after

    backup_path: Path | None = None
    mode = "apply" if args.apply else "dry-run"
    if args.apply:
        backup_dir = out_dir / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / "wiki sample.before-lcdt-module-alias-normalization.xlsx"
        shutil.copy2(WORKBOOK_PATH, backup_path)
        if changes:
            wb.save(WORKBOOK_PATH)

    report = build_report(mode, out_dir, backup_path, changes)
    write_report(out_dir, report)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
