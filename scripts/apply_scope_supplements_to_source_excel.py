#!/usr/bin/env python3
"""Write service-derived scope supplements back to the source workbook.

The script only edits the top-left cell of existing E-column merged ranges and
does not create, delete, split, or re-merge cells.
"""

from __future__ import annotations

import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
SHEET_NAME = "作用域-安全技术服务-安全技术模块映射"
RELATIONS_PATH = ROOT / "data/exports/worker-verify/scope-service-module-mapping-relations.json"
NORMALIZED_ROWS_PATH = ROOT / "data/exports/worker-verify/scope-service-module-mapping-normalized-rows.json"
OUTPUT_DIR = ROOT / "data/exports/worker-verify"
BACKUP_DIR = OUTPUT_DIR / "source-excel-backup"
REPORT_JSON = OUTPUT_DIR / "scope-service-module-scope-writeback-report.json"
REPORT_MD = OUTPUT_DIR / "scope-service-module-scope-writeback-report.md"


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def unique(values: list[str]) -> list[str]:
    return [item for item in dict.fromkeys(value for value in values if value)]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def e_column_merged_ranges(ws) -> list[str]:
    return sorted(
        str(item)
        for item in ws.merged_cells.ranges
        if item.min_col <= 5 <= item.max_col
    )


def source_range_by_context(rows: list[dict[str, Any]]) -> dict[str, str]:
    ranges: dict[str, str] = {}
    for row in rows:
        context_key = text(row.get("contextKey"))
        scope_range = text((row.get("mergedRanges") or {}).get("scope"))
        if context_key and scope_range:
            ranges.setdefault(context_key, scope_range)
    return ranges


def existing_scope_text_by_context(rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    values: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        context_key = text(row.get("contextKey"))
        if not context_key:
            continue
        for scope in row.get("scopes") or []:
            value = text(scope.get("text") or " ".join(part for part in [scope.get("code"), scope.get("title")] if part))
            if value:
                values[context_key].append(value)
    return {key: unique(items) for key, items in values.items()}


def main() -> None:
    rows = load_json(NORMALIZED_ROWS_PATH)
    relations = load_json(RELATIONS_PATH)
    supplements = relations.get("scopeSupplementsFromServices") or []
    source_ranges = source_range_by_context(rows)
    existing_scopes = existing_scope_text_by_context(rows)

    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    ws = wb[SHEET_NAME]
    before_ranges = e_column_merged_ranges(ws)

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = BACKUP_DIR / f"wiki sample.before-scope-writeback.{timestamp}.xlsx"
    shutil.copy2(WORKBOOK_PATH, backup_path)

    supplements_by_context: dict[str, list[str]] = defaultdict(list)
    for item in supplements:
        context_key = text(item.get("objectContextKey"))
        scope = item.get("scope") or {}
        value = text(scope.get("text") or " ".join(part for part in [scope.get("code"), scope.get("title")] if part))
        if context_key and value:
            supplements_by_context[context_key].append(value)

    updates = []
    skipped = []
    for context_key, supplemental_values in sorted(supplements_by_context.items()):
        scope_range = source_ranges.get(context_key)
        if not scope_range:
            skipped.append({"objectContextKey": context_key, "reason": "未找到 E 列合并区域"})
            continue
        merged = next((item for item in ws.merged_cells.ranges if str(item) == scope_range), None)
        if not merged or merged.min_col != 5:
            skipped.append({"objectContextKey": context_key, "scopeRange": scope_range, "reason": "E 列合并区域不匹配"})
            continue
        cell = ws.cell(merged.min_row, merged.min_col)
        before_value = text(cell.value)
        merged_values = unique((existing_scopes.get(context_key) or before_value.splitlines()) + supplemental_values)
        after_value = "\n".join(merged_values)
        if after_value != before_value:
            cell.value = after_value
            updates.append(
                {
                    "objectContextKey": context_key,
                    "scopeRange": scope_range,
                    "cell": cell.coordinate,
                    "before": before_value,
                    "after": after_value,
                    "supplements": unique(supplemental_values),
                }
            )

    after_ranges_before_save = e_column_merged_ranges(ws)
    if before_ranges != after_ranges_before_save:
        raise RuntimeError("E列 merged ranges 在保存前发生变化，已中止写回")

    wb.save(WORKBOOK_PATH)
    wb.close()

    check_wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    check_ws = check_wb[SHEET_NAME]
    after_ranges = e_column_merged_ranges(check_ws)
    check_wb.close()
    if before_ranges != after_ranges:
        shutil.copy2(backup_path, WORKBOOK_PATH)
        raise RuntimeError("E列 merged ranges 保存后发生变化，已从备份恢复")

    report = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sheet": SHEET_NAME,
        "backup": str(backup_path.relative_to(ROOT)),
        "supplementContextCount": len(supplements_by_context),
        "updatedMergedCells": len(updates),
        "skipped": skipped,
        "eColumnMergedRangeCount": len(after_ranges),
        "eColumnMergedRangesPreserved": before_ranges == after_ranges,
        "updates": updates,
    }
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    REPORT_MD.write_text(
        "\n".join(
            [
                "# Scope Writeback Report",
                "",
                f"- workbook: `{report['workbook']}`",
                f"- sheet: `{SHEET_NAME}`",
                f"- backup: `{report['backup']}`",
                f"- supplementContextCount: `{report['supplementContextCount']}`",
                f"- updatedMergedCells: `{report['updatedMergedCells']}`",
                f"- skipped: `{len(skipped)}`",
                f"- eColumnMergedRangeCount: `{report['eColumnMergedRangeCount']}`",
                f"- eColumnMergedRangesPreserved: `{report['eColumnMergedRangesPreserved']}`",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    print(json.dumps({key: report[key] for key in ("workbook", "backup", "supplementContextCount", "updatedMergedCells", "skipped", "eColumnMergedRangesPreserved")}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
