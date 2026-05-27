#!/usr/bin/env python3
"""Audit capability management mappings against the source workbook.

The audit intentionally compares the exported frontend package with the
business meaning of the Excel sheets, including merged-cell inheritance and
explicit placeholder clearing. It catches high-risk ETL regressions such as
leaking a previous row's stakeholder functions into a focus that contains `/`.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from sapd_wiki.parsers import STAKEHOLDER_FUNCTION_ALIASES
from sapd_wiki.transformers import split_multivalue_text


STAKEHOLDER_COLUMNS = ((9, "决策层"), (10, "管理层"), (11, "执行层"), (12, "监督层"))


def merged_anchor_values(ws) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                values[ws.cell(row_index, column_index).coordinate] = anchor_value
    return values


def effective_value(ws, merged_values: dict[str, Any], row_index: int, column_index: int) -> Any:
    cell = ws.cell(row_index, column_index)
    return merged_values.get(cell.coordinate, cell.value)


def split_stakeholders(value: Any) -> list[str]:
    return [
        STAKEHOLDER_FUNCTION_ALIASES.get(item, item)
        for item in split_multivalue_text(value, split_on_ideographic_comma=False)
    ]


def exported_focuses(capability_tree: dict[str, Any]) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for category in capability_tree.get("categories", []):
        for domain in category.get("domains", []):
            for capability in domain.get("capabilities", []):
                for focus in capability.get("focuses", []):
                    code = str(focus.get("code") or "").strip()
                    if code:
                        rows[code] = focus
    return rows


def item_titles(items: list[dict[str, Any]]) -> list[str]:
    return sorted(
        {
            str(item.get("title") or item.get("name") or item.get("code") or "").strip()
            for item in items
            if str(item.get("title") or item.get("name") or item.get("code") or "").strip()
        }
    )


def expected_high_level_rows(workbook) -> dict[str, dict[str, Any]]:
    ws = workbook["安全能力-安全管理元素（high level）"]
    merged_values = merged_anchor_values(ws)
    expected: dict[str, dict[str, Any]] = {}
    for row_index in range(4, ws.max_row + 1):
        code = str(ws.cell(row_index, 5).value or "").strip()
        if not code:
            continue
        process_group = str(effective_value(ws, merged_values, row_index, 7) or "").strip()
        process_refs = split_multivalue_text(
            effective_value(ws, merged_values, row_index, 8),
            split_on_ideographic_comma=False,
        )
        stakeholders = {
            layer: split_stakeholders(effective_value(ws, merged_values, row_index, column_index))
            for column_index, layer in STAKEHOLDER_COLUMNS
        }
        expected[code] = {
            "row": row_index,
            "process_group": process_group,
            "process_refs": process_refs,
            "stakeholders": stakeholders,
        }
    return expected


def expected_security_work_rows(workbook) -> dict[str, set[str]]:
    ws = workbook["安全能力-安全工作"]
    expected: dict[str, set[str]] = {}
    last_focus_code = ""
    for row_index in range(4, ws.max_row + 1):
        if ws.cell(row_index, 5).value:
            last_focus_code = str(ws.cell(row_index, 5).value).strip()
        work_title = str(ws.cell(row_index, 7).value or "").strip()
        if last_focus_code and work_title:
            expected.setdefault(last_focus_code, set()).add(work_title)
    return expected


def audit_high_level(expected: dict[str, dict[str, Any]], actual: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for code, expected_row in expected.items():
        focus = actual.get(code)
        if not focus:
            issues.append({"type": "missing_focus", "code": code, "row": expected_row["row"]})
            continue
        mappings = focus.get("process_mappings") or []
        actual_refs = sorted(
            {
                mapping.get("process_reference", {}).get("title", "")
                for mapping in mappings
                if mapping.get("process_reference")
            }
        )
        expected_refs = sorted(set(expected_row["process_refs"]))
        if actual_refs != expected_refs:
            issues.append(
                {
                    "type": "process_refs_mismatch",
                    "code": code,
                    "row": expected_row["row"],
                    "expected": expected_refs,
                    "actual": actual_refs,
                }
            )
        actual_groups = sorted(
            {
                mapping.get("process_group", {}).get("title", "")
                for mapping in mappings
                if mapping.get("process_group")
            }
        )
        expected_groups = sorted(set([expected_row["process_group"]] if expected_refs and expected_row["process_group"] else []))
        if actual_groups != expected_groups:
            issues.append(
                {
                    "type": "process_group_mismatch",
                    "code": code,
                    "row": expected_row["row"],
                    "expected": expected_groups,
                    "actual": actual_groups,
                }
            )
        if not expected_refs:
            continue
        for _, layer in STAKEHOLDER_COLUMNS:
            actual_stakeholders = item_titles(
                [
                    stakeholder
                    for mapping in mappings
                    for stakeholder in (mapping.get("stakeholders") or {}).get(layer, [])
                ]
            )
            expected_stakeholders = sorted(set(expected_row["stakeholders"][layer]))
            if actual_stakeholders != expected_stakeholders:
                issues.append(
                    {
                        "type": "stakeholder_mismatch",
                        "code": code,
                        "row": expected_row["row"],
                        "layer": layer,
                        "expected": expected_stakeholders,
                        "actual": actual_stakeholders,
                    }
                )
    return issues


def audit_placeholder_leaks(expected: dict[str, dict[str, Any]], actual: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for code, expected_row in expected.items():
        focus = actual.get(code)
        if not focus:
            continue
        mappings = focus.get("process_mappings") or []
        for _, layer in STAKEHOLDER_COLUMNS:
            if expected_row["stakeholders"][layer]:
                continue
            actual_stakeholders = item_titles(
                [
                    stakeholder
                    for mapping in mappings
                    for stakeholder in (mapping.get("stakeholders") or {}).get(layer, [])
                ]
            )
            if actual_stakeholders:
                issues.append(
                    {
                        "type": "placeholder_leak",
                        "code": code,
                        "row": expected_row["row"],
                        "layer": layer,
                        "actual": actual_stakeholders,
                    }
                )
    return issues


def audit_security_works(expected: dict[str, set[str]], actual: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for code, expected_works in expected.items():
        focus = actual.get(code)
        if not focus:
            issues.append({"type": "missing_focus_for_work", "code": code, "expected": sorted(expected_works)})
            continue
        actual_works = {
            str(item.get("title") or item.get("name") or "").strip()
            for item in focus.get("security_works") or []
            if str(item.get("title") or item.get("name") or "").strip()
        }
        if actual_works != expected_works:
            issues.append(
                {
                    "type": "security_work_mismatch",
                    "code": code,
                    "expected": sorted(expected_works),
                    "actual": sorted(actual_works),
                }
            )
    for code, focus in actual.items():
        actual_works = {
            str(item.get("title") or item.get("name") or "").strip()
            for item in focus.get("security_works") or []
            if str(item.get("title") or item.get("name") or "").strip()
        }
        if actual_works and code not in expected:
            issues.append({"type": "unexpected_security_work", "code": code, "actual": sorted(actual_works)})
    return issues


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit capability management mappings against source Excel.")
    parser.add_argument("--workbook", default="data/raw-samples/wiki sample.xlsx")
    parser.add_argument("--capability-tree", default="frontend/capability-browser/public/data/capability-tree.json")
    parser.add_argument("--max-issues", type=int, default=50)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, read_only=False, data_only=True)
    capability_tree = json.loads(Path(args.capability_tree).read_text(encoding="utf-8"))
    actual = exported_focuses(capability_tree)
    expected_high_level = expected_high_level_rows(workbook)
    expected_security_works = expected_security_work_rows(workbook)

    high_level_issues = audit_high_level(expected_high_level, actual)
    placeholder_issues = audit_placeholder_leaks(expected_high_level, actual)
    security_work_issues = audit_security_works(expected_security_works, actual)
    issues = high_level_issues + placeholder_issues + security_work_issues

    summary = {
        "result": "pass" if not issues else "fail",
        "expected_high_level_focus_rows": len(expected_high_level),
        "expected_security_work_focuses": len(expected_security_works),
        "actual_focuses": len(actual),
        "high_level_issue_count": len(high_level_issues),
        "placeholder_leak_count": len(placeholder_issues),
        "security_work_issue_count": len(security_work_issues),
        "issue_count": len(issues),
        "issues": issues[: args.max_issues],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if not issues else 1


if __name__ == "__main__":
    raise SystemExit(main())
