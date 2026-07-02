#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data/raw-samples/wiki sample.xlsx"
DATA_DIR = ROOT / "frontend/capability-browser/public/data"
SECURITY_WORK_SHEET = "安全能力-安全工作"
MANAGEMENT_SHEET = "安全能力-安全管理元素（high level）"


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def read_json(relative: str) -> dict[str, Any]:
    path = DATA_DIR / relative
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def merged_range_for(ws: Any, row: int, column: int) -> str:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return str(merged)
    return ""


def inspect_source(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb[SECURITY_WORK_SHEET]
        rows: list[dict[str, Any]] = []
        issue_rows: list[dict[str, Any]] = []
        for row in range(33, 39):
            item = {
                "row": row,
                "capabilityCode": normalize(ws.cell(row, 4).value),
                "focusCode": normalize(ws.cell(row, 5).value),
                "focusTitle": normalize(ws.cell(row, 6).value),
                "securityWork": normalize(ws.cell(row, 7).value),
                "workMergedRange": merged_range_for(ws, row, 7),
            }
            rows.append(item)
        for index, item in enumerate(rows[:-1]):
            next_item = rows[index + 1]
            same_work_as_next = bool(item["securityWork"]) and item["securityWork"] == next_item["securityWork"]
            next_starts_new_capability = bool(next_item["capabilityCode"])
            current_not_merged = not item["workMergedRange"]
            next_has_work_merge = bool(next_item["workMergedRange"])
            if same_work_as_next and next_starts_new_capability and current_not_merged and next_has_work_merge:
                issue_rows.append(
                    {
                        "row": item["row"],
                        "focusCode": item["focusCode"],
                        "focusTitle": item["focusTitle"],
                        "securityWork": item["securityWork"],
                        "nextRow": next_item["row"],
                        "nextCapabilityCode": next_item["capabilityCode"],
                        "nextWorkMergedRange": next_item["workMergedRange"],
                        "reason": "current row has the same security work as the next capability group's merged work cell",
                    }
                )
        return {
            "workbook": str(workbook_path.relative_to(ROOT)),
            "rows": rows,
            "sourceIssueCount": len(issue_rows),
            "sourceIssues": issue_rows,
        }
    finally:
        wb.close()


def runtime_security_work_issues() -> list[dict[str, Any]]:
    payload = read_json("maintenance/security-works.json")
    issues = []
    for index, item in enumerate(payload.get("security_works") or []):
        title = normalize(item.get("title"))
        for focus in item.get("focuses") or []:
            if focus.get("code") == "T-PD.AC-02" and title == "入侵检测规则持续管理":
                issues.append(
                    {
                        "package": "maintenance/security-works.json",
                        "index": index,
                        "focusCode": focus.get("code"),
                        "focusTitle": focus.get("title"),
                        "securityWork": title,
                    }
                )
    return issues


def capability_workbench_issues() -> list[dict[str, Any]]:
    payload = read_json("capability-workbench.json")
    objects = payload.get("objects") or {}
    focus_id = ""
    for item_id, item in (objects.get("capability_focus") or {}).items():
        if item.get("code") == "T-PD.AC-02":
            focus_id = item_id
            break
    works = objects.get("security_work") or {}
    issues = []
    for index, relation in enumerate(payload.get("relations") or []):
        if relation.get("sourceId") != focus_id or relation.get("type") != "maps_to_work":
            continue
        work = works.get(relation.get("targetId")) or {}
        if normalize(work.get("title") or work.get("name")) == "入侵检测规则持续管理":
            issues.append(
                {
                    "package": "capability-workbench.json",
                    "relationIndex": index,
                    "focusId": focus_id,
                    "targetId": relation.get("targetId"),
                    "securityWork": normalize(work.get("title") or work.get("name")),
                }
            )
    return issues


def split_projection_issues() -> list[dict[str, Any]]:
    index_payload = read_json("capability/index.json")
    projection_path = ""
    for row in index_payload.get("projections") or []:
        if row.get("code") == "T-PD.AC-02":
            projection_path = row.get("projectionPath") or row.get("path") or ""
            break
    if not projection_path:
        return []
    projection = read_json(projection_path)
    text = json.dumps(projection, ensure_ascii=False)
    if "入侵检测规则持续管理" not in text:
        return []
    return [
        {
            "package": projection_path,
            "focusCode": "T-PD.AC-02",
            "securityWork": "入侵检测规则持续管理",
        }
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit capability management source/package alignment.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Workbook path.")
    parser.add_argument("--json", action="store_true", help="Print compact JSON only.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.is_absolute():
        workbook_path = ROOT / workbook_path
    source = inspect_source(workbook_path)
    runtime_issues = runtime_security_work_issues()
    workbench_issues = capability_workbench_issues()
    split_issues = split_projection_issues()
    report = {
        "result": "pass" if not (source["sourceIssueCount"] or runtime_issues or workbench_issues or split_issues) else "fail",
        "source": source,
        "runtimeIssueCount": len(runtime_issues) + len(workbench_issues) + len(split_issues),
        "runtimeIssues": runtime_issues + workbench_issues + split_issues,
    }
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(f"result={report['result']}")
        print(f"sourceIssueCount={source['sourceIssueCount']}")
        print(f"runtimeIssueCount={report['runtimeIssueCount']}")
        for issue in source["sourceIssues"]:
            print(
                "source_issue "
                f"row={issue['row']} focus={issue['focusCode']} work={issue['securityWork']} "
                f"nextRow={issue['nextRow']} nextMerge={issue['nextWorkMergedRange']}"
            )
        for issue in report["runtimeIssues"]:
            print("runtime_issue " + json.dumps(issue, ensure_ascii=False))
    return 0 if report["result"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
