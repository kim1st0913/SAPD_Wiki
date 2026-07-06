#!/usr/bin/env python3
"""Audit security capability mapping integrity across the three views.

This is a source-to-runtime guard for the security capability mapping page.
It intentionally keeps hard failures limited to relationships that can be
proven from controlled sources. Business judgement checks that need a human
review are reported as warnings instead of being auto-fixed or inferred.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data/raw-samples/wiki sample.xlsx"
DATA_ROOT = ROOT / "frontend/capability-browser/public/data"
REPORT_ROOT = ROOT / "data/exports/worker-verify"
REPORT_JSON = REPORT_ROOT / "capability-mapping-integrity-audit.json"
REPORT_MD = REPORT_ROOT / "capability-mapping-integrity-audit.md"

CAPABILITY_CATALOG_SHEET = "安全能力目录"
TECHNICAL_SERVICE_SHEET = "安全能力-安全技术服务"
TECHNICAL_MODULE_SHEET = "安全技术模块清单"
SECURITY_WORK_SHEET = "安全能力-安全工作"
MANAGEMENT_HIGH_LEVEL_SHEET = "安全能力-安全管理元素（high level）"
STANDARD_MAPPING_SHEET = "安全能力-网络安全制度、框架映射"

SERVICE_CODE_RE = re.compile(
    r"(?:I-[A-Z]{2}|ALL|[GM]-[A-Z]{2}\.[A-Z]{2})&[A-ZGM]-[A-Z]{2}\.[A-Z]{2}-\d{2}"
    r"|[GM]-[A-Z]{2}\.[A-Z]{2}-00"
)

STANDARD_COLUMNS = {
    "ISO-IEC-27001-2022": {
        "column": 7,
        "pattern": re.compile(r"(?<![\w.])(\d+\.\d+)(?![\w.])"),
    },
    "NIST-CSF-2.0": {
        "column": 8,
        "pattern": re.compile(r"(?<![A-Z0-9.])[A-Z]{2}\.[A-Z]{2}-\d{2}(?![A-Z0-9-])"),
    },
    "GB-T-22239-2019-L3": {
        "column": 9,
        "pattern": re.compile(r"(?<![\d.])\d+\.\d+\.\d+\.\d+[a-z]?(?![\d.])"),
        "normalize": lambda value: re.sub(r"[a-z]$", "", value),
    },
    "CIS-CSC-V8.1.2": {
        "column": 11,
        "pattern": re.compile(r"(?<![\w.])(\d+\.\d+)(?![\w.])"),
    },
    "CRF-SAFEGUARDS-CORE-2026": {
        "column": 12,
        "pattern": re.compile(r"(?<![A-Z0-9-])[A-Z]{2,4}-\d{2}(?![A-Z0-9-])"),
    },
    "NIST-800-53-REV5": {
        "column": 14,
        "pattern": re.compile(r"(?<![A-Z0-9-])[A-Z]{2,3}-\d+(?:\(\d+\))*(?![A-Z0-9-])"),
    },
}


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").replace("\n", " ").split()).strip()


def item_title(item: dict[str, Any] | None) -> str:
    if not item:
        return ""
    return normalize(item.get("title") or item.get("name") or item.get("code"))


def stable_pair(value: tuple[str, str]) -> str:
    return "::".join(value)


def merged_values(ws: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row_index, column_index)] = anchor_value
    return values


def effective_value(ws: Any, values: dict[tuple[int, int], Any], row: int, column: int) -> Any:
    return values.get((row, column), ws.cell(row, column).value)


def split_lines(value: Any) -> list[str]:
    raw = "" if value is None else str(value)
    parts = re.split(r"[\n;；]+", raw)
    return [normalize(part) for part in parts if normalize(part) and normalize(part) != "/"]


def extract_service_codes(value: Any) -> list[str]:
    return SERVICE_CODE_RE.findall(str(value or ""))


def service_scope_code(service_code: str) -> str:
    if service_code.startswith("ALL&"):
        return "ALL"
    if "&" in service_code:
        return service_code.split("&", 1)[0]
    return ""


def service_focus_code(service_code: str) -> str:
    if "&" in service_code:
        return service_code.split("&", 1)[1]
    if service_code.endswith("-00"):
        return service_code.rsplit("-", 1)[0]
    return ""


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def read_source_focuses(workbook: Any) -> dict[str, dict[str, Any]]:
    ws = workbook[CAPABILITY_CATALOG_SHEET]
    values = merged_values(ws)
    rows: dict[str, dict[str, Any]] = {}
    for row in range(4, ws.max_row + 1):
        code = normalize(effective_value(ws, values, row, 6))
        if not code:
            continue
        rows[code] = {
            "row": row,
            "title": normalize(effective_value(ws, values, row, 7)),
            "description": normalize(effective_value(ws, values, row, 8)),
        }
    return rows


def read_source_focus_services(workbook: Any) -> dict[str, set[str]]:
    ws = workbook[TECHNICAL_SERVICE_SHEET]
    values = merged_values(ws)
    rows: dict[str, set[str]] = defaultdict(set)
    for row in range(4, ws.max_row + 1):
        focus_code = normalize(effective_value(ws, values, row, 5))
        if not focus_code:
            continue
        for column in range(7, ws.max_column + 1):
            rows[focus_code].update(extract_service_codes(ws.cell(row, column).value))
    return dict(rows)


def read_source_module_service_pairs(workbook: Any) -> set[tuple[str, str]]:
    ws = workbook[TECHNICAL_MODULE_SHEET]
    values = merged_values(ws)
    pairs: set[tuple[str, str]] = set()
    for row in range(3, ws.max_row + 1):
        module_title = normalize(effective_value(ws, values, row, 4))
        if not module_title:
            continue
        for service_code in extract_service_codes(ws.cell(row, 6).value):
            pairs.add((service_code, module_title))
    return pairs


def read_source_security_works(workbook: Any) -> dict[str, set[str]]:
    ws = workbook[SECURITY_WORK_SHEET]
    values = merged_values(ws)
    rows: dict[str, set[str]] = defaultdict(set)
    for row in range(4, ws.max_row + 1):
        focus_code = normalize(effective_value(ws, values, row, 5))
        work_title = normalize(effective_value(ws, values, row, 7))
        if focus_code and work_title and work_title != "/":
            rows[focus_code].add(work_title)
    return dict(rows)


def read_source_process_refs(workbook: Any) -> dict[str, set[str]]:
    ws = workbook[MANAGEMENT_HIGH_LEVEL_SHEET]
    values = merged_values(ws)
    rows: dict[str, set[str]] = defaultdict(set)
    for row in range(4, ws.max_row + 1):
        focus_code = normalize(effective_value(ws, values, row, 5))
        if not focus_code:
            continue
        for process_ref in split_lines(effective_value(ws, values, row, 8)):
            rows[focus_code].add(process_ref)
    return dict(rows)


def read_source_standard_mappings(workbook: Any) -> dict[str, set[tuple[str, str]]]:
    ws = workbook[STANDARD_MAPPING_SHEET]
    values = merged_values(ws)
    rows: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for row in range(5, ws.max_row + 1):
        focus_code = normalize(effective_value(ws, values, row, 5))
        if not focus_code:
            continue
        for framework_code, config in STANDARD_COLUMNS.items():
            raw_value = str(ws.cell(row, int(config["column"])).value or "")
            for token in config["pattern"].findall(raw_value):
                if isinstance(token, tuple):
                    token = token[0]
                normalizer = config.get("normalize")
                control_id = normalizer(token) if normalizer else token
                rows[focus_code].add((framework_code, control_id))
    return dict(rows)


class Workbench:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload
        self.objects_by_type: dict[str, dict[str, dict[str, Any]]] = {
            key: value for key, value in (payload.get("objects") or {}).items() if isinstance(value, dict)
        }
        self.by_id: dict[str, dict[str, Any]] = {}
        for group in self.objects_by_type.values():
            self.by_id.update(group)
        self.relations: list[dict[str, Any]] = list(payload.get("relations") or [])

    def objects(self, object_type: str) -> list[dict[str, Any]]:
        return list(self.objects_by_type.get(object_type, {}).values())

    def relation_rows(self, relation_type: str) -> list[dict[str, Any]]:
        return [relation for relation in self.relations if relation.get("type") == relation_type]

    def object_by_code(self, object_type: str) -> dict[str, dict[str, Any]]:
        return {
            normalize(item.get("code")): item
            for item in self.objects(object_type)
            if normalize(item.get("code"))
        }


def actual_focus_services(workbench: Workbench) -> dict[str, set[str]]:
    rows: dict[str, set[str]] = defaultdict(set)
    for relation in workbench.relation_rows("supports_focus"):
        service = workbench.by_id.get(relation.get("sourceId")) or {}
        focus = workbench.by_id.get(relation.get("targetId")) or {}
        if service.get("code") and focus.get("code"):
            rows[focus["code"]].add(service["code"])
    return dict(rows)


def actual_focus_scopes(workbench: Workbench) -> dict[str, set[str]]:
    rows: dict[str, set[str]] = defaultdict(set)
    for relation in workbench.relation_rows("applies_to_scope"):
        source = workbench.by_id.get(relation.get("sourceId")) or {}
        target = workbench.by_id.get(relation.get("targetId")) or {}
        if source.get("type") == "capability_focus" and source.get("code") and target.get("code"):
            rows[source["code"]].add(target["code"])
    return dict(rows)


def actual_service_scopes(workbench: Workbench) -> dict[str, set[str]]:
    rows: dict[str, set[str]] = defaultdict(set)
    for relation in workbench.relation_rows("applies_to_scope"):
        source = workbench.by_id.get(relation.get("sourceId")) or {}
        target = workbench.by_id.get(relation.get("targetId")) or {}
        if source.get("type") == "security_technical_service" and source.get("code") and target.get("code"):
            rows[source["code"]].add(target["code"])
    return dict(rows)


def actual_module_service_pairs(workbench: Workbench) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for relation in workbench.relation_rows("implemented_by_module"):
        service = workbench.by_id.get(relation.get("sourceId")) or {}
        module = workbench.by_id.get(relation.get("targetId")) or {}
        if service.get("code") and item_title(module):
            pairs.add((service["code"], item_title(module)))
    return pairs


def actual_security_works(workbench: Workbench) -> dict[str, set[str]]:
    rows: dict[str, set[str]] = defaultdict(set)
    for relation in workbench.relation_rows("maps_to_work"):
        focus = workbench.by_id.get(relation.get("sourceId")) or {}
        work = workbench.by_id.get(relation.get("targetId")) or {}
        if focus.get("code") and item_title(work):
            rows[focus["code"]].add(item_title(work))
    return dict(rows)


def actual_process_refs(workbench: Workbench) -> dict[str, set[str]]:
    rows: dict[str, set[str]] = defaultdict(set)
    for relation in workbench.relation_rows("maps_to_process"):
        focus = workbench.by_id.get(relation.get("sourceId")) or {}
        process = workbench.by_id.get(relation.get("targetId")) or {}
        if focus.get("code") and item_title(process):
            rows[focus["code"]].add(item_title(process))
    return dict(rows)


def actual_standard_mappings(workbench: Workbench) -> dict[str, set[tuple[str, str]]]:
    rows: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for relation in workbench.relation_rows("maps_to_standard"):
        focus = workbench.by_id.get(relation.get("sourceId")) or {}
        control = workbench.by_id.get(relation.get("targetId")) or {}
        framework_code = normalize(control.get("frameworkCode"))
        control_id = normalize(control.get("originalControlId"))
        if focus.get("code") and framework_code and control_id:
            rows[focus["code"]].add((framework_code, control_id))
    return dict(rows)


def add_diff_issues(
    issues: list[dict[str, Any]],
    issue_type: str,
    expected: dict[str, set[Any]],
    actual: dict[str, set[Any]],
    *,
    sample_limit: int,
    extra_is_hard: bool = True,
) -> tuple[int, int]:
    missing_count = 0
    extra_count = 0
    for focus_code in sorted(set(expected) | set(actual)):
        missing = sorted(expected.get(focus_code, set()) - actual.get(focus_code, set()), key=str)
        extra = sorted(actual.get(focus_code, set()) - expected.get(focus_code, set()), key=str)
        missing_count += len(missing)
        extra_count += len(extra)
        if missing and len(issues) < sample_limit:
            issues.append(
                {
                    "type": f"{issue_type}_missing",
                    "focusCode": focus_code,
                    "count": len(missing),
                    "sample": [stable_pair(item) if isinstance(item, tuple) else item for item in missing[:8]],
                }
            )
        if extra and extra_is_hard and len(issues) < sample_limit:
            issues.append(
                {
                    "type": f"{issue_type}_extra",
                    "focusCode": focus_code,
                    "count": len(extra),
                    "sample": [stable_pair(item) if isinstance(item, tuple) else item for item in extra[:8]],
                }
            )
    return missing_count, extra_count


def validate_relation_shape(workbench: Workbench, sample_limit: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    duplicate_counter = Counter(
        (relation.get("type"), relation.get("sourceId"), relation.get("targetId"))
        for relation in workbench.relations
    )
    duplicate_count = 0
    for key, count in duplicate_counter.items():
        if count <= 1:
            continue
        duplicate_count += count - 1
        if len(issues) < sample_limit:
            issues.append({"type": "duplicate_relation", "relationKey": list(key), "count": count})

    for index, relation in enumerate(workbench.relations):
        source = workbench.by_id.get(relation.get("sourceId"))
        target = workbench.by_id.get(relation.get("targetId"))
        if not source or not target:
            if len(issues) < sample_limit:
                issues.append(
                    {
                        "type": "relation_endpoint_missing",
                        "index": index,
                        "relationType": relation.get("type"),
                        "sourceId": relation.get("sourceId"),
                        "targetId": relation.get("targetId"),
                    }
                )
            continue

        if relation.get("type") == "supports_focus":
            focus_code = normalize(target.get("code"))
            service_code = normalize(source.get("code"))
            expected_focus = service_focus_code(service_code)
            if "&" in service_code and expected_focus != focus_code:
                if len(issues) < sample_limit:
                    issues.append(
                        {
                            "type": "service_focus_code_mismatch",
                            "serviceCode": service_code,
                            "targetFocusCode": focus_code,
                        }
                    )

        if relation.get("type") == "applies_to_scope" and source.get("type") == "security_technical_service":
            scope_code = service_scope_code(normalize(source.get("code")))
            target_scope = normalize(target.get("code"))
            if scope_code not in {"", "ALL"} and target_scope and scope_code != target_scope:
                if len(issues) < sample_limit:
                    issues.append(
                        {
                            "type": "service_scope_mismatch",
                            "serviceCode": source.get("code"),
                            "targetScopeCode": target_scope,
                        }
                    )

    if duplicate_count:
        warnings.append({"type": "duplicate_relation_total", "count": duplicate_count})
    return issues, warnings[:sample_limit]


def validate_duplicates(workbench: Workbench, sample_limit: int) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for object_type, group in workbench.objects_by_type.items():
        codes = Counter(normalize(item.get("code")) for item in group.values() if normalize(item.get("code")))
        for code, count in codes.items():
            if count > 1 and len(issues) < sample_limit:
                issues.append({"type": "duplicate_object_code", "objectType": object_type, "code": code, "count": count})
    return issues


def validate_focus_scopes(
    source_focus_services: dict[str, set[str]],
    actual_scopes: dict[str, set[str]],
    sample_limit: int,
) -> tuple[int, list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    issue_count = 0
    for focus_code, services in sorted(source_focus_services.items()):
        expected_scopes = {service_scope_code(service) for service in services if service_scope_code(service) not in {"", "ALL"}}
        actual = actual_scopes.get(focus_code, set())
        missing = sorted(expected_scopes - actual)
        if missing:
            issue_count += len(missing)
            if len(issues) < sample_limit:
                issues.append(
                    {
                        "type": "focus_scope_missing_from_source_services",
                        "focusCode": focus_code,
                        "missingScopes": missing,
                        "actualScopes": sorted(actual),
                    }
                )
    return issue_count, issues


def validate_service_scopes(
    service_scopes: dict[str, set[str]],
    sample_limit: int,
) -> tuple[int, list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    issue_count = 0
    for service_code, scopes in sorted(service_scopes.items()):
        expected = service_scope_code(service_code)
        if expected in {"", "ALL"}:
            continue
        if expected not in scopes:
            issue_count += 1
            if len(issues) < sample_limit:
                issues.append(
                    {
                        "type": "service_scope_relation_missing_code_scope",
                        "serviceCode": service_code,
                        "expectedScope": expected,
                        "actualScopes": sorted(scopes),
                    }
                )
    return issue_count, issues


def build_report(args: argparse.Namespace) -> dict[str, Any]:
    workbook_path = Path(args.workbook)
    if not workbook_path.is_absolute():
        workbook_path = ROOT / workbook_path

    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        source_focuses = read_source_focuses(workbook)
        source_focus_services = read_source_focus_services(workbook)
        source_module_pairs = read_source_module_service_pairs(workbook)
        source_security_works = read_source_security_works(workbook)
        source_process_refs = read_source_process_refs(workbook)
        source_standard_mappings = read_source_standard_mappings(workbook)
    finally:
        workbook.close()

    workbench = Workbench(read_json(DATA_ROOT / "capability-workbench.json"))
    focus_by_code = workbench.object_by_code("capability_focus")

    issues: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    issue_counts: dict[str, int] = {}
    warning_counts: dict[str, int] = {}

    duplicate_issues = validate_duplicates(workbench, args.max_issues)
    relation_issues, relation_warnings = validate_relation_shape(workbench, args.max_issues)
    issues.extend(duplicate_issues)
    issues.extend(relation_issues)
    warnings.extend(relation_warnings)
    issue_counts["duplicate_object_code"] = len(duplicate_issues)
    issue_counts["relation_shape"] = len(relation_issues)
    warning_counts["relation_review"] = len(relation_warnings)

    source_focus_codes = set(source_focuses)
    actual_focus_codes = set(focus_by_code)
    missing_focuses = sorted(source_focus_codes - actual_focus_codes)
    extra_focuses = sorted(actual_focus_codes - source_focus_codes)
    issue_counts["focus_code_missing"] = len(missing_focuses)
    issue_counts["focus_code_extra"] = len(extra_focuses)
    if missing_focuses:
        issues.append({"type": "focus_code_missing", "count": len(missing_focuses), "sample": missing_focuses[:8]})
    if extra_focuses:
        issues.append({"type": "focus_code_extra", "count": len(extra_focuses), "sample": extra_focuses[:8]})

    focus_service_missing, focus_service_extra = add_diff_issues(
        issues,
        "focus_service_mapping",
        source_focus_services,
        actual_focus_services(workbench),
        sample_limit=args.max_issues,
        extra_is_hard=False,
    )
    issue_counts["focus_service_missing"] = focus_service_missing
    warning_counts["focus_service_extra_review"] = focus_service_extra
    if focus_service_extra:
        warnings.append(
            {
                "type": "focus_service_extra_review",
                "count": focus_service_extra,
                "reason": "current workbench includes focus-service pairs that are not proven by the technical-service source sheet; report only, do not auto-delete or infer fixes",
            }
        )

    source_module_expected = {
        service_code: set()
        for service_code, _ in source_module_pairs
    }
    actual_module_pairs = actual_module_service_pairs(workbench)
    module_missing = sorted(source_module_pairs - actual_module_pairs)
    module_extra = sorted(actual_module_pairs - source_module_pairs)
    issue_counts["service_module_missing"] = len(module_missing)
    issue_counts["service_module_extra"] = len(module_extra)
    if module_missing:
        issues.append(
            {
                "type": "service_module_mapping_missing",
                "count": len(module_missing),
                "sample": [stable_pair(item) for item in module_missing[:8]],
            }
        )
    if module_extra:
        issues.append(
            {
                "type": "service_module_mapping_extra",
                "count": len(module_extra),
                "sample": [stable_pair(item) for item in module_extra[:8]],
            }
        )

    focus_scope_count, focus_scope_issues = validate_focus_scopes(
        source_focus_services,
        actual_focus_scopes(workbench),
        args.max_issues,
    )
    service_scope_count, service_scope_issues = validate_service_scopes(
        actual_service_scopes(workbench),
        args.max_issues,
    )
    issue_counts["focus_scope_missing"] = focus_scope_count
    issue_counts["service_scope_missing"] = service_scope_count
    issues.extend(focus_scope_issues)
    issues.extend(service_scope_issues)

    work_missing, work_extra = add_diff_issues(
        issues,
        "management_security_work",
        source_security_works,
        actual_security_works(workbench),
        sample_limit=args.max_issues,
    )
    process_missing, process_extra = add_diff_issues(
        issues,
        "management_process_ref",
        source_process_refs,
        actual_process_refs(workbench),
        sample_limit=args.max_issues,
    )
    issue_counts["management_work_missing"] = work_missing
    issue_counts["management_work_extra"] = work_extra
    issue_counts["management_process_missing"] = process_missing
    issue_counts["management_process_extra"] = process_extra

    standard_missing, standard_extra = add_diff_issues(
        issues,
        "standard_mapping",
        source_standard_mappings,
        actual_standard_mappings(workbench),
        sample_limit=args.max_issues,
    )
    issue_counts["standard_mapping_missing"] = standard_missing
    issue_counts["standard_mapping_extra"] = standard_extra

    relation_type_counts = Counter(relation.get("type") for relation in workbench.relations)
    object_counts = {key: len(value) for key, value in workbench.objects_by_type.items()}
    hard_issue_count = sum(issue_counts.values())

    report = {
        "result": "pass" if hard_issue_count == 0 else "fail",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "contract": "capability_mapping_integrity_matrix",
        "sourceAuthority": {
            "workbook": str(workbook_path.relative_to(ROOT)),
            "focusCatalog": CAPABILITY_CATALOG_SHEET,
            "focusToTechnicalService": TECHNICAL_SERVICE_SHEET,
            "serviceToModule": TECHNICAL_MODULE_SHEET,
            "focusToSecurityWork": SECURITY_WORK_SHEET,
            "focusToProcess": MANAGEMENT_HIGH_LEVEL_SHEET,
            "focusToStandardControls": STANDARD_MAPPING_SHEET,
        },
        "businessGrain": "capability_focus",
        "forbiddenInference": [
            "front-end ViewModel or fallback cannot create authoritative relationships",
            "L0/L1/L2 aggregate data cannot prove focus-level relations",
            "module global service catalog cannot be projected to a focus unless source service mapping exists",
            "standard titles cannot replace frameworkCode + originalControlId as matching keys",
            "source-defined M-* -00 management service codes cannot be reclassified as generated review warnings",
        ],
        "summary": {
            "sourceFocuses": len(source_focuses),
            "workbenchFocuses": len(focus_by_code),
            "sourceFocusServicePairs": sum(len(value) for value in source_focus_services.values()),
            "workbenchFocusServicePairs": sum(len(value) for value in actual_focus_services(workbench).values()),
            "sourceServiceModulePairs": len(source_module_pairs),
            "workbenchServiceModulePairs": len(actual_module_pairs),
            "sourceSecurityWorkPairs": sum(len(value) for value in source_security_works.values()),
            "workbenchSecurityWorkPairs": sum(len(value) for value in actual_security_works(workbench).values()),
            "sourceProcessPairs": sum(len(value) for value in source_process_refs.values()),
            "workbenchProcessPairs": sum(len(value) for value in actual_process_refs(workbench).values()),
            "sourceStandardMappings": sum(len(value) for value in source_standard_mappings.values()),
            "workbenchStandardMappings": sum(len(value) for value in actual_standard_mappings(workbench).values()),
            "hardIssueCount": hard_issue_count,
            "reviewWarningCount": sum(warning_counts.values()),
        },
        "objectCounts": object_counts,
        "relationTypeCounts": dict(sorted(relation_type_counts.items())),
        "issueCounts": issue_counts,
        "warningCounts": warning_counts,
        "issues": issues[: args.max_issues],
        "warnings": warnings[: args.max_issues],
        "goldenSamples": {
            "technical": [
                "T-AS.AD-01 maps explicit I-* source services and focus scope relations",
                "ALL&T-AD.IR-01 service-module rows must be parsed as valid aggregate technical-service codes",
            ],
            "management": [
                "T-AS.AM-01 keeps both source security works and one source high-level process",
                "G-SP.SM-02 management/governance relationships are validated at focus grain",
                "M-PM.PR-00 / M-SA.RM-00 / M-SA.RE-00 / M-SA.CO-00 / M-SE.PE-00 / M-PS.CT-00 are source-defined focus-service pairs",
            ],
            "standards": [
                "T-AS.AD-01 maps ISO/CSF/MLPS/CIS/CRF/NIST controls from the source row",
                "NIST enhancement PM-7(1) must not be truncated to PM-7",
            ],
        },
    }
    return report


def write_report(report: dict[str, Any]) -> None:
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    summary = report["summary"]
    lines = [
        "# Capability Mapping Integrity Audit",
        "",
        f"- 审计结果：`{report['result']}`",
        f"- 生成时间：`{report['generatedAt']}`",
        f"- 业务粒度：`{report['businessGrain']}`",
        f"- 硬错误：`{summary['hardIssueCount']}`",
        f"- Review 提醒：`{summary['reviewWarningCount']}`",
        "",
        "## 核心计数",
        "",
        f"- 源关注点 / workbench 关注点：`{summary['sourceFocuses']} / {summary['workbenchFocuses']}`",
        f"- 关注点-技术服务：`{summary['sourceFocusServicePairs']} / {summary['workbenchFocusServicePairs']}`",
        f"- 服务-安全技术模块：`{summary['sourceServiceModulePairs']} / {summary['workbenchServiceModulePairs']}`",
        f"- 关注点-安全工作：`{summary['sourceSecurityWorkPairs']} / {summary['workbenchSecurityWorkPairs']}`",
        f"- 关注点-管理流程：`{summary['sourceProcessPairs']} / {summary['workbenchProcessPairs']}`",
        f"- 关注点-标准控制项：`{summary['sourceStandardMappings']} / {summary['workbenchStandardMappings']}`",
        "",
        "## 固化规则",
        "",
        "- `capability_focus` 是三视角关系唯一验收粒度。",
        "- 技术服务、管理工作、管理流程和标准控制项均按源表明确值对账。",
        "- 标准映射使用 `frameworkCode + originalControlId`，不得用标题近似匹配。",
        "- `PM-7(1)` 等 NIST enhancement 必须保留括号后缀，不得截断为基础控制项。",
        "- `M-* -00` 管理类服务按源表明确值验收，不按生成项或作用域推断降级为 review warning。",
        "- Review 提醒只用于源表无法证明的差异，不触发自动修复。",
        "",
    ]
    REPORT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit capability mapping technical, management, and standard relations.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Source workbook path.")
    parser.add_argument("--json", action="store_true", help="Print full report JSON.")
    parser.add_argument("--write-report", action="store_true", help="Write audit report under data/exports/worker-verify.")
    parser.add_argument("--max-issues", type=int, default=50, help="Maximum issue samples to keep.")
    args = parser.parse_args()

    report = build_report(args)
    if args.write_report:
        write_report(report)
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(f"result={report['result']}")
        print(f"hard_issue_count={report['summary']['hardIssueCount']}")
        print(f"review_warning_count={report['summary']['reviewWarningCount']}")
        print(f"source_focuses={report['summary']['sourceFocuses']}")
        print(f"workbench_focuses={report['summary']['workbenchFocuses']}")
        print(f"focus_service_pairs={report['summary']['sourceFocusServicePairs']}/{report['summary']['workbenchFocusServicePairs']}")
        print(f"service_module_pairs={report['summary']['sourceServiceModulePairs']}/{report['summary']['workbenchServiceModulePairs']}")
        print(f"management_work_pairs={report['summary']['sourceSecurityWorkPairs']}/{report['summary']['workbenchSecurityWorkPairs']}")
        print(f"management_process_pairs={report['summary']['sourceProcessPairs']}/{report['summary']['workbenchProcessPairs']}")
        print(f"standard_mappings={report['summary']['sourceStandardMappings']}/{report['summary']['workbenchStandardMappings']}")
        if args.write_report:
            print(f"json={REPORT_JSON}")
            print(f"markdown={REPORT_MD}")
    return 0 if report["result"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
