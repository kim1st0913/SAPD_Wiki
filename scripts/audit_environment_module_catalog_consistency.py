#!/usr/bin/env python3
"""Audit environment module/service/system relations against the module catalog.

This script is read-only for official data. It compares source-derived
normalized rows from ``作用域-安全技术服务-安全技术模块映射`` with the source
``安全技术模块清单`` sheet, preserving merged-cell evidence in both inputs.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from audit_scope_service_module_mapping import (
    OUTPUT_DIR,
    PROJECT_ROOT,
    as_text,
    is_empty_security_system_value,
    service_child_relation_entries,
    service_payload,
    split_relation_titles,
)
from sapd_wiki.transformers import is_blank_or_placeholder, split_multivalue_text


WORKBOOK_PATH = PROJECT_ROOT / "data/raw-samples/wiki sample.xlsx"
MODULE_CATALOG_SHEET = "安全技术模块清单"
NORMALIZED_ROWS_PATH = OUTPUT_DIR / "scope-service-module-mapping-normalized-rows.json"
SCOPE_AUDIT_PATH = OUTPUT_DIR / "scope-service-module-mapping-reimport-audit.json"
WORKBENCH_PATH = PROJECT_ROOT / "frontend/capability-browser/public/data/environment-workbench.json"

AUDIT_JSON_OUT = OUTPUT_DIR / "environment-module-catalog-consistency-audit.json"
AUDIT_MD_OUT = OUTPUT_DIR / "environment-module-catalog-consistency-audit.md"
REVIEW_ROWS_OUT = OUTPUT_DIR / "environment-module-catalog-consistency-review-rows.json"
TRIAGE_JSON_OUT = OUTPUT_DIR / "environment-module-catalog-consistency-triage.json"
TRIAGE_MD_OUT = OUTPUT_DIR / "environment-module-catalog-consistency-triage.md"

CATALOG_FIELD_COLUMNS = {
    "B": "securitySystemCategory",
    "C": "securitySystem",
    "D": "securityTechnologyModule",
    "F": "securityTechnicalService",
    "G": "product",
}


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def merged_index_for_sheet(ws) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for merged_range in sorted(ws.merged_cells.ranges, key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col)):
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        payload = {
            "range": str(merged_range),
            "topLeft": top_left.coordinate,
            "topLeftValue": as_text(top_left.value),
            "rowSpan": merged_range.max_row - merged_range.min_row + 1,
            "columnSpan": merged_range.max_col - merged_range.min_col + 1,
        }
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                index[f"{get_column_letter(column_index)}{row_index}"] = payload
    return index


def catalog_cell(ws, merged_index: dict[str, dict[str, Any]], row_index: int, column_letter: str) -> dict[str, Any]:
    cell = ws[f"{column_letter}{row_index}"]
    merged = merged_index.get(cell.coordinate)
    value = merged["topLeftValue"] if merged else as_text(cell.value)
    return {
        "value": value,
        "sourceCell": cell.coordinate,
        "mergedRange": merged["range"] if merged else None,
        "topLeft": merged["topLeft"] if merged else cell.coordinate,
    }


def service_key(value: Any) -> str:
    service = service_payload(value)
    return service.get("code") or service.get("title") or as_text(value)


def split_catalog_services(value: Any) -> list[str]:
    services: list[str] = []
    for line in split_multivalue_text(value, split_on_ideographic_comma=False):
        item = as_text(line)
        if item and not is_blank_or_placeholder(item):
            services.append(item)
    return list(dict.fromkeys(services))


def read_module_catalog(workbook_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if MODULE_CATALOG_SHEET not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {MODULE_CATALOG_SHEET}")
    ws = workbook[MODULE_CATALOG_SHEET]
    merged_index = merged_index_for_sheet(ws)
    relations: list[dict[str, Any]] = []
    for row_index in range(3, ws.max_row + 1):
        fields = {
            field: catalog_cell(ws, merged_index, row_index, column)
            for column, field in CATALOG_FIELD_COLUMNS.items()
        }
        module_title = as_text(fields["securityTechnologyModule"]["value"])
        system_title = as_text(fields["securitySystem"]["value"])
        category_title = as_text(fields["securitySystemCategory"]["value"])
        service_values = split_catalog_services(fields["securityTechnicalService"]["value"])
        if not module_title or is_blank_or_placeholder(module_title) or not service_values:
            continue
        for service_text in service_values:
            relations.append(
                {
                    "securitySystemCategory": category_title,
                    "securitySystem": system_title,
                    "securityTechnologyModule": module_title,
                    "securityTechnicalService": service_text,
                    "securityTechnicalServiceKey": service_key(service_text),
                    "sourceRow": row_index,
                    "sourceCells": {field: payload["sourceCell"] for field, payload in fields.items()},
                    "mergedRanges": {field: payload["mergedRange"] for field, payload in fields.items() if payload["mergedRange"]},
                }
            )
    return relations, {
        "sheet": MODULE_CATALOG_SHEET,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "mergedRangeCount": len(list(ws.merged_cells.ranges)),
    }


def append_index(index: dict[str, set[str]], key: str, value: str) -> None:
    if key and value:
        index[key].add(value)


def build_catalog_indices(relations: list[dict[str, Any]]) -> dict[str, dict[str, list[str]]]:
    system_to_modules: dict[str, set[str]] = defaultdict(set)
    module_to_services: dict[str, set[str]] = defaultdict(set)
    service_to_modules: dict[str, set[str]] = defaultdict(set)
    module_to_systems: dict[str, set[str]] = defaultdict(set)
    service_to_systems: dict[str, set[str]] = defaultdict(set)
    for relation in relations:
        system = as_text(relation.get("securitySystem"))
        module = as_text(relation.get("securityTechnologyModule"))
        service = as_text(relation.get("securityTechnicalServiceKey"))
        append_index(system_to_modules, system, module)
        append_index(module_to_services, module, service)
        append_index(service_to_modules, service, module)
        append_index(module_to_systems, module, system)
        append_index(service_to_systems, service, system)
    return {
        "securitySystemToModules": {key: sorted(values) for key, values in sorted(system_to_modules.items())},
        "securityTechnologyModuleToServices": {key: sorted(values) for key, values in sorted(module_to_services.items())},
        "securityTechnicalServiceToModules": {key: sorted(values) for key, values in sorted(service_to_modules.items())},
        "securityTechnologyModuleToSystems": {key: sorted(values) for key, values in sorted(module_to_systems.items())},
        "securityTechnicalServiceToSystems": {key: sorted(values) for key, values in sorted(service_to_systems.items())},
    }


def measure_titles_from_workbench(workbench: dict[str, Any] | None) -> set[str]:
    titles: set[str] = set()
    if not isinstance(workbench, dict):
        return titles
    objects = workbench.get("objects") or []
    if isinstance(objects, dict):
        objects = objects.values()
    for obj in objects:
        if not isinstance(obj, dict):
            continue
        if as_text(obj.get("type")) == "security_technical_measure":
            title = as_text(obj.get("title") or obj.get("name"))
            if title:
                titles.add(title)
    return titles


def env_entries(normalized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for row in normalized_rows:
        for entry in service_child_relation_entries(row):
            entries.append({**entry, "sourceRow": row})
    return entries


def make_review_row(entry: dict[str, Any], indices: dict[str, dict[str, list[str]]], module_titles: set[str], measure_titles: set[str]) -> dict[str, Any]:
    source_row = entry["sourceRow"]
    service = as_text(entry["securityTechnicalService"])
    child_type = as_text(entry["childType"])
    child_title = as_text(entry["securityTechnologyModuleOrMeasure"])
    system = as_text(entry["securitySystem"])
    allowed_modules = indices["securityTechnicalServiceToModules"].get(service, [])
    allowed_services = indices["securityTechnologyModuleToServices"].get(child_title, [])
    allowed_systems = indices["securityTechnologyModuleToSystems"].get(child_title, [])
    issues: list[dict[str, Any]] = []
    module_service_match: bool | None = None
    system_module_match: bool | None = None

    if child_type == "module":
        module_service_match = service in set(allowed_services)
        if not module_service_match:
            issues.append(
                {
                    "type": "moduleServiceMismatch",
                    "severity": "high",
                    "reason": "环境映射表中出现的模块-服务关系不在安全技术模块清单中",
                }
            )
        if system:
            system_module_match = child_title in set(indices["securitySystemToModules"].get(system, []))
            if not system_module_match:
                issues.append(
                    {
                        "type": "systemModuleMismatch",
                        "severity": "high",
                        "reason": "环境映射表中出现的系统-模块关系不在安全技术模块清单中",
                    }
                )
        if allowed_systems and (not system or system not in set(allowed_systems)):
            issues.append(
                {
                    "type": "securitySystemCoverageOrMismatch",
                    "severity": "review",
                    "allowedSecuritySystemsFromCatalog": allowed_systems,
                    "reason": "安全系统为空或与安全技术模块清单不完全一致，需人工确认",
                }
            )
        if child_title not in module_titles:
            issues.append(
                {
                    "type": "moduleMeasureClassificationIssue",
                    "severity": "review",
                    "reason": "该值被识别为安全技术模块，但未在安全技术模块清单中找到",
                }
            )
    elif child_type == "measure":
        if measure_titles and child_title not in measure_titles:
            issues.append(
                {
                    "type": "moduleMeasureClassificationIssue",
                    "severity": "review",
                    "reason": "该值被识别为安全技术措施，但未在当前维护数据的措施集合中找到",
                }
            )

    return {
        "excelRow": source_row.get("row"),
        "objectContextKey": entry["objectContextKey"],
        "environment": entry["environment"],
        "environmentSegment": entry["environmentSegment"],
        "informationObject": entry["informationObject"],
        "declaredScopes": [
            scope.get("code") or scope.get("title") or scope.get("text")
            for scope in source_row.get("scopes", [])
            if scope.get("code") or scope.get("title") or scope.get("text")
        ],
        "securityTechnicalService": entry["securityTechnicalServiceRaw"],
        "securityTechnicalServiceKey": service,
        "childType": child_type,
        "securityTechnologyModuleOrMeasure": child_title,
        "securitySystem": system,
        "catalogMatch": {
            "moduleServiceMatch": module_service_match,
            "systemModuleMatch": system_module_match,
            "allowedModulesFromCatalog": allowed_modules,
            "allowedServicesFromCatalog": allowed_services,
            "allowedSystemsFromCatalog": allowed_systems,
        },
        "issues": issues,
        "sourceCells": source_row.get("sourceCells") or {},
        "mergedRanges": source_row.get("mergedRanges") or {},
    }


def coverage_candidates(review_rows: list[dict[str, Any]], indices: dict[str, dict[str, list[str]]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    actual_modules_by_context_service: dict[tuple[str, str], set[str]] = defaultdict(set)
    actual_services_by_context_module: dict[tuple[str, str], set[str]] = defaultdict(set)
    first_row_by_context_service: dict[tuple[str, str], dict[str, Any]] = {}
    first_row_by_context_module: dict[tuple[str, str], dict[str, Any]] = {}
    for row in review_rows:
        if row.get("childType") != "module":
            continue
        context_key = as_text(row.get("objectContextKey"))
        service = as_text(row.get("securityTechnicalServiceKey"))
        module = as_text(row.get("securityTechnologyModuleOrMeasure"))
        actual_modules_by_context_service[(context_key, service)].add(module)
        actual_services_by_context_module[(context_key, module)].add(service)
        first_row_by_context_service.setdefault((context_key, service), row)
        first_row_by_context_module.setdefault((context_key, module), row)

    service_module_gaps: list[dict[str, Any]] = []
    for key, actual_modules in sorted(actual_modules_by_context_service.items()):
        context_key, service = key
        allowed_modules = set(indices["securityTechnicalServiceToModules"].get(service, []))
        missing = sorted(allowed_modules - actual_modules)
        if not missing:
            continue
        first = first_row_by_context_service[key]
        service_module_gaps.append(
            {
                "type": "serviceModuleCoverageGapCandidate",
                "objectContextKey": context_key,
                "securityTechnicalService": service,
                "allowedModulesFromCatalog": sorted(allowed_modules),
                "actualModulesInEnvironmentMapping": sorted(actual_modules),
                "missingModulesCandidate": missing,
                "severity": "review",
                "reason": "安全技术模块清单中该服务关联更多模块；需人工确认环境映射表是否应完整列出",
                "sampleExcelRow": first.get("excelRow"),
            }
        )

    module_service_gaps: list[dict[str, Any]] = []
    for key, actual_services in sorted(actual_services_by_context_module.items()):
        context_key, module = key
        allowed_services = set(indices["securityTechnologyModuleToServices"].get(module, []))
        missing = sorted(allowed_services - actual_services)
        if not missing:
            continue
        first = first_row_by_context_module[key]
        module_service_gaps.append(
            {
                "type": "moduleServiceCoverageGapCandidate",
                "objectContextKey": context_key,
                "securityTechnologyModule": module,
                "allowedServicesFromCatalog": sorted(allowed_services),
                "actualServicesInEnvironmentMapping": sorted(actual_services),
                "missingServicesCandidate": missing,
                "severity": "review",
                "reason": "安全技术模块清单中该模块关联更多服务；需人工确认环境对象是否应包含这些服务",
                "sampleExcelRow": first.get("excelRow"),
            }
        )
    return service_module_gaps, module_service_gaps


def issue_items(review_rows: list[dict[str, Any]], issue_type: str) -> list[dict[str, Any]]:
    items = []
    for row in review_rows:
        for issue in row.get("issues") or []:
            if issue.get("type") == issue_type:
                items.append(
                    {
                        **issue,
                        "excelRow": row.get("excelRow"),
                        "objectContextKey": row.get("objectContextKey"),
                        "securityTechnicalService": row.get("securityTechnicalServiceKey"),
                        "securityTechnologyModuleOrMeasure": row.get("securityTechnologyModuleOrMeasure"),
                        "securitySystem": row.get("securitySystem"),
                    }
                )
    return items


TRIAGE_LABELS = {
    "A": "确认错误：完整重复关系",
    "B": "高优先级核对：模块/措施分类问题",
    "C": "目录不一致：模块-服务关系不在模块清单目录",
    "D": "目录不一致：系统-模块关系不在模块清单目录",
    "E": "覆盖差异候选：目录中有更多模块，但环境映射表可能只是选择性引用",
    "F": "覆盖差异候选：目录中有更多服务，但环境映射表可能只是选择性引用",
    "G": "安全系统候选差异：安全系统为空、多个候选或系统不一致",
    "H": "可能命名/别名问题：空格、标点、全半角、别名、后缀差异",
    "I": "低风险提示：同一服务多模块/措施/系统合法 1:N 展开",
}


ISSUE_TRIAGE_CATEGORY = {
    "duplicateExactServiceChildRelations": "A",
    "moduleMeasureClassificationIssue": "B",
    "moduleServiceMismatch": "C",
    "systemModuleMismatch": "D",
    "serviceModuleCoverageGapCandidate": "E",
    "moduleServiceCoverageGapCandidate": "F",
    "securitySystemCoverageOrMismatch": "G",
    "possibleAliasMatches": "H",
    "repeatedServiceWithDifferentChildren": "I",
}


def sorted_unique(values: list[Any], limit: int | None = None) -> list[str]:
    result = sorted({as_text(value) for value in values if as_text(value)})
    return result[:limit] if limit else result


def row_issue_types(row: dict[str, Any]) -> list[str]:
    return [as_text(issue.get("type") if isinstance(issue, dict) else issue) for issue in row.get("issues") or [] if issue]


def normalize_for_alias(value: Any) -> str:
    text_value = unicodedata.normalize("NFKC", as_text(value)).lower()
    replacements = {
        "（": "(",
        "）": ")",
        "、": ",",
        "，": ",",
        "／": "/",
        "－": "-",
        "—": "-",
        "–": "-",
        "_": "",
        " ": "",
        "\u3000": "",
    }
    for source, target in replacements.items():
        text_value = text_value.replace(source, target)
    text_value = re.sub(r"[()（）,，、/\\\-]+", "", text_value)
    text_value = re.sub(r"\s+", "", text_value)
    return text_value


def strip_common_suffix(value: str) -> str:
    result = value
    for suffix in ("平台", "系统", "管理", "防护", "服务"):
        if result.endswith(suffix) and len(result) > len(suffix) + 1:
            result = result[: -len(suffix)]
    return result


def alias_hint(left: Any, right: Any) -> str:
    if as_text(left) == as_text(right):
        return ""
    left_norm = normalize_for_alias(left)
    right_norm = normalize_for_alias(right)
    if not left_norm or not right_norm:
        return ""
    if left_norm == right_norm:
        return "归一化后完全一致"
    if strip_common_suffix(left_norm) == strip_common_suffix(right_norm):
        return "去除常见后缀后可能一致"
    if left_norm in right_norm or right_norm in left_norm:
        return "一个名称包含另一个名称"
    left_tokens = set(re.findall(r"[\u4e00-\u9fff]{2,}|[a-z0-9]+", left_norm))
    right_tokens = set(re.findall(r"[\u4e00-\u9fff]{2,}|[a-z0-9]+", right_norm))
    if left_tokens and right_tokens:
        overlap = len(left_tokens & right_tokens) / max(1, min(len(left_tokens), len(right_tokens)))
        if overlap >= 0.6:
            return "关键词高度重合"
    return ""


def compact_rows(rows: list[Any], limit: int = 20) -> list[int]:
    result = []
    for row in rows:
        if isinstance(row, int) or str(row).isdigit():
            result.append(int(row))
    return sorted(set(result))[:limit]


def relation_key(system: Any, module: Any, service: Any) -> str:
    return "||".join([as_text(system), as_text(module), as_text(service)])


def compact_source_evidence(rows: list[dict[str, Any]], field_names: list[str], limit: int = 12) -> dict[str, Any]:
    source_rows = compact_rows([row.get("sourceRow") or row.get("excelRow") or row.get("row") for row in rows], limit=limit)
    source_cells: dict[str, list[str]] = defaultdict(list)
    merged_ranges: dict[str, list[str]] = defaultdict(list)
    for row in rows:
      for field_name in field_names:
          cell_value = row.get("sourceCells", {}).get(field_name)
          if isinstance(cell_value, dict):
              cell_value = cell_value.get("sourceCell") or cell_value.get("cell")
          range_value = row.get("mergedRanges", {}).get(field_name)
          if cell_value:
              source_cells[field_name].append(as_text(cell_value))
          if range_value:
              merged_ranges[field_name].append(as_text(range_value))
    return {
        "sourceRows": source_rows,
        "sourceCells": {key: sorted_unique(values, limit=limit) for key, values in source_cells.items()},
        "mergedRanges": {key: sorted_unique(values, limit=limit) for key, values in merged_ranges.items()},
    }


def environment_usage_position(row: dict[str, Any]) -> dict[str, Any]:
    source_row = row.get("sourceRow") or {}
    return {
        "excelRow": row.get("excelRow"),
        "informationEnvironment": row.get("environment"),
        "environmentSegment": row.get("environmentSegment"),
        "informationObject": row.get("informationObject"),
        "objectContextKey": row.get("objectContextKey"),
        "declaredScopes": row.get("declaredScopes") or [],
        "securityTechnicalService": row.get("securityTechnicalService"),
        "securityTechnicalServiceKey": row.get("securityTechnicalServiceKey"),
        "securityTechnologyModuleOrMeasure": row.get("securityTechnologyModuleOrMeasure"),
        "securitySystem": row.get("securitySystem"),
        "issueTypes": row_issue_types(row),
        "sourceCells": source_row.get("sourceCells") or row.get("sourceCells") or {},
        "mergedRanges": source_row.get("mergedRanges") or row.get("mergedRanges") or {},
    }


def relation_difference_types(record: dict[str, Any]) -> list[str]:
    origin = as_text(record.get("relationOrigin"))
    exact_count = int(record.get("exactEnvironmentUsageCount") or 0)
    module_service_match = record.get("moduleServiceConsistent")
    system_module_match = record.get("systemModuleConsistent")
    difference_types: list[str] = []
    if origin == "catalog" and exact_count == 0:
        difference_types.append("catalog_unused")
    if origin == "environmentOnly":
        difference_types.append("environment_only")
    if module_service_match is False:
        difference_types.append("module_service_mismatch")
    if system_module_match is False:
        difference_types.append("system_module_mismatch")
    if record.get("possibleAliasMatches"):
        difference_types.append("possible_alias")
    if origin == "catalog" and exact_count == 0 and not any(item in difference_types for item in ("module_service_mismatch", "system_module_mismatch", "environment_only")):
        difference_types.append("selective_reference_candidate")
    if not difference_types:
        difference_types.append("aligned")
    return difference_types


def review_suggestion(difference_types: list[str]) -> str:
    if "environment_only" in difference_types and "module_service_mismatch" in difference_types and "system_module_mismatch" in difference_types:
        return "环境映射表中该系统-模块-服务三元关系不在目录中；先判断模块清单是否漏维护，还是环境映射表选错服务/模块/系统。"
    if "module_service_mismatch" in difference_types:
        return "核对模块清单是否应补充该模块-服务关系，或环境映射表是否选错模块/服务。"
    if "system_module_mismatch" in difference_types:
        return "核对模块清单是否应补充该系统-模块关系，或环境映射表是否选错安全系统。"
    if "possible_alias" in difference_types:
        return "优先确认是否只是命名、全半角、标点或后缀口径差异；确认后再决定是否建立受控 alias。"
    if "catalog_unused" in difference_types or "selective_reference_candidate" in difference_types:
        return "目录存在但环境映射表未精确引用；按选择性引用候选处理，不默认补到环境映射表。"
    return "目录关系与环境映射表存在精确使用记录，可作为对照样例。"


def build_dual_table_review(audit: dict[str, Any], review_rows: list[dict[str, Any]], alias_matches: list[dict[str, Any]]) -> dict[str, Any]:
    catalog_relations = audit.get("catalogRelations") or []
    module_rows = [row for row in review_rows if row.get("childType") == "module"]
    catalog_by_exact: dict[str, list[dict[str, Any]]] = defaultdict(list)
    catalog_by_module_service: dict[str, list[dict[str, Any]]] = defaultdict(list)
    catalog_by_system_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
    system_categories: dict[str, set[str]] = defaultdict(set)
    for relation in catalog_relations:
        system = as_text(relation.get("securitySystem"))
        module = as_text(relation.get("securityTechnologyModule"))
        service = as_text(relation.get("securityTechnicalServiceKey"))
        exact = relation_key(system, module, service)
        catalog_by_exact[exact].append(relation)
        catalog_by_module_service[relation_key("", module, service)].append(relation)
        catalog_by_system_module[relation_key(system, module, "")].append(relation)
        append_index(system_categories, system, as_text(relation.get("securitySystemCategory")))

    env_by_exact: dict[str, list[dict[str, Any]]] = defaultdict(list)
    env_by_module_service: dict[str, list[dict[str, Any]]] = defaultdict(list)
    env_by_system_module: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in module_rows:
        system = as_text(row.get("securitySystem"))
        module = as_text(row.get("securityTechnologyModuleOrMeasure"))
        service = as_text(row.get("securityTechnicalServiceKey"))
        exact = relation_key(system, module, service)
        env_by_exact[exact].append(row)
        env_by_module_service[relation_key("", module, service)].append(row)
        env_by_system_module[relation_key(system, module, "")].append(row)

    alias_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for match in alias_matches:
        for row_number in match.get("rows") or []:
            if isinstance(row_number, int) or str(row_number).isdigit():
                alias_by_row[int(row_number)].append(match)

    def alias_for_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        result = []
        for row in rows:
            row_number = row.get("excelRow")
            if isinstance(row_number, int) or str(row_number).isdigit():
                result.extend(alias_by_row.get(int(row_number), []))
        seen = set()
        unique = []
        for item in result:
            key = (item.get("environmentValue"), item.get("catalogValue"), tuple(item.get("rows") or []))
            if key in seen:
                continue
            seen.add(key)
            unique.append(item)
        return unique[:12]

    relation_records: list[dict[str, Any]] = []
    for exact, catalog_items in sorted(catalog_by_exact.items()):
        first = catalog_items[0]
        system = as_text(first.get("securitySystem"))
        module = as_text(first.get("securityTechnologyModule"))
        service = as_text(first.get("securityTechnicalServiceKey"))
        related_rows = env_by_exact.get(exact) or []
        module_service_rows = env_by_module_service.get(relation_key("", module, service)) or []
        system_module_rows = env_by_system_module.get(relation_key(system, module, "")) or []
        possible_alias = alias_for_rows(related_rows + module_service_rows + system_module_rows)
        record = {
            "relationKey": f"catalog::{exact}",
            "relationOrigin": "catalog",
            "securitySystemCategory": as_text(first.get("securitySystemCategory")),
            "securitySystem": system,
            "securityTechnologyModule": module,
            "securityTechnicalService": as_text(first.get("securityTechnicalService")),
            "securityTechnicalServiceKey": service,
            "exactEnvironmentUsageCount": len(related_rows),
            "environmentUsageCount": len(related_rows),
            "relatedEnvironmentUsageCount": len(set([row.get("excelRow") for row in module_service_rows + system_module_rows if row.get("excelRow")])),
            "informationEnvironmentCount": len({as_text(row.get("environment")) for row in related_rows if as_text(row.get("environment"))}),
            "environmentSegmentCount": len({as_text(row.get("environmentSegment")) for row in related_rows if as_text(row.get("environmentSegment"))}),
            "informationObjectCount": len({as_text(row.get("objectContextKey")) for row in related_rows if as_text(row.get("objectContextKey"))}),
            "scopeCount": len({scope for row in related_rows for scope in row.get("declaredScopes") or [] if scope}),
            "moduleServiceConsistent": True,
            "systemModuleConsistent": True,
            "possibleAliasMatches": possible_alias,
            "catalogEvidence": compact_source_evidence(
                catalog_items,
                ["securitySystemCategory", "securitySystem", "securityTechnologyModule", "securityTechnicalService", "product"],
            ),
            "environmentUsageRows": [environment_usage_position(row) for row in related_rows],
            "relatedEnvironmentUsageRows": [environment_usage_position(row) for row in (module_service_rows + system_module_rows)[:80]],
        }
        record["differenceTypes"] = relation_difference_types(record)
        record["reviewSuggestion"] = review_suggestion(record["differenceTypes"])
        relation_records.append(record)

    for exact, rows in sorted(env_by_exact.items()):
        if exact in catalog_by_exact:
            continue
        first = rows[0]
        system = as_text(first.get("securitySystem"))
        module = as_text(first.get("securityTechnologyModuleOrMeasure"))
        service = as_text(first.get("securityTechnicalServiceKey"))
        module_service_consistent = bool(catalog_by_module_service.get(relation_key("", module, service)))
        system_module_consistent = bool(catalog_by_system_module.get(relation_key(system, module, "")))
        catalog_candidates = catalog_by_module_service.get(relation_key("", module, service)) or catalog_by_system_module.get(relation_key(system, module, "")) or []
        possible_alias = alias_for_rows(rows)
        category = sorted_unique(list(system_categories.get(system, set())), limit=1)
        record = {
            "relationKey": f"environment::{exact}",
            "relationOrigin": "environmentOnly",
            "securitySystemCategory": category[0] if category else "",
            "securitySystem": system,
            "securityTechnologyModule": module,
            "securityTechnicalService": as_text(first.get("securityTechnicalService")),
            "securityTechnicalServiceKey": service,
            "exactEnvironmentUsageCount": len(rows),
            "environmentUsageCount": len(rows),
            "relatedEnvironmentUsageCount": len(rows),
            "informationEnvironmentCount": len({as_text(row.get("environment")) for row in rows if as_text(row.get("environment"))}),
            "environmentSegmentCount": len({as_text(row.get("environmentSegment")) for row in rows if as_text(row.get("environmentSegment"))}),
            "informationObjectCount": len({as_text(row.get("objectContextKey")) for row in rows if as_text(row.get("objectContextKey"))}),
            "scopeCount": len({scope for row in rows for scope in row.get("declaredScopes") or [] if scope}),
            "moduleServiceConsistent": module_service_consistent,
            "systemModuleConsistent": system_module_consistent,
            "possibleAliasMatches": possible_alias,
            "catalogEvidence": compact_source_evidence(
                catalog_candidates,
                ["securitySystemCategory", "securitySystem", "securityTechnologyModule", "securityTechnicalService", "product"],
            ) if catalog_candidates else {},
            "catalogCandidateRelations": catalog_candidates[:12],
            "environmentUsageRows": [environment_usage_position(row) for row in rows],
            "relatedEnvironmentUsageRows": [environment_usage_position(row) for row in rows],
        }
        record["differenceTypes"] = relation_difference_types(record)
        record["reviewSuggestion"] = review_suggestion(record["differenceTypes"])
        relation_records.append(record)

    relation_records = sorted(
        relation_records,
        key=lambda item: (
            0 if item["relationOrigin"] == "environmentOnly" and "module_service_mismatch" in item["differenceTypes"] else 1,
            0 if item["relationOrigin"] == "environmentOnly" and "system_module_mismatch" in item["differenceTypes"] else 1,
            0 if item["relationOrigin"] == "catalog" and item["exactEnvironmentUsageCount"] == 0 else 1,
            as_text(item.get("securitySystemCategory")),
            as_text(item.get("securitySystem")),
            as_text(item.get("securityTechnologyModule")),
            as_text(item.get("securityTechnicalServiceKey")),
        ),
    )

    def cluster_by(records: list[dict[str, Any]], key_fields: list[str], cluster_type: str) -> list[dict[str, Any]]:
        groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in records:
            key = "||".join(as_text(record.get(field)) for field in key_fields)
            groups[key].append(record)
        clusters = []
        for key, items in sorted(groups.items(), key=lambda item: (-sum(int(row.get("environmentUsageCount") or 0) for row in item[1]), item[0])):
            clusters.append(
                {
                    "clusterKey": key,
                    "clusterType": cluster_type,
                    "relationCount": len(items),
                    "environmentUsageCount": sum(int(row.get("environmentUsageCount") or 0) for row in items),
                    "informationObjectCount": len({usage.get("objectContextKey") for row in items for usage in row.get("environmentUsageRows") or [] if usage.get("objectContextKey")}),
                    "differenceTypes": sorted_unique([diff for row in items for diff in row.get("differenceTypes") or []]),
                    "sampleRelationKeys": [row.get("relationKey") for row in items[:12]],
                }
            )
        return clusters

    tree: dict[str, Any] = {"categories": []}
    category_map: dict[str, dict[str, Any]] = {}
    for record in relation_records:
        category_name = as_text(record.get("securitySystemCategory")) or "未归类"
        system_name = as_text(record.get("securitySystem")) or "未填写安全系统"
        module_name = as_text(record.get("securityTechnologyModule")) or "未填写安全技术模块"
        category_node = category_map.setdefault(category_name, {"title": category_name, "systems": {}, "relationCount": 0})
        system_node = category_node["systems"].setdefault(system_name, {"title": system_name, "modules": {}, "relationCount": 0})
        module_node = system_node["modules"].setdefault(module_name, {"title": module_name, "relations": [], "relationCount": 0})
        module_node["relations"].append(record.get("relationKey"))
        module_node["relationCount"] += 1
        system_node["relationCount"] += 1
        category_node["relationCount"] += 1
    for category_node in category_map.values():
        category_node["systems"] = [
            {**system_node, "modules": list(system_node["modules"].values())}
            for system_node in category_node["systems"].values()
        ]
        for system_node in category_node["systems"]:
            system_node["modules"] = sorted(system_node["modules"], key=lambda item: item["title"])
        category_node["systems"] = sorted(category_node["systems"], key=lambda item: item["title"])
        tree["categories"].append(category_node)
    tree["categories"] = sorted(tree["categories"], key=lambda item: item["title"])

    difference_counts = Counter(diff for record in relation_records for diff in record.get("differenceTypes") or [])
    return {
        "summary": {
            "directoryRelationCount": len(relation_records),
            "catalogRelationCount": len(catalog_relations),
            "environmentModuleRelationCount": len(module_rows),
            "environmentOnlyRelationCount": sum(1 for record in relation_records if record.get("relationOrigin") == "environmentOnly"),
            "catalogUnusedRelationCount": sum(1 for record in relation_records if record.get("relationOrigin") == "catalog" and int(record.get("exactEnvironmentUsageCount") or 0) == 0),
            "moduleServiceMismatchRelationCount": sum(1 for record in relation_records if "module_service_mismatch" in record.get("differenceTypes", [])),
            "systemModuleMismatchRelationCount": sum(1 for record in relation_records if "system_module_mismatch" in record.get("differenceTypes", [])),
            "possibleAliasRelationCount": sum(1 for record in relation_records if record.get("possibleAliasMatches")),
            "selectiveReferenceCandidateCount": sum(1 for record in relation_records if "selective_reference_candidate" in record.get("differenceTypes", [])),
            "differenceTypeCounts": dict(sorted(difference_counts.items())),
        },
        "tree": tree,
        "directoryRelationRows": relation_records,
        "clusters": {
            "bySecuritySystemModuleService": relation_records,
            "byModuleService": cluster_by(relation_records, ["securityTechnologyModule", "securityTechnicalServiceKey"], "moduleService"),
            "bySystemModule": cluster_by(relation_records, ["securitySystem", "securityTechnologyModule"], "systemModule"),
            "environmentOnlyRelations": [record for record in relation_records if record.get("relationOrigin") == "environmentOnly"],
            "catalogUnusedRelations": [record for record in relation_records if record.get("relationOrigin") == "catalog" and int(record.get("exactEnvironmentUsageCount") or 0) == 0],
            "possibleAliasRelationClusters": [record for record in relation_records if record.get("possibleAliasMatches")],
        },
        "differenceTypeLabels": {
            "aligned": "一致",
            "catalog_unused": "目录有，环境未用",
            "environment_only": "环境有，目录没有",
            "module_service_mismatch": "模块-服务不一致",
            "system_module_mismatch": "系统-模块不一致",
            "possible_alias": "可能别名",
            "selective_reference_candidate": "选择性引用候选",
        },
    }


def cluster_review_rows(
    rows: list[dict[str, Any]],
    issue_type: str,
    key_fn,
    pattern_fn,
    category: str,
    suggested_action: str,
) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if issue_type not in row_issue_types(row):
            continue
        groups[key_fn(row)].append(row)
    clusters = []
    for key, items in sorted(groups.items(), key=lambda item: (-len(item[1]), item[0])):
        source_rows = [item.get("excelRow") for item in items]
        clusters.append(
            {
                "patternSignature": pattern_fn(key),
                "issueType": issue_type,
                "triageCategory": category,
                "triageCategoryLabel": TRIAGE_LABELS[category],
                "affectedRows": len(items),
                "affectedObjectContexts": len({as_text(item.get("objectContextKey")) for item in items if as_text(item.get("objectContextKey"))}),
                "securityTechnicalServices": sorted_unique([item.get("securityTechnicalService") for item in items], limit=20),
                "securityTechnologyModuleOrMeasure": key if "::" not in key else key.split("::")[-1],
                "securitySystems": sorted_unique([item.get("securitySystem") for item in items], limit=20),
                "sourceRows": compact_rows(source_rows, limit=50),
                "sampleRows": items[:5],
                "suggestedManualAction": suggested_action,
            }
        )
    return clusters


def cluster_coverage_gaps(items: list[dict[str, Any]], issue_type: str, category: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        if issue_type == "serviceModuleCoverageGapCandidate":
            key = as_text(item.get("securityTechnicalService"))
            pattern = f"coverageGap::serviceModule::{key}"
        else:
            key = as_text(item.get("securityTechnologyModule"))
            pattern = f"coverageGap::moduleService::{key}"
        groups[pattern].append(item)
    clusters = []
    for pattern, group_items in sorted(groups.items(), key=lambda item: (-len(item[1]), item[0])):
        rows = [item.get("sampleExcelRow") for item in group_items]
        clusters.append(
            {
                "patternSignature": pattern,
                "issueType": issue_type,
                "triageCategory": category,
                "triageCategoryLabel": TRIAGE_LABELS[category],
                "affectedRows": len(group_items),
                "affectedObjectContexts": len({as_text(item.get("objectContextKey")) for item in group_items if as_text(item.get("objectContextKey"))}),
                "securityTechnicalServices": sorted_unique([item.get("securityTechnicalService") for item in group_items], limit=20),
                "securityTechnologyModuleOrMeasure": as_text(group_items[0].get("securityTechnologyModule")) if group_items else "",
                "securitySystems": [],
                "sourceRows": compact_rows(rows, limit=50),
                "sampleRows": group_items[:5],
                "suggestedManualAction": "确认环境映射表是否应按目录全量覆盖，还是允许按对象场景选择性引用。",
            }
        )
    return clusters


def cluster_repeated_service_expansion(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    clusters = []
    for item in items:
        context_key = as_text(item.get("objectContextKey"))
        service = as_text(item.get("securityTechnicalService"))
        clusters.append(
            {
                "patternSignature": f"serviceExpansion::{context_key}::{service}",
                "issueType": "repeatedServiceWithDifferentChildren",
                "triageCategory": "I",
                "triageCategoryLabel": TRIAGE_LABELS["I"],
                "affectedRows": len(set(item.get("rows") or [])),
                "affectedObjectContexts": 1 if context_key else 0,
                "securityTechnicalServices": [service] if service else [],
                "securityTechnologyModuleOrMeasure": "",
                "securitySystems": sorted_unique([child.get("securitySystem") for child in item.get("children") or []], limit=20),
                "sourceRows": compact_rows(item.get("rows") or [], limit=50),
                "sampleRows": [item],
                "suggestedManualAction": "作为 1:N 展开提示核对，不按错误处理；仅在业务认为该服务不应展开时再回源表确认。",
            }
        )
    return sorted(clusters, key=lambda item: (-item["affectedRows"], item["patternSignature"]))


def possible_alias_matches(review_rows: list[dict[str, Any]], limit: int = 120) -> list[dict[str, Any]]:
    candidates: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in review_rows:
        issues = row_issue_types(row)
        if not issues:
            continue
        match = row.get("catalogMatch") or {}
        comparisons: list[tuple[str, str, list[str]]] = []
        child = as_text(row.get("securityTechnologyModuleOrMeasure"))
        system = as_text(row.get("securitySystem"))
        allowed_modules = match.get("allowedModulesFromCatalog") or []
        allowed_systems = match.get("allowedSystemsFromCatalog") or []
        child_has_exact_catalog_match = normalize_for_alias(child) in {
            normalize_for_alias(value) for value in allowed_modules
        }
        system_has_exact_catalog_match = normalize_for_alias(system) in {
            normalize_for_alias(value) for value in allowed_systems
        }
        if not child_has_exact_catalog_match:
            for value in allowed_modules:
                comparisons.append((child, value, issues))
        if not system_has_exact_catalog_match:
            for value in allowed_systems:
                comparisons.append((system, value, issues))
        for left, right, issue_types in comparisons:
            hint = alias_hint(left, right)
            if not hint:
                continue
            key = (as_text(left), as_text(right), hint)
            payload = candidates.setdefault(
                key,
                {
                    "environmentValue": as_text(left),
                    "catalogValue": as_text(right),
                    "normalizedEnvironmentValue": normalize_for_alias(left),
                    "normalizedCatalogValue": normalize_for_alias(right),
                    "similarityHint": hint,
                    "issueTypes": [],
                    "rows": [],
                    "patternSignatures": [],
                },
            )
            payload["issueTypes"] = sorted_unique(payload["issueTypes"] + issue_types)
            payload["rows"] = compact_rows(payload["rows"] + [row.get("excelRow")], limit=50)
            for issue in issue_types:
                if issue == "moduleServiceMismatch":
                    payload["patternSignatures"].append(f"moduleServiceMismatch::{child}")
                elif issue == "systemModuleMismatch":
                    payload["patternSignatures"].append(f"systemModuleMismatch::{system}::{child}")
            payload["patternSignatures"] = sorted_unique(payload["patternSignatures"], limit=20)
    return sorted(candidates.values(), key=lambda item: (-len(item["rows"]), item["environmentValue"], item["catalogValue"]))[:limit]


def build_top_manual_review_items(patterns: dict[str, list[dict[str, Any]]], alias_matches: list[dict[str, Any]], limit: int = 50) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []

    def add_pattern(pattern: dict[str, Any], issue_type: str | None = None) -> None:
        sample = (pattern.get("sampleRows") or [{}])[0]
        items.append(
            {
                "priority": 0,
                "issueType": issue_type or pattern.get("issueType"),
                "triageCategory": pattern.get("triageCategory"),
                "triageCategoryLabel": pattern.get("triageCategoryLabel"),
                "patternSignature": pattern.get("patternSignature"),
                "affectedRows": pattern.get("affectedRows", 0),
                "affectedObjectContexts": pattern.get("affectedObjectContexts", 0),
                "securityTechnicalService": " / ".join(pattern.get("securityTechnicalServices") or []),
                "securityTechnologyModuleOrMeasure": pattern.get("securityTechnologyModuleOrMeasure") or as_text(sample.get("securityTechnologyModuleOrMeasure")),
                "securitySystem": " / ".join(pattern.get("securitySystems") or []),
                "catalogEvidence": sample.get("catalogMatch") or {
                    key: sample.get(key)
                    for key in ("allowedModulesFromCatalog", "allowedServicesFromCatalog", "allowedSystemsFromCatalog")
                    if sample.get(key) is not None
                },
                "environmentEvidence": {
                    key: sample.get(key)
                    for key in (
                        "objectContextKey",
                        "securityTechnicalService",
                        "securityTechnicalServiceKey",
                        "securityTechnologyModuleOrMeasure",
                        "securityTechnologyModule",
                        "securitySystem",
                    )
                    if sample.get(key) is not None
                },
                "suggestedManualAction": pattern.get("suggestedManualAction"),
                "sourceRows": pattern.get("sourceRows") or [],
            }
        )

    for pattern in patterns.get("duplicateExactServiceChildRelations", []):
        add_pattern(pattern)
    for pattern in patterns.get("moduleMeasureClassificationIssue", []):
        add_pattern(pattern)
    for pattern in patterns.get("moduleServiceMismatch", [])[:12]:
        add_pattern(pattern)
    for pattern in patterns.get("systemModuleMismatch", [])[:12]:
        add_pattern(pattern)
    for match in alias_matches[:6]:
        items.append(
            {
                "priority": 0,
                "issueType": "possibleAliasMatches",
                "triageCategory": "H",
                "triageCategoryLabel": TRIAGE_LABELS["H"],
                "patternSignature": (match.get("patternSignatures") or ["possibleAliasMatches"])[0],
                "affectedRows": len(match.get("rows") or []),
                "affectedObjectContexts": 0,
                "securityTechnicalService": "",
                "securityTechnologyModuleOrMeasure": match.get("environmentValue", ""),
                "securitySystem": "",
                "catalogEvidence": {"catalogValue": match.get("catalogValue"), "normalizedCatalogValue": match.get("normalizedCatalogValue")},
                "environmentEvidence": {"environmentValue": match.get("environmentValue"), "normalizedEnvironmentValue": match.get("normalizedEnvironmentValue")},
                "suggestedManualAction": "核对是否只是命名、标点、全半角或后缀差异；若确认，后续再单独建立受控 alias 规则。",
                "sourceRows": match.get("rows") or [],
            }
        )
    for pattern in patterns.get("serviceModuleCoverageGapCandidate", [])[:4]:
        add_pattern(pattern)
    for pattern in patterns.get("moduleServiceCoverageGapCandidate", [])[:4]:
        add_pattern(pattern)
    for pattern in patterns.get("repeatedServiceWithDifferentChildren", [])[:4]:
        add_pattern(pattern)

    items = sorted(
        items,
        key=lambda item: (
            {"A": 0, "B": 1, "C": 2, "D": 3, "G": 4, "H": 5, "E": 6, "F": 7, "I": 8}.get(as_text(item.get("triageCategory")), 9),
            -int(item.get("affectedRows") or 0),
            as_text(item.get("patternSignature")),
        ),
    )[:limit]
    for index, item in enumerate(items, start=1):
        item["priority"] = index
    return items


def build_triage(audit: dict[str, Any]) -> dict[str, Any]:
    issues = audit.get("issues") or {}
    review_rows = load_json(REVIEW_ROWS_OUT, []) or []
    duplicate_patterns = []
    for item in audit.get("duplicateExactServiceChildRelations") or []:
        key = "duplicateExactServiceChildRelations::" + "::".join(
            as_text(item.get(field))
            for field in ("objectContextKey", "securityTechnicalService", "childType", "securityTechnologyModuleOrMeasure", "securitySystem")
            if item.get(field) is not None
        )
        duplicate_patterns.append(
            {
                "patternSignature": key,
                "issueType": "duplicateExactServiceChildRelations",
                "triageCategory": "A",
                "triageCategoryLabel": TRIAGE_LABELS["A"],
                "affectedRows": len(set(item.get("rows") or [])),
                "affectedObjectContexts": 1 if item.get("objectContextKey") else 0,
                "securityTechnicalServices": sorted_unique([item.get("securityTechnicalService")]),
                "securityTechnologyModuleOrMeasure": as_text(item.get("securityTechnologyModuleOrMeasure")),
                "securitySystems": sorted_unique([item.get("securitySystem")]),
                "sourceRows": compact_rows(item.get("rows") or [], limit=50),
                "sampleRows": [item],
                "suggestedManualAction": "回到原始 Excel 删除或合并完整重复关系；该类若存在应作为高风险处理。",
            }
        )
    patterns = {
        "duplicateExactServiceChildRelations": duplicate_patterns,
        "moduleMeasureClassificationIssue": cluster_review_rows(
            review_rows,
            "moduleMeasureClassificationIssue",
            lambda row: as_text(row.get("securityTechnologyModuleOrMeasure")),
            lambda key: f"moduleMeasureClassificationIssue::{key}",
            "B",
            "核对该名称应作为安全技术模块还是安全技术措施；必要时修正源表分类或目录。",
        ),
        "moduleServiceMismatch": cluster_review_rows(
            review_rows,
            "moduleServiceMismatch",
            lambda row: as_text(row.get("securityTechnologyModuleOrMeasure")),
            lambda key: f"moduleServiceMismatch::{key}",
            "C",
            "判断环境映射表中的模块-服务关系是否应补进模块清单，或源表是否选错模块/服务。",
        ),
        "systemModuleMismatch": cluster_review_rows(
            review_rows,
            "systemModuleMismatch",
            lambda row: f"{as_text(row.get('securitySystem'))}::{as_text(row.get('securityTechnologyModuleOrMeasure'))}",
            lambda key: f"systemModuleMismatch::{key}",
            "D",
            "判断环境映射表中的系统-模块关系是否应补进模块清单，或源表是否选错安全系统。",
        ),
        "securitySystemCoverageOrMismatch": cluster_review_rows(
            review_rows,
            "securitySystemCoverageOrMismatch",
            lambda row: f"{as_text(row.get('securitySystem'))}::{as_text(row.get('securityTechnologyModuleOrMeasure'))}",
            lambda key: f"securitySystemCoverageOrMismatch::{key}",
            "G",
            "核对安全系统是否为空、候选过多或与模块清单不一致；确认哪个目录是准的。",
        ),
        "serviceModuleCoverageGapCandidate": cluster_coverage_gaps(issues.get("serviceModuleCoverageGapCandidate") or [], "serviceModuleCoverageGapCandidate", "E"),
        "moduleServiceCoverageGapCandidate": cluster_coverage_gaps(issues.get("moduleServiceCoverageGapCandidate") or [], "moduleServiceCoverageGapCandidate", "F"),
        "repeatedServiceWithDifferentChildren": cluster_repeated_service_expansion(audit.get("repeatedServiceWithDifferentChildren") or []),
    }
    alias_matches = possible_alias_matches(review_rows)
    dual_table_review = build_dual_table_review(audit, review_rows, alias_matches)
    category_counts = {
        "A": len(audit.get("duplicateExactServiceChildRelations") or []),
        "B": len(issues.get("moduleMeasureClassificationIssue") or []),
        "C": len(issues.get("moduleServiceMismatch") or []),
        "D": len(issues.get("systemModuleMismatch") or []),
        "E": len(issues.get("serviceModuleCoverageGapCandidate") or []),
        "F": len(issues.get("moduleServiceCoverageGapCandidate") or []),
        "G": len(issues.get("securitySystemCoverageOrMismatch") or []),
        "H": len(alias_matches),
        "I": len(audit.get("repeatedServiceWithDifferentChildren") or []),
    }
    issue_type_counts = {
        "duplicateExactServiceChildRelations": len(audit.get("duplicateExactServiceChildRelations") or []),
        "repeatedServiceWithDifferentChildren": len(audit.get("repeatedServiceWithDifferentChildren") or []),
        **{key: len(value or []) for key, value in issues.items()},
        "possibleAliasMatches": len(alias_matches),
    }
    top_items = build_top_manual_review_items(patterns, alias_matches, limit=50)
    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceAudit": str(AUDIT_JSON_OUT.relative_to(PROJECT_ROOT)),
        "sourceReviewRows": str(REVIEW_ROWS_OUT.relative_to(PROJECT_ROOT)),
        "summary": audit.get("summary") or {},
        "issueTypeCounts": issue_type_counts,
        "triageCategoryLabels": TRIAGE_LABELS,
        "triageCategoryCounts": {key: category_counts.get(key, 0) for key in sorted(TRIAGE_LABELS)},
        "patternClusters": patterns,
        "patternClusterCounts": {key: len(value) for key, value in patterns.items()},
        "possibleAliasMatches": alias_matches,
        "dualTableReview": dual_table_review,
        "topManualReviewItems": top_items,
        "conclusion": {
            "formalDataModified": False,
            "formalUiModified": False,
            "coverageGapsAreErrors": False,
            "serviceExpansionIsError": False,
            "duplicateExactServiceChildRelationCountUsesAuditValue": True,
        },
    }


def build_triage_markdown(triage: dict[str, Any]) -> str:
    summary = triage.get("summary") or {}
    lines = [
        "# Environment Module Catalog Consistency Triage",
        "",
        f"- 生成时间：{triage['generatedAt']}",
        f"- 审计来源：`{triage['sourceAudit']}`",
        f"- 完整重复关系：`{summary.get('duplicateExactServiceChildRelationCount', 0)}`",
        f"- moduleServiceMismatch patterns：`{triage['patternClusterCounts'].get('moduleServiceMismatch', 0)}`",
        f"- systemModuleMismatch patterns：`{triage['patternClusterCounts'].get('systemModuleMismatch', 0)}`",
        f"- possibleAliasMatches：`{len(triage.get('possibleAliasMatches') or [])}`",
        f"- 双表对照目录关系：`{(triage.get('dualTableReview') or {}).get('summary', {}).get('directoryRelationCount', 0)}`",
        f"- Top 人工核对项：`{len(triage.get('topManualReviewItems') or [])}`",
        "",
        "## triageCategory 数量",
        "",
        "| 类别 | 含义 | 数量 |",
        "|---|---|---:|",
    ]
    for category, label in TRIAGE_LABELS.items():
        lines.append(f"| {category} | {label} | {triage['triageCategoryCounts'].get(category, 0)} |")
    lines.extend(["", "## Top 人工核对项", ""])
    lines.append(
        markdown_table(
            triage.get("topManualReviewItems") or [],
            [
                ("priority", "优先级"),
                ("triageCategory", "类别"),
                ("issueType", "问题类型"),
                ("patternSignature", "模式"),
                ("affectedRows", "影响行"),
                ("affectedObjectContexts", "影响上下文"),
                ("suggestedManualAction", "建议动作"),
            ],
            limit=50,
        )
    )
    lines.extend(["", "## moduleServiceMismatch 聚类", ""])
    lines.append(
        markdown_table(
            triage["patternClusters"].get("moduleServiceMismatch") or [],
            [("patternSignature", "模式"), ("affectedRows", "影响行"), ("affectedObjectContexts", "上下文"), ("sourceRows", "样例行")],
            limit=30,
        )
    )
    lines.extend(["", "## systemModuleMismatch 聚类", ""])
    lines.append(
        markdown_table(
            triage["patternClusters"].get("systemModuleMismatch") or [],
            [("patternSignature", "模式"), ("affectedRows", "影响行"), ("affectedObjectContexts", "上下文"), ("sourceRows", "样例行")],
            limit=30,
        )
    )
    lines.extend(
        [
            "",
            "## 治理说明",
            "",
            "- `serviceModuleCoverageGapCandidate` 和 `moduleServiceCoverageGapCandidate` 默认是目录覆盖差异候选，不作为错误。",
            "- `repeatedServiceWithDifferentChildren` 显示为 1:N 展开，不作为错误。",
            "- `possibleAliasMatches` 只提示命名疑似差异，不自动替换，不建立正式 alias 字典。",
        ]
    )
    return "\n".join(lines) + "\n"


def markdown_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], limit: int = 20) -> str:
    if not rows:
        return "_无_"
    lines = ["| " + " | ".join(label for _, label in columns) + " |", "| " + " | ".join("---" for _ in columns) + " |"]
    for row in rows[:limit]:
        lines.append("| " + " | ".join(str(row.get(key, "")).replace("\n", "<br>") for key, _ in columns) + " |")
    if len(rows) > limit:
        lines.append(f"| ... | 还有 {len(rows) - limit} 条 |" + " |" * (len(columns) - 2))
    return "\n".join(lines)


def build_markdown(audit: dict[str, Any]) -> str:
    summary = audit["summary"]
    issues = audit["issues"]
    lines = [
        "# Environment Module Catalog Consistency Audit",
        "",
        f"- 生成时间：{audit['generatedAt']}",
        f"- 原始文件：`{audit['workbook']}`",
        f"- 环境映射 normalized rows：`{summary['normalizedRowCount']}`",
        f"- 安全技术模块清单 catalogRelations：`{summary['catalogRelationCount']}`",
        f"- 完整重复关系：`{summary['duplicateExactServiceChildRelationCount']}`",
        f"- 合法服务多下级展开提示：`{summary['repeatedServiceWithDifferentChildrenCount']}`",
        f"- moduleServiceMismatch：`{summary['moduleServiceMismatchCount']}`",
        f"- systemModuleMismatch：`{summary['systemModuleMismatchCount']}`",
        f"- serviceModuleCoverageGapCandidate：`{summary['serviceModuleCoverageGapCandidateCount']}`",
        f"- moduleServiceCoverageGapCandidate：`{summary['moduleServiceCoverageGapCandidateCount']}`",
        f"- securitySystemCoverageOrMismatch：`{summary['securitySystemCoverageOrMismatchCount']}`",
        f"- moduleMeasureClassificationIssue：`{summary['moduleMeasureClassificationIssueCount']}`",
        f"- scopeCompletenessIssues：`{summary['scopeCompletenessIssueCount']}`",
        "",
        "## 样例问题",
        "",
        "### 模块-服务目录不一致",
        "",
        markdown_table(
            issues["moduleServiceMismatch"],
            [("excelRow", "Excel 行"), ("objectContextKey", "对象上下文"), ("securityTechnicalService", "服务"), ("securityTechnologyModuleOrMeasure", "模块")],
            limit=20,
        ),
        "",
        "### 系统-模块目录不一致",
        "",
        markdown_table(
            issues["systemModuleMismatch"],
            [("excelRow", "Excel 行"), ("objectContextKey", "对象上下文"), ("securitySystem", "安全系统"), ("securityTechnologyModuleOrMeasure", "模块")],
            limit=20,
        ),
        "",
        "### 安全系统缺失或错配候选",
        "",
        markdown_table(
            issues["securitySystemCoverageOrMismatch"],
            [("excelRow", "Excel 行"), ("objectContextKey", "对象上下文"), ("securityTechnologyModuleOrMeasure", "模块"), ("securitySystem", "实际系统")],
            limit=20,
        ),
        "",
        "## 结论",
        "",
        "- 本报告只做跨表一致性审计，不自动补齐、替换或修改正式业务数据。",
        "- `repeatedServiceWithDifferentChildren` 属于 1:N 展开的信息提示，不作为高风险错误。",
        "- `serviceModuleCoverageGapCandidate` / `moduleServiceCoverageGapCandidate` 只表示目录覆盖差异候选，需要人工判断环境映射表是否应完整列出。",
    ]
    return "\n".join(lines) + "\n"


def main() -> int:
    catalog_relations, catalog_sheet_summary = read_module_catalog(WORKBOOK_PATH)
    indices = build_catalog_indices(catalog_relations)
    normalized_rows = load_json(NORMALIZED_ROWS_PATH, [])
    scope_audit = load_json(SCOPE_AUDIT_PATH, {}) or {}
    workbench = load_json(WORKBENCH_PATH, {}) or {}
    module_titles = set(indices["securityTechnologyModuleToServices"].keys())
    measure_titles = measure_titles_from_workbench(workbench)
    review_rows = [
        make_review_row(entry, indices, module_titles, measure_titles)
        for entry in env_entries(normalized_rows)
    ]
    service_module_gaps, module_service_gaps = coverage_candidates(review_rows, indices)
    issues = {
        "moduleServiceMismatch": issue_items(review_rows, "moduleServiceMismatch"),
        "systemModuleMismatch": issue_items(review_rows, "systemModuleMismatch"),
        "serviceModuleCoverageGapCandidate": service_module_gaps,
        "moduleServiceCoverageGapCandidate": module_service_gaps,
        "securitySystemCoverageOrMismatch": issue_items(review_rows, "securitySystemCoverageOrMismatch"),
        "moduleMeasureClassificationIssue": issue_items(review_rows, "moduleMeasureClassificationIssue"),
        "scopeCompletenessIssues": scope_audit.get("scopeCompletenessIssues") or [],
    }
    summary = {
        "workbook": str(WORKBOOK_PATH.relative_to(PROJECT_ROOT)),
        "catalogSheet": catalog_sheet_summary,
        "normalizedRowCount": len(normalized_rows),
        "reviewRowCount": len(review_rows),
        "catalogRelationCount": len(catalog_relations),
        "duplicateExactServiceChildRelationCount": len(scope_audit.get("duplicateExactServiceChildRelations") or []),
        "repeatedServiceWithDifferentChildrenCount": len(scope_audit.get("repeatedServiceWithDifferentChildren") or []),
        "moduleServiceMismatchCount": len(issues["moduleServiceMismatch"]),
        "systemModuleMismatchCount": len(issues["systemModuleMismatch"]),
        "serviceModuleCoverageGapCandidateCount": len(service_module_gaps),
        "moduleServiceCoverageGapCandidateCount": len(module_service_gaps),
        "securitySystemCoverageOrMismatchCount": len(issues["securitySystemCoverageOrMismatch"]),
        "moduleMeasureClassificationIssueCount": len(issues["moduleMeasureClassificationIssue"]),
        "scopeCompletenessIssueCount": len(issues["scopeCompletenessIssues"]),
        "hasIusScopeGap": any("I-US" in item.get("missingScopes", []) for item in issues["scopeCompletenessIssues"]),
    }
    audit = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": summary["workbook"],
        "summary": summary,
        "catalogRelations": catalog_relations,
        "catalogIndices": indices,
        "duplicateExactServiceChildRelations": scope_audit.get("duplicateExactServiceChildRelations") or [],
        "repeatedServiceWithDifferentChildren": scope_audit.get("repeatedServiceWithDifferentChildren") or [],
        "issues": issues,
        "reviewRowsPath": str(REVIEW_ROWS_OUT.relative_to(PROJECT_ROOT)),
        "conclusion": {
            "formalDataModified": False,
            "formalUiModified": False,
            "notes": "本轮只输出跨表一致性审计和临时核对行，不自动补全 environment-workbench 或 node-details。",
        },
    }
    write_json(AUDIT_JSON_OUT, audit)
    write_json(REVIEW_ROWS_OUT, review_rows)
    AUDIT_MD_OUT.write_text(build_markdown(audit), encoding="utf-8")
    triage = build_triage(audit)
    write_json(TRIAGE_JSON_OUT, triage)
    TRIAGE_MD_OUT.write_text(build_triage_markdown(triage), encoding="utf-8")
    print(
        json.dumps(
            {
                "result": "pass",
                **summary,
                "triageJson": str(TRIAGE_JSON_OUT.relative_to(PROJECT_ROOT)),
                "triageMd": str(TRIAGE_MD_OUT.relative_to(PROJECT_ROOT)),
                "topManualReviewItemCount": len(triage.get("topManualReviewItems") or []),
                "moduleServiceMismatchPatternCount": triage.get("patternClusterCounts", {}).get("moduleServiceMismatch", 0),
                "systemModuleMismatchPatternCount": triage.get("patternClusterCounts", {}).get("systemModuleMismatch", 0),
                "possibleAliasMatchCount": len(triage.get("possibleAliasMatches") or []),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
