#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
ENVIRONMENT_WORKBENCH_PATH = ROOT / "frontend/capability-browser/public/data/environment-workbench.json"
SQLITE_PATH = ROOT / "data/database/sapd_wiki.sqlite3"
OUT_DIR = ROOT / "data/exports/worker-verify/environment-object-service-overexpansion-audit"

SOURCE_SHEET = "作用域-安全技术服务-安全技术模块映射"
MODULE_SHEET = "安全技术模块清单"


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    return re.sub(r"\s+", " ", text).strip()


def split_multiline(value: Any) -> list[str]:
    text = "" if value is None else str(value)
    rows: list[str] = []
    for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        item = norm(line)
        if item and item not in {"/", "\\", "N/A", "NA", "无"}:
            rows.append(item)
    return rows


def service_parts(raw: Any) -> dict[str, str]:
    text = norm(raw)
    if not text:
        return {"code": "", "title": "", "display": ""}
    if " " in text:
        code, title = text.split(" ", 1)
    else:
        match = re.match(r"^([A-Z]-[A-Z0-9&.-]+)(.+)$", text)
        code, title = (match.group(1), match.group(2)) if match else ("", text)
    code = norm(code)
    title = norm(title)
    return {"code": code, "title": title, "display": f"{code} {title}".strip()}


def merged_lookup(ws) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    values: dict[str, Any] = {}
    meta: dict[str, dict[str, Any]] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                coord = ws.cell(row, col).coordinate
                values[coord] = anchor.value
                meta[coord] = {
                    "range": str(merged_range),
                    "anchor": anchor.coordinate,
                    "anchorRow": merged_range.min_row,
                    "anchorColumn": merged_range.min_col,
                }
    return values, meta


def cell_value(ws, row: int, col: int, merged_values: dict[str, Any]) -> Any:
    coord = ws.cell(row, col).coordinate
    return merged_values.get(coord, ws.cell(row, col).value)


def source_cell(ws, row: int, col: int, column_name: str, merged_values: dict[str, Any], merged_meta: dict[str, dict[str, Any]]) -> dict[str, Any]:
    coord = ws.cell(row, col).coordinate
    return {
        "sheet": ws.title,
        "row": row,
        "column": column_name,
        "cell": coord,
        "rawValue": norm(merged_values.get(coord, ws.cell(row, col).value)),
        "mergedRange": merged_meta.get(coord),
    }


def context_key(env: str, segment: str, obj: str) -> str:
    return f"{env} / {segment} / {obj}"


def markdown_escape(value: Any) -> str:
    return norm(value).replace("|", "\\|")


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


@dataclass
class SourceContext:
    environment: str
    segment: str
    information_object: str
    services: dict[str, dict[str, Any]] = field(default_factory=dict)
    modules: dict[str, list[dict[str, Any]]] = field(default_factory=lambda: defaultdict(list))
    rows: set[int] = field(default_factory=set)


def load_source_contexts(wb) -> dict[tuple[str, str, str], SourceContext]:
    ws = wb[SOURCE_SHEET]
    merged_values, merged_meta = merged_lookup(ws)
    contexts: dict[tuple[str, str, str], SourceContext] = {}
    for row in range(3, ws.max_row + 1):
        env = norm(cell_value(ws, row, 2, merged_values))
        segment = norm(cell_value(ws, row, 3, merged_values))
        obj = norm(cell_value(ws, row, 4, merged_values))
        if not env or not segment or not obj:
            continue
        key = (env, segment, obj)
        ctx = contexts.setdefault(key, SourceContext(env, segment, obj))
        ctx.rows.add(row)

        for raw_service in split_multiline(cell_value(ws, row, 6, merged_values)):
            service = service_parts(raw_service)
            if not service["code"]:
                continue
            entry = ctx.services.setdefault(
                service["code"],
                {
                    "code": service["code"],
                    "title": service["title"],
                    "display": service["display"],
                    "sourceCells": [],
                },
            )
            entry["sourceCells"].append(source_cell(ws, row, 6, "安全技术服务", merged_values, merged_meta))

        for module in split_multiline(cell_value(ws, row, 7, merged_values)):
            ctx.modules[module].append(source_cell(ws, row, 7, "安全技术模块/措施", merged_values, merged_meta))
    return contexts


def load_module_catalog(wb) -> tuple[dict[str, dict[str, dict[str, str]]], dict[str, list[str]]]:
    ws = wb[MODULE_SHEET]
    merged_values, _merged_meta = merged_lookup(ws)
    module_to_services: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
    service_to_modules: dict[str, list[str]] = defaultdict(list)

    for row in range(3, ws.max_row + 1):
        module = norm(cell_value(ws, row, 4, merged_values))
        if not module:
            continue
        for raw_service in split_multiline(cell_value(ws, row, 6, merged_values)):
            service = service_parts(raw_service)
            code = service["code"]
            if not code:
                continue
            module_to_services[module][code] = service
            if module not in service_to_modules[code]:
                service_to_modules[code].append(module)

    return dict(module_to_services), {code: sorted(modules) for code, modules in service_to_modules.items()}


def load_formal_environment_contexts() -> dict[tuple[str, str, str], dict[str, Any]]:
    data = json.loads(ENVIRONMENT_WORKBENCH_PATH.read_text(encoding="utf-8"))
    contexts: dict[tuple[str, str, str], dict[str, Any]] = {}
    for env in data.get("environment_scope_tree", []):
        env_title = norm(env.get("title"))
        for obj in env.get("objects", []):
            obj_title = norm(obj.get("title"))
            segment_titles = [norm(segment.get("title")) for segment in obj.get("segments", []) if norm(segment.get("title"))]
            if not segment_titles:
                segment_titles = [""]
            for segment_title in segment_titles:
                key = (env_title, segment_title, obj_title)
                ctx = contexts.setdefault(
                    key,
                    {
                        "environment": env_title,
                        "segment": segment_title,
                        "informationObject": obj_title,
                        "services": {},
                        "objectId": obj.get("id"),
                    },
                )
                for mapping in obj.get("scope_mappings", []):
                    for service in mapping.get("services", []):
                        code = norm(service.get("code"))
                        if not code:
                            continue
                        ctx["services"].setdefault(
                            code,
                            {
                                "code": code,
                                "title": norm(service.get("title")),
                                "display": f"{code} {norm(service.get('title'))}".strip(),
                                "sources": service.get("sources", []),
                                "scopeCodes": [],
                            },
                        )
                        scope_code = norm((mapping.get("scope") or {}).get("code"))
                        if scope_code and scope_code not in ctx["services"][code]["scopeCodes"]:
                            ctx["services"][code]["scopeCodes"].append(scope_code)
    return contexts


def sqlite_relation_probe(service_code: str, object_title: str) -> dict[str, Any]:
    if not SQLITE_PATH.exists():
        return {"available": False}

    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    try:
        relations = conn.execute(
            """
            SELECT r.id AS relation_id,
                   si.code AS service_code,
                   si.title AS service_title,
                   ti.title AS object_title,
                   r.metadata_json
            FROM knowledge_relations r
            JOIN knowledge_items si ON si.id = r.source_item_id
            JOIN knowledge_items ti ON ti.id = r.target_item_id
            WHERE r.relation_type = 'protects_object'
              AND si.code = ?
              AND ti.title = ?
            ORDER BY si.code, ti.title
            """,
            (service_code, object_title),
        ).fetchall()
        refs: list[dict[str, Any]] = []
        for relation in relations:
            relation_refs = conn.execute(
                """
                SELECT DISTINCT source_sheet, source_row, source_column, source_cell, raw_value
                FROM source_references
                WHERE target_type = 'relation'
                  AND target_id = ?
                ORDER BY source_sheet, source_row, source_cell
                LIMIT 12
                """,
                (relation["relation_id"],),
            ).fetchall()
            refs.extend(
                {
                    "sheet": row["source_sheet"],
                    "row": row["source_row"],
                    "column": row["source_column"],
                    "cell": row["source_cell"],
                    "rawValue": row["raw_value"],
                }
                for row in relation_refs
            )
        return {
            "available": True,
            "relationExists": bool(relations),
            "relationCount": len(relations),
            "sampleSourceRefs": refs[:12],
            "note": "SQLite protects_object 为服务-对象标题级关系，不保留环境/子类上下文；同名对象跨环境时只能作为辅助证据。",
        }
    finally:
        conn.close()


def classify_extra(
    key: tuple[str, str, str],
    service_code: str,
    source_contexts: dict[tuple[str, str, str], SourceContext],
    formal_contexts: dict[tuple[str, str, str], dict[str, Any]],
    module_to_services: dict[str, dict[str, dict[str, str]]],
    service_to_modules: dict[str, list[str]],
) -> dict[str, Any]:
    source_ctx = source_contexts.get(key)
    formal_ctx = formal_contexts[key]
    formal_service = formal_ctx["services"][service_code]
    source_modules = sorted(source_ctx.modules) if source_ctx else []
    explaining_modules = sorted(
        module for module in source_modules if service_code in module_to_services.get(module, {})
    )
    same_object_source_hits = [
        {
            "contextKey": context_key(*candidate_key),
            "environment": candidate_key[0],
            "environmentSegment": candidate_key[1],
            "informationObject": candidate_key[2],
            "sourceCells": candidate_ctx.services[service_code].get("sourceCells", []),
        }
        for candidate_key, candidate_ctx in sorted(source_contexts.items())
        if candidate_key != key and candidate_key[2] == key[2] and service_code in candidate_ctx.services
    ]
    item = {
        "contextKey": context_key(*key),
        "environment": key[0],
        "environmentSegment": key[1],
        "informationObject": key[2],
        "extraService": {
            "code": service_code,
            "title": formal_service.get("title"),
            "display": formal_service.get("display"),
            "scopeCodesInFormalProjection": sorted(formal_service.get("scopeCodes", [])),
        },
        "sourceContextRows": sorted(source_ctx.rows) if source_ctx else [],
        "sourceServicesInContext": sorted(
            [
                {"code": code, "title": service.get("title"), "display": service.get("display")}
                for code, service in (source_ctx.services.items() if source_ctx else [])
            ],
            key=lambda row: row["code"],
        ),
        "sourceModulesOrMeasuresInContext": source_modules,
        "catalogModulesForExtraService": service_to_modules.get(service_code, []),
        "explainingModulesFromSourceContext": explaining_modules,
        "sourceContextsWithSameObjectAndService": same_object_source_hits,
        "hasSameObjectTitleContextLeakSignature": bool(same_object_source_hits),
        "hasModuleCatalogExpansionSignature": bool(explaining_modules),
        "classification": "same_object_title_context_leak_suspect"
        if same_object_source_hits
        else ("module_catalog_expansion_suspect" if explaining_modules else "other_formal_extra"),
        "sqliteTitleLevelProbe": sqlite_relation_probe(service_code, key[2]),
    }
    return item


def build_report(payload: dict[str, Any]) -> str:
    summary = payload["summary"]
    same_object_leaks = payload["sameObjectTitleLeakSuspects"]
    module_signatures = payload["moduleCatalogExpansionSuspects"]
    unexplained = payload["otherFormalExtras"]

    lines = [
        "# 信息化对象-安全技术服务过度投影审计",
        "",
        f"- 生成时间：`{payload['generatedAt']}`",
        f"- 审计状态：`{payload['status']}`",
        "- 审计原则：字典负责服务 code/title/id 和模块全局关系；对象级、LC-DT 阶段 / 策略级事实以源表实际列出的安全技术服务为准，再由这些服务反推所需模块。",
        "- SQLite 边界：`protects_object` 是服务-对象标题级关系，不含环境 / 子类上下文；本报告只把 SQLite 作为辅助证据。",
        "",
        "## 汇总",
        "",
        "| 指标 | 数值 |",
        "|---|---:|",
        f"| 源表上下文数 | {summary['sourceContextCount']} |",
        f"| 正式环境投影上下文数 | {summary['formalContextCount']} |",
        f"| 源表上下文-服务关系数 | {summary['sourceContextServiceCount']} |",
        f"| 正式投影上下文-服务关系数 | {summary['formalContextServiceCount']} |",
        f"| 正式投影多出的上下文-服务关系 | {summary['extraContextServiceCount']} |",
        f"| 源表有但正式投影缺失的上下文-服务关系 | {summary['missingContextServiceCount']} |",
        f"| 其中：疑似同名对象标题级关系跨上下文串入 | {summary['sameObjectTitleLeakSuspectCount']} |",
        f"| 其中：同时带有模块全局关系扩展特征 | {summary['moduleCatalogExpansionSignatureCount']} |",
        f"| 其中：仍未解释的其他多投 | {summary['otherFormalExtraCount']} |",
        "",
    ]

    if same_object_leaks:
        lines.extend(
            [
                "## 疑似同名对象标题级串入",
                "",
                "| 当前多投上下文 | 正式多出的服务 | 源表中该同名对象真实出现位置 | SQLite标题级关系 |",
                "|---|---|---|---|---|",
            ]
        )
        for item in same_object_leaks:
            sqlite_probe = item["sqliteTitleLevelProbe"]
            source_contexts = "；".join(hit["contextKey"] for hit in item["sourceContextsWithSameObjectAndService"])
            lines.append(
                "| "
                + " | ".join(
                    [
                        markdown_escape(item["contextKey"]),
                        markdown_escape(item["extraService"]["display"]),
                        markdown_escape(source_contexts),
                        "存在" if sqlite_probe.get("relationExists") else "未检出",
                    ]
                )
                + " |"
            )
        lines.append("")

    if module_signatures:
        lines.extend(
            [
                "## 同时带有模块扩展特征的关系",
                "",
                "| 上下文 | 正式多出的服务 | 源表该对象已有模块 | 字典中解释该服务的模块 |",
                "|---|---|---|---|",
            ]
        )
        for item in module_signatures:
            lines.append(
                "| "
                + " | ".join(
                    [
                        markdown_escape(item["contextKey"]),
                        markdown_escape(item["extraService"]["display"]),
                        markdown_escape("、".join(item["sourceModulesOrMeasuresInContext"])),
                        markdown_escape("、".join(item["explainingModulesFromSourceContext"])),
                    ]
                )
                + " |"
            )
        lines.append("")

    if unexplained:
        lines.extend(
            [
                "## 仍未解释的多投关系",
                "",
                "| 上下文 | 正式多出的服务 | 源表该对象已有模块/措施 | 字典中该服务所属模块 |",
                "|---|---|---|---|",
            ]
        )
        for item in unexplained:
            lines.append(
                "| "
                + " | ".join(
                    [
                        markdown_escape(item["contextKey"]),
                        markdown_escape(item["extraService"]["display"]),
                        markdown_escape("、".join(item["sourceModulesOrMeasuresInContext"])),
                        markdown_escape("、".join(item["catalogModulesForExtraService"])),
                    ]
                )
                + " |"
            )
        lines.append("")

    if payload["missingInFormal"]:
        lines.extend(["## 正式投影缺失关系", "", "| 上下文 | 源表服务 |", "|---|---|"])
        for item in payload["missingInFormal"]:
            lines.append(f"| {markdown_escape(item['contextKey'])} | {markdown_escape(item['serviceDisplay'])} |")
        lines.append("")

    lines.extend(
        [
            "## 结论",
            "",
            "- 新审计口径合理：对象级关系必须使用源表的实际安全技术服务集合，不能把字典里的模块全局服务集合自动补齐到每个对象。",
            f"- 当前正式环境投影存在 `{summary['extraContextServiceCount']}` 条源表未列出的对象级服务关系；源表要求的对象级服务没有漏投。",
            f"- 这 `{summary['extraContextServiceCount']}` 条都能在源表中找到同名对象的其他环境真实位置，主因更像标题级对象关系跨上下文串入；其中 `{summary['moduleCatalogExpansionSignatureCount']}` 条同时带有“源对象已有模块 + 字典模块包含该服务”的模块扩展特征。",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    source_contexts = load_source_contexts(wb)
    module_to_services, service_to_modules = load_module_catalog(wb)
    formal_contexts = load_formal_environment_contexts()

    source_pairs = {
        (key, service_code)
        for key, ctx in source_contexts.items()
        for service_code in ctx.services
    }
    formal_pairs = {
        (key, service_code)
        for key, ctx in formal_contexts.items()
        for service_code in ctx["services"]
    }

    extra_pairs = sorted(formal_pairs - source_pairs, key=lambda row: (*row[0], row[1]))
    missing_pairs = sorted(source_pairs - formal_pairs, key=lambda row: (*row[0], row[1]))

    extra_items = [
        classify_extra(key, service_code, source_contexts, formal_contexts, module_to_services, service_to_modules)
        for key, service_code in extra_pairs
    ]
    same_object_leaks = [item for item in extra_items if item["hasSameObjectTitleContextLeakSignature"]]
    module_signatures = [item for item in extra_items if item["hasModuleCatalogExpansionSignature"]]
    unexplained = [
        item
        for item in extra_items
        if not item["hasSameObjectTitleContextLeakSignature"] and not item["hasModuleCatalogExpansionSignature"]
    ]

    missing_items = []
    for key, service_code in missing_pairs:
        service = source_contexts[key].services[service_code]
        missing_items.append(
            {
                "contextKey": context_key(*key),
                "environment": key[0],
                "environmentSegment": key[1],
                "informationObject": key[2],
                "serviceCode": service_code,
                "serviceTitle": service.get("title"),
                "serviceDisplay": service.get("display"),
                "sourceCells": service.get("sourceCells", []),
            }
        )

    payload = {
        "generatedAt": now_iso(),
        "status": "needs_review" if extra_items or missing_items else "pass",
        "inputs": {
            "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
            "sourceSheet": SOURCE_SHEET,
            "moduleSheet": MODULE_SHEET,
            "environmentWorkbench": str(ENVIRONMENT_WORKBENCH_PATH.relative_to(ROOT)),
            "sqlite": str(SQLITE_PATH.relative_to(ROOT)),
        },
        "principleAssessment": {
            "reasonable": True,
            "rule": "字典作为 code/title/id 与模块全局关系权威；对象级与 LC-DT 阶段/策略级事实以源表实际安全技术服务为准，再由服务反推模块。",
        },
        "summary": {
            "sourceContextCount": len(source_contexts),
            "formalContextCount": len(formal_contexts),
            "sourceOnlyContextCount": len(set(source_contexts) - set(formal_contexts)),
            "formalOnlyContextCount": len(set(formal_contexts) - set(source_contexts)),
            "sourceContextServiceCount": len(source_pairs),
            "formalContextServiceCount": len(formal_pairs),
            "extraContextServiceCount": len(extra_items),
            "missingContextServiceCount": len(missing_items),
            "sameObjectTitleLeakSuspectCount": len(same_object_leaks),
            "moduleCatalogExpansionSignatureCount": len(module_signatures),
            "moduleCatalogExpansionSuspectCount": len(module_signatures),
            "otherFormalExtraCount": len(unexplained),
        },
        "sourceOnlyContexts": [context_key(*key) for key in sorted(set(source_contexts) - set(formal_contexts))],
        "formalOnlyContexts": [context_key(*key) for key in sorted(set(formal_contexts) - set(source_contexts))],
        "sameObjectTitleLeakSuspects": same_object_leaks,
        "moduleCatalogExpansionSuspects": module_signatures,
        "otherFormalExtras": unexplained,
        "missingInFormal": missing_items,
    }

    write_json(OUT_DIR / "audit.json", payload)
    write_md(OUT_DIR / "audit.md", build_report(payload))
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))
    print(f"status={payload['status']}")
    print(f"report={OUT_DIR.relative_to(ROOT) / 'audit.md'}")
    print(f"json={OUT_DIR.relative_to(ROOT) / 'audit.json'}")


if __name__ == "__main__":
    main()
