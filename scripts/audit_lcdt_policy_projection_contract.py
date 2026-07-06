#!/usr/bin/env python3
"""Audit LC-DT policy row projections against the source Excel mapping sheet."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "data/raw-samples/wiki sample.xlsx"
DATA_DIR = ROOT / "frontend/capability-browser/public/data"
MAPPING_SHEET = "LC-DT 安全技术服务、模块、策略映射表"


def norm(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split()).strip()


def split_multiline(value: Any) -> list[str]:
    raw = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
    return [norm(part) for part in raw.split("\n") if norm(part) and norm(part) != "\\"]


def service_code(value: str) -> str:
    match = re.match(r"^([A-Z]-[A-Z0-9&]+-[A-Z]{2}\.[A-Z]{2}-\d{2})\s+(.+)$", value)
    return match.group(1) if match else value.split(" ", 1)[0].strip()


def stage_title(value: Any) -> str:
    return re.sub(r"（[A-Z]{2}）$", "", norm(value)).strip()


def source_rows() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    try:
        worksheet = workbook[MAPPING_SHEET]
        rows: dict[str, dict[str, Any]] = {}
        last_stage = ""
        last_category = ""
        for row_number, row in enumerate(worksheet.iter_rows(min_row=4), start=4):
            stage = stage_title(row[1].value if len(row) > 1 else None) or last_stage
            if stage:
                last_stage = stage
            category = norm(row[2].value if len(row) > 2 else None) or last_category
            if category:
                last_category = category
            sequence = norm(row[3].value if len(row) > 3 else None)
            services = split_multiline(row[12].value if len(row) > 12 else None)
            modules = split_multiline(row[13].value if len(row) > 13 else None)
            if not stage or not sequence:
                continue
            rows[f"data-policy:{stage}:{row_number}"] = {
                "excelRow": row_number,
                "stage": stage,
                "category": category,
                "sequence": sequence,
                "serviceCodes": [service_code(item) for item in services],
                "moduleTitles": modules,
            }
        return rows
    finally:
        workbook.close()


def read_json(name: str) -> dict[str, Any]:
    return json.loads((DATA_DIR / name).read_text(encoding="utf-8"))


def lifecycle_rows(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for process in payload.get("data_lifecycle", {}).get("processes", []):
        for row in process.get("data_policy_rows", []):
            rows[str(row.get("id") or "")] = row
    return rows


def workbench_rows(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for stage in payload.get("objects", {}).get("lifecycle_stage", {}).values():
        for row in stage.get("dataPolicyRows", []):
            rows[str(row.get("id") or "")] = row
    return rows


def projected_service_codes(row: dict[str, Any]) -> list[str]:
    return [norm(item.get("code")) for item in row.get("technical_services", []) if norm(item.get("code"))]


def projected_module_titles(row: dict[str, Any]) -> list[str]:
    items = row.get("module_or_measure_items") or row.get("technology_modules") or row.get("technical_measures") or []
    return [norm(item.get("title") or item.get("name")) for item in items if norm(item.get("title") or item.get("name"))]


def compare_projection(label: str, expected: dict[str, dict[str, Any]], actual: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    failures: list[dict[str, Any]] = []
    for row_id, source in expected.items():
        row = actual.get(row_id)
        if not row:
            failures.append({"package": label, "rowId": row_id, "reason": "missing_policy_row", "source": source})
            continue
        service_codes = projected_service_codes(row)
        module_titles = projected_module_titles(row)
        if service_codes != source["serviceCodes"]:
            failures.append(
                {
                    "package": label,
                    "rowId": row_id,
                    "reason": "service_codes_mismatch",
                    "expected": source["serviceCodes"],
                    "actual": service_codes,
                }
            )
        if module_titles != source["moduleTitles"]:
            failures.append(
                {
                    "package": label,
                    "rowId": row_id,
                    "reason": "module_titles_mismatch",
                    "expected": source["moduleTitles"],
                    "actual": module_titles,
                }
            )
    return failures


def main() -> int:
    expected = source_rows()
    lifecycle = lifecycle_rows(read_json("lifecycle-knowledge.json"))
    workbench = workbench_rows(read_json("lifecycle-workbench.json"))
    failures = [
        *compare_projection("lifecycle-knowledge.json", expected, lifecycle),
        *compare_projection("lifecycle-workbench.json", expected, workbench),
    ]
    critical = {
        "data-policy:加工/使用:40": ["I-AP&T-PD.DP-01", "I-OS&T-PD.DP-01"],
        "data-policy:提供:55": ["I-DI&T-PD.DP-01"],
    }
    for row_id, expected_codes in critical.items():
        for label, rows in (("lifecycle-knowledge.json", lifecycle), ("lifecycle-workbench.json", workbench)):
            actual = projected_service_codes(rows.get(row_id, {}))
            if actual != expected_codes:
                failures.append({"package": label, "rowId": row_id, "reason": "critical_watermark_scope_mismatch", "expected": expected_codes, "actual": actual})
    report = {
        "result": "fail" if failures else "pass",
        "sourceRows": len(expected),
        "checkedRows": len(expected) * 2,
        "failureCount": len(failures),
        "failures": failures[:50],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
