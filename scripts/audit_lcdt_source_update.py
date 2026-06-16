#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
OUT_DIR = ROOT / "data/exports/worker-verify/lcdt-source-update"

LIFECYCLE_SHEET = "LC-DT 数据生命周期"
MAPPING_SHEET = "LC-DT 安全技术服务、模块、策略映射表"

SERVICES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/services.json"
MODULES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/modules.json"
MEASURES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/measures.json"
MAINTENANCE_KNOWLEDGE_PATH = ROOT / "frontend/capability-browser/public/data/maintenance-knowledge.json"
MODULE_SOURCE_SHEET = "安全技术模块清单"

EXPECTED_SYSTEM_CORRECTIONS = {
    "知情同意管理": "数据安全管理与运营",
    "隐私安全影响评估": "数据安全管理与运营",
}


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    return re.sub(r"[ \t]+", " ", text).strip()


def norm_key(value: Any) -> str:
    return norm(value).replace("（", "(").replace("）", ")")


def split_multiline(value: Any) -> list[str]:
    text = norm(value)
    if not text or text in {"/", "\\", "N/A", "NA", "无"}:
        return []
    parts: list[str] = []
    for line in str(value).replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        item = norm(line)
        if item and item not in {"/", "\\", "N/A", "NA", "无"}:
            parts.append(item)
    return parts


def stage_title(value: Any) -> str:
    text = norm(value)
    if not text:
        return ""
    first_line = text.splitlines()[0].strip()
    return re.sub(r"\s*[（(][^)）]+[)）]\s*$", "", first_line).strip()


def lifecycle_code(order: Any) -> str:
    text = norm(order)
    if not text:
        return ""
    try:
        return f"DT-{int(float(text)):02d}"
    except ValueError:
        return f"DT-{text}"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_md(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def merged_lookup(ws) -> tuple[dict[str, Any], dict[str, dict[str, Any]]]:
    values: dict[str, Any] = {}
    meta: dict[str, dict[str, Any]] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                coord = ws.cell(row, col).coordinate
                values[coord] = anchor.value
                meta[coord] = {
                    "range": str(merged_range),
                    "anchor": anchor.coordinate,
                    "anchorRow": merged_range.min_row,
                    "anchorColumn": merged_range.min_col,
                    "rowSpan": merged_range.max_row - merged_range.min_row + 1,
                    "columnSpan": merged_range.max_col - merged_range.min_col + 1,
                }
    return values, meta


def cell(ws, row: int, col: int, values: dict[str, Any]) -> Any:
    coord = ws.cell(row, col).coordinate
    if coord in values:
        return values[coord]
    return ws.cell(row, col).value


def source_cell(ws, row: int, col: int, values: dict[str, Any], meta: dict[str, dict[str, Any]], column_name: str) -> dict[str, Any]:
    c = ws.cell(row, col)
    coord = c.coordinate
    return {
        "sheet": ws.title,
        "row": row,
        "column": column_name,
        "cell": coord,
        "rawValue": norm(values.get(coord, c.value)),
        "mergedRange": meta.get(coord),
    }


def service_parts(raw: str) -> dict[str, str]:
    text = norm(raw)
    if not text:
        return {"code": "", "title": "", "display": ""}
    if " " in text:
        code, title = text.split(" ", 1)
    else:
        match = re.match(r"^([A-Z0-9&.-]+)(.+)$", text)
        code, title = (match.group(1), match.group(2)) if match else ("", text)
    return {"code": norm(code), "title": norm(title), "display": f"{norm(code)} {norm(title)}".strip()}


def service_display(item: dict[str, Any]) -> str:
    code = norm(item.get("code"))
    title = norm(item.get("title"))
    return f"{code} {title}".strip()


def compact_item(item: dict[str, Any], keys: tuple[str, ...]) -> dict[str, Any]:
    return {key: item.get(key) for key in keys if item.get(key) not in (None, [], "")}


@dataclass
class Dictionaries:
    services_by_code: dict[str, dict[str, Any]]
    services_by_display: dict[str, dict[str, Any]]
    service_titles: set[str]
    modules_by_title: dict[str, dict[str, Any]]
    measures_by_title: dict[str, dict[str, Any]]
    source_module_systems_by_title: dict[str, dict[str, Any]]


def load_source_service_candidates() -> list[dict[str, Any]]:
    try:
        from audit_security_technical_service_dictionary_update import (
            build_mapping,
            load_current_services,
            parse_source_services,
            source_candidate_services,
        )
    except ImportError:
        return []
    source_services, _merged_ranges, invalid_rows, _candidate_sheets = parse_source_services()
    if not source_services or invalid_rows:
        return []
    current_services, _raw_service_entries = load_current_services()
    mapping = build_mapping(source_services, current_services)
    return source_candidate_services(source_services, current_services, mapping)


def load_source_module_systems() -> dict[str, dict[str, Any]]:
    if not WORKBOOK_PATH.exists():
        return {}
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    if MODULE_SOURCE_SHEET not in wb.sheetnames:
        return {}
    ws = wb[MODULE_SOURCE_SHEET]
    merged_values, merged_meta = merged_lookup(ws)
    result: dict[str, dict[str, Any]] = {}
    for row in range(3, ws.max_row + 1):
        title = norm(cell(ws, row, 4, merged_values))
        if not title:
            continue
        result[title] = {
            "moduleTitle": title,
            "category": norm(cell(ws, row, 2, merged_values)),
            "securitySystem": norm(cell(ws, row, 3, merged_values)),
            "source": source_cell(ws, row, 3, merged_values, merged_meta, "安全系统"),
        }
    return result


def build_dictionaries() -> Dictionaries:
    services = []
    services_data = load_json(SERVICES_PATH)
    for row in services_data.get("security_technical_services", []):
        service = row.get("service") if isinstance(row.get("service"), dict) else row
        if isinstance(service, dict):
            services.append(service)
    maintenance = load_json(MAINTENANCE_KNOWLEDGE_PATH)
    for service in maintenance.get("security_technical_services", []):
        if isinstance(service, dict):
            services.append(service)
    services.extend(load_source_service_candidates())

    services_by_code: dict[str, dict[str, Any]] = {}
    services_by_display: dict[str, dict[str, Any]] = {}
    service_titles: set[str] = set()
    for service in services:
        code = norm(service.get("code"))
        title = norm(service.get("title"))
        if not code or not title:
            continue
        services_by_code[code] = service
        services_by_display[norm_key(f"{code} {title}")] = service
        service_titles.add(title)

    modules_data = load_json(MODULES_PATH)
    modules_by_title = {
        norm(module.get("title")): module
        for module in modules_data.get("security_technology_modules", [])
        if norm(module.get("title"))
    }
    measures_data = load_json(MEASURES_PATH)
    measures_by_title = {
        norm(measure.get("name") or measure.get("title")): measure
        for measure in measures_data.get("security_technical_measures", [])
        if norm(measure.get("name") or measure.get("title"))
    }
    return Dictionaries(
        services_by_code=services_by_code,
        services_by_display=services_by_display,
        service_titles=service_titles,
        modules_by_title=modules_by_title,
        measures_by_title=measures_by_title,
        source_module_systems_by_title=load_source_module_systems(),
    )


def classify_service(raw: str, dictionaries: Dictionaries) -> dict[str, Any]:
    parts = service_parts(raw)
    code = parts["code"]
    title = parts["title"]
    display_key = norm_key(parts["display"])
    by_display = dictionaries.services_by_display.get(display_key)
    by_code = dictionaries.services_by_code.get(code)
    if by_display:
        status = "matched"
        issue = None
        canonical = by_display
    elif by_code:
        canonical_title = norm(by_code.get("title"))
        status = "title_mismatch" if title and title != canonical_title else "matched_by_code"
        issue = f"源表名称 `{title}` 与字典 `{canonical_title}` 不一致" if status == "title_mismatch" else None
        canonical = by_code
    else:
        status = "unknown_service"
        issue = "未在安全技术服务字典中找到该服务编号"
        canonical = None
    possible_matches = []
    if status == "unknown_service" and title:
        possible_matches = [
            compact_item(service, ("id", "type", "code", "title", "category"))
            for service in dictionaries.services_by_code.values()
            if norm(service.get("title")) == title
        ]
        if possible_matches:
            issue = f"未在安全技术服务字典中找到编号 `{code}`；但存在同名服务：{', '.join(service_display(item) for item in possible_matches)}"
    return {
        "raw": raw,
        "code": code,
        "title": title,
        "display": parts["display"],
        "validationStatus": status,
        "issue": issue,
        "canonical": compact_item(canonical or {}, ("id", "type", "code", "title", "category")),
        "possibleMatches": possible_matches,
    }


def classify_module_or_measure(raw: str, dictionaries: Dictionaries) -> dict[str, Any]:
    title = norm(raw)
    if title in dictionaries.modules_by_title:
        module = dictionaries.modules_by_title[title]
        return {
            "raw": raw,
            "title": title,
            "kind": "security_technology_module",
            "validationStatus": "matched_module_catalog",
            "canonical": compact_item(module, ("id", "type", "title", "category")),
            "systems": [compact_item(system, ("id", "type", "title", "category")) for system in module.get("systems", [])],
        }
    alias_title = resolve_module_alias(title, dictionaries.modules_by_title)
    if alias_title:
        module = dictionaries.modules_by_title[alias_title]
        return {
            "raw": raw,
            "title": title,
            "kind": "security_technology_module",
            "validationStatus": "module_title_alias",
            "issue": f"源表模块/措施值 `{title}` 未完全等于模块清单标题，建议归一为 `{alias_title}`。",
            "canonical": compact_item(module, ("id", "type", "title", "category")),
            "resolvedTitle": alias_title,
            "systems": [compact_item(system, ("id", "type", "title", "category")) for system in module.get("systems", [])],
        }
    if title in dictionaries.measures_by_title:
        measure = dictionaries.measures_by_title[title]
        return {
            "raw": raw,
            "title": title,
            "kind": "security_technical_measure",
            "validationStatus": "matched_existing_measure",
            "canonical": compact_item(measure, ("id", "type", "name", "category")),
            "systems": [],
        }
    return {
        "raw": raw,
        "title": title,
        "kind": "security_technical_measure_candidate",
        "validationStatus": "not_in_module_catalog_or_measure_list",
        "canonical": {},
        "systems": [],
    }


def resolve_module_alias(title: str, modules_by_title: dict[str, dict[str, Any]]) -> str | None:
    if not title:
        return None
    alias_rules = [
        ("API安全防护" in title, "API安全防护"),
        ("主机安全管理" in title or "主机系统安全管理" in title or title == "应用程序控制", "主机安全管理"),
        ("文件完整性监控" in title or title == "主机入侵防御", "主机入侵防御（HIPS）"),
        ("终端安全检测与响应" in title, "终端安全检测与响应（EDR）"),
        ("终端恶意代码防护" in title, "终端恶意代码防护(EPP)"),
        ("终端数据防泄露" in title, "终端数据防泄露（EDLP）"),
        ("移动安全管理" in title, "移动安全管理(MTD)"),
        ("Web应用防火墙" in title, "Web应用防火墙（WAF）"),
        ("运行时应用自防护" in title, "运行时应用自防护（RASP）"),
        ("安全接入网关" in title, "安全接入网关（VPN）"),
        ("网络准入控制" in title, "网络准入控制（NAC）"),
        ("数据加密" in title, "数据加密和令牌化"),
        ("数据水印溯源" in title, "数据水印溯源"),
        ("数据脱敏" in title, "数据脱敏(去标识化)"),
        ("零信任访问代理" in title, "零信任访问代理"),
        ("零信任访问控制台" in title, "零信任访问控制台"),
        (title == "安全工作区", "终端安全工作区"),
    ]
    for matched, candidate in alias_rules:
        if matched and candidate in modules_by_title:
            return candidate
    contains_matches = [candidate for candidate in modules_by_title if candidate and candidate in title]
    if len(contains_matches) == 1:
        return contains_matches[0]
    loose = norm_key(title)
    for candidate in modules_by_title:
        if norm_key(candidate) == loose:
            return candidate
    return None


def parse_lifecycle_table(workbook, dictionaries: Dictionaries) -> list[dict[str, Any]]:
    ws = workbook[LIFECYCLE_SHEET]
    values, meta = merged_lookup(ws)
    rows: list[dict[str, Any]] = []
    by_stage: dict[str, dict[str, Any]] = {}

    for row in range(4, ws.max_row + 1):
        stage = norm(cell(ws, row, 3, values))
        if not stage:
            continue
        order = cell(ws, row, 2, values)
        code = lifecycle_code(order)
        bucket = by_stage.setdefault(
            stage,
            {
                "sourceTable": LIFECYCLE_SHEET,
                "stageId": code,
                "stageOrder": int(float(order)) if norm(order).replace(".", "", 1).isdigit() else norm(order),
                "stageName": stage,
                "excelRows": [],
                "sceneRows": [],
                "securityTechnicalServices": [],
                "securityTechnologyModulesOrMeasures": [],
                "sourceCells": {},
            },
        )
        bucket["excelRows"].append(row)
        scene_code = norm(cell(ws, row, 5, values))
        scene_title = norm(cell(ws, row, 6, values))
        if scene_code or scene_title:
            bucket["sceneRows"].append(
                {
                    "excelRow": row,
                    "sceneCode": scene_code,
                    "sceneTitle": scene_title,
                    "sceneDescription": norm(cell(ws, row, 7, values)),
                }
            )
        if "stage" not in bucket["sourceCells"]:
            bucket["sourceCells"]["stage"] = source_cell(ws, row, 3, values, meta, "数据处理")
            bucket["sourceCells"]["services"] = source_cell(ws, row, 8, values, meta, "安全技术服务")
            bucket["sourceCells"]["modules"] = source_cell(ws, row, 9, values, meta, "安全技术模块")

        if not bucket["securityTechnicalServices"]:
            bucket["securityTechnicalServices"] = [
                classify_service(item, dictionaries) for item in split_multiline(cell(ws, row, 8, values))
            ]
        if not bucket["securityTechnologyModulesOrMeasures"]:
            bucket["securityTechnologyModulesOrMeasures"] = [
                classify_module_or_measure(item, dictionaries) for item in split_multiline(cell(ws, row, 9, values))
            ]

    for stage in sorted(by_stage.values(), key=lambda row: row["stageOrder"] if isinstance(row["stageOrder"], int) else 999):
        stage["excelRowRange"] = f"{min(stage['excelRows'])}-{max(stage['excelRows'])}"
        rows.append(stage)
    return rows


def parse_mapping_table(workbook, dictionaries: Dictionaries) -> list[dict[str, Any]]:
    ws = workbook[MAPPING_SHEET]
    values, meta = merged_lookup(ws)
    rows: list[dict[str, Any]] = []
    last_stage = ""
    last_category = ""
    stage_order = {
        "收集/采集": 1,
        "存储": 2,
        "传输": 3,
        "加工/使用": 4,
        "提供": 5,
        "公开": 6,
        "删除": 7,
    }

    for row in range(6, ws.max_row + 1):
        stage = stage_title(cell(ws, row, 2, values)) or last_stage
        category = norm(cell(ws, row, 3, values)) or last_category
        if stage:
            last_stage = stage
        if category:
            last_category = category
        if not stage:
            continue
        services = [classify_service(item, dictionaries) for item in split_multiline(cell(ws, row, 13, values))]
        modules = [classify_module_or_measure(item, dictionaries) for item in split_multiline(cell(ws, row, 14, values))]
        policy_refs = []
        for col, name in [
            (5, "数据安全法"),
            (6, "个人信息保护法"),
            (7, "网络安全法"),
            (8, "GB/T 35273"),
            (9, "GB/T 37988"),
            (10, "GB/T 39335"),
            (11, "GB/T 43697"),
            (12, "其他策略"),
        ]:
            value = norm(cell(ws, row, col, values))
            if value:
                policy_refs.append({"column": name, "value": value})
        rows.append(
            {
                "sourceTable": MAPPING_SHEET,
                "excelRow": row,
                "stageId": lifecycle_code(stage_order.get(stage, "")),
                "stageName": stage,
                "category": category,
                "policySequence": norm(cell(ws, row, 4, values)),
                "policyReferences": policy_refs,
                "securityTechnicalServices": services,
                "securityTechnologyModulesOrMeasures": modules,
                "sourceCells": {
                    "stage": source_cell(ws, row, 2, values, meta, "阶段"),
                    "category": source_cell(ws, row, 3, values, meta, "类别"),
                    "services": source_cell(ws, row, 13, values, meta, "安全技术服务"),
                    "modules": source_cell(ws, row, 14, values, meta, "安全技术模块"),
                },
            }
        )
    return rows


def stage_sets_from_lifecycle(rows: list[dict[str, Any]]) -> dict[str, dict[str, set[str]]]:
    result: dict[str, dict[str, set[str]]] = {}
    for row in rows:
        result[row["stageName"]] = {
            "services": {item["display"] for item in row["securityTechnicalServices"]},
            "modulesOrMeasures": {item["title"] for item in row["securityTechnologyModulesOrMeasures"]},
        }
    return result


def stage_sets_from_mapping(rows: list[dict[str, Any]]) -> dict[str, dict[str, set[str]]]:
    result: dict[str, dict[str, set[str]]] = defaultdict(lambda: {"services": set(), "modulesOrMeasures": set()})
    for row in rows:
        stage = row["stageName"]
        result[stage]["services"].update(item["display"] for item in row["securityTechnicalServices"])
        result[stage]["modulesOrMeasures"].update(item["title"] for item in row["securityTechnologyModulesOrMeasures"])
    return result


def build_consistency(lifecycle_rows: list[dict[str, Any]], mapping_rows: list[dict[str, Any]]) -> dict[str, Any]:
    lifecycle_sets = stage_sets_from_lifecycle(lifecycle_rows)
    mapping_sets = stage_sets_from_mapping(mapping_rows)
    all_stages = sorted(
        set(lifecycle_sets) | set(mapping_sets),
        key=lambda s: next((r["stageOrder"] for r in lifecycle_rows if r["stageName"] == s), 999),
    )
    stage_results = []
    counters = Counter()
    for stage in all_stages:
        left = lifecycle_sets.get(stage, {"services": set(), "modulesOrMeasures": set()})
        right = mapping_sets.get(stage, {"services": set(), "modulesOrMeasures": set()})
        service_only_lifecycle = sorted(left["services"] - right["services"])
        service_only_mapping = sorted(right["services"] - left["services"])
        module_only_lifecycle = sorted(left["modulesOrMeasures"] - right["modulesOrMeasures"])
        module_only_mapping = sorted(right["modulesOrMeasures"] - left["modulesOrMeasures"])
        passed = not (service_only_lifecycle or service_only_mapping or module_only_lifecycle or module_only_mapping)
        counters["stageCompared"] += 1
        counters["stagePassed" if passed else "stageMismatch"] += 1
        counters["serviceOnlyLifecycle"] += len(service_only_lifecycle)
        counters["serviceOnlyMapping"] += len(service_only_mapping)
        counters["moduleOnlyLifecycle"] += len(module_only_lifecycle)
        counters["moduleOnlyMapping"] += len(module_only_mapping)
        stage_results.append(
            {
                "stageName": stage,
                "status": "pass" if passed else "mismatch",
                "serviceOnlyInLifecycleTable": service_only_lifecycle,
                "serviceOnlyInMappingTable": service_only_mapping,
                "moduleOrMeasureOnlyInLifecycleTable": module_only_lifecycle,
                "moduleOrMeasureOnlyInMappingTable": module_only_mapping,
            }
        )
    return {"summary": dict(counters), "stageResults": stage_results}


def build_dictionary_report(lifecycle_rows: list[dict[str, Any]], mapping_rows: list[dict[str, Any]]) -> dict[str, Any]:
    service_items = []
    module_items = []
    for table, rows in ((LIFECYCLE_SHEET, lifecycle_rows), (MAPPING_SHEET, mapping_rows)):
        for row in rows:
            row_ref = row.get("excelRowRange") or row.get("excelRow")
            for item in row["securityTechnicalServices"]:
                service_items.append({"sourceTable": table, "stageName": row["stageName"], "excelRow": row_ref, **item})
            for item in row["securityTechnologyModulesOrMeasures"]:
                module_items.append({"sourceTable": table, "stageName": row["stageName"], "excelRow": row_ref, **item})

    service_status = Counter(item["validationStatus"] for item in service_items)
    module_status = Counter(item["validationStatus"] for item in module_items)
    new_measure_candidates = sorted(
        {item["title"] for item in module_items if item["kind"] == "security_technical_measure_candidate"}
    )
    return {
        "summary": {
            "serviceReferenceCount": len(service_items),
            "serviceValidationStatus": dict(service_status),
            "moduleOrMeasureReferenceCount": len(module_items),
            "moduleOrMeasureValidationStatus": dict(module_status),
            "newMeasureCandidateCount": len(new_measure_candidates),
        },
        "serviceIssues": [
            item
            for item in service_items
            if item["validationStatus"] in {"unknown_service", "title_mismatch"}
        ],
        "moduleOrMeasureIssues": [
            item
            for item in module_items
            if item["kind"] == "security_technical_measure_candidate"
        ],
        "moduleAliasIssues": [
            item
            for item in module_items
            if item["validationStatus"] == "module_title_alias"
        ],
        "existingMeasureReferences": [
            item for item in module_items if item["kind"] == "security_technical_measure"
        ],
        "newMeasureCandidates": new_measure_candidates,
    }


def build_system_corrections(dictionaries: Dictionaries) -> list[dict[str, Any]]:
    corrections = []
    for title, target_system in EXPECTED_SYSTEM_CORRECTIONS.items():
        source_module = dictionaries.source_module_systems_by_title.get(title)
        module = dictionaries.modules_by_title.get(title)
        if source_module and source_module.get("securitySystem") == target_system:
            corrections.append(
                {
                    "moduleId": module.get("id") if module else None,
                    "moduleTitle": title,
                    "moduleCategory": source_module.get("category") or (module or {}).get("category"),
                    "currentSecuritySystems": [source_module.get("securitySystem")],
                    "targetSecuritySystem": target_system,
                    "status": "info_only",
                    "action": "no_op_source_already_correct",
                    "reason": "原始安全技术模块清单已将该模块归入数据安全管理与运营，本轮不再生成系统归属更正候选。",
                    "source": source_module.get("source"),
                }
            )
            continue
        if not module:
            corrections.append(
                {
                    "moduleTitle": title,
                    "targetSecuritySystem": target_system,
                    "status": "blocking_issue",
                    "issue": "模块未在安全技术模块清单中找到，无法生成系统归属候选更正。",
                }
            )
            continue
        current_systems = [norm(system.get("title")) for system in module.get("systems", []) if norm(system.get("title"))]
        if target_system in current_systems:
            status = "info_only"
            action = "no_op_already_correct"
        else:
            status = "confirmed_change"
            action = "add_security_system_relation"
        corrections.append(
            {
                "moduleId": module.get("id"),
                "moduleTitle": title,
                "moduleCategory": module.get("category"),
                "currentSecuritySystems": current_systems,
                "targetSecuritySystem": target_system,
                "status": status,
                "action": action,
                "reason": "用户截图确认该模块不应位于未分组安全系统，应归入数据安全管理与运营。",
            }
        )
    return corrections


def build_issues(consistency: dict[str, Any], dictionary_report: dict[str, Any], corrections: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for item in dictionary_report["serviceIssues"]:
        status = "blocking_issue" if item["validationStatus"] == "unknown_service" else "needs_user_confirmation"
        issues.append(
            {
                "id": f"LCDT-SERVICE-{len(issues)+1:03d}",
                "status": status,
                "type": item["validationStatus"],
                "sourceTable": item["sourceTable"],
                "stageName": item["stageName"],
                "excelRow": item["excelRow"],
                "value": item["raw"],
                "message": item.get("issue") or "安全技术服务需人工确认。",
            }
        )
    for title in dictionary_report["newMeasureCandidates"]:
        refs = [
            item
            for item in dictionary_report["moduleOrMeasureIssues"]
            if item["title"] == title
        ]
        issues.append(
            {
                "id": f"LCDT-MEASURE-{len(issues)+1:03d}",
                "status": "needs_user_confirmation",
                "type": "security_technical_measure_candidate",
                "value": title,
                "referenceCount": len(refs),
                "references": [
                    {
                        "sourceTable": item["sourceTable"],
                        "stageName": item["stageName"],
                        "excelRow": item["excelRow"],
                    }
                    for item in refs
                ],
                "message": "该值未命中安全技术模块清单，也未命中现有安全技术措施清单；建议作为安全技术措施候选项等待确认。",
            }
        )
    alias_groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in dictionary_report.get("moduleAliasIssues", []):
        alias_groups[(item["title"], item.get("resolvedTitle") or item["canonical"].get("title", ""))].append(item)
    for (source_title, resolved_title), refs in sorted(alias_groups.items()):
        issues.append(
            {
                "id": f"LCDT-MODULE-ALIAS-{len(issues)+1:03d}",
                "status": "needs_user_confirmation",
                "type": "module_title_alias",
                "value": source_title,
                "resolvedTitle": resolved_title,
                "referenceCount": len(refs),
                "references": [
                    {
                        "sourceTable": item["sourceTable"],
                        "stageName": item["stageName"],
                        "excelRow": item["excelRow"],
                    }
                    for item in refs
                ],
                "message": f"该值不是安全技术模块清单的精确标题，但可按别名/短名归一到 `{resolved_title}`；请确认是否按该模块处理。",
            }
        )
    for stage in consistency["stageResults"]:
        if stage["status"] == "pass":
            continue
        issues.append(
            {
                "id": f"LCDT-CONSISTENCY-{len(issues)+1:03d}",
                "status": "needs_user_confirmation",
                "type": "stage_table_mismatch",
                "stageName": stage["stageName"],
                "serviceOnlyInLifecycleTable": stage["serviceOnlyInLifecycleTable"],
                "serviceOnlyInMappingTable": stage["serviceOnlyInMappingTable"],
                "moduleOrMeasureOnlyInLifecycleTable": stage["moduleOrMeasureOnlyInLifecycleTable"],
                "moduleOrMeasureOnlyInMappingTable": stage["moduleOrMeasureOnlyInMappingTable"],
                "message": "LC-DT 数据生命周期表与策略映射表在该阶段的服务或模块/措施集合不一致。",
            }
        )
    for correction in corrections:
        if correction.get("status") == "info_only":
            continue
        issues.append(
            {
                "id": f"LCDT-SYSTEM-{len(issues)+1:03d}",
                "status": correction["status"],
                "type": "security_system_category_correction",
                "moduleTitle": correction["moduleTitle"],
                "currentSecuritySystems": correction.get("currentSecuritySystems", []),
                "targetSecuritySystem": correction["targetSecuritySystem"],
                "message": correction.get("reason") or correction.get("issue") or "安全系统归属候选更正。",
            }
        )
    return issues


def render_consistency_md(payload: dict[str, Any]) -> str:
    lines = ["# LC-DT 双表一致性审计", "", "## 摘要", ""]
    for key, value in payload["summary"].items():
        lines.append(f"- `{key}`: {value}")
    lines.extend(["", "## 阶段差异", ""])
    for row in payload["stageResults"]:
        lines.append(f"### {row['stageName']} - {row['status']}")
        for key, label in [
            ("serviceOnlyInLifecycleTable", "仅生命周期表服务"),
            ("serviceOnlyInMappingTable", "仅策略映射表服务"),
            ("moduleOrMeasureOnlyInLifecycleTable", "仅生命周期表模块/措施"),
            ("moduleOrMeasureOnlyInMappingTable", "仅策略映射表模块/措施"),
        ]:
            values = row[key]
            lines.append(f"- {label}: {len(values)}")
            for value in values[:20]:
                lines.append(f"  - {value}")
    return "\n".join(lines)


def render_dictionary_md(payload: dict[str, Any]) -> str:
    lines = ["# LC-DT 字典校验报告", "", "## 摘要", ""]
    for key, value in payload["summary"].items():
        lines.append(f"- `{key}`: {value}")
    lines.extend(["", "## 待确认新增安全技术措施候选", ""])
    if payload["newMeasureCandidates"]:
        for title in payload["newMeasureCandidates"]:
            lines.append(f"- {title}")
    else:
        lines.append("- 无")
    lines.extend(["", "## 安全技术服务问题", ""])
    if payload["serviceIssues"]:
        for item in payload["serviceIssues"]:
            lines.append(f"- `{item['sourceTable']}` {item['stageName']} 行 {item['excelRow']}: {item['raw']} - {item['validationStatus']}")
    else:
        lines.append("- 无")
    return "\n".join(lines)


def render_issues_md(issues: list[dict[str, Any]]) -> str:
    lines = ["# LC-DT 待用户确认问题清单", ""]
    if not issues:
        return "# LC-DT 待用户确认问题清单\n\n无待确认问题。"
    by_status = defaultdict(list)
    for issue in issues:
        by_status[issue["status"]].append(issue)
    for status in ["blocking_issue", "needs_user_confirmation", "confirmed_change", "info_only"]:
        rows = by_status.get(status, [])
        if not rows:
            continue
        lines.extend([f"## {status}", ""])
        for issue in rows:
            title = issue.get("value") or issue.get("moduleTitle") or issue.get("stageName") or issue["type"]
            lines.append(f"- `{issue['id']}` {issue['type']}：{title}")
            lines.append(f"  - 说明：{issue['message']}")
    return "\n".join(lines)


def audit() -> dict[str, Any]:
    dictionaries = build_dictionaries()
    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    lifecycle_rows = parse_lifecycle_table(workbook, dictionaries)
    mapping_rows = parse_mapping_table(workbook, dictionaries)
    consistency = build_consistency(lifecycle_rows, mapping_rows)
    dictionary_report = build_dictionary_report(lifecycle_rows, mapping_rows)
    corrections = build_system_corrections(dictionaries)
    issues = build_issues(consistency, dictionary_report, corrections)

    generated_at = now_iso()
    lifecycle_payload = {
        "version": 1,
        "generatedAt": generated_at,
        "sourceWorkbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sourceSheet": LIFECYCLE_SHEET,
        "rowCount": len(lifecycle_rows),
        "rows": lifecycle_rows,
    }
    mapping_payload = {
        "version": 1,
        "generatedAt": generated_at,
        "sourceWorkbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sourceSheet": MAPPING_SHEET,
        "rowCount": len(mapping_rows),
        "rows": mapping_rows,
    }
    consistency_payload = {"version": 1, "generatedAt": generated_at, **consistency}
    dictionary_payload = {"version": 1, "generatedAt": generated_at, **dictionary_report}
    issues_payload = {
        "version": 1,
        "generatedAt": generated_at,
        "requiresUserConfirmation": bool(issues),
        "issueCount": len(issues),
        "issueStatusCounts": dict(Counter(issue["status"] for issue in issues)),
        "issues": issues,
    }

    write_json(OUT_DIR / "lcdt-lifecycle-table-normalized.json", lifecycle_payload)
    write_json(OUT_DIR / "lcdt-service-module-policy-mapping-normalized.json", mapping_payload)
    write_json(OUT_DIR / "lcdt-table-consistency-audit.json", consistency_payload)
    write_json(OUT_DIR / "lcdt-dictionary-validation-report.json", dictionary_payload)
    write_json(OUT_DIR / "lcdt-issues-for-user-confirmation.json", issues_payload)
    write_md(OUT_DIR / "lcdt-table-consistency-audit.md", render_consistency_md(consistency_payload))
    write_md(OUT_DIR / "lcdt-dictionary-validation-report.md", render_dictionary_md(dictionary_payload))
    write_md(OUT_DIR / "lcdt-issues-for-user-confirmation.md", render_issues_md(issues))

    summary = {
        "status": "ready_with_user_confirmation" if issues else "ready",
        "outputDir": str(OUT_DIR.relative_to(ROOT)),
        "lifecycleStageCount": len(lifecycle_rows),
        "mappingRowCount": len(mapping_rows),
        "consistencySummary": consistency["summary"],
        "dictionarySummary": dictionary_report["summary"],
        "issueStatusCounts": issues_payload["issueStatusCounts"],
    }
    return summary


def main() -> None:
    summary = audit()
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
