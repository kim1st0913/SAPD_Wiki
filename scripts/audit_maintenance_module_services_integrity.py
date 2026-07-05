#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
SQLITE_PATH = ROOT / "data/database/sapd_wiki.sqlite3"
MAINTENANCE_PATH = ROOT / "frontend/capability-browser/public/data/maintenance-knowledge.json"
MAINTENANCE_MODULES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/modules.json"
MAINTENANCE_SECTION_MODULES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/sections/modules.json"
MAINTENANCE_SERVICES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/services.json"
MAINTENANCE_SECTION_SERVICES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/sections/services.json"
SHARED_LOOKUPS_PATH = ROOT / "frontend/capability-browser/public/data/shared-lookups.json"
SHARED_SERVICE_MODULE_INDEX_PATH = ROOT / "frontend/capability-browser/public/data/shared-lookups/service-module-index.json"

SERVICE_SHEET = "安全能力-安全技术服务"
MODULE_SHEET = "安全技术模块清单"
TARGET_MODULES = ("数据水印溯源", "数据脱敏(去标识化)")
TARGET_SERVICE_CODES = (
    "I-DI&T-PD.DP-01",
    "I-AP&T-PD.DP-01",
    "I-OS&T-PD.DP-01",
    "I-AP&T-PD.DP-02",
    "I-NT&T-PD.DP-02",
    "I-DI&T-PD.DP-02",
)
EMPTY_MARKERS = {"/", "\\", "N/A", "NA", "无", "-"}


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    return re.sub(r"\s+", " ", text).strip()


def split_multiline(value: Any) -> list[str]:
    text = "" if value is None else str(value)
    rows: list[str] = []
    for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        item = norm(line)
        if item and item not in EMPTY_MARKERS:
            rows.append(item)
    return rows


def service_parts(raw: Any) -> dict[str, str]:
    text = norm(raw)
    if not text:
        return {"code": "", "title": "", "display": ""}
    if " " in text:
        code, title = text.split(" ", 1)
    else:
        match = re.match(r"^([A-Z]+-[A-Z0-9&.-]+)(.+)$", text)
        code, title = (match.group(1), match.group(2)) if match else ("", text)
    code = norm(code)
    title = norm(title)
    return {"code": code, "title": title, "display": f"{code} {title}".strip()}


def service_key(service: dict[str, Any]) -> str:
    code = norm(service.get("code"))
    title = norm(service.get("title"))
    return f"{code} {title}".strip()


def merged_values(ws) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[ws.cell(row, col).coordinate] = anchor.value
    return values


def cell_value(ws, row: int, col: int, merged: dict[str, Any]) -> Any:
    return merged.get(ws.cell(row, col).coordinate, ws.cell(row, col).value)


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def add_relation(target: dict[str, set[str]], module_title: str, service: dict[str, str]) -> None:
    module_title = norm(module_title)
    key = service_key(service)
    if module_title and key:
        target[module_title].add(key)


def load_source_service_dictionary(wb) -> dict[str, Any]:
    ws = wb[SERVICE_SHEET]
    records: list[dict[str, Any]] = []
    pair_counter: Counter[tuple[str, str]] = Counter()
    code_titles: dict[str, set[str]] = defaultdict(set)
    title_codes: dict[str, set[str]] = defaultdict(set)

    for row in range(4, ws.max_row + 1):
        for col in range(7, ws.max_column + 1):
            for raw in split_multiline(ws.cell(row, col).value):
                service = service_parts(raw)
                if not service["code"]:
                    continue
                records.append({"code": service["code"], "title": service["title"], "cell": ws.cell(row, col).coordinate})
                pair_counter[(service["code"], service["title"])] += 1
                code_titles[service["code"]].add(service["title"])
                title_codes[service["title"]].add(service["code"])

    return summarize_service_dictionary("source_excel", records, pair_counter, code_titles, title_codes)


def summarize_service_dictionary(
    name: str,
    records: list[dict[str, Any]],
    pair_counter: Counter[tuple[str, str]],
    code_titles: dict[str, set[str]],
    title_codes: dict[str, set[str]],
) -> dict[str, Any]:
    duplicate_code_title = [
        {"code": code, "title": title, "count": count}
        for (code, title), count in sorted(pair_counter.items())
        if count > 1
    ]
    same_code_different_title = [
        {"code": code, "titles": sorted(titles)}
        for code, titles in sorted(code_titles.items())
        if len(titles) > 1
    ]
    same_title_different_code = [
        {"title": title, "codes": sorted(codes)}
        for title, codes in sorted(title_codes.items())
        if title and len(codes) > 1
    ]
    return {
        "name": name,
        "recordCount": len(records),
        "uniqueCodeTitleCount": len(pair_counter),
        "uniqueCodeCount": len(code_titles),
        "duplicateCodeTitleCount": len(duplicate_code_title),
        "sameCodeDifferentTitleCount": len(same_code_different_title),
        "sameTitleDifferentCodeCount": len(same_title_different_code),
        "duplicateCodeTitle": duplicate_code_title[:20],
        "sameCodeDifferentTitle": same_code_different_title[:20],
        "sameTitleDifferentCode": same_title_different_code[:20],
    }


def load_json_service_dictionary(name: str, services: list[dict[str, Any]]) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    pair_counter: Counter[tuple[str, str]] = Counter()
    code_titles: dict[str, set[str]] = defaultdict(set)
    title_codes: dict[str, set[str]] = defaultdict(set)
    for item in services:
        service = item.get("service") if isinstance(item.get("service"), dict) else item
        code = norm(service.get("code"))
        title = norm(service.get("title"))
        if not code:
            continue
        records.append({"code": code, "title": title})
        pair_counter[(code, title)] += 1
        code_titles[code].add(title)
        title_codes[title].add(code)
    return summarize_service_dictionary(name, records, pair_counter, code_titles, title_codes)


def load_sqlite_service_dictionary() -> dict[str, Any]:
    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT code, title
            FROM knowledge_items
            WHERE type = 'security_technical_service'
              AND status = 'active'
              AND COALESCE(code, '') <> ''
            """
        ).fetchall()
    finally:
        conn.close()
    records = [{"code": norm(row["code"]), "title": norm(row["title"])} for row in rows]
    pair_counter: Counter[tuple[str, str]] = Counter((row["code"], row["title"]) for row in records)
    code_titles: dict[str, set[str]] = defaultdict(set)
    title_codes: dict[str, set[str]] = defaultdict(set)
    for row in records:
        code_titles[row["code"]].add(row["title"])
        title_codes[row["title"]].add(row["code"])
    return summarize_service_dictionary("sqlite", records, pair_counter, code_titles, title_codes)


def load_source_module_relations(wb) -> dict[str, Any]:
    ws = wb[MODULE_SHEET]
    merged = merged_values(ws)
    relations: dict[str, set[str]] = defaultdict(set)
    relation_counter: dict[str, Counter[str]] = defaultdict(Counter)

    for row in range(3, ws.max_row + 1):
        module_title = norm(cell_value(ws, row, 4, merged))
        if not module_title:
            continue
        for raw in split_multiline(cell_value(ws, row, 6, merged)):
            service = service_parts(raw)
            key = service_key(service)
            if not key:
                continue
            relations[module_title].add(key)
            relation_counter[module_title][key] += 1

    duplicates = [
        {"module": module_title, "service": key, "count": count}
        for module_title, counter in sorted(relation_counter.items())
        for key, count in sorted(counter.items())
        if count > 1
    ]
    return {"relations": dict(relations), "duplicates": duplicates}


def load_sqlite_module_relations() -> dict[str, set[str]]:
    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT m.title AS module_title,
                   s.code AS service_code,
                   s.title AS service_title
            FROM knowledge_relations r
            JOIN knowledge_items m ON m.id = r.source_item_id
            JOIN knowledge_items s ON s.id = r.target_item_id
            WHERE r.relation_type = 'implements_service'
              AND m.type = 'security_technology_module'
              AND s.type = 'security_technical_service'
              AND m.status = 'active'
              AND s.status = 'active'
            """
        ).fetchall()
    finally:
        conn.close()
    relations: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        add_relation(relations, row["module_title"], {"code": row["service_code"], "title": row["service_title"]})
    return dict(relations)


def load_maintenance_module_relations(path: Path) -> dict[str, set[str]]:
    data = read_json(path)
    relations: dict[str, set[str]] = defaultdict(set)
    for module in data.get("security_technology_modules", []):
        module_title = norm(module.get("title"))
        for service in module.get("services", []):
            add_relation(relations, module_title, {"code": service.get("code"), "title": service.get("title")})
    return dict(relations)


def load_shared_module_relations(path: Path) -> dict[str, set[str]]:
    data = read_json(path)
    entries = data.get("service_module_index") or data.get("serviceModuleIndex") or []
    relations: dict[str, set[str]] = defaultdict(set)
    for entry in entries:
        service = entry.get("service") or {}
        service_record = {"code": service.get("code"), "title": service.get("title")}
        for module in entry.get("modules", []):
            add_relation(relations, module.get("title"), service_record)
    return dict(relations)


def compare_relation_maps(expected: dict[str, set[str]], actual: dict[str, set[str]], name: str) -> dict[str, Any]:
    expected_modules = set(expected)
    actual_modules = set(actual)
    missing_modules = sorted(expected_modules - actual_modules)
    extra_modules = sorted(actual_modules - expected_modules)
    mismatches = []
    for module_title in sorted(expected_modules | actual_modules):
        missing = sorted(expected.get(module_title, set()) - actual.get(module_title, set()))
        extra = sorted(actual.get(module_title, set()) - expected.get(module_title, set()))
        if missing or extra:
            mismatches.append({"module": module_title, "missing": missing, "extra": extra})
    return {
        "name": name,
        "moduleCount": len(actual_modules),
        "relationCount": sum(len(values) for values in actual.values()),
        "missingModuleCount": len(missing_modules),
        "extraModuleCount": len(extra_modules),
        "mismatchCount": len(mismatches),
        "missingModules": missing_modules[:20],
        "extraModules": extra_modules[:20],
        "mismatches": mismatches[:20],
    }


def target_module_summary(relation_maps: dict[str, dict[str, set[str]]]) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for module_title in TARGET_MODULES:
        summary[module_title] = {
            name: sorted(relations.get(module_title, set()))
            for name, relations in relation_maps.items()
        }
    return summary


def target_service_module_summary(shared_relations: dict[str, set[str]]) -> dict[str, list[str]]:
    by_service: dict[str, list[str]] = defaultdict(list)
    for module_title, services in shared_relations.items():
        for service in services:
            code = service.split(" ", 1)[0]
            if code in TARGET_SERVICE_CODES:
                by_service[service].append(module_title)
    return {service: sorted(modules) for service, modules in sorted(by_service.items())}


def relation_duplicate_summary(relations: dict[str, set[str]]) -> dict[str, Any]:
    service_counts = [len(values) for values in relations.values()]
    return {
        "moduleCount": len(relations),
        "relationCount": sum(service_counts),
        "maxServicesPerModule": max(service_counts) if service_counts else 0,
    }


def main() -> int:
    errors: list[str] = []
    wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)

    source_services = load_source_service_dictionary(wb)
    maintenance = read_json(MAINTENANCE_PATH)
    services_json = read_json(MAINTENANCE_SERVICES_PATH)
    section_services_json = read_json(MAINTENANCE_SECTION_SERVICES_PATH)
    service_dictionary_summaries = [
        source_services,
        load_sqlite_service_dictionary(),
        load_json_service_dictionary("maintenance_knowledge", maintenance.get("security_technical_services", [])),
        load_json_service_dictionary("maintenance_services", services_json.get("security_technical_services", [])),
        load_json_service_dictionary("maintenance_sections_services", section_services_json.get("security_technical_services", [])),
    ]

    source_module_payload = load_source_module_relations(wb)
    source_relations = source_module_payload["relations"]
    relation_maps = {
        "source_excel": source_relations,
        "sqlite": load_sqlite_module_relations(),
        "shared_lookups": load_shared_module_relations(SHARED_LOOKUPS_PATH),
        "shared_split_service_module_index": load_shared_module_relations(SHARED_SERVICE_MODULE_INDEX_PATH),
        "maintenance_knowledge_module_services": load_maintenance_module_relations(MAINTENANCE_PATH),
        "maintenance_modules_module_services": load_maintenance_module_relations(MAINTENANCE_MODULES_PATH),
        "maintenance_sections_modules_module_services": load_maintenance_module_relations(MAINTENANCE_SECTION_MODULES_PATH),
    }

    relation_comparisons = [
        compare_relation_maps(source_relations, relations, name)
        for name, relations in relation_maps.items()
        if name != "source_excel"
    ]

    if source_module_payload["duplicates"]:
        errors.append("source module catalog has duplicate services within at least one module")
    for summary in service_dictionary_summaries:
        if summary["duplicateCodeTitleCount"]:
            errors.append(f"{summary['name']} has duplicate code+title services")
        if summary["sameCodeDifferentTitleCount"]:
            errors.append(f"{summary['name']} has same code with different titles")
        if summary["sameTitleDifferentCodeCount"]:
            errors.append(f"{summary['name']} has same title with different codes")
    for comparison in relation_comparisons:
        if comparison["missingModuleCount"] or comparison["extraModuleCount"] or comparison["mismatchCount"]:
            errors.append(f"{comparison['name']} module-service relations do not match source Excel")

    result = {
        "status": "pass" if not errors else "fail",
        "source": {
            "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
            "moduleSheet": MODULE_SHEET,
            "serviceSheet": SERVICE_SHEET,
        },
        "serviceDictionaryUniqueness": service_dictionary_summaries,
        "moduleServiceUniqueness": {
            **relation_duplicate_summary(source_relations),
            "duplicateWithinModuleCount": len(source_module_payload["duplicates"]),
            "duplicates": source_module_payload["duplicates"][:20],
        },
        "moduleServiceComparisons": relation_comparisons,
        "targetModules": target_module_summary(relation_maps),
        "targetServiceModuleRefsFromSharedLookups": target_service_module_summary(relation_maps["shared_lookups"]),
        "errors": errors,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not errors else 1


if __name__ == "__main__":
    sys.exit(main())
