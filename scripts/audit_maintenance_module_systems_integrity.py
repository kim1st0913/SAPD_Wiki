#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
SQLITE_PATH = ROOT / "data/database/sapd_wiki.sqlite3"
MAINTENANCE_PATH = ROOT / "frontend/capability-browser/public/data/maintenance-knowledge.json"
MAINTENANCE_MODULES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/modules.json"
MAINTENANCE_SECTION_MODULES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/sections/modules.json"

MODULE_SHEET = "安全技术模块清单"
TARGET_MODULES = ("运维访问管理", "特权账号管理")
EMPTY_MARKERS = {"/", "\\", "N/A", "NA", "无", "-"}


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ").replace("\u3000", " ")
    return re.sub(r"\s+", " ", text).strip()


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def merged_values(ws) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[ws.cell(row, col).coordinate] = anchor.value
    return values


def cell_value(ws, row: int, col: int, merged: dict[str, Any]) -> Any:
    return merged.get(ws.cell(row, col).coordinate, ws.cell(row, col).value)


def load_source_module_systems() -> dict[str, set[str]]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True, read_only=False)
    try:
        ws = workbook[MODULE_SHEET]
        merged = merged_values(ws)
        relations: dict[str, set[str]] = defaultdict(set)
        for row in range(3, ws.max_row + 1):
            module_title = norm(cell_value(ws, row, 4, merged))
            system_title = norm(cell_value(ws, row, 3, merged))
            definition = norm(cell_value(ws, row, 5, merged))
            if not module_title or module_title in EMPTY_MARKERS or module_title.isdigit():
                continue
            if not definition and system_title.isdigit():
                continue
            if system_title and system_title not in EMPTY_MARKERS:
                relations[module_title].add(system_title)
        return dict(relations)
    finally:
        workbook.close()


def load_sqlite_module_systems() -> dict[str, set[str]]:
    conn = sqlite3.connect(SQLITE_PATH)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT module.title AS module_title,
                   system.title AS system_title
            FROM knowledge_relations AS relation
            JOIN knowledge_items AS module ON module.id = relation.source_item_id
            JOIN knowledge_items AS system ON system.id = relation.target_item_id
            WHERE relation.relation_type = 'part_of_system'
              AND module.type = 'security_technology_module'
              AND system.type = 'security_system'
              AND module.status = 'active'
              AND system.status = 'active'
            """
        ).fetchall()
    finally:
        conn.close()
    relations: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        module_title = norm(row["module_title"])
        system_title = norm(row["system_title"])
        if module_title and system_title:
            relations[module_title].add(system_title)
    return dict(relations)


def load_maintenance_module_systems(path: Path) -> dict[str, set[str]]:
    data = read_json(path)
    relations: dict[str, set[str]] = defaultdict(set)
    for module in data.get("security_technology_modules", []):
        module_title = norm(module.get("title") or module.get("name"))
        if not module_title:
            continue
        for system in module.get("systems", []):
            system_title = norm(system.get("title") or system.get("name"))
            if system_title:
                relations[module_title].add(system_title)
    return dict(relations)


def compare_relation_maps(expected: dict[str, set[str]], actual: dict[str, set[str]], name: str) -> dict[str, Any]:
    expected_modules = set(expected)
    actual_modules = set(actual)
    mismatches: list[dict[str, Any]] = []
    for module_title in sorted(expected_modules | actual_modules):
        missing = sorted(expected.get(module_title, set()) - actual.get(module_title, set()))
        extra = sorted(actual.get(module_title, set()) - expected.get(module_title, set()))
        if missing or extra:
            mismatches.append({"module": module_title, "missing": missing, "extra": extra})
    return {
        "name": name,
        "moduleCount": len(actual_modules),
        "relationCount": sum(len(values) for values in actual.values()),
        "missingModuleCount": len(sorted(expected_modules - actual_modules)),
        "extraModuleCount": len(sorted(actual_modules - expected_modules)),
        "mismatchCount": len(mismatches),
        "mismatches": mismatches[:30],
    }


def target_module_summary(relation_maps: dict[str, dict[str, set[str]]]) -> dict[str, Any]:
    return {
        module_title: {
            name: sorted(relations.get(module_title, set()))
            for name, relations in relation_maps.items()
        }
        for module_title in TARGET_MODULES
    }


def main() -> int:
    errors: list[str] = []
    relation_maps = {
        "source_excel": load_source_module_systems(),
        "sqlite": load_sqlite_module_systems(),
        "maintenance_knowledge_module_systems": load_maintenance_module_systems(MAINTENANCE_PATH),
        "maintenance_modules_module_systems": load_maintenance_module_systems(MAINTENANCE_MODULES_PATH),
        "maintenance_sections_modules_module_systems": load_maintenance_module_systems(MAINTENANCE_SECTION_MODULES_PATH),
    }
    source_relations = relation_maps["source_excel"]
    comparisons = [
        compare_relation_maps(source_relations, relations, name)
        for name, relations in relation_maps.items()
        if name != "source_excel"
    ]
    for comparison in comparisons:
        if comparison["missingModuleCount"] or comparison["extraModuleCount"] or comparison["mismatchCount"]:
            errors.append(f"{comparison['name']} module-system relations do not match source Excel")

    result = {
        "status": "pass" if not errors else "fail",
        "source": {
            "workbook": str(WORKBOOK_PATH.relative_to(ROOT)),
            "moduleSheet": MODULE_SHEET,
        },
        "moduleSystemSource": {
            "moduleCount": len(source_relations),
            "relationCount": sum(len(values) for values in source_relations.values()),
        },
        "moduleSystemComparisons": comparisons,
        "targetModules": target_module_summary(relation_maps),
        "errors": errors,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if not errors else 1


if __name__ == "__main__":
    sys.exit(main())
