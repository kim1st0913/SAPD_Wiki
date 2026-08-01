#!/usr/bin/env python3
from __future__ import annotations

import base64
import copy
import json
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from sapd_wiki.maturity import (  # noqa: E402
    build_maturity_workspace,
    export_maturity_score_exchange,
    export_maturity_template_exchange,
    import_maturity_score_exchange,
    import_maturity_template_exchange,
)


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def workbook(package: dict):
    return load_workbook(BytesIO(base64.b64decode(package["workbookBase64"])), data_only=False)


def mutate(package: dict, mutation) -> dict:
    value = workbook(package)
    mutation(value)
    output = BytesIO()
    value.save(output)
    changed = copy.deepcopy(package)
    changed["workbookBase64"] = base64.b64encode(output.getvalue()).decode("ascii")
    return changed


def scored_entries(template: dict) -> list[dict]:
    return [
        {
            "scoreItemId": item["id"],
            "isApplicable": True,
            "elements": {"organization": "L2", "process": "L2", "tool": "L2", "data": "L2"},
            "dimensionNotes": {"organization": "当前组织说明", "process": "当前流程说明", "tool": "当前工具说明", "data": "当前数据说明"},
            "reviewElements": {},
            "targetElements": {"organization": "L3", "process": "L4", "tool": "L3", "data": "L4"},
            "targetDimensionNotes": {"organization": "目标组织说明", "process": "目标流程说明", "tool": "目标工具说明", "data": "目标数据说明"},
            "status": "scored",
        }
        for item in template["scoreItems"]
    ]


def main() -> int:
    capability = json.loads((ROOT / "frontend" / "capability-browser" / "public" / "data" / "capability-workbench.json").read_text(encoding="utf-8"))
    component = (ROOT / "frontend" / "capability-browser" / "components" / "MaturityAssessmentWorkbench.js").read_text(encoding="utf-8")
    workspace = build_maturity_workspace(capability)
    template = workspace["template"]
    project = next(iter(workspace["projectDetails"].values()))["project"]

    standard_export = export_maturity_template_exchange({"template": template})
    require(standard_export["ok"] and standard_export["fileName"].endswith(".xlsx"), "standard template must export as XLSX")
    standard_book = workbook(standard_export["package"])
    require(standard_book.sheetnames == ["模板基础信息", "评估模板"], "workbook must contain exactly two business sheets")
    require(len(standard_book["评估模板"].merged_cells.ranges) > 20, "hierarchy relationships must be represented by merged business cells")
    visible = " ".join(str(cell.value or "") for sheet in standard_book.worksheets for row in sheet.iter_rows() for cell in row)
    require(all(token not in visible for token in ("模板 ID", "快照 ID", "结构哈希", "客户名称", project["id"])), "workbook must not expose developer or customer fields")
    require("文件用途" not in visible, "basic information must not expose the internal file-purpose field")
    require("评分标题和评分列必须保留" in visible, "template workbook must explain that score headers stay while score cells remain empty")
    require(any("标准模板,自定义模板" in str(rule.formula1) and "B3" in str(rule.sqref) for rule in standard_book["模板基础信息"].data_validations.dataValidation), "template type must provide a standard/custom dropdown")
    require(not import_maturity_template_exchange({"exchange": standard_export["package"]})["ok"], "standard template must not be accepted as a custom-template import")

    def convert_standard_to_custom(value):
        info = value["模板基础信息"]
        info["B2"] = "业务自定义模板"
        info["B3"] = "自定义模板"
        info["B5"] = "由业务人员维护的自定义能力、关注点和服务关系。"

    converted_custom = import_maturity_template_exchange({"exchange": mutate(standard_export["package"], convert_standard_to_custom)})
    require(converted_custom["ok"], "standard workbook must become importable after selecting custom type and filling a new name and description")
    converted_item = converted_custom["template"]["scoreItems"][0]
    require(
        converted_custom["template"]["rubricVersion"] == "sapd-maturity-custom-generic-rubric-v3-2026-07-30"
        and len(converted_item["rubricEntries"]) == 20
        and {entry["sourceType"] for entry in converted_item["rubricEntries"]} == {"CUSTOM_GENERIC_FALLBACK"},
        "custom workbook import must create the versioned four-dimension by five-level generic rubric",
    )

    custom_template = copy.deepcopy(template)
    custom_template.update({"type": "custom", "readOnly": False, "structureMutable": True, "weightMutable": True, "name": "业务自定义模板", "description": "由业务人员维护的成熟度评估模板。"})
    custom_export = export_maturity_template_exchange({"template": custom_template})
    custom_book = workbook(custom_export["package"])
    require(all(custom_book["评估模板"].cell(3, column).value in {None, ""} for column in range(8, 26)), "custom template export must keep all current/target score fields empty")
    custom_import = import_maturity_template_exchange({"exchange": custom_export["package"]})
    require(custom_import["ok"], "valid custom business workbook must import")
    require(custom_import["template"]["stats"]["scoreItems"] == template["stats"]["scoreItems"], "custom template round-trip must preserve assessment-point grain")
    scored_template = mutate(custom_export["package"], lambda value: setattr(value["评估模板"]["I3"], "value", "L3"))
    scored_template_result = import_maturity_template_exchange({"exchange": scored_template})
    require(any(error.get("code") == "template_contains_scores" for error in scored_template_result["rowErrors"]), "template import must reject score data")

    score_export = export_maturity_score_exchange({"project": project, "template": template, "scoreEntries": scored_entries(template)})
    score_import = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": score_export["package"]})
    require(score_import["ok"] and score_import["batch"]["status"] == "success" and score_import["batch"]["successCount"] == template["stats"]["scoreItems"], "complete score XLSX must import every assessment point")
    first_round_trip = next(item for item in score_import["scoreEntries"] if item["scoreItemId"] == template["scoreItems"][0]["id"])
    require(first_round_trip["targetElements"] == {"organization": "L3", "process": "L4", "tool": "L3", "data": "L4"} and first_round_trip["targetDimensionNotes"]["data"] == "目标数据说明" and first_round_trip["dimensionNotes"]["organization"] == "当前组织说明", "score XLSX must round-trip four current/target dimensions and their explanations")
    def make_first_target_equal_current(value):
        sheet = value["评估模板"]
        for current_column, target_column in zip(("I", "J", "K", "L"), ("Q", "R", "S", "T"), strict=True):
            sheet[f"{target_column}3"] = sheet[f"{current_column}3"].value
    equal_target_import = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": mutate(score_export["package"], make_first_target_equal_current)})
    equal_target_entry = next(item for item in equal_target_import["scoreEntries"] if item["scoreItemId"] == template["scoreItems"][0]["id"])
    require(equal_target_import["ok"] and equal_target_entry["targetElements"] == equal_target_entry["elements"], "score XLSX must accept target dimensions equal to current dimensions")
    def convert_to_legacy_v21(value):
        sheet = value["评估模板"]
        for merged in list(sheet.merged_cells.ranges):
            if merged.max_row <= 2:
                sheet.unmerge_cells(str(merged))
        for column, label in enumerate(("组织与角色", "制度与流程", "平台与工具", "数据与信息", "目标等级", "评估说明"), start=9):
            sheet.cell(1, column, label)
            sheet.cell(2, column, "")
        for row in range(3, sheet.max_row + 1):
            sheet.cell(row, 13, "L4")
            sheet.cell(row, 14, "旧目标说明")
    legacy_import = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": mutate(score_export["package"], convert_to_legacy_v21)})
    legacy_entry = next(item for item in legacy_import["scoreEntries"] if item["scoreItemId"] == template["scoreItems"][0]["id"])
    require(legacy_import["ok"] and legacy_entry["targetElements"] == {key: "L4" for key in ("organization", "process", "tool", "data")} and set(legacy_entry["targetDimensionNotes"].values()) == {"旧目标说明"}, "legacy V2.1 XLSX targets and notes must expand into four target dimensions")
    not_applicable_with_scores = mutate(score_export["package"], lambda value: setattr(value["评估模板"]["H3"], "value", "不适用"))
    not_applicable_result = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": not_applicable_with_scores})
    first_entry = next(item for item in not_applicable_result["scoreEntries"] if item["scoreItemId"] == template["scoreItems"][0]["id"])
    require(not_applicable_result["ok"] and not_applicable_result["batch"]["status"] == "success" and first_entry["status"] == "not_applicable" and first_entry["elements"] == {} and first_entry["targetLevel"] is None, "scores on a not-applicable row must be ignored without creating blockers")
    require(all(token in component for token in ("SCORE_FILE_IMPORTED", "scoreImportNotice", "评分文件上传成功", "评分标题和评分列")), "frontend must persist score-upload success feedback and history while explaining the blank-score contract")
    changed_structure = mutate(score_export["package"], lambda value: setattr(value["评估模板"]["E3"], "value", "非法结构修改"))
    changed_result = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": changed_structure})
    require(not changed_result["ok"] and changed_result["dataState"] == "invalid_structure", "score import must reject changed business structure")
    def lower_target(value):
        for column in ("Q", "R", "S", "T"):
            value["评估模板"][f"{column}3"] = "L1"
    invalid_target = mutate(score_export["package"], lower_target)
    target_result = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": invalid_target})
    require(any(error.get("code") == "target_below_current" for error in target_result["rowErrors"]), "score import must create a blocker when target is below current maturity")

    def lower_one_target_dimension(value):
        sheet = value["评估模板"]
        for column, level in zip(("I", "J", "K", "L"), ("L4", "L1", "L1", "L1"), strict=True):
            sheet[f"{column}3"] = level
        for column, level in zip(("Q", "R", "S", "T"), ("L3", "L3", "L1", "L1"), strict=True):
            sheet[f"{column}3"] = level

    dimension_target_result = import_maturity_score_exchange({"project": project, "template": template, "scoreEntries": [], "exchange": mutate(score_export["package"], lower_one_target_dimension)})
    dimension_error = next((error for error in dimension_target_result["rowErrors"] if error.get("code") == "target_below_current"), {})
    require("组织与角色（当前 L4，目标 L3）" in dimension_error.get("message", ""), "score import must reject a same-dimension target regression even when the weighted target is not lower")

    print(json.dumps({"result": "pass", "checks": 23, "rows": template["stats"]["scoreItems"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
