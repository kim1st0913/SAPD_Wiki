#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sqlite3
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from sapd_wiki.exports import _scene_measure_candidates_from_xlsx  # noqa: E402


WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
DB_PATH = ROOT / "data/database/sapd_wiki.sqlite3"
SCENE_SHEET = "作用域-安全技术服务-安全技术模块映射"
MODULE_CATALOG_SHEET = "安全技术模块清单"
EXPECTED_ACTIVE_MEASURE_COUNT = 32
MODULE_TITLE_ALIASES = {
    "安全工作区": "终端安全工作区",
}
MODULE_ONLY_TITLES = {
    "主机防火墙",
    "主机恶意代码防护",
    "主机入侵防御（HIPS）",
    "终端安全工作区",
}


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.replace("\xa0", " ").replace("\u3000", " ")).strip()


def merged_value_lookup(ws) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, col)] = anchor_value
    return values


def cell_text(ws, merged_values: dict[tuple[int, int], Any], row: int, col: int) -> str:
    return norm(merged_values.get((row, col), ws.cell(row, col).value))


def canonical_module_title(title: str, module_titles: set[str]) -> str:
    text = norm(title)
    if text in module_titles:
        return text
    alias = MODULE_TITLE_ALIASES.get(text)
    if alias in module_titles:
        return alias
    return text


def authoritative_module_titles(wb) -> set[str]:
    if MODULE_CATALOG_SHEET not in wb.sheetnames:
        return set()
    ws = wb[MODULE_CATALOG_SHEET]
    merged_values = merged_value_lookup(ws)
    titles: set[str] = set()
    for row in range(3, ws.max_row + 1):
        title = cell_text(ws, merged_values, row, 4)
        if title and not title.replace(".", "", 1).isdigit():
            titles.add(title)
    return titles


def fill_signature(cell) -> dict[str, Any]:
    color = getattr(getattr(cell, "fill", None), "fgColor", None)
    return {
        "type": None if getattr(color, "type", None) is None else str(getattr(color, "type")),
        "theme": str(getattr(color, "theme", "")),
        "tint": getattr(color, "tint", None),
        "styleId": getattr(cell, "style_id", None),
    }


def fill_tint_matches(fill: dict[str, Any], *, theme: str, tint: float) -> bool:
    if fill.get("type") != "theme" or str(fill.get("theme")) != theme:
        return False
    try:
        return abs(float(fill.get("tint") or 0) - tint) < 0.0001
    except (TypeError, ValueError):
        return False


def is_module_style(fill: dict[str, Any]) -> bool:
    return fill_tint_matches(fill, theme="8", tint=0.7999816888943144)


def is_measure_style(fill: dict[str, Any]) -> bool:
    return fill_tint_matches(fill, theme="6", tint=0.5999938962981048) or (
        fill.get("type") == "theme" and str(fill.get("theme")) == "0"
    )


def active_sqlite_measure_names() -> set[str]:
    with sqlite3.connect(DB_PATH) as connection:
        rows = connection.execute(
            """
            SELECT title
            FROM knowledge_items
            WHERE type = 'security_technical_measure'
              AND status = 'active'
            """
        ).fetchall()
    return {norm(row[0]) for row in rows if norm(row[0])}


def main() -> None:
    errors: list[str] = []
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    if SCENE_SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {SCENE_SHEET}")
    scene_ws = wb[SCENE_SHEET]
    scene_merged_values = merged_value_lookup(scene_ws)
    module_titles = authoritative_module_titles(wb)

    if not MODULE_ONLY_TITLES <= module_titles:
        missing = sorted(MODULE_ONLY_TITLES - module_titles)
        errors.append(f"module-only titles missing from module catalog: {', '.join(missing)}")

    module_only_rows: list[dict[str, Any]] = []
    measure_styled_module_rows: list[dict[str, Any]] = []
    not_module_styled_rows: list[dict[str, Any]] = []
    for row in range(3, scene_ws.max_row + 1):
        raw = cell_text(scene_ws, scene_merged_values, row, 7)
        canonical = canonical_module_title(raw, module_titles)
        if canonical not in MODULE_ONLY_TITLES:
            continue
        cell = scene_ws.cell(row, 7)
        fill = fill_signature(cell)
        item = {
            "row": row,
            "cell": cell.coordinate,
            "raw": raw,
            "canonical": canonical,
            "fill": fill,
        }
        module_only_rows.append(item)
        if is_measure_style(fill):
            measure_styled_module_rows.append(item)
        if raw and not is_module_style(fill):
            not_module_styled_rows.append(item)

    wb.close()

    scene_candidates = _scene_measure_candidates_from_xlsx(WORKBOOK_PATH)
    candidate_leaks = sorted(
        {norm(candidate.get("name")) for candidate in scene_candidates if norm(candidate.get("name")) in MODULE_ONLY_TITLES}
    )
    active_measures = active_sqlite_measure_names()
    sqlite_leaks = sorted(active_measures & MODULE_ONLY_TITLES)

    if measure_styled_module_rows:
        errors.append("scene mapping G-column module-only titles still use technical-measure style")
    if candidate_leaks:
        errors.append(f"scene mapping measure candidates leaked module-only titles: {', '.join(candidate_leaks)}")
    if len(active_measures) != EXPECTED_ACTIVE_MEASURE_COUNT:
        errors.append(f"sqlite active security_technical_measure count should be {EXPECTED_ACTIVE_MEASURE_COUNT}, got {len(active_measures)}")
    if sqlite_leaks:
        errors.append(f"sqlite active measures contain module-only titles: {', '.join(sqlite_leaks)}")

    report = {
        "status": "fail" if errors else "pass",
        "sourceWorkbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sheet": SCENE_SHEET,
        "moduleOnlyRowCount": len(module_only_rows),
        "measureStyledModuleRowCount": len(measure_styled_module_rows),
        "nonMeasureOtherStyleModuleRowCount": len(not_module_styled_rows),
        "sceneMeasureCandidateCount": len(scene_candidates),
        "candidateLeakCount": len(candidate_leaks),
        "sqliteActiveMeasureCount": len(active_measures),
        "sqliteLeakCount": len(sqlite_leaks),
        "candidateLeaks": candidate_leaks,
        "sqliteLeaks": sqlite_leaks,
        "manualFixRows": [item for item in module_only_rows if item["row"] in {127, 128, 129, 809}],
        "measureStyledModuleRows": measure_styled_module_rows,
        "errors": errors,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
