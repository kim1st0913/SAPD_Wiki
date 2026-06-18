#!/usr/bin/env python3
"""Audit the scope-service-module mapping sheet from the original workbook.

This script reads the source .xlsx with openpyxl, preserves explicit merged-cell
semantics, records style evidence, rebuilds normalized relations, and compares
them with the current frontend JSON packages. It intentionally does not replace
public data packages.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from sapd_wiki.candidates import normalize_text
from sapd_wiki.exports import _measure_category_status, _split_measure_names, _technical_measure_style
from sapd_wiki.parsers import _is_scene_module_fill, _security_module_titles_from_catalog_alias
from sapd_wiki.transformers import is_blank_or_placeholder, service_parts, split_multivalue_text, split_scope_values

SHEET_NAME = "作用域-安全技术服务-安全技术模块映射"
OUTPUT_DIR = PROJECT_ROOT / "data" / "exports" / "worker-verify"
FIELD_COLUMNS = {
    "B": ("informationEnvironment", "信息化环境"),
    "C": ("environmentSegment", "环境子类"),
    "D": ("informationObject", "信息化对象"),
    "E": ("scope", "作用域"),
    "F": ("securityTechnicalService", "安全技术服务"),
    "G": ("moduleOrMeasureRaw", "安全技术模块/措施"),
    "H": ("securitySystem", "安全系统"),
}
FIELD_BY_INDEX = {index: FIELD_COLUMNS[column] for column, index in zip(FIELD_COLUMNS, range(2, 9))}


def as_text(value: Any) -> str:
    return normalize_text(value)


def json_scalar(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def color_payload(color: Any) -> dict[str, Any]:
    if not color:
        return {}
    payload: dict[str, Any] = {"type": getattr(color, "type", None)}
    for attr in ("rgb", "indexed", "theme", "tint"):
        try:
            value = getattr(color, attr)
        except (AttributeError, TypeError, ValueError):
            continue
        if value is not None:
            payload[attr] = json_scalar(value)
    return {key: value for key, value in payload.items() if value not in (None, "")}


def fill_payload(cell: Any) -> dict[str, Any]:
    fill = getattr(cell, "fill", None)
    if not fill:
        return {}
    fg_color = getattr(fill, "fgColor", None)
    return {
        "patternType": json_scalar(getattr(fill, "patternType", None)),
        "fgColor": color_payload(fg_color),
        "bgColor": color_payload(getattr(fill, "bgColor", None)),
    }


def fill_dict_for_measure_style(cell: Any) -> dict[str, str]:
    fg_color = getattr(getattr(cell, "fill", None), "fgColor", None)
    if not fg_color:
        return {}
    payload: dict[str, str] = {}
    try:
        if fg_color.type:
            payload["type"] = str(fg_color.type)
        if fg_color.theme is not None:
            payload["theme"] = str(fg_color.theme)
        if fg_color.tint is not None:
            payload["tint"] = str(fg_color.tint)
        if fg_color.rgb:
            payload["rgb"] = str(fg_color.rgb)
    except (TypeError, ValueError):
        return payload
    return payload


def style_payload(cell: Any) -> dict[str, Any]:
    font = getattr(cell, "font", None)
    border = getattr(cell, "border", None)
    alignment = getattr(cell, "alignment", None)
    return {
        "fill": fill_payload(cell),
        "font": {
            "bold": json_scalar(getattr(font, "bold", None)),
            "italic": json_scalar(getattr(font, "italic", None)),
            "color": color_payload(getattr(font, "color", None)),
        },
        "border": {
            "left": json_scalar(getattr(getattr(border, "left", None), "style", None)),
            "right": json_scalar(getattr(getattr(border, "right", None), "style", None)),
            "top": json_scalar(getattr(getattr(border, "top", None), "style", None)),
            "bottom": json_scalar(getattr(getattr(border, "bottom", None), "style", None)),
        },
        "alignment": {
            "horizontal": json_scalar(getattr(alignment, "horizontal", None)),
            "vertical": json_scalar(getattr(alignment, "vertical", None)),
            "wrapText": json_scalar(getattr(alignment, "wrap_text", None)),
        },
    }


def field_for_column(column_index: int) -> tuple[str, str] | None:
    return FIELD_BY_INDEX.get(column_index)


def merged_ranges_for_sheet(ws) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    audits: list[dict[str, Any]] = []
    by_coordinate: dict[str, dict[str, Any]] = {}
    for merged_range in sorted(ws.merged_cells.ranges, key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col)):
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        fields = [
            field_for_column(column_index)
            for column_index in range(merged_range.min_col, merged_range.max_col + 1)
            if field_for_column(column_index)
        ]
        field_names = [field[0] for field in fields if field]
        audit = {
            "range": str(merged_range),
            "topLeftCell": top_left.coordinate,
            "topLeftValue": as_text(top_left.value),
            "rowSpan": merged_range.max_row - merged_range.min_row + 1,
            "columnSpan": merged_range.max_col - merged_range.min_col + 1,
            "field": field_names[0] if len(set(field_names)) == 1 else None,
            "fields": sorted(set(field_names)),
            "sourceColumn": get_column_letter(merged_range.min_col),
            "styleEvidence": style_payload(top_left),
        }
        audits.append(audit)
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                by_coordinate[ws.cell(row_index, column_index).coordinate] = audit
    return audits, by_coordinate


def effective_cell(ws, merged_by_coordinate: dict[str, dict[str, Any]], row_index: int, column_letter: str) -> dict[str, Any]:
    cell = ws[f"{column_letter}{row_index}"]
    merged = merged_by_coordinate.get(cell.coordinate)
    source_cell = ws[merged["topLeftCell"]] if merged else cell
    return {
        "value": as_text(source_cell.value),
        "rawValue": source_cell.value,
        "cell": cell.coordinate,
        "sourceCell": source_cell.coordinate,
        "mergedRange": merged["range"] if merged else None,
        "styleEvidence": style_payload(source_cell),
        "sourceCellObject": source_cell,
    }


def source_cells_payload(row_values: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        field: {
            "cell": payload["cell"],
            "sourceCell": payload["sourceCell"],
            "value": payload["value"],
        }
        for field, payload in row_values.items()
    }


def merged_payload(row_values: dict[str, dict[str, Any]]) -> dict[str, str | None]:
    return {field: payload["mergedRange"] for field, payload in row_values.items()}


def style_evidence_payload(row_values: dict[str, dict[str, Any]]) -> dict[str, Any]:
    return {field: payload["styleEvidence"] for field, payload in row_values.items()}


def split_systems(value: Any) -> list[str]:
    text = as_text(value)
    if text == "监测、研判和响应":
        return [text]
    return split_multivalue_text(value, split_on_ideographic_comma=True)


def service_payload(value: Any) -> dict[str, Any]:
    parts = service_parts(value)
    return {
        "raw": as_text(value),
        "code": parts.get("code") or "",
        "title": parts.get("title") or as_text(value),
        "scopeCode": parts.get("scope_code") or "",
        "capabilityFocusCode": parts.get("capability_focus_code") or "",
    }


def relation_key(*parts: Any) -> str:
    return "||".join(as_text(part) for part in parts)


def object_context_key(environment: Any, segment: Any, information_object: Any) -> str:
    return relation_key(environment, segment, information_object)


def add_unique_relation(bucket: dict[str, dict[str, Any]], key: str, payload: dict[str, Any], row_index: int) -> None:
    if key not in bucket:
        bucket[key] = {**payload, "rows": []}
    if row_index not in bucket[key]["rows"]:
        bucket[key]["rows"].append(row_index)


def is_empty_security_system_value(value: Any) -> bool:
    text = as_text(value).upper()
    return is_blank_or_placeholder(value) or text in {"N/A", "NA"}


def module_or_measure_kind(
    cell: Any,
    value: str,
    *,
    system_value: str = "",
    authoritative_module_titles: set[str] | None = None,
) -> tuple[str, str | None]:
    if not value or is_blank_or_placeholder(value):
        return "placeholder", None
    if re.match(r"(?i)^N/A\s*[（(]", value):
        return "measure", None
    module_titles = _security_module_titles_from_catalog_alias(value, authoritative_module_titles or set())
    if _is_scene_module_fill(cell) or module_titles:
        return "module", None
    measure_style = _technical_measure_style(fill_dict_for_measure_style(cell))
    if measure_style == "measure":
        return "measure", None
    if measure_style == "measure_note":
        return "measure", None
    if is_empty_security_system_value(system_value):
        return "measure", None
    return "unknown", "G列非空但样式无法判定为模块或措施"


def load_authoritative_module_titles(workbook) -> set[str]:
    if "安全技术模块清单" not in workbook.sheetnames:
        return set()
    ws = workbook["安全技术模块清单"]
    titles: set[str] = set()
    last_module = ""
    for row_index in range(3, ws.max_row + 1):
        raw = as_text(ws.cell(row_index, 4).value)
        if raw and not re.fullmatch(r"\d+(?:\.\d+)?", raw):
            last_module = raw
        if last_module:
            titles.add(last_module)
    return titles


def normalize_scope_items(value: str) -> list[dict[str, str]]:
    return [
        {"code": code or "", "title": title or code or "", "text": " ".join(item for item in [code, title] if item)}
        for code, title in split_scope_values(value)
        if code or title
    ]


def scope_title_lookup(normalized_rows: list[dict[str, Any]]) -> dict[str, str]:
    titles: dict[str, str] = {}
    for row in normalized_rows:
        for scope in row.get("scopes") or []:
            code = as_text(scope.get("code"))
            title = as_text(scope.get("title"))
            if code and title and title != code:
                titles.setdefault(code, title)
    return titles


def augment_missing_scopes_from_services(
    normalized_rows: list[dict[str, Any]],
    object_scope_relations: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    title_by_code = scope_title_lookup(normalized_rows)
    rows_by_context: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in normalized_rows:
        context_key = as_text(row.get("contextKey"))
        if context_key:
            rows_by_context[context_key].append(row)

    supplements: list[dict[str, Any]] = []
    for context_key, rows in sorted(rows_by_context.items(), key=lambda item: min(int(row.get("row")) for row in item[1] if str(row.get("row")).isdigit())):
        declared_codes = {
            as_text(scope.get("code"))
            for row in rows
            for scope in row.get("scopes", [])
            if as_text(scope.get("code"))
        }
        required: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            scope_code = service_scope_code_from_row(row)
            if not scope_code or scope_code == "ALL":
                continue
            required[scope_code].append(
                {
                    "row": row.get("row"),
                    "service": as_text(row.get("securityTechnicalService")),
                    "sourceCell": (row.get("sourceCells") or {}).get("securityTechnicalService", {}).get("sourceCell"),
                }
            )
        missing_codes = sorted(set(required) - declared_codes)
        if not missing_codes:
            continue

        anchor_row = min(rows, key=lambda item: int(item.get("row")) if str(item.get("row")).isdigit() else 10**9)
        for code in missing_codes:
            title = title_by_code.get(code, code)
            scope = {
                "code": code,
                "title": title,
                "text": " ".join(part for part in [code, title if title != code else ""] if part),
                "source": "service_scope_reverse_check",
                "supplemented": True,
            }
            if not any(as_text(item.get("code")) == code for item in anchor_row.get("scopes", [])):
                anchor_row.setdefault("scopes", []).append(scope)
            anchor_row["scope"] = " / ".join(
                unique_scope_text
                for unique_scope_text in dict.fromkeys(
                    as_text(item.get("text") or " ".join(part for part in [item.get("code"), item.get("title")] if part))
                    for item in anchor_row.get("scopes", [])
                )
                if unique_scope_text
            )

            first = rows[0]
            relation_rows = [
                int(item.get("row"))
                for item in rows
                if service_scope_code_from_row(item) == code and str(item.get("row")).isdigit()
            ]
            if not relation_rows:
                relation_rows = [int(anchor_row.get("row"))] if str(anchor_row.get("row")).isdigit() else []
            for row_number in relation_rows or [anchor_row.get("row")]:
                add_unique_relation(
                    object_scope_relations,
                    relation_key(first.get("informationEnvironment"), first.get("environmentSegment"), first.get("informationObject"), scope["code"], scope["title"]),
                    {
                        "informationEnvironment": first.get("informationEnvironment"),
                        "environmentSegment": first.get("environmentSegment"),
                        "informationObject": first.get("informationObject"),
                        "scope": scope,
                        "objectMergedRange": (first.get("mergedRanges") or {}).get("informationObject"),
                        "scopeMergedRange": None,
                        "supplementedFromServices": True,
                        "supplementReason": "根据当前对象上下文下安全技术服务编号反查作用域后补充",
                        "serviceEvidence": required.get(code, []),
                    },
                    int(row_number) if str(row_number).isdigit() else 0,
                )
            supplements.append(
                {
                    "objectContextKey": context_key,
                    "environment": first.get("informationEnvironment"),
                    "environmentSegment": first.get("environmentSegment"),
                    "informationObject": first.get("informationObject"),
                    "scope": scope,
                    "evidence": required.get(code, []),
                    "reason": "原表作用域单元格未包含该服务编号反查出的作用域，已在导入投影中补充",
                }
            )
    return supplements


def normalize_rows(workbook, ws, merged_by_coordinate: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    authoritative_module_titles = load_authoritative_module_titles(workbook)
    normalized_rows: list[dict[str, Any]] = []
    relations = {
        "objectScopeRelations": {},
        "objectServiceRelations": {},
        "serviceModuleRelations": {},
        "serviceMeasureRelations": {},
        "moduleSystemRelations": {},
        "pendingRelations": {},
        "invalidRows": {},
    }

    for row_index in range(3, ws.max_row + 1):
        row_values = {
            field: effective_cell(ws, merged_by_coordinate, row_index, column)
            for column, (field, _label) in FIELD_COLUMNS.items()
        }
        raw_values = {field: payload["value"] for field, payload in row_values.items()}
        if not any(raw_values.values()):
            continue

        module_cell = row_values["moduleOrMeasureRaw"]["sourceCellObject"]
        module_raw = raw_values["moduleOrMeasureRaw"]
        module_kind, pending_reason = module_or_measure_kind(
            module_cell,
            module_raw,
            system_value=raw_values["securitySystem"],
            authoritative_module_titles=authoritative_module_titles,
        )
        scopes = normalize_scope_items(raw_values["scope"])
        systems = split_systems(raw_values["securitySystem"])
        service = service_payload(raw_values["securityTechnicalService"]) if not is_blank_or_placeholder(raw_values["securityTechnicalService"]) else None
        context_key = object_context_key(
            raw_values["informationEnvironment"],
            raw_values["environmentSegment"],
            raw_values["informationObject"],
        )
        row_payload = {
            "row": row_index,
            "informationEnvironment": raw_values["informationEnvironment"],
            "environmentSegment": raw_values["environmentSegment"],
            "informationObject": raw_values["informationObject"],
            "scope": raw_values["scope"],
            "scopes": scopes,
            "securityTechnicalService": raw_values["securityTechnicalService"],
            "securityTechnologyModule": "",
            "securityTechnicalMeasure": "",
            "moduleOrMeasureRaw": module_raw,
            "moduleOrMeasureKind": module_kind,
            "securitySystem": raw_values["securitySystem"],
            "securitySystems": systems,
            "contextKey": context_key,
            "sourceCells": source_cells_payload(row_values),
            "mergedRanges": merged_payload(row_values),
            "styleEvidence": style_evidence_payload(row_values),
        }

        if not raw_values["informationEnvironment"] or not raw_values["informationObject"]:
            if any(raw_values.values()):
                add_unique_relation(
                    relations["invalidRows"],
                    relation_key("missing_context", row_index),
                    {
                        "type": "missing_context",
                        "row": row_index,
                        "reason": "缺少信息化环境或信息化对象，不能形成上下文关系",
                        "rowData": row_payload,
                    },
                    row_index,
                )
            normalized_rows.append(row_payload)
            continue

        for scope in scopes:
            add_unique_relation(
                relations["objectScopeRelations"],
                relation_key(raw_values["informationEnvironment"], raw_values["environmentSegment"], raw_values["informationObject"], scope["code"], scope["title"]),
                {
                    "informationEnvironment": raw_values["informationEnvironment"],
                    "environmentSegment": raw_values["environmentSegment"],
                    "informationObject": raw_values["informationObject"],
                    "scope": scope,
                    "objectMergedRange": row_values["informationObject"]["mergedRange"],
                    "scopeMergedRange": row_values["scope"]["mergedRange"],
                },
                row_index,
            )

        if service:
            add_unique_relation(
                relations["objectServiceRelations"],
                relation_key(raw_values["informationEnvironment"], raw_values["environmentSegment"], raw_values["informationObject"], service["code"] or service["title"]),
                {
                    "informationEnvironment": raw_values["informationEnvironment"],
                    "environmentSegment": raw_values["environmentSegment"],
                    "informationObject": raw_values["informationObject"],
                    "scope": raw_values["scope"],
                    "service": service,
                    "objectMergedRange": row_values["informationObject"]["mergedRange"],
                    "serviceMergedRange": row_values["securityTechnicalService"]["mergedRange"],
                },
                row_index,
            )

        if module_kind == "placeholder":
            row_payload["securityTechnologyModule"] = ""
            row_payload["securityTechnicalMeasure"] = ""
        elif module_kind == "pending":
            add_unique_relation(
                relations["pendingRelations"],
                relation_key("pending_module_or_measure", row_index, module_raw),
                {
                    "type": "pending_module_or_measure",
                    "row": row_index,
                    "reason": pending_reason or "待确认模块 / 措施说明",
                    "informationObject": raw_values["informationObject"],
                    "service": service,
                    "raw": module_raw,
                },
                row_index,
            )
        elif not service:
            add_unique_relation(
                relations["invalidRows"],
                relation_key("module_without_service", row_index, module_raw),
                {
                    "type": "module_without_service",
                    "row": row_index,
                    "reason": "G列存在模块 / 措施，但当前行没有显式或合并单元格服务",
                    "rowData": row_payload,
                },
                row_index,
            )
        elif module_kind == "module":
            module_titles = _security_module_titles_from_catalog_alias(module_raw, authoritative_module_titles)
            if not module_titles:
                module_titles = split_multivalue_text(module_raw, split_on_ideographic_comma=False) or [module_raw]
                add_unique_relation(
                    relations["pendingRelations"],
                    relation_key("unmatched_module_alias", row_index, module_raw),
                    {
                        "type": "unmatched_module_alias",
                        "row": row_index,
                        "reason": "模块文本未匹配安全技术模块清单别名规则，按原文保留",
                        "raw": module_raw,
                        "service": service,
                    },
                    row_index,
                )
            row_payload["securityTechnologyModule"] = " / ".join(module_titles)
            for module_title in module_titles:
                add_unique_relation(
                    relations["serviceModuleRelations"],
                    relation_key(raw_values["informationEnvironment"], raw_values["environmentSegment"], raw_values["informationObject"], service["code"] or service["title"], module_title),
                    {
                        "informationEnvironment": raw_values["informationEnvironment"],
                        "environmentSegment": raw_values["environmentSegment"],
                        "informationObject": raw_values["informationObject"],
                        "service": service,
                        "module": {"title": module_title, "raw": module_raw},
                        "moduleMergedRange": row_values["moduleOrMeasureRaw"]["mergedRange"],
                    },
                    row_index,
                )
                for system_title in systems:
                    add_unique_relation(
                        relations["moduleSystemRelations"],
                        relation_key(raw_values["informationEnvironment"], raw_values["environmentSegment"], raw_values["informationObject"], module_title, system_title),
                        {
                            "informationEnvironment": raw_values["informationEnvironment"],
                            "environmentSegment": raw_values["environmentSegment"],
                            "informationObject": raw_values["informationObject"],
                            "module": {"title": module_title, "raw": module_raw},
                            "securitySystem": {"title": system_title},
                            "moduleMergedRange": row_values["moduleOrMeasureRaw"]["mergedRange"],
                            "systemMergedRange": row_values["securitySystem"]["mergedRange"],
                        },
                        row_index,
                    )
        elif module_kind == "measure":
            measure_items = _split_measure_names(module_raw)
            normal_measures = []
            for measure_title, was_note_wrapper in measure_items:
                _category, status = _measure_category_status(measure_title, "measure", was_note_wrapper)
                if status != "normal" and not was_note_wrapper:
                    add_unique_relation(
                        relations["pendingRelations"],
                        relation_key("pending_measure", row_index, measure_title),
                        {
                            "type": "pending_measure",
                            "row": row_index,
                            "reason": "N/A 或说明类措施，不作为正常关系",
                            "raw": module_raw,
                            "measure": measure_title,
                            "service": service,
                        },
                        row_index,
                    )
                    continue
                normal_measures.append(measure_title)
                add_unique_relation(
                    relations["serviceMeasureRelations"],
                    relation_key(raw_values["informationEnvironment"], raw_values["environmentSegment"], raw_values["informationObject"], service["code"] or service["title"], measure_title),
                    {
                        "informationEnvironment": raw_values["informationEnvironment"],
                        "environmentSegment": raw_values["environmentSegment"],
                        "informationObject": raw_values["informationObject"],
                        "service": service,
                        "measure": {"title": measure_title, "raw": module_raw},
                        "measureMergedRange": row_values["moduleOrMeasureRaw"]["mergedRange"],
                    },
                    row_index,
                )
            row_payload["securityTechnicalMeasure"] = " / ".join(normal_measures)
        else:
            add_unique_relation(
                relations["invalidRows"],
                relation_key("unknown_module_or_measure_style", row_index, module_raw),
                {
                    "type": "unknown_module_or_measure_style",
                    "row": row_index,
                    "reason": pending_reason or "G列样式未知",
                    "rowData": row_payload,
                },
                row_index,
            )
        normalized_rows.append(row_payload)

    scope_supplements = augment_missing_scopes_from_services(normalized_rows, relations["objectScopeRelations"])
    relation_lists = {name: list(bucket.values()) for name, bucket in relations.items()}
    duplicate_contexts, same_name_different_contexts = duplicate_information_object_contexts(normalized_rows)
    relation_lists["scopeSupplementsFromServices"] = scope_supplements
    return normalized_rows, relation_lists, duplicate_contexts, same_name_different_contexts


def service_scope_code_from_row(row: dict[str, Any]) -> str:
    service = service_payload(row.get("securityTechnicalService"))
    return as_text(service.get("scopeCode"))


def unique_nonempty(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        item = as_text(value)
        if not item or item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def split_relation_titles(value: Any) -> list[str]:
    titles: list[str] = []
    for line in split_multivalue_text(value, split_on_ideographic_comma=False):
        for part in re.split(r"\s+/\s+", line):
            item = as_text(part)
            if item and not is_blank_or_placeholder(item):
                titles.append(item)
    return unique_nonempty(titles)


def normal_security_system_titles(row: dict[str, Any]) -> list[str]:
    source = row.get("securitySystems")
    values = source if isinstance(source, list) else split_systems(row.get("securitySystem"))
    systems = [as_text(value) for value in values if not is_empty_security_system_value(value)]
    return unique_nonempty(systems) or [""]


def service_child_relation_entries(row: dict[str, Any]) -> list[dict[str, Any]]:
    context_key = as_text(row.get("contextKey"))
    service_text = as_text(row.get("securityTechnicalService"))
    if not context_key or not service_text or is_blank_or_placeholder(service_text):
        return []
    service = service_payload(service_text)
    service_key = service.get("code") or service.get("title") or service_text
    merged_ranges = row.get("mergedRanges") or {}
    source_cells = row.get("sourceCells") or {}
    entries: list[dict[str, Any]] = []
    child_groups = [
        ("module", split_relation_titles(row.get("securityTechnologyModule"))),
        ("measure", split_relation_titles(row.get("securityTechnicalMeasure"))),
    ]
    for child_type, children in child_groups:
        for child_title in children:
            for system_title in normal_security_system_titles(row):
                entries.append(
                    {
                        "objectContextKey": context_key,
                        "environment": row.get("informationEnvironment") or "",
                        "environmentSegment": row.get("environmentSegment") or "",
                        "informationObject": row.get("informationObject") or "",
                        "securityTechnicalService": service_key,
                        "securityTechnicalServiceRaw": service_text,
                        "childType": child_type,
                        "securityTechnologyModuleOrMeasure": child_title,
                        "securitySystem": system_title,
                        "row": row.get("row"),
                        "mergedRanges": {
                            "securityTechnicalService": merged_ranges.get("securityTechnicalService"),
                            "moduleOrMeasureRaw": merged_ranges.get("moduleOrMeasureRaw"),
                            "securitySystem": merged_ranges.get("securitySystem"),
                        },
                        "sourceCells": {
                            "securityTechnicalService": (source_cells.get("securityTechnicalService") or {}).get("sourceCell"),
                            "moduleOrMeasureRaw": (source_cells.get("moduleOrMeasureRaw") or {}).get("sourceCell"),
                            "securitySystem": (source_cells.get("securitySystem") or {}).get("sourceCell"),
                        },
                    }
                )
    return entries


def duplicate_services_in_object_context(normalized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in normalized_rows:
        for entry in service_child_relation_entries(row):
            grouped[
                (
                    entry["objectContextKey"],
                    entry["securityTechnicalService"],
                    entry["childType"],
                    entry["securityTechnologyModuleOrMeasure"],
                    entry["securitySystem"],
                )
            ].append(entry)

    duplicates: list[dict[str, Any]] = []
    for (context_key, service_key, child_type, child_title, security_system), rows in sorted(grouped.items(), key=lambda item: (item[0][0], min(row["row"] for row in item[1]))):
        if len(rows) <= 1:
            continue
        source_ranges = sorted({
            " / ".join(
                part
                for part in [
                    row.get("mergedRanges", {}).get("securityTechnicalService") or row.get("sourceCells", {}).get("securityTechnicalService"),
                    row.get("mergedRanges", {}).get("moduleOrMeasureRaw") or row.get("sourceCells", {}).get("moduleOrMeasureRaw"),
                    row.get("mergedRanges", {}).get("securitySystem") or row.get("sourceCells", {}).get("securitySystem"),
                ]
                if part
            )
            or f"row:{row.get('row')}"
            for row in rows
        })
        first = rows[0]
        duplicates.append(
            {
                "objectContextKey": context_key,
                "environment": first.get("environment") or "",
                "environmentSegment": first.get("environmentSegment") or "",
                "informationObject": first.get("informationObject") or "",
                "securityTechnicalService": service_key,
                "childType": child_type,
                "securityTechnologyModuleOrMeasure": child_title,
                "securitySystem": security_system,
                "occurrences": len(rows),
                "rows": [row["row"] for row in rows],
                "mergedRanges": source_ranges,
                "severity": "high",
                "reason": "同一对象上下文下出现完全重复的 服务-模块/措施-安全系统 关系",
            }
        )
    return duplicates


def repeated_services_with_different_children(normalized_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in normalized_rows:
        for entry in service_child_relation_entries(row):
            grouped[(entry["objectContextKey"], entry["securityTechnicalService"])].append(entry)

    repeated: list[dict[str, Any]] = []
    for (context_key, service_key), entries in sorted(grouped.items(), key=lambda item: (item[0][0], min(entry["row"] for entry in item[1]))):
        children: dict[tuple[str, str, str], dict[str, Any]] = {}
        for entry in entries:
            child_key = (entry["childType"], entry["securityTechnologyModuleOrMeasure"], entry["securitySystem"])
            children.setdefault(
                child_key,
                {
                    "childType": entry["childType"],
                    "securityTechnologyModuleOrMeasure": entry["securityTechnologyModuleOrMeasure"],
                    "securitySystem": entry["securitySystem"],
                    "rows": [],
                },
            )
            if entry["row"] not in children[child_key]["rows"]:
                children[child_key]["rows"].append(entry["row"])
        if len(children) <= 1:
            continue
        first = entries[0]
        repeated.append(
            {
                "objectContextKey": context_key,
                "environment": first.get("environment") or "",
                "environmentSegment": first.get("environmentSegment") or "",
                "informationObject": first.get("informationObject") or "",
                "securityTechnicalService": service_key,
                "children": sorted(children.values(), key=lambda item: (item["childType"], item["securityTechnologyModuleOrMeasure"], item["securitySystem"])),
                "rows": sorted({entry["row"] for entry in entries}),
                "severity": "info",
                "reason": "同一安全技术服务映射多个模块/措施/系统，属于可能合法的 1:N 展开",
            }
        )
    return repeated


def scope_completeness_issues(normalized_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_context: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in normalized_rows:
        context_key = as_text(row.get("contextKey"))
        if context_key:
            by_context[context_key].append(row)

    issues: list[dict[str, Any]] = []
    unknown: list[dict[str, Any]] = []
    for context_key, rows in sorted(by_context.items(), key=lambda item: min(row["row"] for row in item[1])):
        declared_scopes = sorted(
            {
                scope.get("code") or scope.get("title") or scope.get("text")
                for row in rows
                for scope in row.get("scopes", [])
                if scope.get("code") or scope.get("title") or scope.get("text")
            }
        )
        required_scopes: set[str] = set()
        unknown_services: list[dict[str, Any]] = []
        services_evidence: list[dict[str, Any]] = []
        for row in rows:
            service_text = as_text(row.get("securityTechnicalService"))
            if not service_text or is_blank_or_placeholder(service_text):
                continue
            service = service_payload(service_text)
            scope_code = as_text(service.get("scopeCode"))
            evidence = {
                "row": row["row"],
                "service": service_text,
                "scopeCode": scope_code,
                "sourceCell": row.get("sourceCells", {}).get("securityTechnicalService", {}).get("sourceCell"),
            }
            if scope_code and scope_code != "ALL":
                required_scopes.add(scope_code)
                services_evidence.append(evidence)
            else:
                unknown_services.append(evidence)
        missing_scopes = sorted(required_scopes - set(declared_scopes))
        first = rows[0]
        if missing_scopes:
            issues.append(
                {
                    "objectContextKey": context_key,
                    "environment": first.get("informationEnvironment") or "",
                    "environmentSegment": first.get("environmentSegment") or "",
                    "informationObject": first.get("informationObject") or "",
                    "declaredScopes": declared_scopes,
                    "requiredScopesFromServices": sorted(required_scopes),
                    "missingScopes": missing_scopes,
                    "servicesEvidence": services_evidence,
                    "severity": "medium",
                    "reason": "根据安全技术服务反查，当前对象作用域疑似缺失",
                }
            )
        if unknown_services:
            unknown.append(
                {
                    "objectContextKey": context_key,
                    "environment": first.get("informationEnvironment") or "",
                    "environmentSegment": first.get("environmentSegment") or "",
                    "informationObject": first.get("informationObject") or "",
                    "unknownScopeEvidence": unknown_services,
                    "severity": "info",
                    "reason": "部分安全技术服务无法从编号反查作用域，未自动推断",
                }
            )
    return issues, unknown


def duplicate_information_object_contexts(normalized_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_context: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_name: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)
    for row in normalized_rows:
        information_object = row.get("informationObject") or ""
        if not information_object:
            continue
        context_key = object_context_key(row.get("informationEnvironment"), row.get("environmentSegment"), information_object)
        by_context[context_key].append(row)
        by_name[information_object][context_key] = {
            "environment": row.get("informationEnvironment") or "",
            "environmentSegment": row.get("environmentSegment") or "",
            "contextKey": context_key,
        }

    duplicate_contexts: list[dict[str, Any]] = []
    for context_key, rows in sorted(by_context.items()):
        if not rows:
            continue
        first = rows[0]
        merged_ranges = sorted({row.get("mergedRanges", {}).get("informationObject") or row.get("sourceCells", {}).get("informationObject", {}).get("sourceCell") for row in rows})
        legal_single_merged_range = len(merged_ranges) == 1
        if legal_single_merged_range:
            continue
        duplicate_contexts.append(
            {
                "contextKey": context_key,
                "environment": first.get("informationEnvironment") or "",
                "environmentSegment": first.get("environmentSegment") or "",
                "informationObject": first.get("informationObject") or "",
                "occurrences": len(rows),
                "rows": [row["row"] for row in rows],
                "mergedRanges": merged_ranges,
                "scopes": sorted({row.get("scope") for row in rows if row.get("scope")}),
                "services": sorted({row.get("securityTechnicalService") for row in rows if row.get("securityTechnicalService")}),
                "severity": "high",
                "reason": "同一环境-环境子类-信息化对象上下文被解析为多个独立对象上下文，疑似合并单元格格式丢失或错误展开",
            }
        )

    same_name_different_contexts: list[dict[str, Any]] = []
    for information_object, contexts in sorted(by_name.items()):
        if len(contexts) <= 1:
            continue
        same_name_different_contexts.append(
            {
                "informationObject": information_object,
                "contexts": sorted(contexts.values(), key=lambda item: item["contextKey"]),
                "severity": "info",
                "reason": "同名信息化对象出现在不同环境或环境子类下，按业务规则视为不同对象实例",
            }
        )
    return duplicate_contexts, same_name_different_contexts


def load_json(path: Path) -> dict[str, Any] | list[Any] | None:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def compact_title(item: dict[str, Any] | None) -> str:
    if not item:
        return ""
    return as_text(item.get("title") or item.get("name") or item.get("code") or item.get("id"))


def current_environment_relations(environment_workbench: dict[str, Any] | None) -> dict[str, list[dict[str, Any]]]:
    buckets = {
        "objectScopeRelations": {},
        "objectServiceRelations": {},
        "serviceModuleRelations": {},
        "serviceMeasureRelations": {},
        "moduleSystemRelations": {},
    }
    if not environment_workbench:
        return {name: [] for name in buckets}
    for environment in environment_workbench.get("environment_scope_tree") or environment_workbench.get("environmentScopeTree") or []:
        environment_title = compact_title(environment)
        for obj in environment.get("objects") or []:
            object_title = compact_title(obj)
            segments = obj.get("segments") or [{}]
            segment_titles = [compact_title(segment) for segment in segments] or [""]
            for segment_title in segment_titles:
                for mapping in obj.get("scope_mappings") or []:
                    scope = mapping.get("scope") or {}
                    scope_code = as_text(scope.get("code"))
                    scope_title = compact_title(scope)
                    add_unique_relation(
                        buckets["objectScopeRelations"],
                        relation_key(environment_title, segment_title, object_title, scope_code, scope_title),
                        {
                            "informationEnvironment": environment_title,
                            "environmentSegment": segment_title,
                            "informationObject": object_title,
                            "scope": {"code": scope_code, "title": scope_title},
                        },
                        0,
                    )
                    for service in mapping.get("services") or []:
                        service_code = as_text(service.get("code"))
                        service_title = compact_title(service)
                        add_unique_relation(
                            buckets["objectServiceRelations"],
                            relation_key(environment_title, segment_title, object_title, service_code or service_title),
                            {
                                "informationEnvironment": environment_title,
                                "environmentSegment": segment_title,
                                "informationObject": object_title,
                                "scope": scope_title,
                                "service": {"code": service_code, "title": service_title},
                            },
                            0,
                        )
                        for module in service.get("modules") or []:
                            module_title = compact_title(module)
                            add_unique_relation(
                                buckets["serviceModuleRelations"],
                                relation_key(environment_title, segment_title, object_title, service_code or service_title, module_title),
                                {
                                    "informationEnvironment": environment_title,
                                    "environmentSegment": segment_title,
                                    "informationObject": object_title,
                                    "service": {"code": service_code, "title": service_title},
                                    "module": {"title": module_title},
                                },
                                0,
                            )
                            for system in module.get("systems") or []:
                                system_title = compact_title(system)
                                add_unique_relation(
                                    buckets["moduleSystemRelations"],
                                    relation_key(environment_title, segment_title, object_title, module_title, system_title),
                                    {
                                        "informationEnvironment": environment_title,
                                        "environmentSegment": segment_title,
                                        "informationObject": object_title,
                                        "module": {"title": module_title},
                                        "securitySystem": {"title": system_title},
                                    },
                                    0,
                                )
                        for measure in service.get("measures") or []:
                            measure_title = compact_title(measure)
                            add_unique_relation(
                                buckets["serviceMeasureRelations"],
                                relation_key(environment_title, segment_title, object_title, service_code or service_title, measure_title),
                                {
                                    "informationEnvironment": environment_title,
                                    "environmentSegment": segment_title,
                                    "informationObject": object_title,
                                    "service": {"code": service_code, "title": service_title},
                                    "measure": {"title": measure_title},
                                },
                                0,
                            )
    return {name: list(bucket.values()) for name, bucket in buckets.items()}


def relation_set(relations: dict[str, list[dict[str, Any]]], name: str) -> set[str]:
    return {relation_identity(name, row) for row in relations.get(name, [])}


def relation_identity(name: str, row: dict[str, Any]) -> str:
    if name == "objectScopeRelations":
        scope = row.get("scope") or {}
        return relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), scope.get("code"), scope.get("title"))
    if name == "objectServiceRelations":
        service = row.get("service") or {}
        return relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), service.get("code") or service.get("title"))
    if name == "serviceModuleRelations":
        service = row.get("service") or {}
        module = row.get("module") or {}
        return relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), service.get("code") or service.get("title"), module.get("title"))
    if name == "serviceMeasureRelations":
        service = row.get("service") or {}
        measure = row.get("measure") or {}
        return relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), service.get("code") or service.get("title"), measure.get("title"))
    if name == "moduleSystemRelations":
        module = row.get("module") or {}
        system = row.get("securitySystem") or {}
        return relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), module.get("title"), system.get("title"))
    return json.dumps(row, ensure_ascii=False, sort_keys=True)


def relation_by_identity(relations: dict[str, list[dict[str, Any]]], name: str) -> dict[str, dict[str, Any]]:
    return {relation_identity(name, row): row for row in relations.get(name, [])}


def find_current_same_name_different_contexts(current_relations: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    contexts: dict[str, set[str]] = defaultdict(set)
    for row in current_relations.get("objectScopeRelations", []):
        object_title = row.get("informationObject") or ""
        if not object_title:
            continue
        contexts[object_title].add(relation_key(row.get("informationEnvironment"), row.get("environmentSegment")))
    return [
        {
            "informationObject": object_title,
            "contextCount": len(values),
            "contexts": sorted(values),
            "severity": "info",
            "reason": "当前 workbench JSON 中同名信息化对象对应多个环境 / 环境子类上下文，按上下文唯一键不直接视为重复",
        }
        for object_title, values in sorted(contexts.items())
        if len(values) > 1
    ]


def compare_relations(expected: dict[str, list[dict[str, Any]]], actual: dict[str, list[dict[str, Any]]], max_items: int = 500) -> dict[str, Any]:
    comparison: dict[str, Any] = {}
    for name in ("objectScopeRelations", "objectServiceRelations", "serviceModuleRelations", "serviceMeasureRelations", "moduleSystemRelations"):
        expected_map = relation_by_identity(expected, name)
        actual_map = relation_by_identity(actual, name)
        expected_keys = set(expected_map)
        actual_keys = set(actual_map)
        missing = sorted(expected_keys - actual_keys)
        unexpected = sorted(actual_keys - expected_keys)
        comparison[name] = {
            "expectedCount": len(expected_keys),
            "actualCount": len(actual_keys),
            "missingInCurrentJsonCount": len(missing),
            "unexpectedInCurrentJsonCount": len(unexpected),
            "missingInCurrentJson": [expected_map[key] for key in missing[:max_items]],
            "unexpectedInCurrentJson": [actual_map[key] for key in unexpected[:max_items]],
        }

    expected_modules = relation_set(expected, "serviceModuleRelations")
    actual_modules = relation_set(actual, "serviceModuleRelations")
    expected_measures = relation_set(expected, "serviceMeasureRelations")
    actual_measures = relation_set(actual, "serviceMeasureRelations")
    comparison["moduleMeasureConfusions"] = {
        "currentModuleButExpectedMeasure": sorted(actual_modules.intersection(expected_measures))[:max_items],
        "currentMeasureButExpectedModule": sorted(actual_measures.intersection(expected_modules))[:max_items],
    }

    expected_module_by_service = defaultdict(set)
    actual_module_by_service = defaultdict(set)
    for row in expected.get("serviceModuleRelations", []):
        service = row.get("service") or {}
        expected_module_by_service[relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), service.get("code") or service.get("title"))].add((row.get("module") or {}).get("title"))
    for row in actual.get("serviceModuleRelations", []):
        service = row.get("service") or {}
        actual_module_by_service[relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"), service.get("code") or service.get("title"))].add((row.get("module") or {}).get("title"))

    over_aggregated = []
    for service_key, actual_modules_for_service in sorted(actual_module_by_service.items()):
        expected_modules_for_service = expected_module_by_service.get(service_key, set())
        extra_modules = sorted(actual_modules_for_service - expected_modules_for_service)
        if extra_modules:
            over_aggregated.append(
                {
                    "serviceContext": service_key,
                    "extraModules": extra_modules,
                    "expectedModules": sorted(expected_modules_for_service),
                    "actualModules": sorted(actual_modules_for_service),
                    "severity": "high",
                    "reason": "当前 JSON 在该对象 / 服务上下文中包含重导入结果没有的模块，疑似过度继承或聚合",
                }
            )
    comparison["overAggregatedServiceModules"] = over_aggregated[:max_items]
    comparison["missingSecuritySystems"] = comparison["moduleSystemRelations"]["missingInCurrentJson"]
    comparison["crossObjectInheritanceSuspicions"] = cross_object_inheritance_suspicions(expected, actual, max_items=max_items)
    comparison["duplicateInformationObjectsInCurrentJson"] = []
    comparison["sameNameDifferentContextsInCurrentJson"] = find_current_same_name_different_contexts(actual)
    return comparison


def cross_object_inheritance_suspicions(expected: dict[str, list[dict[str, Any]]], actual: dict[str, list[dict[str, Any]]], max_items: int) -> list[dict[str, Any]]:
    expected_by_service_module = defaultdict(set)
    for row in expected.get("serviceModuleRelations", []):
        service = row.get("service") or {}
        module = row.get("module") or {}
        expected_by_service_module[relation_key(service.get("code") or service.get("title"), module.get("title"))].add(
            relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"))
        )
    expected_full = relation_set(expected, "serviceModuleRelations")
    suspicions = []
    for row in actual.get("serviceModuleRelations", []):
        identity = relation_identity("serviceModuleRelations", row)
        if identity in expected_full:
            continue
        service = row.get("service") or {}
        module = row.get("module") or {}
        service_module_key = relation_key(service.get("code") or service.get("title"), module.get("title"))
        expected_contexts = expected_by_service_module.get(service_module_key, set())
        if not expected_contexts:
            continue
        current_context = relation_key(row.get("informationEnvironment"), row.get("environmentSegment"), row.get("informationObject"))
        suspicions.append(
            {
                "currentContext": current_context,
                "service": service,
                "module": module,
                "expectedContexts": sorted(expected_contexts),
                "severity": "high",
                "reason": "同一服务-模块关系存在于其他对象上下文，但不在当前对象上下文，疑似跨对象继承",
            }
        )
        if len(suspicions) >= max_items:
            break
    return suspicions


def package_summary(path: Path) -> dict[str, Any]:
    data = load_json(path)
    if data is None:
        return {"path": str(path.relative_to(PROJECT_ROOT)), "exists": False}
    type_counts: Counter[str] = Counter()
    key_counts: Counter[str] = Counter()

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            if value.get("type"):
                type_counts[as_text(value.get("type"))] += 1
            for key in ("modules", "measures", "securitySystems", "security_systems", "systems", "scope_mappings"):
                if key in value and isinstance(value[key], list):
                    key_counts[key] += len(value[key])
            for child in value.values():
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(data)
    return {
        "path": str(path.relative_to(PROJECT_ROOT)),
        "exists": True,
        "topLevelKeys": sorted(data.keys()) if isinstance(data, dict) else [],
        "typeCounts": dict(type_counts),
        "relationLikeCounts": dict(key_counts),
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def markdown_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], limit: int = 20) -> str:
    if not rows:
        return "无\n"
    header = "| " + " | ".join(label for _key, label in columns) + " |\n"
    divider = "| " + " | ".join("---" for _key, _label in columns) + " |\n"
    body = []
    for row in rows[:limit]:
        body.append("| " + " | ".join(as_text(row.get(key)).replace("\n", "<br>") for key, _label in columns) + " |")
    return header + divider + "\n".join(body) + "\n"


def build_markdown_report(audit: dict[str, Any]) -> str:
    summary = audit["summary"]
    comparison = audit["currentJsonComparison"]
    duplicate_contexts = audit.get("duplicateInformationObjectContexts", [])
    same_name_different_contexts = audit.get("sameNameDifferentContexts", [])
    lines = [
        "# Scope-Service-Module Mapping Reimport Audit",
        "",
        f"- 生成时间：{audit['generatedAt']}",
        f"- Workbook：`{audit['workbook']}`",
        f"- Sheet：`{audit['sheet']}`",
        f"- merged ranges：{summary['mergedRangeCount']}",
        f"- normalized rows：{summary['normalizedRowCount']}",
        "",
        "## 关系统计",
        "",
        f"- object-scope：{summary['objectScopeRelationCount']}",
        f"- object-service：{summary['objectServiceRelationCount']}",
        f"- service-module：{summary['serviceModuleRelationCount']}",
        f"- service-measure：{summary['serviceMeasureRelationCount']}",
        f"- module-system：{summary['moduleSystemRelationCount']}",
        f"- pending：{summary['pendingRelationCount']}",
        f"- invalid rows：{summary['invalidRowCount']}",
        "",
        "## merged ranges by field",
        "",
    ]
    lines.append(markdown_table(
        [{"field": key, "count": value} for key, value in sorted(summary["mergedRangesByField"].items())],
        [("field", "字段"), ("count", "合并区域数")],
        limit=50,
    ))
    lines.extend([
        "",
        "## 信息化对象唯一性检查",
        "",
        f"- 唯一键：`信息化环境 + 环境子类 + 信息化对象`",
        f"- duplicateInformationObjectContexts：{len(duplicate_contexts)}",
        f"- sameNameDifferentContexts：{len(same_name_different_contexts)}",
    ])
    if duplicate_contexts:
        lines.append(markdown_table(
            [
                {
                    "contextKey": item["contextKey"],
                    "occurrences": item["occurrences"],
                    "mergedRanges": ", ".join(item["mergedRanges"]),
                    "rows": ", ".join(str(row) for row in item["rows"][:12]),
                    "reason": item["reason"],
                }
                for item in duplicate_contexts
            ],
            [("contextKey", "上下文唯一键"), ("occurrences", "行数"), ("mergedRanges", "merged ranges"), ("rows", "样例行"), ("reason", "原因")],
            limit=30,
        ))
    else:
        lines.append("未发现完整上下文重复且跨多个 object merged range 的高风险对象实例。")
    if same_name_different_contexts:
        lines.extend(["", "### 同名但不同上下文的信息提示", ""])
        lines.append(markdown_table(
            [
                {
                    "informationObject": item["informationObject"],
                    "contexts": "<br>".join(context["contextKey"] for context in item["contexts"][:12]),
                    "reason": item["reason"],
                }
                for item in same_name_different_contexts
            ],
            [("informationObject", "信息化对象名称"), ("contexts", "上下文"), ("reason", "说明")],
            limit=30,
        ))
    lines.extend([
        "",
        "## 完整重复服务-模块/措施-安全系统关系检查",
        "",
        f"- duplicateExactServiceChildRelations：{len(audit.get('duplicateExactServiceChildRelations', []))}",
        f"- repeatedServiceWithDifferentChildren：{len(audit.get('repeatedServiceWithDifferentChildren', []))}",
    ])
    if audit.get("duplicateExactServiceChildRelations"):
        lines.append(markdown_table(
            [
                {
                    "objectContextKey": item["objectContextKey"],
                    "securityTechnicalService": item["securityTechnicalService"],
                    "child": f"{item.get('childType', '')}:{item.get('securityTechnologyModuleOrMeasure', '')}",
                    "securitySystem": item.get("securitySystem", ""),
                    "occurrences": item["occurrences"],
                    "rows": ", ".join(str(row) for row in item["rows"][:12]),
                    "mergedRanges": ", ".join(item["mergedRanges"][:8]),
                }
                for item in audit.get("duplicateExactServiceChildRelations", [])
            ],
            [("objectContextKey", "对象上下文"), ("securityTechnicalService", "服务"), ("child", "模块/措施"), ("securitySystem", "安全系统"), ("occurrences", "次数"), ("rows", "行号"), ("mergedRanges", "merged ranges")],
            limit=30,
        ))
    lines.extend([
        "",
        "## 服务反查作用域完整性",
        "",
        f"- scopeCompletenessIssues：{len(audit.get('scopeCompletenessIssues', []))}",
        f"- unknownScopeEvidence：{len(audit.get('unknownScopeEvidence', []))}",
    ])
    if audit.get("scopeCompletenessIssues"):
        lines.append(markdown_table(
            [
                {
                    "objectContextKey": item["objectContextKey"],
                    "declaredScopes": ", ".join(item.get("declaredScopes", [])),
                    "requiredScopesFromServices": ", ".join(item.get("requiredScopesFromServices", [])),
                    "missingScopes": ", ".join(item.get("missingScopes", [])),
                    "reason": item["reason"],
                }
                for item in audit.get("scopeCompletenessIssues", [])
            ],
            [("objectContextKey", "对象上下文"), ("declaredScopes", "原表作用域"), ("requiredScopesFromServices", "服务反查作用域"), ("missingScopes", "缺失作用域"), ("reason", "说明")],
            limit=40,
        ))
    lines.extend(["", "## 当前 JSON 对比摘要", ""])
    for name in ("objectScopeRelations", "objectServiceRelations", "serviceModuleRelations", "serviceMeasureRelations", "moduleSystemRelations"):
        item = comparison[name]
        lines.extend(
            [
                f"### {name}",
                "",
                f"- 重导入关系数：{item['expectedCount']}",
                f"- 当前 JSON 关系数：{item['actualCount']}",
                f"- 当前 JSON 缺失：{item['missingInCurrentJsonCount']}",
                f"- 当前 JSON 多出：{item['unexpectedInCurrentJsonCount']}",
                "",
            ]
        )
    high_risk_count = (
        len(comparison["overAggregatedServiceModules"])
        + len(comparison["crossObjectInheritanceSuspicions"])
        + len(duplicate_contexts)
        + len(audit.get("duplicateExactServiceChildRelations", []))
        + len(audit.get("scopeCompletenessIssues", []))
    )
    lines.extend([
        "## 高风险问题",
        "",
        f"- 高风险项数量：{high_risk_count}",
        f"- 过度聚合样例：{len(comparison['overAggregatedServiceModules'])}",
        f"- 跨对象继承疑似样例：{len(comparison['crossObjectInheritanceSuspicions'])}",
        f"- 当前 JSON 同名不同上下文提示：{len(comparison.get('sameNameDifferentContextsInCurrentJson', []))}",
        "",
    ])
    if comparison["overAggregatedServiceModules"]:
        lines.extend(["### 过度聚合样例", ""])
        lines.append(markdown_table(
            [
                {
                    "serviceContext": item["serviceContext"],
                    "extraModules": ", ".join(item["extraModules"]),
                    "expectedModules": ", ".join(item["expectedModules"]),
                }
                for item in comparison["overAggregatedServiceModules"][:20]
            ],
            [("serviceContext", "服务上下文"), ("extraModules", "JSON 多出的模块"), ("expectedModules", "重导入模块")],
            limit=20,
        ))
    lines.extend([
        "",
        "## 结论",
        "",
        f"- 当前视图数据可信结论：{audit['conclusion']['currentViewDataTrust']}",
        f"- 是否建议替换正式数据包：{audit['conclusion']['recommendReplacePublicData']}",
        f"- 下一步：{audit['conclusion']['nextStep']}",
        "",
    ])
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit scope-service-module mapping sheet from source xlsx.")
    parser.add_argument("--workbook", default="data/raw-samples/wiki sample.xlsx")
    parser.add_argument("--sheet", default=SHEET_NAME)
    parser.add_argument("--output-dir", default=str(OUTPUT_DIR))
    parser.add_argument("--max-differences", type=int, default=500)
    args = parser.parse_args()

    workbook_path = (PROJECT_ROOT / args.workbook).resolve() if not Path(args.workbook).is_absolute() else Path(args.workbook)
    output_dir = (PROJECT_ROOT / args.output_dir).resolve() if not Path(args.output_dir).is_absolute() else Path(args.output_dir)
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if args.sheet not in workbook.sheetnames:
        candidates = [name for name in workbook.sheetnames if "作用域" in name or "安全技术服务" in name or "模块" in name]
        raise SystemExit(f"Sheet not found: {args.sheet}. Candidates: {candidates}")
    ws = workbook[args.sheet]
    merged_ranges, merged_by_coordinate = merged_ranges_for_sheet(ws)
    normalized_rows, relations, duplicate_contexts, same_name_different_contexts = normalize_rows(workbook, ws, merged_by_coordinate)
    duplicate_services = duplicate_services_in_object_context(normalized_rows)
    repeated_services = repeated_services_with_different_children(normalized_rows)
    scope_issues, unknown_scope_evidence = scope_completeness_issues(normalized_rows)

    field_counts = Counter()
    for item in merged_ranges:
        for field in item.get("fields") or ["__non_core__"]:
            field_counts[field] += 1

    package_paths = {
        "environmentWorkbench": PROJECT_ROOT / "frontend/capability-browser/public/data/environment-workbench.json",
        "capabilityWorkbench": PROJECT_ROOT / "frontend/capability-browser/public/data/capability-workbench.json",
        "maintenanceKnowledge": PROJECT_ROOT / "frontend/capability-browser/public/data/maintenance-knowledge.json",
        "managementKnowledge": PROJECT_ROOT / "frontend/capability-browser/public/data/management-knowledge.json",
    }
    environment_workbench = load_json(package_paths["environmentWorkbench"])
    current_relations = current_environment_relations(environment_workbench if isinstance(environment_workbench, dict) else None)
    comparison = compare_relations(relations, current_relations, max_items=args.max_differences)
    package_summaries = {name: package_summary(path) for name, path in package_paths.items()}

    high_risk_count = (
        len(duplicate_contexts)
        + len(duplicate_services)
        + len(scope_issues)
        + len(comparison["overAggregatedServiceModules"])
        + len(comparison["crossObjectInheritanceSuspicions"])
    )
    current_view_data_trust = "不可信，需要以审计关系重建候选数据包后再验收" if high_risk_count or comparison["serviceModuleRelations"]["missingInCurrentJsonCount"] or comparison["moduleSystemRelations"]["missingInCurrentJsonCount"] else "暂未发现关系级高风险差异"

    summary = {
        "workbook": str(workbook_path.relative_to(PROJECT_ROOT) if workbook_path.is_relative_to(PROJECT_ROOT) else workbook_path),
        "sheet": args.sheet,
        "sheetMaxRow": ws.max_row,
        "sheetMaxColumn": ws.max_column,
        "mergedRangeCount": len(merged_ranges),
        "mergedRangesByField": dict(sorted(field_counts.items())),
        "normalizedRowCount": len(normalized_rows),
        "objectScopeRelationCount": len(relations["objectScopeRelations"]),
        "objectServiceRelationCount": len(relations["objectServiceRelations"]),
        "serviceModuleRelationCount": len(relations["serviceModuleRelations"]),
        "serviceMeasureRelationCount": len(relations["serviceMeasureRelations"]),
        "moduleSystemRelationCount": len(relations["moduleSystemRelations"]),
        "pendingRelationCount": len(relations["pendingRelations"]),
        "invalidRowCount": len(relations["invalidRows"]),
        "duplicateInformationObjectContextCount": len(duplicate_contexts),
        "sameNameDifferentContextCount": len(same_name_different_contexts),
        "duplicateExactServiceChildRelationCount": len(duplicate_services),
        "duplicateServicesInObjectContextCount": len(duplicate_services),
        "repeatedServiceWithDifferentChildrenCount": len(repeated_services),
        "scopeCompletenessIssueCount": len(scope_issues),
        "unknownScopeEvidenceCount": len(unknown_scope_evidence),
    }
    audit = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": summary["workbook"],
        "sheet": args.sheet,
        "summary": summary,
        "mergedRanges": merged_ranges,
        "duplicateInformationObjects": [],
        "duplicateInformationObjectContexts": duplicate_contexts,
        "sameNameDifferentContexts": same_name_different_contexts,
        "duplicateExactServiceChildRelations": duplicate_services,
        "repeatedServiceWithDifferentChildren": repeated_services,
        "duplicateServicesInObjectContext": duplicate_services,
        "scopeCompletenessIssues": scope_issues,
        "unknownScopeEvidence": unknown_scope_evidence,
        "currentJsonComparison": comparison,
        "packageSummaries": package_summaries,
        "conclusion": {
            "currentViewDataTrust": current_view_data_trust,
            "recommendReplacePublicData": "否。本轮只生成审计输出；建议先生成 candidate 并人工验收差异。",
            "nextStep": "使用 normalized relations 生成 environment-workbench candidate，并对环境底图浮层与归纳表格做源表行级验收。",
        },
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_json(output_dir / "scope-service-module-mapping-normalized-rows.json", normalized_rows)
    write_json(output_dir / "scope-service-module-mapping-relations.json", relations)
    write_json(output_dir / "scope-service-module-mapping-reimport-audit.json", audit)
    (output_dir / "scope-service-module-mapping-reimport-audit.md").write_text(build_markdown_report(audit), encoding="utf-8")

    print(json.dumps({"result": "pass", **summary, "currentViewDataTrust": current_view_data_trust}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
