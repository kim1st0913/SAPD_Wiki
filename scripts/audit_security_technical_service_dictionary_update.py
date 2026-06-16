#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
DB_PATH = ROOT / "data/database/sapd_wiki.sqlite3"
OUT_DIR = ROOT / "data/exports/worker-verify/security-technical-service-update"

SERVICE_SHEET = "安全能力-安全技术服务"
FORMAL_PACKAGES = [
    ROOT / "frontend/capability-browser/public/data/maintenance-knowledge.json",
    ROOT / "frontend/capability-browser/public/data/maintenance-index.json",
    ROOT / "frontend/capability-browser/public/data/maintenance/services.json",
    ROOT / "frontend/capability-browser/public/data/maintenance/modules.json",
    ROOT / "frontend/capability-browser/public/data/maintenance/measures.json",
    ROOT / "frontend/capability-browser/public/data/source-evidence/maintenance/services.sources.json",
    ROOT / "frontend/capability-browser/public/data/capability-tree.json",
    ROOT / "frontend/capability-browser/public/data/capability-workbench.json",
    ROOT / "frontend/capability-browser/public/data/lifecycle-workbench.json",
    ROOT / "frontend/capability-browser/public/data/environment-workbench.json",
    ROOT / "frontend/capability-browser/generated/environmentBasemap.node-details.json",
]

SERVICE_CODE_RE = re.compile(r"^(?:(?:ALL|I-[A-Z]{2})&T-[A-Z]{2}\.[A-Z]{2}-\d{2}|M-[A-Z]{2}\.[A-Z]{2}-00)$")


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def text(value: Any) -> str:
    raw = "" if value is None else str(value)
    raw = raw.replace("\xa0", " ").replace("\u3000", " ")
    return re.sub(r"[ \t]+", " ", raw).strip()


def split_code_title(value: Any) -> tuple[str, str]:
    value_text = text(value)
    if not value_text:
        return "", ""
    parts = value_text.split(maxsplit=1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[1]


def normalize_scope_code(value: Any) -> str:
    return text(value).replace("_", "-")


def service_id_for_code(code: str) -> str:
    digest = hashlib.sha1(code.encode("utf-8")).hexdigest()[:16]
    return f"security_technical_service:{digest}"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_md(path: Path, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body.rstrip() + "\n", encoding="utf-8")


def merged_meta(ws) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                result[ws.cell(row, col).coordinate] = {
                    "range": str(merged_range),
                    "topLeftCell": anchor.coordinate,
                    "topLeftValue": text(anchor.value),
                    "rowSpan": merged_range.max_row - merged_range.min_row + 1,
                    "columnSpan": merged_range.max_col - merged_range.min_col + 1,
                }
    return result


def parse_source_services() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    wb = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    candidates = [
        ws.title
        for ws in wb.worksheets
        if "安全技术服务" in ws.title or "技术服务" in ws.title
    ]
    if SERVICE_SHEET not in wb.sheetnames:
        return [], [], [{"issue": "service_sheet_not_found", "candidateSheets": candidates}], candidates

    ws = wb[SERVICE_SHEET]
    merged = merged_meta(ws)
    services: list[dict[str, Any]] = []
    invalid_rows: list[dict[str, Any]] = []
    scope_headers = {col: split_code_title(ws.cell(3, col).value) for col in range(7, 14)}
    last_category = ""
    last_domain = ""
    last_capability = ""

    for row in range(4, ws.max_row + 1):
        if text(ws.cell(row, 2).value):
            last_category = text(ws.cell(row, 2).value)
        if text(ws.cell(row, 3).value):
            last_domain = text(ws.cell(row, 3).value)
        if text(ws.cell(row, 4).value):
            last_capability = text(ws.cell(row, 4).value)
        focus_code = text(ws.cell(row, 5).value)
        focus_title = text(ws.cell(row, 6).value)
        if not focus_code:
            continue
        for col in range(7, 14):
            raw = text(ws.cell(row, col).value)
            if not raw or raw in {"/", "\\", "N/A", "NA", "无"}:
                continue
            scope_code, scope_title = scope_headers[col]
            scope_code = normalize_scope_code(scope_code)
            code, title = split_code_title(raw)
            if not SERVICE_CODE_RE.match(code):
                invalid_rows.append(
                    {
                        "excelRow": row,
                        "excelColumn": col,
                        "cell": ws.cell(row, col).coordinate,
                        "rawValue": raw,
                        "issue": "invalid_or_missing_service_code",
                    }
                )
                continue
            code_scope = code.split("&", 1)[0] if "&" in code else scope_code
            capability_focus_code = code.split("&", 1)[1] if "&" in code else ""
            if scope_code and code_scope not in {scope_code, "ALL"} and not code.startswith("M-"):
                invalid_rows.append(
                    {
                        "excelRow": row,
                        "excelColumn": col,
                        "cell": ws.cell(row, col).coordinate,
                        "rawValue": raw,
                        "issue": "service_scope_code_mismatch_with_column_header",
                        "expectedScopeCode": scope_code,
                        "actualScopeCode": code_scope,
                    }
                )
            if capability_focus_code and capability_focus_code != focus_code:
                invalid_rows.append(
                    {
                        "excelRow": row,
                        "excelColumn": col,
                        "cell": ws.cell(row, col).coordinate,
                        "rawValue": raw,
                        "issue": "service_focus_code_mismatch_with_row_focus",
                        "expectedFocusCode": focus_code,
                        "actualFocusCode": capability_focus_code,
                    }
                )
            services.append(
                {
                    "id": None,
                    "type": "security_technical_service",
                    "code": code,
                    "title": title,
                    "description": None,
                    "category": code_scope,
                    "metadata": {
                        "scope_code": code_scope,
                        "capability_focus_code": capability_focus_code,
                        "capability_category": last_category or None,
                        "capability_domain": last_domain or None,
                        "capability": last_capability or None,
                        "source_sheet": SERVICE_SHEET,
                        "source_row": row,
                        "source_cell": ws.cell(row, col).coordinate,
                    },
                    "source": {
                        "sheet": SERVICE_SHEET,
                        "row": row,
                        "column": f"作用域列{col}",
                        "cell": ws.cell(row, col).coordinate,
                        "rawValue": raw,
                        "mergedRange": merged.get(ws.cell(row, col).coordinate),
                    },
                }
            )
    return services, list(merged.values()), invalid_rows, candidates


def load_current_services() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    services_data = load_json(ROOT / "frontend/capability-browser/public/data/maintenance/services.json")
    services = []
    for entry in services_data.get("security_technical_services", []):
        service = dict(entry.get("service") if isinstance(entry.get("service"), dict) else entry)
        services.append(service)
    return services, services_data.get("security_technical_services", [])


def load_sqlite_services() -> list[dict[str, Any]]:
    if not DB_PATH.exists():
        return []
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    try:
        return [dict(row) for row in con.execute("select * from knowledge_items where type='security_technical_service'")]
    finally:
        con.close()


def index_by(items: list[dict[str, Any]], key: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in items:
        value = text(item.get(key))
        if value:
            result[value] = item
    return result


def duplicates(items: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        value = text(item.get(key))
        if value:
            grouped[value].append(item)
    return [
        {
            key: value,
            "count": len(rows),
            "services": [{"code": row.get("code"), "title": row.get("title"), "id": row.get("id")} for row in rows],
        }
        for value, rows in sorted(grouped.items())
        if len(rows) > 1
    ]


def build_mapping(source_services: list[dict[str, Any]], current_services: list[dict[str, Any]]) -> dict[str, Any]:
    source_by_code = index_by(source_services, "code")
    current_by_code = index_by(current_services, "code")
    source_by_title: dict[str, list[dict[str, Any]]] = defaultdict(list)
    current_by_title: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in source_services:
        source_by_title[text(item.get("title"))].append(item)
    for item in current_services:
        current_by_title[text(item.get("title"))].append(item)

    removed_codes = sorted(set(current_by_code) - set(source_by_code))
    added_codes = sorted(set(source_by_code) - set(current_by_code))
    code_changes = []
    consumed_added: set[str] = set()
    for old_code in removed_codes:
        current = current_by_code[old_code]
        title = text(current.get("title"))
        candidates = [item for item in source_by_title.get(title, []) if item["code"] in added_codes]
        if len(candidates) == 1:
            new = candidates[0]
            code_changes.append(
                {
                    "id": current.get("id"),
                    "title": title,
                    "oldCode": old_code,
                    "newCode": new["code"],
                    "oldCategory": current.get("category"),
                    "newCategory": new.get("category"),
                    "source": new.get("source"),
                }
            )
            consumed_added.add(new["code"])

    added_services = [
        source_by_code[code]
        for code in added_codes
        if code not in consumed_added
    ]
    removed_services = [
        current_by_code[code]
        for code in removed_codes
        if code not in {change["oldCode"] for change in code_changes}
    ]
    renamed = []
    category_changed = []
    for code in sorted(set(source_by_code) & set(current_by_code)):
        source = source_by_code[code]
        current = current_by_code[code]
        if text(source.get("title")) != text(current.get("title")):
            renamed.append(
                {
                    "id": current.get("id"),
                    "code": code,
                    "oldTitle": current.get("title"),
                    "newTitle": source.get("title"),
                    "source": source.get("source"),
                }
            )
        if text(source.get("category")) != text(current.get("category")):
            category_changed.append(
                {
                    "id": current.get("id"),
                    "code": code,
                    "title": source.get("title"),
                    "oldCategory": current.get("category"),
                    "newCategory": source.get("category"),
                    "source": source.get("source"),
                }
            )

    return {
        "addedServices": added_services,
        "removedServices": removed_services,
        "renamedServicesBySameId": renamed,
        "codeChangedServicesBySameName": code_changes,
        "definitionChangedServices": [],
        "categoryChangedServices": category_changed,
        "duplicateServiceIds": duplicates(current_services, "id"),
        "duplicateServiceNames": duplicates(source_services, "title"),
    }


def source_candidate_services(source_services: list[dict[str, Any]], current_services: list[dict[str, Any]], mapping: dict[str, Any]) -> list[dict[str, Any]]:
    current_by_code = index_by(current_services, "code")
    id_by_new_code = {change["newCode"]: change["id"] for change in mapping.get("codeChangedServicesBySameName", [])}
    candidate = []
    for service in source_services:
        item = dict(service)
        code = text(item.get("code"))
        if code in current_by_code:
            item["id"] = current_by_code[code].get("id")
        elif code in id_by_new_code:
            item["id"] = id_by_new_code[code]
        else:
            item["id"] = service_id_for_code(code)
        candidate.append(item)
    return candidate


def choose_canonical_for_reference(ref: dict[str, Any], candidate_by_code: dict[str, dict[str, Any]], candidate_by_title: dict[str, list[dict[str, Any]]]) -> tuple[dict[str, Any] | None, str, list[dict[str, Any]]]:
    code = text(ref.get("code"))
    title = text(ref.get("title") or ref.get("name"))
    if code in candidate_by_code:
        candidate = candidate_by_code[code]
        if title and title != text(candidate.get("title")):
            return candidate, "code_matched_title_stale", []
        return candidate, "code_matched", []
    title_matches = candidate_by_title.get(title, [])
    if title and len(title_matches) == 1:
        candidate = title_matches[0]
        if code and candidate.get("code") != code:
            return candidate, "title_unique_code_changed", []
    if title and len(title_matches) > 1:
        return None, "ambiguous_title", title_matches
    if title_matches:
        return title_matches[0], "title_matched", []
    return None, "unknown", []


def scan_json_references(candidate_services: list[dict[str, Any]]) -> dict[str, Any]:
    candidate_by_code = index_by(candidate_services, "code")
    candidate_by_title: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in candidate_services:
        candidate_by_title[text(item.get("title"))].append(item)
    current_services, _ = load_current_services()
    current_by_code = index_by(current_services, "code")
    known_codes = set(candidate_by_code) | set(current_by_code)

    references_by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    references_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    stale = []
    unknown = []
    ambiguous = []
    packages_to_update: set[str] = set()
    packages_no_change: set[str] = set()

    skip_keys = {"source", "sources", "sourceCells", "evidenceRefs", "sourceReferences", "rawValue", "raw_value"}

    def walk(node: Any, package: Path, path: str, skip: bool = False) -> None:
        nonlocal packages_to_update
        if isinstance(node, dict):
            if node.get("type") == "security_technical_service" and not skip:
                ref = {"id": node.get("id"), "code": node.get("code"), "title": node.get("title") or node.get("name")}
                canonical, status, choices = choose_canonical_for_reference(ref, candidate_by_code, candidate_by_title)
                entry = {
                    "package": str(package.relative_to(ROOT)),
                    "jsonPath": path,
                    "id": ref.get("id"),
                    "code": ref.get("code"),
                    "title": ref.get("title"),
                    "status": status,
                    "canonical": {
                        "id": canonical.get("id"),
                        "code": canonical.get("code"),
                        "title": canonical.get("title"),
                        "category": canonical.get("category"),
                    }
                    if canonical
                    else None,
                }
                if ref.get("id"):
                    references_by_id[str(ref["id"])].append(entry)
                if ref.get("title"):
                    references_by_name[str(ref["title"])].append(entry)
                if status in {"title_unique_code_changed", "code_matched_title_stale"}:
                    stale.append(entry)
                    packages_to_update.add(str(package.relative_to(ROOT)))
                elif status == "unknown":
                    unknown.append(entry)
                elif status == "ambiguous_title":
                    entry["choices"] = [
                        {"id": item.get("id"), "code": item.get("code"), "title": item.get("title")}
                        for item in choices
                    ]
                    ambiguous.append(entry)
            for key, value in node.items():
                walk(value, package, f"{path}.{key}" if path else key, skip or key in skip_keys)
        elif isinstance(node, list):
            for index, item in enumerate(node):
                walk(item, package, f"{path}[{index}]", skip)
        elif isinstance(node, str) and not skip:
            if node in candidate_by_title or node in known_codes:
                references_by_name[node].append({"package": str(package.relative_to(ROOT)), "jsonPath": path, "value": node, "status": "string_reference"})

    for package in FORMAL_PACKAGES:
        if not package.exists():
            continue
        payload = load_json(package)
        before_stale = len(stale)
        walk(payload, package, "")
        if len(stale) == before_stale:
            packages_no_change.add(str(package.relative_to(ROOT)))

    return {
        "referencesByServiceId": dict(references_by_id),
        "referencesByServiceName": dict(references_by_name),
        "staleReferences": stale,
        "unknownServiceReferences": unknown,
        "ambiguousReferences": ambiguous,
        "packagesToUpdate": sorted(packages_to_update),
        "packagesNoChange": sorted(packages_no_change - packages_to_update),
    }


def sqlite_diff(source_services: list[dict[str, Any]], candidate_services: list[dict[str, Any]]) -> dict[str, Any]:
    sqlite_services = load_sqlite_services()
    sqlite_by_code = index_by(sqlite_services, "code")
    candidate_by_code = index_by(candidate_services, "code")
    return {
        "sqliteServiceCount": len(sqlite_services),
        "candidateServiceCount": len(candidate_services),
        "missingInSqlite": [
            {"code": code, "title": item.get("title")}
            for code, item in sorted(candidate_by_code.items())
            if code not in sqlite_by_code
        ],
        "extraInSqlite": [
            {"code": code, "title": item.get("title"), "id": item.get("id")}
            for code, item in sorted(sqlite_by_code.items())
            if code not in candidate_by_code
        ],
        "titleChangedInSqlite": [
            {
                "code": code,
                "sqliteTitle": sqlite_by_code[code].get("title"),
                "candidateTitle": item.get("title"),
                "id": sqlite_by_code[code].get("id"),
            }
            for code, item in sorted(candidate_by_code.items())
            if code in sqlite_by_code and text(sqlite_by_code[code].get("title")) != text(item.get("title"))
        ],
    }


def render_diff_md(payload: dict[str, Any]) -> str:
    diff = payload["diff"]
    lines = ["# 安全技术服务字典 diff", "", "## 摘要", ""]
    for key in [
        "sourceServiceCount",
        "currentJsonServiceCount",
        "currentSqliteServiceCount",
        "addedCount",
        "removedCount",
        "renamedCount",
        "codeChangedCount",
        "categoryChangedCount",
        "requiresUserConfirmationCount",
    ]:
        lines.append(f"- `{key}`: {payload['summary'].get(key)}")
    for key, label in [
        ("addedServices", "新增服务"),
        ("removedServices", "移除服务"),
        ("renamedServicesBySameId", "同 code 改名"),
        ("codeChangedServicesBySameName", "同名改 code"),
        ("categoryChangedServices", "作用域 / 分类变化"),
    ]:
        lines.extend(["", f"## {label}", ""])
        rows = diff.get(key, [])
        if not rows:
            lines.append("- 无")
        for row in rows:
            lines.append(f"- `{row.get('code') or row.get('oldCode')}` {row.get('title') or row.get('oldTitle', '')} -> {row.get('newCode') or row.get('newTitle') or row.get('code')}")
    return "\n".join(lines)


def render_impact_md(payload: dict[str, Any]) -> str:
    lines = ["# 安全技术服务引用影响分析", "", "## 摘要", ""]
    for key, value in payload["summary"].items():
        lines.append(f"- `{key}`: {value}")
    lines.extend(["", "## 需要同步的数据包", ""])
    if payload["packagesToUpdate"]:
        for package in payload["packagesToUpdate"]:
            lines.append(f"- `{package}`")
    else:
        lines.append("- 无")
    lines.extend(["", "## stale references", ""])
    if payload["staleReferences"]:
        for ref in payload["staleReferences"][:80]:
            c = ref.get("canonical") or {}
            lines.append(f"- `{ref['package']}` `{ref['jsonPath']}`: `{ref.get('code')} {ref.get('title')}` -> `{c.get('code')} {c.get('title')}`")
    else:
        lines.append("- 无")
    return "\n".join(lines)


def audit(post_apply: bool = False) -> dict[str, Any]:
    source_services, merged_ranges, invalid_rows, candidate_sheets = parse_source_services()
    current_services, _raw_service_entries = load_current_services()
    sqlite_services = load_sqlite_services()
    mapping = build_mapping(source_services, current_services)
    candidate_services = source_candidate_services(source_services, current_services, mapping)
    impact = scan_json_references(candidate_services)
    sqlite_plan = sqlite_diff(source_services, candidate_services)

    requires_confirmation = []
    if mapping["removedServices"]:
        requires_confirmation.extend({"type": "removed_service", **item} for item in mapping["removedServices"])
    if mapping["duplicateServiceIds"] or mapping["duplicateServiceNames"]:
        requires_confirmation.append({"type": "duplicate_services", "duplicateServiceIds": mapping["duplicateServiceIds"], "duplicateServiceNames": mapping["duplicateServiceNames"]})
    if invalid_rows:
        requires_confirmation.extend({"type": "invalid_source_row", **item} for item in invalid_rows)
    if impact["ambiguousReferences"] or impact["unknownServiceReferences"]:
        requires_confirmation.append({"type": "reference_resolution_issue", "ambiguousCount": len(impact["ambiguousReferences"]), "unknownCount": len(impact["unknownServiceReferences"])})

    summary = {
        "sourceServiceCount": len(source_services),
        "currentJsonServiceCount": len(current_services),
        "currentSqliteServiceCount": len(sqlite_services),
        "addedCount": len(mapping["addedServices"]),
        "removedCount": len(mapping["removedServices"]),
        "renamedCount": len(mapping["renamedServicesBySameId"]),
        "codeChangedCount": len(mapping["codeChangedServicesBySameName"]),
        "categoryChangedCount": len(mapping["categoryChangedServices"]),
        "requiresUserConfirmationCount": len(requires_confirmation),
    }
    generated_at = now_iso()
    diff_payload = {
        "version": 1,
        "generatedAt": generated_at,
        "postApply": post_apply,
        "sourceWorkbook": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sourceSheet": SERVICE_SHEET if source_services else None,
        "candidateServiceSheets": candidate_sheets,
        "summary": summary,
        "diff": {
            **mapping,
            "invalidRows": invalid_rows,
            "mergedRangeIssues": [],
            "requiresUserConfirmation": requires_confirmation,
        },
        "sourceMergedRanges": merged_ranges,
        "sourceServices": source_services,
        "candidateServices": candidate_services,
        "sqliteDiff": sqlite_plan,
    }
    impact_payload = {
        "version": 1,
        "generatedAt": generated_at,
        "summary": {
            "staleReferenceCount": len(impact["staleReferences"]),
            "unknownServiceReferenceCount": len(impact["unknownServiceReferences"]),
            "ambiguousReferenceCount": len(impact["ambiguousReferences"]),
            "packagesToUpdateCount": len(impact["packagesToUpdate"]),
        },
        **impact,
    }
    write_json(OUT_DIR / "security-technical-service-dictionary-diff.json", diff_payload)
    write_md(OUT_DIR / "security-technical-service-dictionary-diff.md", render_diff_md(diff_payload))
    write_json(OUT_DIR / "security-technical-service-reference-impact.json", impact_payload)
    write_md(OUT_DIR / "security-technical-service-reference-impact.md", render_impact_md(impact_payload))
    return {
        "status": "needs_user_confirmation" if requires_confirmation else "ready_for_candidate",
        "outputDir": str(OUT_DIR.relative_to(ROOT)),
        **summary,
        "referenceImpact": impact_payload["summary"],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--post-apply", action="store_true")
    args = parser.parse_args()
    print(json.dumps(audit(post_apply=args.post_apply), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
