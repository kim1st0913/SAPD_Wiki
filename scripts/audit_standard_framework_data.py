#!/usr/bin/env python3
"""Audit standard/framework workbook sheets against split frontend data packages."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = PROJECT_ROOT / "data/raw-samples/wiki sample.xlsx"
DEFAULT_STANDARDS_DIR = PROJECT_ROOT / "frontend/capability-browser/public/data/standards"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data/processed/reviews"
DEFAULT_DATE = datetime.now().strftime("%Y%m%d")

FOCUS_CODE_RE = re.compile(r"\b[TMG]-[A-Z]{2}(?:&[A-Z]{2})?\.[A-Z]{2}-\d{2}\b")
# Do not require a trailing \b: Python treats Chinese characters as word
# characters, so values such as "8.1.1.1物理位置" would otherwise be truncated.
NUMBERED_ID_RE = re.compile(r"(?<![\w.])\d+(?:\.\d+)+(?:[a-z])?", flags=re.IGNORECASE)
CSF_ID_RE = re.compile(r"\b[A-Z]{2}\.[A-Z]{2}-\d{2}\b")
CRF_ID_RE = re.compile(r"\b[A-Z]{2,4}-\d{2}\b")
NIST_ID_RE = re.compile(r"\b[A-Z]{2}-\d+(?:\(\d+\))?")
DSP_ID_RE = re.compile(r"\b[A-Z]{2,4}-\d+(?:\.\d+)?\b")

NIST_SECURITY_TYPE_KEY = "安全类型（O=组织层面控制，S=系统层面控制，O/S=组织和系统均涉及）"


STANDARD_PACKAGES = {
    "GB-T-22239-2019-L3": {
        "label": "等级保护三级",
        "sheet": "等保三级测评清单",
        "json": ["mlps-level-3.json"],
        "id_field": "等保三级控制要求",
        "id_regex": NUMBERED_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
    "CIS-CSC-V8.1.2": {
        "label": "CIS CSC v8",
        "sheet": "CIS CSC V8",
        "json": ["cis-csc-v8.json"],
        "id_field": "保护措施编号",
        "id_regex": NUMBERED_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
    "NIST-CSF-2.0": {
        "label": "NIST CSF 2.0",
        "sheet": "CSF2.0",
        "json": ["nist-csf-2/csf-core.json"],
        "tier_json": ["nist-csf-2/csf-tiers.json"],
        "id_field": "分类标识符说明",
        "id_regex": CSF_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
    "ISO-IEC-27001-2022": {
        "label": "ISO 27001:2022",
        "sheet": "27001-2022",
        "json": ["iso-27001-2022.json"],
        "id_field": "控制编号",
        "id_regex": NUMBERED_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
    "DSP-SCF-2026": {
        "label": "DSP SCF 2026",
        "sheet": "DSP策略清单（2026）",
        "json": ["dsp-level-2/dsp-scf-controls-2026.json"],
        "maturity_json": ["dsp-level-2/dsp-scf-maturity-2026.json"],
        "id_field": "SCF编号",
        "id_regex": DSP_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
    "CRF-SAFEGUARDS-CORE-2026": {
        "label": "CRF Safeguards Core 2026",
        "sheet": "CRF Safeguards Core 2026",
        "json": ["crf/crf-safeguards-core-2026.json"],
        "id_field": "Safeguard ID",
        "id_regex": CRF_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
    "CRF-MATURITY-MODEL-2026": {
        "label": "CRF Maturity Model 2026",
        "sheet": "CRF Maturity Model 2026",
        "json": ["crf/crf-maturity-model-2026.json"],
        "id_field": "等级编号",
        "id_regex": re.compile(r"\bLevel\s+\d+\b", flags=re.IGNORECASE),
        "related_field": "",
    },
    "NIST-800-53-REV5": {
        "label": "NIST SP 800-53 Rev.5",
        "sheet": "NIST 800-53rev5",
        "json": ["nist-800-53-rev5.json"],
        "id_field": "安全策略编号",
        "id_regex": NIST_ID_RE,
        "related_field": "关联安全能力/关注点",
    },
}

MAPPING_COLUMNS = (
    ("ISO-IEC-27001-2022", 7, "ISO 27001:2022"),
    ("NIST-CSF-2.0", 8, "CSF 2.0"),
    ("GB-T-22239-2019-L3", 9, "等级保护3级通用要求"),
    ("CIS-CSC-V8.1.2", 11, "CIS CSC V8"),
    ("CRF-SAFEGUARDS-CORE-2026", 12, "CRF"),
    ("NIST-800-53-REV5", 14, "NIST 800-53 rev5"),
)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value)).strip()


def cell(ws: Any, row: int, col: int) -> str:
    return compact(ws.cell(row, col).value)


def extract_first(pattern: re.Pattern[str], value: Any) -> str:
    match = pattern.search(text(value))
    return match.group(0) if match else ""


def extract_focus_codes(value: Any) -> list[str]:
    return list(dict.fromkeys(FOCUS_CODE_RE.findall(text(value))))


def extract_mapping_ids(framework_code: str, value: Any) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    if framework_code == "NIST-CSF-2.0":
        matches = CSF_ID_RE.findall(raw)
    elif framework_code == "CRF-SAFEGUARDS-CORE-2026":
        matches = CRF_ID_RE.findall(raw)
    elif framework_code == "NIST-800-53-REV5":
        matches = NIST_ID_RE.findall(raw)
    else:
        matches = NUMBERED_ID_RE.findall(raw)
    return list(dict.fromkeys(matches))


def normalize_mapping_control_id(framework_code: str, control_id: str) -> str:
    normalized = text(control_id)
    if framework_code == "GB-T-22239-2019-L3":
        return re.sub(r"[a-z]$", "", normalized, flags=re.IGNORECASE)
    return normalized


def issue(
    issues: list[dict[str, Any]],
    severity: str,
    framework: str,
    issue_type: str,
    message: str,
    *,
    source: str = "",
    row: int | None = None,
    control_id: str = "",
    field: str = "",
    expected: Any = "",
    actual: Any = "",
) -> None:
    issues.append(
        {
            "severity": severity,
            "framework": framework,
            "issue_type": issue_type,
            "source": source,
            "row": row or "",
            "control_id": control_id,
            "field": field,
            "expected": expected,
            "actual": actual,
            "message": message,
        }
    )


def duplicate_values(values: Iterable[str]) -> list[str]:
    counts = Counter(v for v in values if v)
    return sorted([value for value, count in counts.items() if count > 1])


def source_record(framework: str, row: int, control_id: str, fields: dict[str, str], related: str = "") -> dict[str, Any]:
    return {
        "framework_code": framework,
        "source_row": row,
        "control_id": control_id,
        "fields": fields,
        "related_focus_codes": extract_focus_codes(related),
        "related_raw": text(related),
    }


def parse_source_records(workbook_path: Path, issues: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    records: dict[str, list[dict[str, Any]]] = defaultdict(list)

    expected_sheets = {config["sheet"] for config in STANDARD_PACKAGES.values()}
    expected_sheets.add("安全能力-网络安全制度、框架映射")
    for sheet in sorted(expected_sheets):
        if sheet not in wb.sheetnames:
            issue(issues, "error", "ALL", "missing_sheet", f"原始 workbook 缺少 Sheet：{sheet}", source=str(workbook_path))

    if "等保三级测评清单" in wb.sheetnames:
        ws = wb["等保三级测评清单"]
        for row in range(3, ws.max_row + 1):
            requirement_text = cell(ws, row, 5)
            if not requirement_text:
                continue
            control_id = extract_first(NUMBERED_ID_RE, requirement_text)
            rec = source_record(
                "GB-T-22239-2019-L3",
                row,
                control_id,
                {
                    "等级保护": cell(ws, row, 2),
                    "等保要求": cell(ws, row, 3),
                    "等保控制项": cell(ws, row, 4),
                    "等保三级控制要求": requirement_text,
                },
                cell(ws, row, 7),
            )
            records["GB-T-22239-2019-L3"].append(rec)
            if not control_id:
                issue(issues, "error", "GB-T-22239-2019-L3", "missing_control_id", "等保要求文本未解析出控制编号", source=ws.title, row=row, field="等保三级控制要求", actual=requirement_text[:120])
            for field in ("等级保护", "等保要求", "等保控制项"):
                if not rec["fields"][field]:
                    issue(issues, "warn", "GB-T-22239-2019-L3", "missing_inherited_context", f"等保数据行缺少层级字段：{field}", source=ws.title, row=row, control_id=control_id, field=field)

    if "CIS CSC V8" in wb.sheetnames:
        ws = wb["CIS CSC V8"]
        current_control_id = current_control_name = current_control_description = ""
        for row in range(3, ws.max_row + 1):
            if cell(ws, row, 2):
                current_control_id = cell(ws, row, 2)
                current_control_name = cell(ws, row, 3)
                current_control_description = cell(ws, row, 4)
            safeguard_id = cell(ws, row, 5)
            if not safeguard_id:
                continue
            rec = source_record(
                "CIS-CSC-V8.1.2",
                row,
                safeguard_id,
                {
                    "安全控制项": current_control_id,
                    "安全控制项名称": current_control_name,
                    "控制项描述": current_control_description,
                    "保护措施编号": safeguard_id,
                    "名称": cell(ws, row, 6),
                    "资产类型": cell(ws, row, 7),
                    "实施组": cell(ws, row, 8),
                    "安全功能": cell(ws, row, 9),
                    "描述": cell(ws, row, 10),
                },
                cell(ws, row, 11),
            )
            records["CIS-CSC-V8.1.2"].append(rec)
            if not NUMBERED_ID_RE.fullmatch(safeguard_id):
                issue(issues, "error", "CIS-CSC-V8.1.2", "invalid_control_id", "CIS 保护措施编号格式异常", source=ws.title, row=row, control_id=safeguard_id, field="保护措施编号", actual=safeguard_id)
            for field in ("安全控制项", "安全控制项名称", "控制项描述"):
                if not rec["fields"][field]:
                    issue(issues, "error", "CIS-CSC-V8.1.2", "inheritance_missing", f"CIS 保护措施缺少继承字段：{field}", source=ws.title, row=row, control_id=safeguard_id, field=field)

    if "CSF2.0" in wb.sheetnames:
        ws = wb["CSF2.0"]
        current_function = current_category = current_category_id = ""
        for row in range(3, min(ws.max_row, 108) + 1):
            if cell(ws, row, 2):
                current_function = cell(ws, row, 2)
            if cell(ws, row, 3):
                current_category = cell(ws, row, 3)
            if cell(ws, row, 4):
                current_category_id = cell(ws, row, 4)
            subcategory_text = cell(ws, row, 5)
            if not subcategory_text:
                continue
            control_id = extract_first(CSF_ID_RE, subcategory_text)
            rec = source_record(
                "NIST-CSF-2.0",
                row,
                control_id,
                {
                    "功能": current_function,
                    "分类": current_category,
                    "分类标识符": current_category_id,
                    "分类标识符说明": subcategory_text,
                },
                cell(ws, row, 6),
            )
            records["NIST-CSF-2.0"].append(rec)
            if not control_id:
                issue(issues, "error", "NIST-CSF-2.0", "missing_control_id", "CSF 子类说明未解析出编号", source=ws.title, row=row, field="分类标识符说明", actual=subcategory_text[:120])
            for field in ("功能", "分类", "分类标识符"):
                if not rec["fields"][field]:
                    issue(issues, "error", "NIST-CSF-2.0", "inheritance_missing", f"CSF 子类缺少继承字段：{field}", source=ws.title, row=row, control_id=control_id, field=field)

        for row in range(112, min(ws.max_row, 115) + 1):
            tier = cell(ws, row, 2)
            if tier:
                records["NIST-CSF-2.0:tiers"].append(source_record("NIST-CSF-2.0:tiers", row, tier, {"层级": tier, "治理": cell(ws, row, 3), "风险管理": cell(ws, row, 4)}))

    if "27001-2022" in wb.sheetnames:
        ws = wb["27001-2022"]
        current_category = ""
        for row in range(4, ws.max_row + 1):
            if cell(ws, row, 2):
                current_category = cell(ws, row, 2)
            control_id = cell(ws, row, 3)
            if not control_id:
                continue
            rec = source_record(
                "ISO-IEC-27001-2022",
                row,
                control_id,
                {
                    "控制类别": current_category,
                    "控制编号": control_id,
                    "控制名称": cell(ws, row, 4),
                    "控制描述": cell(ws, row, 5),
                    "控制类型": cell(ws, row, 6),
                    "信息安全特性": cell(ws, row, 7),
                    "网络安全概念": cell(ws, row, 8),
                    "运营能力": cell(ws, row, 9),
                    "安全域": cell(ws, row, 10),
                },
                cell(ws, row, 11),
            )
            records["ISO-IEC-27001-2022"].append(rec)
            if not NUMBERED_ID_RE.fullmatch(control_id):
                issue(issues, "error", "ISO-IEC-27001-2022", "invalid_control_id", "ISO 控制编号格式异常", source=ws.title, row=row, control_id=control_id, field="控制编号", actual=control_id)
            if not current_category:
                issue(issues, "error", "ISO-IEC-27001-2022", "inheritance_missing", "ISO 控制项缺少控制类别继承", source=ws.title, row=row, control_id=control_id, field="控制类别")

    if "DSP策略清单（2026）" in wb.sheetnames:
        ws = wb["DSP策略清单（2026）"]
        current_domain = current_principle = current_intent = ""
        for row in range(3, ws.max_row + 1):
            if cell(ws, row, 2):
                current_domain = cell(ws, row, 2)
            if cell(ws, row, 3):
                current_principle = cell(ws, row, 3)
            if cell(ws, row, 4):
                current_intent = cell(ws, row, 4)
            control_id = cell(ws, row, 5)
            if not control_id:
                continue
            rec = source_record(
                "DSP-SCF-2026",
                row,
                control_id,
                {
                    "SCF域": current_domain,
                    "策略原则": current_principle,
                    "策略意图": current_intent,
                    "SCF编号": control_id,
                    "SCF控制项": cell(ws, row, 6),
                    "SCF控制项描述": cell(ws, row, 7),
                    "安全策略项": cell(ws, row, 8),
                    "NIST CSF功能分组": cell(ws, row, 9),
                    "SCR-CMM 0级 未执行": cell(ws, row, 10),
                    "SCR-CMM 1级 非正式执行": cell(ws, row, 11),
                    "SCR-CMM 2级 已计划并跟踪": cell(ws, row, 12),
                    "SCR-CMM 3级 定义良好": cell(ws, row, 13),
                    "SCR-CMM 4级 量化控制": cell(ws, row, 14),
                    "SCR-CMM 5级 持续改进": cell(ws, row, 15),
                },
                "",
            )
            records["DSP-SCF-2026"].append(rec)
            if not DSP_ID_RE.fullmatch(control_id):
                issue(issues, "error", "DSP-SCF-2026", "invalid_control_id", "DSP SCF 编号格式异常", source=ws.title, row=row, control_id=control_id, field="SCF编号", actual=control_id)
            for field in ("SCF域", "策略原则", "策略意图"):
                if not rec["fields"][field]:
                    issue(issues, "error", "DSP-SCF-2026", "inheritance_missing", f"DSP 控制项缺少继承字段：{field}", source=ws.title, row=row, control_id=control_id, field=field)

    if "CRF Safeguards Core 2026" in wb.sheetnames:
        ws = wb["CRF Safeguards Core 2026"]
        for row in range(2, ws.max_row + 1):
            control_id = cell(ws, row, 4)
            if not control_id:
                continue
            rec = source_record(
                "CRF-SAFEGUARDS-CORE-2026",
                row,
                control_id,
                {
                    "保障措施分类": cell(ws, row, 1),
                    "保障措施域": cell(ws, row, 2),
                    "CRF成熟度等级": cell(ws, row, 3),
                    "Safeguard ID": control_id,
                    "保障措施描述": cell(ws, row, 5),
                    "保障措施系统": cell(ws, row, 6),
                },
                cell(ws, row, 7),
            )
            records["CRF-SAFEGUARDS-CORE-2026"].append(rec)
            if not CRF_ID_RE.fullmatch(control_id):
                issue(issues, "error", "CRF-SAFEGUARDS-CORE-2026", "invalid_control_id", "CRF Safeguard ID 格式异常", source=ws.title, row=row, control_id=control_id, field="Safeguard ID", actual=control_id)

    if "CRF Maturity Model 2026" in wb.sheetnames:
        ws = wb["CRF Maturity Model 2026"]
        for row in range(2, ws.max_row + 1):
            level_id = cell(ws, row, 1)
            if not level_id:
                continue
            records["CRF-MATURITY-MODEL-2026"].append(
                source_record(
                    "CRF-MATURITY-MODEL-2026",
                    row,
                    level_id,
                    {
                        "等级编号": level_id,
                        "成熟度等级": cell(ws, row, 2),
                        "英文等级": cell(ws, row, 3),
                        "等级定义": cell(ws, row, 4),
                        "高层特征": cell(ws, row, 5),
                        "边界说明": cell(ws, row, 6),
                    },
                )
            )

    if "NIST 800-53rev5" in wb.sheetnames:
        ws = wb["NIST 800-53rev5"]
        current_family = ""
        for row in range(5, ws.max_row + 1):
            if cell(ws, row, 2):
                current_family = cell(ws, row, 2)
            control_id = cell(ws, row, 3)
            if not control_id:
                continue
            rec = source_record(
                "NIST-800-53-REV5",
                row,
                control_id,
                {
                    "安全控制类": current_family,
                    "安全策略编号": control_id,
                    "英文名称": cell(ws, row, 4),
                    "安全级别": cell(ws, row, 5),
                    "安全类型": cell(ws, row, 6),
                    "中文名称": cell(ws, row, 7),
                    "控制描述": cell(ws, row, 8),
                },
                cell(ws, row, 9),
            )
            records["NIST-800-53-REV5"].append(rec)
            if not NIST_ID_RE.fullmatch(control_id):
                issue(issues, "error", "NIST-800-53-REV5", "invalid_control_id", "NIST 安全策略编号格式异常", source=ws.title, row=row, control_id=control_id, field="安全策略编号", actual=control_id)
            if not current_family:
                issue(issues, "error", "NIST-800-53-REV5", "inheritance_missing", "NIST 控制项缺少安全控制类继承", source=ws.title, row=row, control_id=control_id, field="安全控制类")
            if "(" in control_id and ")" in control_id:
                base_id = control_id.split("(", 1)[0]
                rec["base_control_id"] = base_id

    wb.close()

    nist_ids = {rec["control_id"] for rec in records.get("NIST-800-53-REV5", [])}
    for rec in records.get("NIST-800-53-REV5", []):
        base_id = rec.get("base_control_id")
        if base_id and base_id not in nist_ids:
            issue(issues, "error", "NIST-800-53-REV5", "missing_base_control", "NIST 3级策略项找不到对应 2级基础策略项", source="NIST 800-53rev5", row=rec["source_row"], control_id=rec["control_id"], expected=base_id)

    return records


def load_json_rows(standards_dir: Path, rel_path: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    path = standards_dir / rel_path
    data = json.loads(path.read_text(encoding="utf-8"))
    return data, data.get("rows") or []


def json_control_id(framework_code: str, row: dict[str, Any], config: dict[str, Any]) -> str:
    field = config["id_field"]
    raw = row.get(field, "")
    if framework_code == "NIST-CSF-2.0":
        return extract_first(CSF_ID_RE, raw)
    if framework_code == "GB-T-22239-2019-L3":
        return extract_first(NUMBERED_ID_RE, raw)
    return text(raw)


def load_package_records(standards_dir: Path, issues: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    records: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for framework_code, config in STANDARD_PACKAGES.items():
        for rel_path in config["json"]:
            path = standards_dir / rel_path
            if not path.exists():
                issue(issues, "error", framework_code, "missing_json", f"拆包 JSON 不存在：{rel_path}", source=str(path))
                continue
            data, rows = load_json_rows(standards_dir, rel_path)
            if data.get("data_state") != "ready":
                issue(issues, "error", framework_code, "data_state_not_ready", "拆包 JSON data_state 不是 ready", source=rel_path, expected="ready", actual=data.get("data_state"))
            for idx, row in enumerate(rows, start=1):
                control_id = json_control_id(framework_code, row, config)
                related_field = config.get("related_field") or ""
                records[framework_code].append(
                    {
                        "framework_code": framework_code,
                        "json_path": rel_path,
                        "json_row": idx,
                        "control_id": control_id,
                        "row": row,
                        "related_focus_codes": extract_focus_codes(row.get(related_field, "")) if related_field else [],
                        "related_raw": text(row.get(related_field, "")) if related_field else "",
                    }
                )
                if not control_id:
                    issue(issues, "error", framework_code, "missing_control_id", "JSON 行未解析出控制编号", source=rel_path, row=idx, field=config["id_field"])
                if "\n" in text(row.get(config["id_field"], "")) or "\r" in text(row.get(config["id_field"], "")):
                    issue(issues, "warn", framework_code, "linebreak_in_identifier_field", "编号字段中存在换行，需确认是否为错误换行", source=rel_path, row=idx, control_id=control_id, field=config["id_field"], actual=row.get(config["id_field"], ""))

        for group_key in ("tier_json", "maturity_json"):
            for rel_path in config.get(group_key, []):
                path = standards_dir / rel_path
                if not path.exists():
                    issue(issues, "error", framework_code, "missing_json", f"拆包 JSON 不存在：{rel_path}", source=str(path))
    return records


def compare_records(
    source_records: dict[str, list[dict[str, Any]]],
    package_records: dict[str, list[dict[str, Any]]],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    summary: dict[str, Any] = {}
    for framework_code, config in STANDARD_PACKAGES.items():
        src = source_records.get(framework_code, [])
        pkg = package_records.get(framework_code, [])
        src_ids = [rec["control_id"] for rec in src if rec["control_id"]]
        pkg_ids = [rec["control_id"] for rec in pkg if rec["control_id"]]
        src_set = set(src_ids)
        pkg_set = set(pkg_ids)

        for control_id in duplicate_values(src_ids):
            issue(issues, "error", framework_code, "duplicate_source_id", "原始 Sheet 控制编号重复", source=config["sheet"], control_id=control_id)
        for control_id in duplicate_values(pkg_ids):
            issue(issues, "error", framework_code, "duplicate_json_id", "拆包 JSON 控制编号重复", control_id=control_id)

        for control_id in sorted(src_set - pkg_set):
            source_row = next((rec["source_row"] for rec in src if rec["control_id"] == control_id), "")
            issue(issues, "error", framework_code, "missing_in_json", "原始 Sheet 有记录，但拆包 JSON 缺失", source=config["sheet"], row=source_row, control_id=control_id)
        for control_id in sorted(pkg_set - src_set):
            json_row = next((rec["json_row"] for rec in pkg if rec["control_id"] == control_id), "")
            issue(issues, "error", framework_code, "extra_in_json", "拆包 JSON 有记录，但原始 Sheet 未找到", row=json_row, control_id=control_id)

        if len(src) != len(pkg):
            issue(issues, "error", framework_code, "row_count_mismatch", "原始 Sheet 与拆包 JSON 行数不一致", expected=len(src), actual=len(pkg))

        summary[framework_code] = {
            "label": config["label"],
            "source_rows": len(src),
            "json_rows": len(pkg),
            "source_unique_ids": len(src_set),
            "json_unique_ids": len(pkg_set),
            "missing_in_json": len(src_set - pkg_set),
            "extra_in_json": len(pkg_set - src_set),
            "duplicate_source_ids": len(duplicate_values(src_ids)),
            "duplicate_json_ids": len(duplicate_values(pkg_ids)),
        }
    return summary


def audit_required_fields(
    source_records: dict[str, list[dict[str, Any]]],
    package_records: dict[str, list[dict[str, Any]]],
    issues: list[dict[str, Any]],
) -> None:
    required_by_framework = {
        "GB-T-22239-2019-L3": ["等保三级控制要求"],
        "CIS-CSC-V8.1.2": ["保护措施编号", "名称", "描述", "安全控制项", "安全控制项名称", "控制项描述"],
        "NIST-CSF-2.0": ["功能", "分类", "分类标识符", "分类标识符说明"],
        "ISO-IEC-27001-2022": ["控制类别", "控制编号", "控制名称", "控制描述"],
        "DSP-SCF-2026": ["SCF域", "策略原则", "策略意图", "SCF编号", "SCF控制项", "SCF控制项描述"],
        "CRF-SAFEGUARDS-CORE-2026": ["保障措施分类", "保障措施域", "Safeguard ID", "保障措施描述"],
        "CRF-MATURITY-MODEL-2026": ["等级编号", "成熟度等级", "等级定义"],
        "NIST-800-53-REV5": ["安全控制类", "安全策略编号", "安全控制项", "控制描述"],
    }
    for framework_code, records in package_records.items():
        for rec in records:
            row = rec["row"]
            for field in required_by_framework.get(framework_code, []):
                if not text(row.get(field, "")):
                    issue(issues, "error", framework_code, "blank_required_field", f"JSON 必填业务字段为空：{field}", source=rec["json_path"], row=rec["json_row"], control_id=rec["control_id"], field=field)

            related_raw = rec.get("related_raw", "")
            if related_raw and not rec.get("related_focus_codes"):
                issue(issues, "warn", framework_code, "unparsed_related_focus", "关联安全能力/关注点字段非空，但未解析出合法关注点编号", source=rec["json_path"], row=rec["json_row"], control_id=rec["control_id"], field="关联安全能力/关注点", actual=related_raw[:160])

    source_required_text_fields = {
        "GB-T-22239-2019-L3": ["等保三级控制要求"],
        "CIS-CSC-V8.1.2": ["名称", "描述"],
        "NIST-CSF-2.0": ["分类标识符说明"],
        "ISO-IEC-27001-2022": ["控制名称", "控制描述"],
        "DSP-SCF-2026": ["SCF控制项", "SCF控制项描述"],
        "CRF-SAFEGUARDS-CORE-2026": ["保障措施描述"],
        "CRF-MATURITY-MODEL-2026": ["成熟度等级", "等级定义"],
        "NIST-800-53-REV5": ["中文名称", "控制描述"],
    }
    for framework_code, records in source_records.items():
        base_framework = framework_code.replace(":tiers", "")
        for rec in records:
            fields = rec.get("fields", {})
            for field in source_required_text_fields.get(base_framework, []):
                if field in fields and not text(fields.get(field)):
                    issue(issues, "error", base_framework, "blank_source_field", f"原始 Sheet 关键业务字段为空：{field}", row=rec["source_row"], control_id=rec["control_id"], field=field)


def projection_pairs(records: dict[str, list[dict[str, Any]]]) -> set[tuple[str, str, str]]:
    pairs: set[tuple[str, str, str]] = set()
    for framework_code, framework_records in records.items():
        if framework_code.endswith(":tiers") or framework_code in {"DSP-SCF-2026", "CRF-MATURITY-MODEL-2026"}:
            continue
        for rec in framework_records:
            control_id = normalize_mapping_control_id(framework_code, rec.get("control_id", ""))
            for focus_code in rec.get("related_focus_codes", []):
                pairs.add((framework_code, control_id, focus_code))
    return pairs


def mapping_source_pairs(workbook_path: Path, issues: list[dict[str, Any]]) -> set[tuple[str, str, str]]:
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    pairs: set[tuple[str, str, str]] = set()
    sheet_name = "安全能力-网络安全制度、框架映射"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return pairs
    ws = wb[sheet_name]
    for row in range(5, ws.max_row + 1):
        focus_raw = cell(ws, row, 5)
        focus_codes = extract_focus_codes(focus_raw)
        if not focus_codes and focus_raw:
            issue(issues, "warn", "ALL", "mapping_focus_code_unparsed", "能力映射表 E 列非空但未解析出关注点编号", source=sheet_name, row=row, field="关注点编号", actual=focus_raw)
            continue
        for framework_code, col_idx, label in MAPPING_COLUMNS:
            raw_value = cell(ws, row, col_idx)
            ids = extract_mapping_ids(framework_code, raw_value)
            if raw_value and not ids:
                issue(issues, "warn", framework_code, "mapping_control_id_unparsed", f"能力映射表 {label} 列非空但未解析出控制编号", source=sheet_name, row=row, field=label, actual=raw_value[:160])
            for focus_code in focus_codes:
                for control_id in ids:
                    pairs.add((framework_code, normalize_mapping_control_id(framework_code, control_id), focus_code))
    wb.close()
    return pairs


def audit_mapping_consistency(
    workbook_path: Path,
    package_records: dict[str, list[dict[str, Any]]],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    source_pairs = mapping_source_pairs(workbook_path, issues)
    package_pairs = projection_pairs(package_records)

    framework_ids = {
        framework_code: {normalize_mapping_control_id(framework_code, rec.get("control_id", "")) for rec in records}
        for framework_code, records in package_records.items()
    }

    for framework_code, control_id, focus_code in sorted(source_pairs):
        if control_id not in framework_ids.get(framework_code, set()):
            issue(issues, "error", framework_code, "mapping_control_missing_in_standard", "能力映射表引用的控制项不在当前标准主数据中", control_id=control_id, field=focus_code)

    missing_in_projection = source_pairs - package_pairs
    extra_in_projection = package_pairs - source_pairs

    for framework_code, control_id, focus_code in sorted(missing_in_projection)[:200]:
        issue(issues, "error", framework_code, "mapping_missing_in_standard_projection", "能力映射表存在该关系，但标准页关联关注点字段缺失", control_id=control_id, field=focus_code)
    for framework_code, control_id, focus_code in sorted(extra_in_projection)[:200]:
        issue(issues, "error", framework_code, "mapping_extra_in_standard_projection", "标准页关联关注点字段存在该关系，但能力映射表未找到", control_id=control_id, field=focus_code)

    by_framework: dict[str, dict[str, int]] = {}
    for framework_code, _, _ in source_pairs | package_pairs:
        src = {pair for pair in source_pairs if pair[0] == framework_code}
        pkg = {pair for pair in package_pairs if pair[0] == framework_code}
        by_framework[framework_code] = {
            "mapping_source_pairs": len(src),
            "standard_projection_pairs": len(pkg),
            "missing_in_standard_projection": len(src - pkg),
            "extra_in_standard_projection": len(pkg - src),
        }

    return {
        "mapping_source_pairs": len(source_pairs),
        "standard_projection_pairs": len(package_pairs),
        "missing_in_standard_projection": len(missing_in_projection),
        "extra_in_standard_projection": len(extra_in_projection),
        "by_framework": by_framework,
        "issue_limit_note": "mapping pair issue rows are capped at 200 per side in CSV/JSON details",
    }


def audit_index(standards_dir: Path, issues: list[dict[str, Any]]) -> dict[str, Any]:
    index_path = standards_dir.parent / "standards-index.json"
    compat_path = standards_dir.parent / "standards-data.json"
    result: dict[str, Any] = {"index_exists": index_path.exists(), "compat_index_exists": compat_path.exists(), "missing_data_paths": []}

    for path in (index_path, compat_path):
        if not path.exists():
            issue(issues, "error", "ALL", "missing_index", f"标准索引不存在：{path.name}", source=str(path))
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        if data.get("rows"):
            issue(issues, "error", "ALL", "full_rows_in_index", f"{path.name} 不应包含全量 rows", source=str(path))
        frameworks = data.get("frameworks") or []
        result[f"{path.name}_frameworks"] = len(frameworks)
        for framework in frameworks:
            data_paths = []
            if framework.get("dataPath"):
                data_paths.append(framework["dataPath"])
            for tab in framework.get("tabs", []):
                if tab.get("dataPath"):
                    data_paths.append(tab["dataPath"])
            for data_path in data_paths:
                rel_path = (
                    data_path.replace("./public/data/standards/", "")
                    .replace("/public/data/standards/", "")
                    .replace("/data/standards/", "")
                    .replace("data/standards/", "")
                )
                full_path = standards_dir / rel_path
                if not full_path.exists():
                    result["missing_data_paths"].append(data_path)
                    issue(issues, "error", "ALL", "missing_split_data_path", "索引指向的拆包 JSON 不存在", source=path.name, field=data_path)
    return result


def severity_counts(issues: list[dict[str, Any]]) -> dict[str, int]:
    counts = Counter(issue["severity"] for issue in issues)
    return {key: counts.get(key, 0) for key in ("error", "warn", "info")}


def write_outputs(
    output_dir: Path,
    stamp: str,
    report: dict[str, Any],
    issues: list[dict[str, Any]],
) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / f"standard-framework-full-data-audit-{stamp}.json"
    csv_path = output_dir / f"standard-framework-full-data-audit-issues-{stamp}.csv"
    md_path = output_dir / f"standard-framework-full-data-audit-{stamp}.md"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    columns = ["severity", "framework", "issue_type", "source", "row", "control_id", "field", "expected", "actual", "message"]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for item in issues:
            writer.writerow({key: item.get(key, "") for key in columns})

    lines = [
        f"# 标准 / 框架数据全量审计（{stamp}）",
        "",
        "## 审计范围",
        "",
        f"- 原始 workbook：`{report['source_workbook']}`",
        f"- 标准拆包目录：`{report['standards_dir']}`",
        "- 覆盖页面：等保三级、NIST CSF 2.0、ISO 27001:2022、DSP SCF 2026、CIS CSC v8、CRF、NIST SP 800-53 Rev.5",
        "",
        "## 总体结论",
        "",
        f"- 错误：{report['issue_counts']['error']}",
        f"- 警告：{report['issue_counts']['warn']}",
        f"- 索引缺失 dataPath：{len(report['index']['missing_data_paths'])}",
        f"- 标准映射源关系：{report['mapping']['mapping_source_pairs']}",
        f"- 标准页投影关系：{report['mapping']['standard_projection_pairs']}",
        f"- 映射缺失：{report['mapping']['missing_in_standard_projection']}",
        f"- 映射额外：{report['mapping']['extra_in_standard_projection']}",
        "",
        "## 逐框架行数对账",
        "",
        "| 框架 | 原始行 | JSON行 | 原始唯一编号 | JSON唯一编号 | JSON缺失 | JSON额外 |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for framework_code, item in report["framework_summary"].items():
        lines.append(
            f"| {item['label']} | {item['source_rows']} | {item['json_rows']} | {item['source_unique_ids']} | {item['json_unique_ids']} | {item['missing_in_json']} | {item['extra_in_json']} |"
        )
    lines.extend(["", "## 映射一致性", "", "| 框架 | 映射源关系 | 标准页投影 | 缺失 | 额外 |", "| --- | ---: | ---: | ---: | ---: |"])
    for framework_code, item in sorted(report["mapping"]["by_framework"].items()):
        label = STANDARD_PACKAGES.get(framework_code, {}).get("label", framework_code)
        lines.append(
            f"| {label} | {item['mapping_source_pairs']} | {item['standard_projection_pairs']} | {item['missing_in_standard_projection']} | {item['extra_in_standard_projection']} |"
        )
    lines.extend(
        [
            "",
            "## 高频问题类型",
            "",
            "| 严重级别 | 问题类型 | 数量 |",
            "| --- | --- | ---: |",
        ]
    )
    issue_type_counts = Counter((item["severity"], item["issue_type"]) for item in issues)
    for (severity, issue_type), count in issue_type_counts.most_common(30):
        lines.append(f"| {severity} | {issue_type} | {count} |")
    lines.extend(["", f"完整问题明细见 `{csv_path.name}`。", ""])
    md_path.write_text("\n".join(lines), encoding="utf-8")
    return {"json": str(json_path), "csv": str(csv_path), "markdown": str(md_path)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--standards-dir", type=Path, default=DEFAULT_STANDARDS_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--stamp", default=DEFAULT_DATE)
    args = parser.parse_args()

    issues: list[dict[str, Any]] = []
    source_records = parse_source_records(args.workbook, issues)
    package_records = load_package_records(args.standards_dir, issues)
    framework_summary = compare_records(source_records, package_records, issues)
    audit_required_fields(source_records, package_records, issues)
    mapping_summary = audit_mapping_consistency(args.workbook, package_records, issues)
    index_summary = audit_index(args.standards_dir, issues)

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_workbook": str(args.workbook),
        "standards_dir": str(args.standards_dir),
        "framework_summary": framework_summary,
        "mapping": mapping_summary,
        "index": index_summary,
        "issue_counts": severity_counts(issues),
        "issues": issues,
    }
    outputs = write_outputs(args.output_dir, args.stamp, report, issues)

    print(f"audit_status={'pass' if not issues else 'issues_found'}")
    print(f"errors={report['issue_counts']['error']} warnings={report['issue_counts']['warn']}")
    for framework_code, item in framework_summary.items():
        print(
            f"{framework_code}: source={item['source_rows']} json={item['json_rows']} "
            f"missing={item['missing_in_json']} extra={item['extra_in_json']} dupSource={item['duplicate_source_ids']} dupJson={item['duplicate_json_ids']}"
        )
    print(
        "mapping: "
        f"sourcePairs={mapping_summary['mapping_source_pairs']} "
        f"projectionPairs={mapping_summary['standard_projection_pairs']} "
        f"missing={mapping_summary['missing_in_standard_projection']} "
        f"extra={mapping_summary['extra_in_standard_projection']}"
    )
    for name, path in outputs.items():
        print(f"{name}={path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
