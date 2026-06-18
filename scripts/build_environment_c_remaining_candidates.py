#!/usr/bin/env python3
"""Build C-class remaining environment mapping candidate tables.

Read-only candidate generator. It does not edit the source workbook or formal
runtime packages.
"""

from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "data/raw-samples/wiki sample.xlsx"
REVIEW_ROWS_PATH = PROJECT_ROOT / "data/exports/worker-verify/environment-module-catalog-consistency-review-rows.json"
MEASURES_PATH = PROJECT_ROOT / "frontend/capability-browser/public/data/maintenance/measures.json"
OUT_DIR = PROJECT_ROOT / "data/exports/worker-verify/environment-c-remaining-candidates"
SHEET_NAME = "作用域-安全技术服务-安全技术模块映射"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def source_info(review_by_row: dict[int, dict[str, Any]], row: int, ws) -> dict[str, Any]:
    review = review_by_row.get(row) or {}
    source_cells = review.get("sourceCells") or {}
    merged_ranges = review.get("mergedRanges") or {}
    return {
        "excelRow": row,
        "context": review.get("objectContextKey") or "",
        "environment": review.get("environment") or text(ws.cell(row, 2).value),
        "environmentSegment": review.get("environmentSegment") or text(ws.cell(row, 3).value),
        "informationObject": review.get("informationObject") or text(ws.cell(row, 4).value),
        "declaredScopes": review.get("declaredScopes") or [],
        "serviceCell": (source_cells.get("securityTechnicalService") or {}).get("sourceCell") or f"F{row}",
        "moduleCell": (source_cells.get("moduleOrMeasureRaw") or {}).get("sourceCell") or f"G{row}",
        "systemCell": (source_cells.get("securitySystem") or {}).get("sourceCell") or f"H{row}",
        "serviceMergedRange": merged_ranges.get("securityTechnicalService"),
        "moduleMergedRange": merged_ranges.get("moduleOrMeasureRaw"),
        "systemMergedRange": merged_ranges.get("securitySystem"),
    }


def row_value(ws, row: int, col: int) -> str:
    return text(ws.cell(row, col).value)


def measure_names() -> set[str]:
    if not MEASURES_PATH.exists():
        return set()
    data = load_json(MEASURES_PATH)
    names = set()
    for item in data.get("security_technical_measures") or []:
        name = text(item.get("name") or item.get("title"))
        if name:
            names.add(name)
    return names


def add_candidate(
    rows: list[dict[str, Any]],
    *,
    issue_no: str,
    issue_name: str,
    action_type: str,
    row: int,
    ws,
    review_by_row: dict[int, dict[str, Any]],
    current_service: str | None = None,
    current_child_type: str = "module",
    current_child: str | None = None,
    current_system: str | None = None,
    proposed_service: str | None = None,
    proposed_child_type: str = "module",
    proposed_child: str | None = None,
    proposed_system: str | None = None,
    writeback_plan: str = "",
    risk: str = "medium",
    note: str = "",
    measure_exists: bool | None = None,
) -> None:
    info = source_info(review_by_row, row, ws)
    rows.append(
        {
            "issueNo": issue_no,
            "issueName": issue_name,
            "actionType": action_type,
            "excelRow": row,
            "context": info["context"],
            "environment": info["environment"],
            "environmentSegment": info["environmentSegment"],
            "informationObject": info["informationObject"],
            "declaredScopes": " / ".join(info["declaredScopes"]),
            "currentService": current_service if current_service is not None else row_value(ws, row, 6),
            "currentChildType": current_child_type,
            "currentChild": current_child if current_child is not None else row_value(ws, row, 7),
            "currentSecuritySystem": current_system if current_system is not None else row_value(ws, row, 8),
            "proposedService": proposed_service if proposed_service is not None else row_value(ws, row, 6),
            "proposedChildType": proposed_child_type,
            "proposedChild": proposed_child if proposed_child is not None else row_value(ws, row, 7),
            "proposedSecuritySystem": proposed_system if proposed_system is not None else row_value(ws, row, 8),
            "serviceCell": info["serviceCell"],
            "moduleCell": info["moduleCell"],
            "systemCell": info["systemCell"],
            "serviceMergedRange": info["serviceMergedRange"],
            "moduleMergedRange": info["moduleMergedRange"],
            "systemMergedRange": info["systemMergedRange"],
            "writebackPlan": writeback_plan,
            "risk": risk,
            "measureExistsInMaintenanceCatalog": measure_exists,
            "note": note,
        }
    )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_md(path: Path, title: str, summary: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    lines = [f"# {title}", ""]
    lines.extend(
        [
            f"- generatedAt：`{summary['generatedAt']}`",
            f"- workbook：`{summary['workbook']}`",
            f"- candidateCount：`{summary['candidateCount']}`",
            f"- affectedExcelRowCount：`{summary['affectedExcelRowCount']}`",
            f"- actionTypeCounts：`{summary['actionTypeCounts']}`",
            "",
            "## 候选明细",
            "",
            "| # | 问题 | 行 | 当前服务 | 当前模块/措施 | 候选服务 | 候选模块/措施 | 写回方式 | 风险 | 备注 |",
            "|---:|---|---:|---|---|---|---|---|---|---|",
        ]
    )
    for index, row in enumerate(rows, 1):
        lines.append(
            "| {index} | {issue} | {excel_row} | {current_service} | {current_child} | {proposed_service} | {proposed_child} | {plan} | {risk} | {note} |".format(
                index=index,
                issue=f"{row['issueNo']} {row['issueName']}",
                excel_row=row["excelRow"],
                current_service=row["currentService"].replace("\n", "<br>"),
                current_child=f"{row['currentChildType']}：{row['currentChild']}".replace("\n", "<br>"),
                proposed_service=row["proposedService"].replace("\n", "<br>"),
                proposed_child=f"{row['proposedChildType']}：{row['proposedChild']}".replace("\n", "<br>"),
                plan=row["writebackPlan"],
                risk=row["risk"],
                note=row["note"].replace("\n", "<br>"),
            )
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    review_rows = load_json(REVIEW_ROWS_PATH)
    review_by_row = {int(row["excelRow"]): row for row in review_rows}
    mismatch_rows = {
        int(row["excelRow"])
        for row in review_rows
        if (row.get("catalogMatch") or {}).get("moduleServiceMatch") is False
    }
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    ws = workbook[SHEET_NAME]
    measures = measure_names()
    env_rows: list[dict[str, Any]] = []
    measure_rows: list[dict[str, Any]] = []

    # 2/3. 零信任访问代理 / 控制台拆分：补全整个合并区域的候选归属，
    # 包含当前未报错但拆分时必须保留的正确行。
    zero_trust_groups = [
        range(144, 155),
        range(238, 248),
        range(319, 329),
        range(399, 410),
    ]
    zero_trust_console = {
        "I-US&T-AS.IA-01 用户身份管理",
        "I-AP&T-AS.IA-01 应用身份管理",
        "I-US&T-AS.IA-04 用户凭证管理",
        "I-AP&T-AS.IA-04 应用凭证管理",
        "I-US&T-AS.IA-02 用户认证",
        "I-AP&T-AS.IA-02 应用身份认证",
        "I-US&T-PD.AC-02 用户信誉评估",
        "I-AP&T-PD.AC-02 应用软件信誉评估",
        "I-OS&T-PD.AC-02 主机/终端信誉评估",
        "I-NT&T-AS.IA-03 网络资源授权",
        "I-AP&T-AS.IA-03 应用资源授权",
        "I-NT&T-PD.AC-01 网络访问控制",
        "I-AP&T-PD.AC-01 应用访问控制",
        "I-US&T-AS.LA-03 用户行为审计",
        "I-AP&T-AS.LA-03 应用操作审计",
    }
    zero_trust_agent = {
        "I-NT&T-PD.PP-01 网络安全接入",
        "I-AP&T-PD.PP-01 应用安全访问",
        "I-NT&T-PD.AC-01 网络访问控制",
        "I-NT&T-AS.IA-03 网络资源授权",
        "I-AP&T-AS.IA-03 应用资源授权",
        "I-NT&T-AS.CG-01 网络传输通道加解密",
        "I-AP&T-PD.AC-01 应用访问控制",
        "I-AP&T-AS.LA-01 应用日志记录",
    }
    for group in zero_trust_groups:
        if not any(row in mismatch_rows for row in group):
            continue
        for row in group:
            service = row_value(ws, row, 6)
            proposed_modules = []
            if service in zero_trust_agent:
                proposed_modules.append("零信任访问代理")
            if service in zero_trust_console:
                proposed_modules.append("零信任访问控制台")
            if not proposed_modules:
                continue
            proposed = "\n".join(proposed_modules)
            add_candidate(
                env_rows,
                issue_no="2/3",
                issue_name="零信任访问代理 / 零信任访问控制台拆分",
                action_type="split_merged_module_cell",
                row=row,
                ws=ws,
                review_by_row=review_by_row,
                current_child="零信任访问代理 / 零信任访问控制台（当前合并写法）",
                proposed_child=proposed,
                writeback_plan="拆分 G 列合并区域，按模块清单允许关系逐行写入代理 / 控制台；同一服务允许两个模块时保留两个模块关系",
                risk="high",
                note="候选覆盖整段合并区域，包含当前未报错但拆分时必须保留的正确行；不处理同一信息化对象内重复安全技术服务。",
            )

    # 5. 安全Web浏览器：每个上下文补齐模块清单 3 条服务。
    browser_groups = [(472, 473), (539, 540), (615, 616)]
    browser_services = [
        "I-AP&T-PD.PP-01 应用安全访问",
        "I-AP&T-AS.IA-03 应用资源授权",
        "I-AP&T-PD.TP-03 应用攻击入侵防御",
    ]
    for first, second in browser_groups:
        current_services = [row_value(ws, first, 6), row_value(ws, second, 6)]
        add_candidate(
            env_rows,
            issue_no="5",
            issue_name="安全Web浏览器服务集合替换",
            action_type="replace_and_insert_service_rows",
            row=first,
            ws=ws,
            review_by_row=review_by_row,
            current_service=" / ".join(current_services),
            current_child="安全Web浏览器",
            proposed_service=" / ".join(browser_services),
            proposed_child="安全Web浏览器",
            writeback_plan="将当前 2 行服务调整为模块清单 3 行服务；需插入 1 行并维护 G 列合并区域",
            risk="high",
            note="保留现有“应用攻击入侵防御”，把“应用身份认证”替换并补齐“应用安全访问 / 应用资源授权”。",
        )

    # 6. 网络准入控制（NAC）
    for row in [526, 602, 695]:
        add_candidate(
            env_rows,
            issue_no="6",
            issue_name="网络准入控制（NAC）硬件凭证管理笔误",
            action_type="replace_service",
            row=row,
            ws=ws,
            review_by_row=review_by_row,
            proposed_service="I-HD&T-AS.IA-04 硬件凭证管理",
            proposed_child="网络准入控制（NAC）",
            writeback_plan="直接替换 F 列服务值，G/H 合并区域保持",
            risk="low",
            note="按用户判断：硬件资源授权为笔误，应为硬件凭证管理。",
        )

    # 10. 安全接入网关（VPN）：当前 5 行调整为 6 行。
    vpn_services = [
        "I-NT&T-AS.CG-01 网络传输通道加解密",
        "I-US&T-AS.IA-02 用户认证",
        "I-HD&T-AS.IA-02 设备身份认证",
        "I-NT&T-PD.PP-01 网络安全接入",
        "I-NT&T-AS.IA-03 网络资源授权",
        "I-NT&T-PD.AC-01 网络访问控制",
    ]
    add_candidate(
        env_rows,
        issue_no="10",
        issue_name="安全接入网关（VPN）服务集合替换",
        action_type="replace_and_insert_service_rows",
        row=6,
        ws=ws,
        review_by_row=review_by_row,
        current_service=" / ".join(row_value(ws, row, 6) for row in range(5, 10)),
        current_child="安全接入网关（VPN）",
        proposed_service=" / ".join(vpn_services),
        proposed_child="安全接入网关（VPN）",
        writeback_plan="将当前 5 行服务调整为模块清单 6 行服务；需替换用户身份管理并插入设备身份认证",
        risk="high",
        note="当前同一 G 合并区域 F5:F9 内已有 4 条正确服务，候选为最小集合替换；主问题行为 F6。",
    )

    # 11. 容器网络防火墙
    add_candidate(
        env_rows,
        issue_no="11",
        issue_name="容器网络防火墙服务替换",
        action_type="replace_service",
        row=79,
        ws=ws,
        review_by_row=review_by_row,
        proposed_service="I-NT&T-PD.AC-01 网络访问控制",
        proposed_child="容器网络防火墙",
        writeback_plan="直接替换 F 列服务值，G/H 合并区域保持",
        risk="low",
        note="保持模块清单中容器网络防火墙映射不变，只修环境表服务。",
    )

    # 4. API安全防护：非 API 防护服务剥离为措施 API网关。
    api_gateway = "API网关"
    for row in [174, 175, 176, 177, 179, 180]:
        add_candidate(
            measure_rows,
            issue_no="4",
            issue_name="API安全防护非本模块服务改为措施",
            action_type="module_to_measure",
            row=row,
            ws=ws,
            review_by_row=review_by_row,
            current_child="API安全防护",
            current_system="应用安全防护",
            proposed_child_type="measure",
            proposed_child=api_gateway,
            proposed_system="",
            writeback_plan="拆分 G/H 合并区域；G 列改为措施，H 列对措施行置空或按措施规则处理",
            risk="high",
            measure_exists=api_gateway in measures,
            note="现有维护措施清单未命中 API网关，正式写回前需先确认是否新增安全技术措施。",
        )

    # 7. 数据加密和令牌化：应用层数据加解密改为应用自身措施。
    app_crypto_measure = "应用自身数据加解密模块，数据加密后传至数据库存储"
    for row in [194, 423]:
        add_candidate(
            measure_rows,
            issue_no="7",
            issue_name="应用层数据加解密改为应用自身措施",
            action_type="module_to_measure",
            row=row,
            ws=ws,
            review_by_row=review_by_row,
            current_child="数据加密和令牌化",
            current_system="数据安全防护",
            proposed_child_type="measure",
            proposed_child=app_crypto_measure,
            proposed_system="",
            writeback_plan="G 列改为措施，H 列对措施行置空或按措施规则处理",
            risk="medium",
            measure_exists=app_crypto_measure in measures,
            note="现有维护措施清单未命中该措施名，正式写回前需先确认是否新增安全技术措施。",
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for stem, title, rows in [
        ("environment-mapping-candidates", "C类环境映射表修正候选", env_rows),
        ("module-to-measure-candidates", "C类模块改措施候选", measure_rows),
    ]:
        summary = {
            "generatedAt": generated_at,
            "workbook": str(WORKBOOK_PATH.relative_to(PROJECT_ROOT)),
            "sheet": SHEET_NAME,
            "candidateCount": len(rows),
            "affectedExcelRowCount": len({row["excelRow"] for row in rows}),
            "actionTypeCounts": dict(Counter(row["actionType"] for row in rows)),
            "formalDataChanged": False,
            "sourceWorkbookChanged": False,
        }
        payload = {"summary": summary, "candidates": rows}
        write_json(OUT_DIR / f"{stem}.json", payload)
        write_csv(OUT_DIR / f"{stem}.csv", rows)
        write_md(OUT_DIR / f"{stem}.md", title, summary, rows)

    index = {
        "generatedAt": generated_at,
        "formalDataChanged": False,
        "sourceWorkbookChanged": False,
        "outputs": {
            "environmentMappingCandidates": {
                "json": str((OUT_DIR / "environment-mapping-candidates.json").relative_to(PROJECT_ROOT)),
                "csv": str((OUT_DIR / "environment-mapping-candidates.csv").relative_to(PROJECT_ROOT)),
                "md": str((OUT_DIR / "environment-mapping-candidates.md").relative_to(PROJECT_ROOT)),
                "candidateCount": len(env_rows),
            },
            "moduleToMeasureCandidates": {
                "json": str((OUT_DIR / "module-to-measure-candidates.json").relative_to(PROJECT_ROOT)),
                "csv": str((OUT_DIR / "module-to-measure-candidates.csv").relative_to(PROJECT_ROOT)),
                "md": str((OUT_DIR / "module-to-measure-candidates.md").relative_to(PROJECT_ROOT)),
                "candidateCount": len(measure_rows),
            },
        },
    }
    write_json(OUT_DIR / "index.json", index)
    print(json.dumps(index, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
