#!/usr/bin/env python3
"""Build a manual review checklist for environment mapping reimport data.

This script is intentionally read-only for official data packages. It compares
source-derived normalized rows against the current official workbench and
node-details, then writes human-review artifacts under worker-verify.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency exists in project env
    load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "data/exports/worker-verify"
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
SHEET_NAME = "作用域-安全技术服务-安全技术模块映射"
NORMALIZED_ROWS_PATH = OUTPUT_DIR / "scope-service-module-mapping-normalized-rows.json"
RELATIONS_PATH = OUTPUT_DIR / "scope-service-module-mapping-relations.json"
AUDIT_PATH = OUTPUT_DIR / "scope-service-module-mapping-reimport-audit.json"
CONSISTENCY_AUDIT_PATH = OUTPUT_DIR / "environment-module-catalog-consistency-audit.json"
CONSISTENCY_REVIEW_ROWS_PATH = OUTPUT_DIR / "environment-module-catalog-consistency-review-rows.json"
CONSISTENCY_TRIAGE_PATH = OUTPUT_DIR / "environment-module-catalog-consistency-triage.json"
WORKBENCH_PATH = ROOT / "frontend/capability-browser/public/data/environment-workbench.json"
NODE_DETAILS_PATH = ROOT / "frontend/capability-browser/generated/environmentBasemap.node-details.json"
MAINTENANCE_SERVICES_PATH = ROOT / "frontend/capability-browser/public/data/maintenance/services.json"

JSON_OUT = OUTPUT_DIR / "environment-manual-review-checklist.json"
CSV_OUT = OUTPUT_DIR / "environment-manual-review-checklist.csv"
MD_OUT = OUTPUT_DIR / "environment-manual-review-checklist.md"
PUBLIC_REVIEW_OUT = ROOT / "frontend/capability-browser/public/data/review/environment-manual-review-checklist.json"

TARGET_OBJECTS = [
    "PC终端设备",
    "PC终端操作系统",
    "PC终端软件应用",
    "移动终端软件应用",
    "业务应用",
    "园区网出口边界",
    "互联网入口边界",
    "容器",
    "物理主机硬件",
    "物理主机操作系统",
    "办公人员",
]

SERVICE_COUNT_REVIEW_THRESHOLD = 25
MODULE_MEASURE_COUNT_REVIEW_THRESHOLD = 25
SERVICE_CODE_RE = re.compile(r"(I-[A-Z]{2}&T-[A-Z]{2}\.[A-Z]{2}-\d{2})(?:\s+([^,，;；|\\n]+))?")
KNOWN_SERVICE_CODE_MIGRATIONS = {
    "I-OS&T-AS.DS-01": "I-AP&T-AS.DS-01",
    "I-OS&T-AS.DS-02": "I-AP&T-AS.DS-02",
    "I-OS&T-AS.DS-03": "I-AP&T-AS.DS-03",
    "I-OS&T-AS.DS-04": "I-AP&T-AS.DS-04",
    "I-OS&T-AS.DS-05": "I-AP&T-AS.DS-05",
    "I-OS&T-AS.DS-06": "I-AP&T-AS.DS-06",
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_json_if_exists(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return load_json(path)


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def as_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def service_catalog_by_code() -> dict[str, dict[str, Any]]:
    payload = load_json(MAINTENANCE_SERVICES_PATH)
    items = payload.get("security_technical_services") or []
    index: dict[str, dict[str, Any]] = {}
    for item in items:
        service = item.get("service") if isinstance(item, dict) else None
        if not isinstance(service, dict):
            service = item if isinstance(item, dict) else {}
        code = text(service.get("code"))
        title = text(service.get("title") or service.get("name"))
        if code and title:
            index[code] = {
                "id": text(service.get("id")),
                "code": code,
                "title": title,
                "category": text(service.get("category")),
            }
    return index


def canonical_service_for_code(code: str, catalog: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    canonical_code = KNOWN_SERVICE_CODE_MIGRATIONS.get(code, code)
    return catalog.get(canonical_code)


def canonicalize_service_text(value: str, catalog: dict[str, dict[str, Any]]) -> str:
    def replace(match: re.Match[str]) -> str:
        code = match.group(1)
        title = text(match.group(2))
        service = canonical_service_for_code(code, catalog)
        if not service:
            return match.group(0)
        if title:
            return f"{service['code']} {service['title']}"
        return service["code"]

    return SERVICE_CODE_RE.sub(replace, value)


def canonicalize_review_service_refs(value: Any, catalog: dict[str, dict[str, Any]]) -> Any:
    if isinstance(value, str):
        return canonicalize_service_text(value, catalog)
    if isinstance(value, list):
        return [canonicalize_review_service_refs(item, catalog) for item in value]
    if not isinstance(value, dict):
        return value

    normalized = {key: canonicalize_review_service_refs(item, catalog) for key, item in value.items()}
    code = text(normalized.get("code"))
    service = canonical_service_for_code(code, catalog) if code else None
    if service:
        normalized["code"] = service["code"]
        if "title" in normalized:
            normalized["title"] = service["title"]
        if "name" in normalized:
            normalized["name"] = service["title"]
        if "category" in normalized:
            normalized["category"] = service["category"]
        if "id" in normalized and service["id"]:
            normalized["id"] = service["id"]
    return normalized


def relation_rows(rel: dict[str, Any]) -> set[int]:
    rows = rel.get("sourceRows") or rel.get("payload", {}).get("rows") or []
    return {int(row) for row in rows if isinstance(row, int) or str(row).isdigit()}


def relation_key(rel: dict[str, Any]) -> tuple[str, str]:
    return (text(rel.get("workerVerifyType")), text(rel.get("objectContextKey")))


def build_relation_row_index(workbench: dict[str, Any]) -> dict[tuple[str, str], set[int]]:
    index: dict[tuple[str, str], set[int]] = defaultdict(set)
    for rel in workbench.get("relations") or []:
        key = relation_key(rel)
        if not key[0] or not key[1]:
            continue
        index[key].update(relation_rows(rel))
    return index


def flatten_source_cells(source_cells: dict[str, Any]) -> str:
    parts = []
    for key in [
        "informationEnvironment",
        "environmentSegment",
        "informationObject",
        "scope",
        "securityTechnicalService",
        "moduleOrMeasureRaw",
        "securitySystem",
    ]:
        cell = source_cells.get(key) or {}
        source_cell = text(cell.get("sourceCell") or cell.get("cell"))
        value = text(cell.get("value"))
        if source_cell or value:
            parts.append(f"{key}:{source_cell}:{value}")
    return " | ".join(parts)


def flatten_merged_ranges(merged_ranges: dict[str, Any]) -> str:
    parts = []
    for key, value in merged_ranges.items():
        if value:
            parts.append(f"{key}:{value}")
    return " | ".join(parts)


def scope_text(row: dict[str, Any]) -> str:
    scopes = row.get("scopes") or []
    if scopes:
        return " / ".join(text(scope.get("text") or scope.get("title") or scope.get("code")) for scope in scopes if scope)
    return text(row.get("scope"))


def security_system_text(row: dict[str, Any]) -> str:
    systems = [text(item) for item in as_list(row.get("securitySystems")) if text(item)]
    return " / ".join(systems) or text(row.get("securitySystem"))


def is_normal_relation_value(value: Any) -> bool:
    normalized = text(value)
    return bool(normalized) and normalized != "/" and not normalized.startswith("N/A")


def security_system_count_text(row: dict[str, Any]) -> str:
    systems = [text(item) for item in as_list(row.get("securitySystems")) if is_normal_relation_value(item)]
    fallback = text(row.get("securitySystem"))
    if systems:
        return " / ".join(systems)
    return fallback if is_normal_relation_value(fallback) else ""


def node_detail_service_count(detail: dict[str, Any]) -> int:
    count = 0
    for group in detail.get("directScopeGroups") or []:
        count += len(group.get("services") or [])
    return count


def build_node_detail_index(node_details: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for detail in (node_details.get("nodeDetailsByMxId") or {}).values():
        key = text(detail.get("matchedObjectContextKey") or detail.get("objectContextKey"))
        if key:
            index[key].append(detail)
    return index


def classify_target(row: dict[str, Any]) -> str:
    obj = text(row.get("informationObject"))
    seg = text(row.get("environmentSegment"))
    if obj in TARGET_OBJECTS:
        return obj
    if seg == "业务应用":
        return "业务应用（环境子类）"
    return ""


def include_row(row: dict[str, Any]) -> bool:
    return bool(classify_target(row))


def unique_nonempty(values: list[str]) -> list[str]:
    seen = set()
    result = []
    for value in values:
        value = text(value)
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def module_measure_text(row: dict[str, Any]) -> str:
    module = text(row.get("securityTechnologyModule"))
    measure = text(row.get("securityTechnicalMeasure"))
    if module:
        return module
    if measure:
        return measure
    return ""


def module_measure_kind(row: dict[str, Any]) -> str:
    if text(row.get("securityTechnologyModule")):
        return "module"
    if text(row.get("securityTechnicalMeasure")):
        return "measure"
    kind = text(row.get("moduleOrMeasureKind"))
    return kind if kind and kind != "placeholder" else ""


def split_relation_titles(value: str) -> list[str]:
    return unique_nonempty([part.strip() for part in text(value).split(" / ")])


def build_modules_with_systems(source_relations: dict[str, Any]) -> set[str]:
    modules = set()
    relations = source_relations.get("moduleSystemRelations") or {}
    relation_values = relations.values() if isinstance(relations, dict) else relations
    for relation in relation_values:
        module_title = text((relation.get("module") or {}).get("title"))
        system_title = text((relation.get("securitySystem") or {}).get("title"))
        if module_title and is_normal_relation_value(system_title):
            modules.add(module_title)
    return modules


def missing_required_system_modules(row: dict[str, Any], modules_with_systems: set[str]) -> list[str]:
    if security_system_count_text(row):
        return []
    module_titles = split_relation_titles(text(row.get("securityTechnologyModule")))
    return [module for module in module_titles if module in modules_with_systems]


def issue_index(audit: dict[str, Any]) -> dict[str, Any]:
    duplicate_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    duplicate_items = audit.get("duplicateExactServiceChildRelations") or audit.get("duplicateServicesInObjectContext") or []
    for item in duplicate_items:
        for row in item.get("rows") or []:
            if isinstance(row, int) or str(row).isdigit():
                duplicate_by_row[int(row)].append(item)

    repeated_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for item in audit.get("repeatedServiceWithDifferentChildren") or []:
        for row in item.get("rows") or []:
            if isinstance(row, int) or str(row).isdigit():
                repeated_by_row[int(row)].append(item)

    scope_by_context: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in audit.get("scopeCompletenessIssues") or []:
        scope_by_context[text(item.get("objectContextKey"))].append(item)

    unknown_scope_by_context: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in audit.get("unknownScopeEvidence") or []:
        unknown_scope_by_context[text(item.get("objectContextKey"))].append(item)

    return {
        "duplicateByRow": duplicate_by_row,
        "repeatedByRow": repeated_by_row,
        "scopeByContext": scope_by_context,
        "unknownScopeByContext": unknown_scope_by_context,
    }


def consistency_issue_index(consistency_audit: dict[str, Any], review_rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    by_context: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for review_row in review_rows:
        row_number = review_row.get("excelRow")
        if not (isinstance(row_number, int) or str(row_number).isdigit()):
            continue
        for issue in review_row.get("issues") or []:
            payload = {
                "type": text(issue.get("type")),
                "severity": text(issue.get("severity")),
                "reason": text(issue.get("reason")),
                "objectContextKey": text(review_row.get("objectContextKey")),
            }
            by_row[int(row_number)].append(payload)
            by_context[payload["objectContextKey"]].append(payload)
    for issue_type in ("serviceModuleCoverageGapCandidate", "moduleServiceCoverageGapCandidate"):
        for issue in (consistency_audit.get("issues") or {}).get(issue_type) or []:
            row_number = issue.get("sampleExcelRow")
            if not (isinstance(row_number, int) or str(row_number).isdigit()):
                continue
            payload = {
                "type": issue_type,
                "severity": text(issue.get("severity") or "review"),
                "reason": text(issue.get("reason")),
                "objectContextKey": text(issue.get("objectContextKey")),
            }
            by_row[int(row_number)].append(payload)
            by_context[payload["objectContextKey"]].append(payload)
    return {"byRow": by_row, "byContext": by_context}


def triage_index(triage: dict[str, Any]) -> dict[str, Any]:
    by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    top_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    alias_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    pattern_by_signature: dict[str, dict[str, Any]] = {}

    for clusters in (triage.get("patternClusters") or {}).values():
        for pattern in clusters or []:
            signature = text(pattern.get("patternSignature"))
            if signature:
                pattern_by_signature[signature] = pattern
            for row in pattern.get("sourceRows") or []:
                if isinstance(row, int) or str(row).isdigit():
                    by_row[int(row)].append(pattern)

    for item in triage.get("topManualReviewItems") or []:
        for row in item.get("sourceRows") or []:
            if isinstance(row, int) or str(row).isdigit():
                top_by_row[int(row)].append(item)

    for item in triage.get("possibleAliasMatches") or []:
        for row in item.get("rows") or []:
            if isinstance(row, int) or str(row).isdigit():
                alias_by_row[int(row)].append(item)

    return {
        "byRow": by_row,
        "topByRow": top_by_row,
        "aliasByRow": alias_by_row,
        "patternBySignature": pattern_by_signature,
    }


def declared_scope_text(rows: list[dict[str, Any]]) -> str:
    values = []
    for row in rows:
        for scope in row.get("scopes") or []:
            value = text(scope.get("text") or " ".join(part for part in [scope.get("code"), scope.get("title")] if part))
            if value:
                values.append(value)
    return "\n".join(unique_nonempty(values))


def merged_map_for_row(row: dict[str, Any]) -> dict[str, Any]:
    return row.get("mergedRanges") or {}


def source_cells_for_row(row: dict[str, Any]) -> dict[str, Any]:
    return row.get("sourceCells") or {}


def workbook_summary() -> dict[str, Any]:
    if load_workbook is None:
        return {"path": str(WORKBOOK_PATH.relative_to(ROOT)), "sheet": SHEET_NAME, "available": False}
    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    worksheet = workbook[SHEET_NAME]
    return {
        "path": str(WORKBOOK_PATH.relative_to(ROOT)),
        "sheet": SHEET_NAME,
        "available": True,
        "maxRow": worksheet.max_row,
        "maxColumn": worksheet.max_column,
        "mergedRangeCount": len(list(worksheet.merged_cells.ranges)),
    }


def build_checklist() -> dict[str, Any]:
    normalized_rows = load_json(NORMALIZED_ROWS_PATH)
    source_relations = load_json(RELATIONS_PATH)
    audit = load_json(AUDIT_PATH)
    consistency_audit = load_json_if_exists(CONSISTENCY_AUDIT_PATH, {})
    consistency_review_rows = load_json_if_exists(CONSISTENCY_REVIEW_ROWS_PATH, [])
    consistency_triage = load_json_if_exists(CONSISTENCY_TRIAGE_PATH, {})
    workbench = load_json(WORKBENCH_PATH)
    node_details = load_json(NODE_DETAILS_PATH)

    relation_index = build_relation_row_index(workbench)
    detail_index = build_node_detail_index(node_details)
    issues = issue_index(audit if isinstance(audit, dict) else {})
    consistency_issues = consistency_issue_index(
        consistency_audit if isinstance(consistency_audit, dict) else {},
        consistency_review_rows if isinstance(consistency_review_rows, list) else [],
    )
    triage_lookup = triage_index(consistency_triage if isinstance(consistency_triage, dict) else {})
    modules_with_systems = build_modules_with_systems(source_relations)
    same_name_contexts: dict[str, set[str]] = defaultdict(set)
    for row in normalized_rows:
        same_name_contexts[text(row.get("informationObject"))].add(text(row.get("contextKey")))

    selected_rows = [row for row in normalized_rows if text(row.get("contextKey"))]
    context_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in selected_rows:
        context_rows[text(row.get("contextKey"))].append(row)

    rows_out: list[dict[str, Any]] = []
    context_summaries: list[dict[str, Any]] = []

    context_order = {
        context_key: min(int(row.get("row")) for row in rows if str(row.get("row")).isdigit())
        for context_key, rows in context_rows.items()
    }

    for context_key, rows in sorted(context_rows.items(), key=lambda item: context_order.get(item[0], 10**9)):
        first = rows[0]
        target = classify_target(first) or text(first.get("informationObject")) or "全量核对"
        service_values = unique_nonempty([text(row.get("securityTechnicalService")) for row in rows])
        module_values = unique_nonempty([text(row.get("securityTechnologyModule")) for row in rows])
        measure_values = unique_nonempty([text(row.get("securityTechnicalMeasure")) for row in rows])
        system_values = unique_nonempty([security_system_count_text(row) for row in rows])
        scope_values = unique_nonempty(declared_scope_text(rows).splitlines())
        detail_matches = detail_index.get(context_key, [])
        detail_service_counts = [node_detail_service_count(detail) for detail in detail_matches]
        scope_issues = issues["scopeByContext"].get(context_key, [])
        unknown_scope_issues = issues["unknownScopeByContext"].get(context_key, [])
        duplicate_service_rows_in_context = [
            int(row.get("row"))
            for row in rows
            if str(row.get("row")).isdigit() and issues["duplicateByRow"].get(int(row.get("row")))
        ]
        consistency_issue_rows_in_context = [
            int(row.get("row"))
            for row in rows
            if str(row.get("row")).isdigit() and consistency_issues["byRow"].get(int(row.get("row")))
        ]

        context_prompts = []
        context_risk = "low"
        if duplicate_service_rows_in_context:
            context_prompts.append(f"同一对象上下文下存在完整重复服务-模块/措施-安全系统关系({len(duplicate_service_rows_in_context)} 行)")
            context_risk = "high"
        if consistency_issue_rows_in_context:
            high_consistency = [
                issue
                for item in consistency_issues["byContext"].get(context_key, [])
                for issue in [item]
                if item.get("severity") == "high"
            ]
            context_prompts.append(f"跨表目录一致性待核对({len(consistency_issue_rows_in_context)} 行)")
            if high_consistency:
                context_risk = "high"
            elif context_risk != "high":
                context_risk = "medium"
        if len(service_values) > SERVICE_COUNT_REVIEW_THRESHOLD:
            context_prompts.append(f"服务数量较多({len(service_values)})，建议核对是否符合原表合并范围")
            if context_risk != "high":
                context_risk = "medium"
        if len(module_values) + len(measure_values) > MODULE_MEASURE_COUNT_REVIEW_THRESHOLD:
            context_prompts.append(f"模块/措施数量较多({len(module_values) + len(measure_values)})，建议核对是否过度展开")
            if context_risk != "high":
                context_risk = "medium"
        missing_system_modules_by_context = unique_nonempty(
            [
                module
                for row in rows
                for module in missing_required_system_modules(row, modules_with_systems)
            ]
        )
        if missing_system_modules_by_context:
            context_prompts.append(
                "安全技术模块存在但安全系统为空，且该模块在原表其他位置有关联系统："
                + "、".join(missing_system_modules_by_context)
            )
            context_risk = "high"
        if scope_issues:
            missing = unique_nonempty([scope for item in scope_issues for scope in item.get("missingScopes", [])])
            context_prompts.append(f"根据安全技术服务反查，作用域疑似缺失：{', '.join(missing)}")
            if context_risk != "high":
                context_risk = "medium"
        if unknown_scope_issues:
            context_prompts.append("部分服务无法从编号反查作用域，已保留 unknownScopeEvidence")
        if not detail_matches:
            context_prompts.append("当前底图 node-details 未直接匹配该 objectContextKey")
        elif sum(detail_service_counts) == 0 and service_values:
            context_prompts.append("底图节点能映射但详情关系为空或未展开")
            context_risk = "high"

        missing_rows = []
        for row in sorted(rows, key=lambda item: int(item.get("row"))):
            row_number = int(row.get("row"))
            context = text(row.get("contextKey"))
            expected = {
                "object_service": True,
                "service_module": bool(text(row.get("securityTechnologyModule"))),
                "service_measure": bool(text(row.get("securityTechnicalMeasure"))),
                "module_system": bool(security_system_count_text(row) and text(row.get("securityTechnologyModule"))),
            }
            found = {
                key: row_number in relation_index.get((key, context), set())
                for key in expected
            }
            row_prompts = list(context_prompts)
            row_issue_types = []
            row_risk = context_risk
            duplicate_service_items = issues["duplicateByRow"].get(row_number, [])
            if duplicate_service_items:
                row_prompts.append("同一对象上下文下完整重复服务-模块/措施-安全系统关系")
                row_issue_types.append("duplicate_exact_service_child_relation")
                row_risk = "high"
            repeated_service_items = issues["repeatedByRow"].get(row_number, [])
            if repeated_service_items:
                row_prompts.append("同一安全技术服务映射多个模块/措施/系统，按 1:N 展开展示")
            for consistency_issue in consistency_issues["byRow"].get(row_number, []):
                issue_type = text(consistency_issue.get("type"))
                if issue_type:
                    row_issue_types.append(issue_type)
                reason = text(consistency_issue.get("reason"))
                if reason:
                    row_prompts.append(reason)
                if consistency_issue.get("severity") == "high":
                    row_risk = "high"
                elif row_risk != "high":
                    row_risk = "medium"
            triage_patterns = triage_lookup["byRow"].get(row_number, [])
            top_items_for_row = triage_lookup["topByRow"].get(row_number, [])
            alias_items_for_row = triage_lookup["aliasByRow"].get(row_number, [])
            triage_categories = unique_nonempty([text(item.get("triageCategory")) for item in triage_patterns])
            pattern_signatures = unique_nonempty([text(item.get("patternSignature")) for item in triage_patterns])
            if top_items_for_row:
                row_prompts.append("Top 人工核对项")
            if any(item.get("issueType") == "repeatedServiceWithDifferentChildren" for item in triage_patterns):
                row_prompts.append("服务多模块/措施展开，按 1:N 展开展示")
            if alias_items_for_row:
                row_prompts.append("可能存在命名/别名/标点差异，需人工确认")
            if scope_issues:
                missing = unique_nonempty([scope for item in scope_issues for scope in item.get("missingScopes", [])])
                row_prompts.append(f"缺少作用域：{', '.join(missing)}")
                row_issue_types.append("missing_scope_by_service_reverse_check")
            for key, should_exist in expected.items():
                if should_exist and not found.get(key):
                    row_prompts.append(f"正式 workbench 缺少 {key} 行级关系")
                    row_issue_types.append(f"missing_{key}")
                    row_risk = "high"
                    missing_rows.append(row_number)
            if text(row.get("moduleOrMeasureRaw")) == "/":
                row_prompts.append("原表为单独 /，应不生成模块/措施关系")
                row_issue_types.append("slash_placeholder")
            if "N/A(" in text(row.get("moduleOrMeasureRaw")) and not module_measure_text(row):
                row_prompts.append("原表为 N/A(...)，应作为待确认说明")
                row_issue_types.append("na_pending_note")
            row_missing_system_modules = missing_required_system_modules(row, modules_with_systems)
            if row_missing_system_modules:
                row_prompts.append(
                    "安全技术模块存在但安全系统为空，且该模块在原表其他位置有关联系统："
                    + "、".join(row_missing_system_modules)
                )
                row_issue_types.append("missing_security_system_for_known_module")
                row_risk = "high"
            if text(row.get("securityTechnicalService")) and not module_measure_text(row) and text(row.get("moduleOrMeasureRaw")) not in {"", "/"}:
                row_prompts.append("服务存在但模块/措施未形成正常关系")
                row_issue_types.append("missing_module_or_measure")
            if not detail_matches:
                row_issue_types.append("node_details_missing")
            output_row = {
                "reviewTarget": target,
                "riskLevel": row_risk,
                "reviewPrompts": "；".join(unique_nonempty(row_prompts)),
                "issueTypes": unique_nonempty(row_issue_types),
                "informationEnvironment": text(row.get("informationEnvironment")),
                "environmentSegment": text(row.get("environmentSegment")),
                "informationObject": text(row.get("informationObject")),
                "objectContextKey": context,
                "scope": scope_text(row),
                "declaredScopeCell": declared_scope_text(rows),
                "securityTechnicalService": text(row.get("securityTechnicalService")),
                "securityTechnologyModule": text(row.get("securityTechnologyModule")),
                "securityTechnicalMeasure": text(row.get("securityTechnicalMeasure")),
                "moduleOrMeasure": module_measure_text(row),
                "moduleOrMeasureKind": module_measure_kind(row),
                "moduleOrMeasureRaw": text(row.get("moduleOrMeasureRaw")),
                "securitySystem": security_system_text(row),
                "excelRow": row_number,
                "mergedRanges": flatten_merged_ranges(row.get("mergedRanges") or {}),
                "mergedRangesMap": merged_map_for_row(row),
                "sourceCells": flatten_source_cells(row.get("sourceCells") or {}),
                "sourceCellsMap": source_cells_for_row(row),
                "nodeDetailsContains": "是" if detail_matches else "否",
                "nodeDetailMxIds": " / ".join(text(detail.get("mxId")) for detail in detail_matches),
                "sameNameDifferentContext": "是" if len(same_name_contexts[text(row.get("informationObject"))]) > 1 else "否",
                "workbenchObjectServiceRelation": "是" if found["object_service"] else "否",
                "workbenchServiceModuleRelation": "不适用" if not expected["service_module"] else ("是" if found["service_module"] else "否"),
                "workbenchServiceMeasureRelation": "不适用" if not expected["service_measure"] else ("是" if found["service_measure"] else "否"),
                "workbenchModuleSystemRelation": "不适用" if not expected["module_system"] else ("是" if found["module_system"] else "否"),
                "triageCategories": triage_categories,
                "triageCategoryLabels": unique_nonempty([text(item.get("triageCategoryLabel")) for item in triage_patterns]),
                "patternSignatures": pattern_signatures,
                "isTopManualReviewItem": bool(top_items_for_row),
                "topManualReviewPriority": min([int(item.get("priority")) for item in top_items_for_row if str(item.get("priority")).isdigit()] or [0]),
                "possibleAliasMatches": alias_items_for_row,
                "directoryEvidence": {
                    "patterns": triage_patterns[:5],
                    "topManualReviewItems": top_items_for_row[:5],
                    "environmentRelation": {
                        "objectContextKey": context,
                        "securityTechnicalService": text(row.get("securityTechnicalService")),
                        "childType": module_measure_kind(row),
                        "securityTechnologyModuleOrMeasure": module_measure_text(row),
                        "securitySystem": security_system_text(row),
                        "excelRow": row_number,
                    },
                    "suggestedManualAction": "；".join(
                        unique_nonempty([text(item.get("suggestedManualAction")) for item in (top_items_for_row or triage_patterns)])
                    ),
                },
            }
            rows_out.append(output_row)

        context_summaries.append(
            {
                "riskLevel": "high" if missing_rows else context_risk,
                "reviewTarget": target,
                "informationEnvironment": text(first.get("informationEnvironment")),
                "environmentSegment": text(first.get("environmentSegment")),
                "informationObject": text(first.get("informationObject")),
                "objectContextKey": context_key,
                "excelRows": [int(row.get("row")) for row in rows],
                "scopeCount": len(scope_values),
                "serviceCount": len(service_values),
                "securitySystemCount": len(system_values),
                "moduleCount": len(module_values),
                "measureCount": len(measure_values),
                "duplicateServiceCount": len(duplicate_service_rows_in_context),
                "missingScopesByServiceReverseCheck": unique_nonempty([scope for item in scope_issues for scope in item.get("missingScopes", [])]),
                "nodeDetailsContains": bool(detail_matches),
                "nodeDetailMxIds": [detail.get("mxId") for detail in detail_matches],
                "sameNameDifferentContext": len(same_name_contexts[text(first.get("informationObject"))]) > 1,
                "missingOfficialRelationRows": missing_rows,
                "reviewPrompts": unique_nonempty(context_prompts + ([f"存在 {len(missing_rows)} 行正式关系缺失"] if missing_rows else [])),
            }
        )

    risk_counter = Counter(row["riskLevel"] for row in rows_out)
    high_risk_contexts = [item for item in context_summaries if item["riskLevel"] == "high"]
    medium_risk_contexts = [item for item in context_summaries if item["riskLevel"] == "medium"]
    top_priority = sorted(
        context_summaries,
        key=lambda item: (
            {"high": 0, "medium": 1, "low": 2}.get(item["riskLevel"], 3),
            -item["serviceCount"],
            item["objectContextKey"],
        ),
    )[:20]

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "task": "Environment Mapping Manual Data Review Pack 1.0",
        "source": {
            "workbook": workbook_summary(),
            "normalizedRows": str(NORMALIZED_ROWS_PATH.relative_to(ROOT)),
            "relations": str(RELATIONS_PATH.relative_to(ROOT)),
            "audit": str(AUDIT_PATH.relative_to(ROOT)),
            "consistencyAudit": str(CONSISTENCY_AUDIT_PATH.relative_to(ROOT)) if CONSISTENCY_AUDIT_PATH.exists() else "",
            "consistencyReviewRows": str(CONSISTENCY_REVIEW_ROWS_PATH.relative_to(ROOT)) if CONSISTENCY_REVIEW_ROWS_PATH.exists() else "",
            "consistencyTriage": str(CONSISTENCY_TRIAGE_PATH.relative_to(ROOT)) if CONSISTENCY_TRIAGE_PATH.exists() else "",
            "officialWorkbench": str(WORKBENCH_PATH.relative_to(ROOT)),
            "officialNodeDetails": str(NODE_DETAILS_PATH.relative_to(ROOT)),
        },
        "targets": TARGET_OBJECTS,
        "summary": {
            "selectedContextCount": len(context_summaries),
            "selectedRowCount": len(rows_out),
            "highRiskContextCount": len(high_risk_contexts),
            "mediumRiskContextCount": len(medium_risk_contexts),
            "rowRiskDistribution": dict(risk_counter),
            "duplicateExactServiceChildRelationRowCount": sum(1 for row in rows_out if "duplicate_exact_service_child_relation" in row.get("issueTypes", [])),
            "duplicateServiceRowCount": sum(1 for row in rows_out if "duplicate_exact_service_child_relation" in row.get("issueTypes", [])),
            "moduleServiceMismatchRowCount": sum(1 for row in rows_out if "moduleServiceMismatch" in row.get("issueTypes", [])),
            "systemModuleMismatchRowCount": sum(1 for row in rows_out if "systemModuleMismatch" in row.get("issueTypes", [])),
            "topManualReviewRowCount": sum(1 for row in rows_out if row.get("isTopManualReviewItem")),
            "possibleAliasRowCount": sum(1 for row in rows_out if row.get("possibleAliasMatches")),
            "scopeCompletenessIssueContextCount": len(audit.get("scopeCompletenessIssues") or []) if isinstance(audit, dict) else 0,
            "unknownScopeEvidenceContextCount": len(audit.get("unknownScopeEvidence") or []) if isinstance(audit, dict) else 0,
            "shouldPauseUiAlignment": bool(high_risk_contexts),
        },
        "sourceRelationStats": {
            key: len(source_relations.get(key) or [])
            for key in [
                "objectScopeRelations",
                "objectServiceRelations",
                "serviceModuleRelations",
                "serviceMeasureRelations",
                "moduleSystemRelations",
                "pendingRelations",
                "invalidRows",
            ]
        },
        "contextSummaries": context_summaries,
        "topPriorityReviewItems": (consistency_triage.get("topManualReviewItems") or top_priority)[:50]
        if isinstance(consistency_triage, dict)
        else top_priority,
        "triageSummary": {
            "triageReport": str(CONSISTENCY_TRIAGE_PATH.relative_to(ROOT)) if CONSISTENCY_TRIAGE_PATH.exists() else "",
            "issueTypeCounts": consistency_triage.get("issueTypeCounts") or {},
            "triageCategoryCounts": consistency_triage.get("triageCategoryCounts") or {},
            "triageCategoryLabels": consistency_triage.get("triageCategoryLabels") or {},
            "patternClusterCounts": consistency_triage.get("patternClusterCounts") or {},
            "dualTableReviewSummary": (consistency_triage.get("dualTableReview") or {}).get("summary") or {},
            "possibleAliasMatchCount": len(consistency_triage.get("possibleAliasMatches") or []) if isinstance(consistency_triage, dict) else 0,
            "topManualReviewItemCount": len(consistency_triage.get("topManualReviewItems") or []) if isinstance(consistency_triage, dict) else 0,
        },
        "dualTableReview": consistency_triage.get("dualTableReview") or {} if isinstance(consistency_triage, dict) else {},
        "possibleAliasMatches": consistency_triage.get("possibleAliasMatches") or [] if isinstance(consistency_triage, dict) else [],
        "rows": rows_out,
    }


def write_outputs(payload: dict[str, Any]) -> None:
    payload = canonicalize_review_service_refs(payload, service_catalog_by_code())
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    PUBLIC_REVIEW_OUT.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_REVIEW_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    fieldnames = [
        "informationEnvironment",
        "environmentSegment",
        "informationObject",
        "objectContextKey",
        "declaredScopeCell",
        "securityTechnicalService",
        "moduleOrMeasure",
        "moduleOrMeasureKind",
        "securitySystem",
        "riskLevel",
        "issueTypes",
        "reviewPrompts",
        "excelRow",
        "mergedRanges",
        "sourceCells",
        "securityTechnologyModule",
        "securityTechnicalMeasure",
        "moduleOrMeasureRaw",
        "nodeDetailsContains",
        "nodeDetailMxIds",
        "sameNameDifferentContext",
        "workbenchObjectServiceRelation",
        "workbenchServiceModuleRelation",
        "workbenchServiceMeasureRelation",
        "workbenchModuleSystemRelation",
        "triageCategories",
        "patternSignatures",
        "isTopManualReviewItem",
        "topManualReviewPriority",
    ]
    with CSV_OUT.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in payload["rows"]:
            writer.writerow({key: " / ".join(row.get(key, [])) if isinstance(row.get(key), list) else row.get(key, "") for key in fieldnames})

    md = []
    summary = payload["summary"]
    md.append("# Environment Mapping Manual Data Review Checklist")
    md.append("")
    md.append(f"- 生成时间：{payload['generatedAt']}")
    md.append(f"- 原始文件：`{payload['source']['workbook']['path']}`")
    md.append(f"- Sheet：`{payload['source']['workbook']['sheet']}`")
    md.append(f"- merged ranges：`{payload['source']['workbook'].get('mergedRangeCount', 'unknown')}`")
    md.append(f"- 抽查对象上下文数：`{summary['selectedContextCount']}`")
    md.append(f"- 抽查行数：`{summary['selectedRowCount']}`")
    md.append(f"- 高风险核对项：`{summary['highRiskContextCount']}`")
    md.append(f"- 中风险核对项：`{summary['mediumRiskContextCount']}`")
    md.append(f"- 完整重复服务-模块/措施-安全系统行数：`{summary.get('duplicateExactServiceChildRelationRowCount', 0)}`")
    md.append(f"- 模块-服务目录不一致行数：`{summary.get('moduleServiceMismatchRowCount', 0)}`")
    md.append(f"- 系统-模块目录不一致行数：`{summary.get('systemModuleMismatchRowCount', 0)}`")
    md.append(f"- 服务反查作用域缺漏上下文数：`{summary.get('scopeCompletenessIssueContextCount', 0)}`")
    md.append(f"- 是否建议暂停 UI 改造：`{'是' if summary['shouldPauseUiAlignment'] else '否'}`")
    md.append("")
    md.append("## Top 人工核对")
    md.append("")
    md.append("| 优先级 | 类别 | issueType | patternSignature | 影响行 | 影响上下文 | 建议动作 |")
    md.append("|---:|---|---|---|---:|---:|---|")
    for index, item in enumerate(payload["topPriorityReviewItems"], start=1):
        if "patternSignature" not in item:
            prompts = "；".join(item.get("reviewPrompts") or [])
            md.append(
                f"| {index} | {item.get('riskLevel', '')} | contextSummary | {item.get('objectContextKey', '')} | "
                f"{len(item.get('excelRows') or [])} | 1 | {prompts} |"
            )
            continue
        md.append(
            f"| {item.get('priority', index)} | {item.get('triageCategory', '')} | {item.get('issueType', '')} | "
            f"{item.get('patternSignature', '')} | {item.get('affectedRows', 0)} | {item.get('affectedObjectContexts', 0)} | "
            f"{item.get('suggestedManualAction', '')} |"
        )
    md.append("")
    md.append("## 抽查对象上下文汇总")
    md.append("")
    md.append("| 风险 | 目标 | 信息化环境 | 环境子类 | 信息化对象 | 行范围 | 作用域 | 服务 | 模块 | 措施 | 安全系统 | node-details | 提示 |")
    md.append("|---|---|---|---|---|---|---:|---:|---:|---:|---:|---|---|")
    for item in payload["contextSummaries"]:
        rows = item["excelRows"]
        row_range = f"{min(rows)}-{max(rows)}" if rows else ""
        prompts = "；".join(item.get("reviewPrompts") or [])
        node_text = "是" if item.get("nodeDetailsContains") else "否"
        md.append(
            f"| {item['riskLevel']} | {item['reviewTarget']} | {item['informationEnvironment']} | {item['environmentSegment']} | "
            f"{item['informationObject']} | {row_range} | {item['scopeCount']} | {item['serviceCount']} | "
            f"{item['moduleCount']} | {item['measureCount']} | {item['securitySystemCount']} | {node_text} | {prompts} |"
        )
    md.append("")
    md.append("## 使用方式")
    md.append("")
    md.append("- CSV 是逐 Excel 行核对，适合筛选 `riskLevel`、`objectContextKey`、`excelRow`。")
    md.append("- JSON 保留完整对象上下文、Top 20 和行级字段。")
    md.append("- `nodeDetailsContains=否` 表示底图当前没有直接对应该 objectContextKey，不等于 workbench 数据缺失。")
    md.append("- `业务应用（环境子类）` 表示源表中 `业务应用` 是环境子类，不是信息化对象名。")
    MD_OUT.write_text("\n".join(md) + "\n", encoding="utf-8")


def main() -> None:
    payload = build_checklist()
    write_outputs(payload)
    print(
        json.dumps(
            {
                "result": "pass",
                "outputs": [str(JSON_OUT.relative_to(ROOT)), str(CSV_OUT.relative_to(ROOT)), str(MD_OUT.relative_to(ROOT))],
                "publicReviewData": str(PUBLIC_REVIEW_OUT.relative_to(ROOT)),
                "selectedContextCount": payload["summary"]["selectedContextCount"],
                "selectedRowCount": payload["summary"]["selectedRowCount"],
                "highRiskContextCount": payload["summary"]["highRiskContextCount"],
                "mediumRiskContextCount": payload["summary"]["mediumRiskContextCount"],
                "duplicateServiceRowCount": payload["summary"].get("duplicateServiceRowCount", 0),
                "duplicateExactServiceChildRelationRowCount": payload["summary"].get("duplicateExactServiceChildRelationRowCount", 0),
                "moduleServiceMismatchRowCount": payload["summary"].get("moduleServiceMismatchRowCount", 0),
                "systemModuleMismatchRowCount": payload["summary"].get("systemModuleMismatchRowCount", 0),
                "scopeCompletenessIssueContextCount": payload["summary"].get("scopeCompletenessIssueContextCount", 0),
                "shouldPauseUiAlignment": payload["summary"]["shouldPauseUiAlignment"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
