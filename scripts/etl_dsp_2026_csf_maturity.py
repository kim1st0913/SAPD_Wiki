#!/usr/bin/env python3
"""Populate DSP 2026 CSF mapping and Chinese SCR-CMM maturity descriptions."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.styles import Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent))
from translate_dsp_2026_sheet import normalize_cn


TARGET_SHEET = "DSP策略清单（2026）"
SOURCE_SHEET = "SCF 2026.1"

SOURCE_DEFAULT = (
    "/Users/kim1st/Documents/work@qax/01.战略咨询规划部/技术架构组/安全规划方法论设计研究/"
    "架构参考材料/网络安全框架材料/ComplianceForge/SCF 2026/"
    "Secure Controls Framework (SCF) - 2026.1.1.xlsx"
)

TARGET_HEADERS = {
    9: "NIST CSF\n功能分组",
    10: "SCR-CMM 0级\n未执行",
    11: "SCR-CMM 1级\n非正式执行",
    12: "SCR-CMM 2级\n已计划并跟踪",
    13: "SCR-CMM 3级\n定义良好",
    14: "SCR-CMM 4级\n量化控制",
    15: "SCR-CMM 5级\n持续改进",
}

CSF_TRANSLATIONS = {
    "Govern": "治理",
    "Identify": "识别",
    "Protect": "保护",
    "Detect": "检测",
    "Respond": "响应",
    "Recover": "恢复",
}

MATURITY_GLOSSARY = {
    "网络安全和数据保护": "网络安全与数据保护",
    "网络安全和数据隐私": "网络安全与数据隐私",
    "数据隐私": "数据隐私",
    "数据保护": "数据保护",
    "安全性": "安全",
    "合规性": "合规",
    "弹性": "韧性",
    "韧性和合规": "韧性与合规",
    "安全，合规": "安全、合规",
    "安全、合规和韧性": "安全、合规与韧性",
    "治理、风险和合规": "治理、风险与合规",
    "技术资产、应用程序、服务和/或数据": "技术资产、应用、服务和/或数据",
    "技术资产，应用程序，服务和/或数据": "技术资产、应用、服务和/或数据",
    "技术资产、应用、服务和/或数据": "技术资产、应用、服务和/或数据",
    "技术资产、应用程序和/或服务": "技术资产、应用和/或服务",
    "应用程序": "应用",
    "受管制": "受监管",
    "控制措施": "控制",
    "关键绩效指标": "关键绩效指标",
    "关键风险指标": "关键风险指标",
    "业务流程所有者": "业务流程负责人",
    "决策者": "决策人员",
    "第三方": "第三方",
    "正式记录": "正式记录",
    "持续改进": "持续改进",
    "世界一流": "世界级",
    "一个理智的人会得出这样的结论：": "合理判断可认为：",
    "一个理智的人会得出这样的结论": "合理判断可认为",
    "一个理智的人会得出结论，": "合理判断可认为，",
    "一个理智的人会得出结论": "合理判断可认为",
    "实践是不存在的": "实践不存在",
    "控制没有被执行": "该控制未执行",
    "实施和操作能力": "实施和运行能力",
    "操作能力": "运行能力",
    "利用 SCR-CMM": "采用 SCR-CMM",
    "组织特定来定义": "由组织自行定义",
    "明确定义": "定义良好",
    "功能是": "能力是",
    "功能在": "能力在",
    "功能除了": "能力除了",
    "级功能": "级能力",
    "功能应": "能力应",
    "响应能力": "响应能力",
    "域功能": "域能力",
    "将由由组织自行定义": "将由组织自行定义",
    "标准将由由组织自行定义": "标准将由组织自行定义",
    "机器学习": "机器学习",
    "人工智能": "人工智能",
    "IT /网络": "IT/网络",
    "IT / 网络": "IT/网络",
    "GRC 团队": "GRC 团队",
    "GRC团队": "GRC 团队",
    "SCR-CMM级别": "SCR-CMM 级别",
    "级别5": "5级",
    "4级": "4级",
    "3级": "3级",
}


def cleanup_maturity(text: str | None) -> str | None:
    if text is None:
        return None
    value = normalize_cn(str(text)) or ""
    for old, new in MATURITY_GLOSSARY.items():
        value = value.replace(old, new)
    value = value.replace("▪", "▪")
    value = value.replace("“", "\"").replace("”", "\"")
    value = re.sub(r"\s+([，。；：？）])", r"\1", value)
    value = re.sub(r"([\u4e00-\u9fff])\s+([\u4e00-\u9fff])", r"\1\2", value)
    value = re.sub(r"SCR-CMM\s*Level\s*(\d)", r"SCR-CMM \1级", value, flags=re.I)
    value = re.sub(r"Level\s*(\d)", r"\1级", value, flags=re.I)
    value = re.sub(r"与([A-Z]{3}) 域", r"与 \1 域", value)
    value = re.sub(r"([A-Z]{3})域", r"\1 域", value)
    return value.strip()


def load_cache(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_cache(path: Path, cache: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def google_translate_batch(texts: list[str]) -> list[str]:
    markers = [f"@@SAPD_MATURITY_SPLIT_{index:04d}@@" for index in range(1, len(texts))]
    parts: list[str] = []
    for index, text in enumerate(texts):
        if index:
            parts.append(markers[index - 1])
        parts.append(text)
    query = "\n".join(parts)
    url = "https://translate.googleapis.com/translate_a/single?" + urlencode(
        {"client": "gtx", "sl": "en", "tl": "zh-CN", "dt": "t", "q": query}
    )
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=40) as response:
        data = json.loads(response.read().decode("utf-8"))
    translated = "".join(part[0] for part in data[0])
    results = [translated]
    for marker in markers:
        if marker not in results[-1]:
            raise RuntimeError(f"Translation marker lost: {marker}")
        head, tail = results[-1].split(marker, 1)
        results[-1] = head
        results.append(tail)
    return [cleanup_maturity(item) or "" for item in results]


def build_batches(items: list[str], max_chars: int) -> list[list[str]]:
    batches: list[list[str]] = []
    current: list[str] = []
    current_len = 0
    for item in items:
        projected = current_len + len(item) + 32
        if current and projected > max_chars:
            batches.append(current)
            current = []
            current_len = 0
        current.append(item)
        current_len += len(item) + 32
    if current:
        batches.append(current)
    return batches


def translate_missing(cache: dict[str, str], segments: list[str], max_chars: int) -> None:
    missing = [segment for segment in segments if segment and segment not in cache]
    batches = build_batches(missing, max_chars)
    print(f"missing_segments={len(missing)} batches={len(batches)}", flush=True)
    for batch_index, batch in enumerate(batches, start=1):
        for attempt in range(1, 4):
            try:
                translated = google_translate_batch(batch)
                break
            except Exception as exc:
                if attempt == 3:
                    raise
                print(f"batch {batch_index}/{len(batches)} retry {attempt}: {exc}", flush=True)
                time.sleep(2 * attempt)
        for original, zh in zip(batch, translated, strict=True):
            cache[original] = zh
        if batch_index % 20 == 0 or batch_index == len(batches):
            print(f"translated_batches={batch_index}/{len(batches)}", flush=True)
        time.sleep(0.12)


def split_segments(text: str | None) -> list[str]:
    if not text:
        return []
    return [line.strip() for line in str(text).split("\n") if line.strip()]


def render_translated_text(text: str | None, cache: dict[str, str]) -> str | None:
    if not text:
        return None
    translated_lines = []
    for segment in split_segments(text):
        translated_lines.append(cleanup_maturity(cache.get(segment, segment)) or "")
    return "\n".join(translated_lines)


def translate_csf(value: str | None) -> str | None:
    if value is None:
        return None
    return CSF_TRANSLATIONS.get(str(value).strip(), str(value).strip())


def collect_source_records(source_path: Path) -> tuple[dict[str, dict[int, str | None]], list[str]]:
    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    source_ws = source_wb[SOURCE_SHEET]
    by_code: dict[str, dict[int, str | None]] = {}
    segments: set[str] = set()
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        code = str(row[2] or "").strip()
        if not code:
            continue
        record = {9: row[14]}
        for target_col, source_index in zip(range(10, 16), range(18, 24), strict=True):
            value = row[source_index]
            record[target_col] = value
            segments.update(split_segments(value))
        by_code[code] = record
    return by_code, sorted(segments)


def apply(args: argparse.Namespace) -> None:
    workbook_path = Path(args.workbook)
    source_path = Path(args.source)
    cache_path = Path(args.cache)

    backup_dir = workbook_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{workbook_path.stem}.before-dsp-2026-csf-maturity-{datetime.now():%Y%m%d%H%M%S}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup_path)
    print(f"backup={backup_path}", flush=True)

    source_by_code, segments = collect_source_records(source_path)
    cache = load_cache(cache_path)
    translate_missing(cache, segments, args.max_chars)
    save_cache(cache_path, cache)

    wb = load_workbook(workbook_path)
    ws = wb[TARGET_SHEET]
    for col, header in TARGET_HEADERS.items():
        ws.cell(2, col).value = header
        ws.cell(2, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    stats = {"rows": 0, "missing_codes": 0, "maturity_cells": 0, "csf_cells": 0}
    for row_index in range(3, ws.max_row + 1):
        code = str(ws.cell(row_index, 5).value or "").strip()
        record = source_by_code.get(code)
        if not record:
            stats["missing_codes"] += 1
            continue
        stats["rows"] += 1
        ws.cell(row_index, 9).value = translate_csf(record.get(9))
        stats["csf_cells"] += 1 if record.get(9) else 0
        for col in range(10, 16):
            value = render_translated_text(record.get(col), cache)
            ws.cell(row_index, col).value = value
            if value:
                stats["maturity_cells"] += 1
            ws.cell(row_index, col).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row_index, 9).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb.save(workbook_path)
    save_cache(cache_path, cache)
    print(json.dumps(stats, ensure_ascii=False), flush=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default="data/raw-samples/wiki sample.xlsx")
    parser.add_argument("--source", default=SOURCE_DEFAULT)
    parser.add_argument("--cache", default="data/processed/translation-cache/dsp-2026-maturity-google-zh.json")
    parser.add_argument("--max-chars", type=int, default=3000)
    args = parser.parse_args()
    apply(args)


if __name__ == "__main__":
    main()
