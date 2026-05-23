#!/usr/bin/env python3
"""Translate the DSP 2026 sheet with an external draft translator and local glossary."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


SHEET_NAME = "DSP策略清单（2026）"
SOURCE_2024_PATH = Path(
    "/Users/kim1st/Documents/work@qax/01.战略咨询规划部/技术架构组/安全规划方法论设计研究/架构参考材料/网络安全框架材料/ComplianceForge/SCF 2024/secure-controls-framework-scf-2024-3.xlsx"
)

DOMAIN_TRANSLATIONS = {
    "Security, Compliance & Resilience Governance": "安全、合规与韧性治理（Security, Compliance & Resilience Governance）",
    "Artificial Intelligence & Autonomous Technologies": "人工智能与自主技术（Artificial Intelligence & Autonomous Technologies）",
    "Asset Management": "资产管理（Asset Management）",
    "Business Continuity & Disaster Recovery": "业务连续性与灾难恢复（Business Continuity & Disaster Recovery）",
    "Capacity & Performance Planning": "容量与性能规划（Capacity & Performance Planning）",
    "Change Management": "变更管理（Change Management）",
    "Embedded Technology": "嵌入式技术（Embedded Technology）",
    "Cloud Security": "云安全（Cloud Security）",
    "Compliance": "合规（Compliance）",
    "Configuration Management": "配置管理（Configuration Management）",
    "Continuous Monitoring": "持续监测（Continuous Monitoring）",
    "Cryptographic Protections": "密码保护（Cryptographic Protections）",
    "Data Classification & Handling": "数据分类与处理（Data Classification & Handling）",
    "Endpoint Security": "终端安全（Endpoint Security）",
    "Human Resources Security": "人力资源安全（Human Resources Security）",
    "Identification & Authentication": "身份识别与认证（Identification & Authentication）",
    "Incident Response": "事件响应（Incident Response）",
    "Information Assurance": "信息保障（Information Assurance）",
    "Maintenance": "维护（Maintenance）",
    "Mobile Device Management": "移动设备管理（Mobile Device Management）",
    "Network Security": "网络安全（Network Security）",
    "Physical & Environmental Security": "物理与环境安全（Physical & Environmental Security）",
    "Data Privacy": "数据隐私（Data Privacy）",
    "Project & Resource Management": "项目与资源管理（Project & Resource Management）",
    "Risk Management": "风险管理（Risk Management）",
    "Secure Engineering & Architecture": "安全工程与架构（Secure Engineering & Architecture）",
    "Security Operations": "安全运营（Security Operations）",
    "Security Awareness & Training": "安全意识与培训（Security Awareness & Training）",
    "Technology Development & Acquisition": "技术开发与采购（Technology Development & Acquisition）",
    "Third-Party Management": "第三方管理（Third-Party Management）",
    "Threat Management": "威胁管理（Threat Management）",
    "Vulnerability & Patch Management": "漏洞与补丁管理（Vulnerability & Patch Management）",
    "Web Security": "Web安全（Web Security）",
}

POST_GLOSSARY = {
    "合规性": "合规",
    "弹性": "韧性",
    "安全、合规和韧性": "安全、合规与韧性",
    "安全性、合规性和弹性": "安全、合规与韧性",
    "安全、合规性和弹性": "安全、合规与韧性",
    "安全、合规性和韧性": "安全、合规与韧性",
    "安全性、合规和/或韧性": "安全、合规和/或韧性",
    "安全性、合规和韧性": "安全、合规与韧性",
    "安全性、合规与韧性": "安全、合规与韧性",
    "网络安全和数据保护": "网络安全与数据保护",
    "网络安全和数据隐私": "网络安全与数据隐私",
    "指导委员会或咨询委员会": "指导委员会或顾问委员会",
    "咨询委员会": "顾问委员会",
    "行动计划和里程碑": "行动计划与里程碑",
    "技术资产、应用程序和/或服务": "技术资产、应用和/或服务",
    "技术资产，应用和/或服务": "技术资产、应用和/或服务",
    "应用程序": "应用",
    "访问控制": "访问控制",
    "控制措施": "控制",
    "控制项项": "控制项",
    "Web 安全": "Web安全",
    "AI ": "AI ",
    "管理机构": "治理机构",
    "高管": "负责人",
    "受管制": "受监管",
    "结清总结": "关闭总结",
    "如 果": "如果",
    "Indicators of Compromise": "失陷指标",
}

ACRONYM_FIXES = {
    "scrp": "SCRP",
    "taas": "TAAS",
    "ai": "AI",
    "api": "API",
    "sdlc": "SDLC",
    "devsecops": "DevSecOps",
    "sbom": "SBOM",
    "mfa": "MFA",
    "sso": "SSO",
    "vpn": "VPN",
    "dns": "DNS",
    "http": "HTTP",
    "https": "HTTPS",
    "tls": "TLS",
    "ip": "IP",
    "url": "URL",
    "iot": "IoT",
    "ot": "OT",
}


def normalize_cn(text: str | None) -> str | None:
    if text is None:
        return None
    value = str(text).strip()
    for old, new in POST_GLOSSARY.items():
        value = value.replace(old, new)
    for old, new in ACRONYM_FIXES.items():
        value = re.sub(rf"\b{old}\b", new, value, flags=re.IGNORECASE)
    value = value.replace(" ;", "；").replace(" :", "：")
    value = value.replace(",", "，").replace(";", "；").replace("?", "？")
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"([\u4e00-\u9fff])\s+([\u4e00-\u9fff])", r"\1\2", value)
    value = re.sub(r"([\u4e00-\u9fff])\s+([，。；：？）])", r"\1\2", value)
    value = re.sub(r"([（])\s+([\u4e00-\u9fffA-Za-z0-9])", r"\1\2", value)
    value = re.sub(r"促进(.+?)的实施", r"促进\1实施", value)
    return value.strip()


def load_cache(cache_path: Path) -> dict[str, str]:
    if not cache_path.exists():
        return {}
    return json.loads(cache_path.read_text(encoding="utf-8"))


def save_cache(cache_path: Path, cache: dict[str, str]) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def google_translate_batch(texts: list[str]) -> list[str]:
    if not texts:
        return []
    markers = [f"@@SAPD_SPLIT_{i:04d}@@" for i in range(1, len(texts))]
    payload_parts: list[str] = []
    for i, text in enumerate(texts):
        if i:
            payload_parts.append(markers[i - 1])
        payload_parts.append(text)
    query = "\n".join(payload_parts)
    params = urlencode(
        {"client": "gtx", "sl": "en", "tl": "zh-CN", "dt": "t", "q": query}
    )
    url = f"https://translate.googleapis.com/translate_a/single?{params}"
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=30) as response:
        data = json.loads(response.read().decode("utf-8"))
    translated = "".join(part[0] for part in data[0])
    pieces = [translated]
    for marker in markers:
        if marker not in pieces[-1]:
            raise RuntimeError(f"Batch marker lost during translation: {marker}")
        head, tail = pieces[-1].split(marker, 1)
        pieces[-1] = head
        pieces.append(tail)
    return [normalize_cn(piece) or "" for piece in pieces]


def build_batches(texts: list[str], max_chars: int) -> list[list[str]]:
    batches: list[list[str]] = []
    current: list[str] = []
    current_len = 0
    for text in texts:
        projected = current_len + len(text) + 24
        if current and projected > max_chars:
            batches.append(current)
            current = []
            current_len = 0
        current.append(text)
        current_len += len(text) + 24
    if current:
        batches.append(current)
    return batches


def translate_missing(cache: dict[str, str], texts: list[str], max_chars: int) -> None:
    missing = [text for text in texts if text and text not in cache]
    batches = build_batches(missing, max_chars=max_chars)
    print(f"missing={len(missing)} batches={len(batches)}", flush=True)
    for index, batch in enumerate(batches, start=1):
        for attempt in range(1, 4):
            try:
                translated = google_translate_batch(batch)
                break
            except Exception as exc:
                if attempt == 3:
                    raise
                print(f"batch {index}/{len(batches)} retry {attempt}: {exc}", flush=True)
                time.sleep(2 * attempt)
        for original, zh in zip(batch, translated, strict=True):
            cache[original] = zh
        if index % 10 == 0 or index == len(batches):
            print(f"translated_batches={index}/{len(batches)}", flush=True)
        time.sleep(0.15)


def load_2024_exact_cn(wb, source_2024_path: Path) -> dict[str, dict[str, str]]:
    ws_cn = wb["DSP2级策略清单（2024年版本）"]
    cn_by_code: dict[str, dict[str, str]] = {}
    for row in range(3, ws_cn.max_row + 1):
        code = ws_cn.cell(row, 5).value
        if code:
            cn_by_code[str(code).strip()] = {
                "control": ws_cn.cell(row, 6).value,
                "desc": ws_cn.cell(row, 7).value,
            }

    source_wb = load_workbook(source_2024_path, read_only=True, data_only=True)
    source_ws = source_wb["SCF 2024.3"]
    exact: dict[str, dict[str, str]] = {}
    for record in source_ws.iter_rows(min_row=2, values_only=True):
        code = str(record[2] or "").strip()
        if not code or "." in code or code not in cn_by_code:
            continue
        exact[code] = {
            "en_control": str(record[1] or "").strip(),
            "en_desc": str(record[3] or "").strip(),
            "cn_control": cn_by_code[code]["control"],
            "cn_desc": cn_by_code[code]["desc"],
        }
    return exact


def clause_for_description(text: str | None) -> tuple[str, str] | None:
    if not text:
        return None
    value = str(text).strip()
    match = re.match(r"^Mechanisms exist to (.+?)\.?$", value, flags=re.IGNORECASE)
    if match:
        return ("mechanism_to", match.group(1))
    match = re.match(r"^Mechanisms exist for (.+?)\.?$", value, flags=re.IGNORECASE)
    if match:
        return ("mechanism_for", match.group(1))
    return None


def render_description(text: str | None, cache: dict[str, str]) -> str | None:
    clause = clause_for_description(text)
    if clause:
        kind, body = clause
        zh = normalize_cn(cache.get(body, body))
        if kind == "mechanism_to":
            return f"具有{zh}的机制。"
        return f"具有用于{zh}的机制。"
    if not text:
        return None
    value = normalize_cn(cache.get(str(text).strip(), str(text).strip()))
    if value and not value.endswith(("。", "？")):
        value += "。"
    return value


def clause_for_question(text: str | None) -> str | None:
    if not text:
        return None
    match = re.match(r"^Does the organization (.+?)\?$", str(text).strip(), flags=re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def render_question(text: str | None, cache: dict[str, str]) -> str | None:
    clause = clause_for_question(text)
    if clause:
        body = normalize_cn(cache.get(clause, clause))
        return f"组织是否{body}？"
    if not text:
        return None
    value = normalize_cn(cache.get(str(text).strip(), str(text).strip()))
    if value and not value.endswith("？"):
        value = value.rstrip("。") + "？"
    return value


def collect_translation_units(ws) -> list[str]:
    units: set[str] = set()
    for row in range(3, ws.max_row + 1):
        for col in (3, 4, 6):
            value = ws.cell(row, col).value
            if value:
                units.add(str(value).strip())
        desc = ws.cell(row, 7).value
        clause = clause_for_description(desc)
        units.add(clause[1] if clause else str(desc or "").strip())
        question = ws.cell(row, 9).value
        q_clause = clause_for_question(question)
        units.add(q_clause if q_clause else str(question or "").strip())
    return sorted(unit for unit in units if unit)


def apply_translation(args: argparse.Namespace) -> None:
    workbook_path = Path(args.workbook)
    cache_path = Path(args.cache)
    backup_dir = workbook_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{workbook_path.stem}.before-dsp-2026-external-translation-{datetime.now():%Y%m%d%H%M%S}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup_path)
    print(f"backup={backup_path}", flush=True)

    wb = load_workbook(workbook_path)
    ws = wb[SHEET_NAME]
    if args.postprocess_only:
        for row in range(3, ws.max_row + 1):
            for col in (2, 3, 4, 6, 7, 8, 9):
                value = ws.cell(row, col).value
                if isinstance(value, str):
                    ws.cell(row, col).value = normalize_cn(value)
        wb.save(workbook_path)
        print(json.dumps({"postprocessed_rows": ws.max_row - 2}, ensure_ascii=False), flush=True)
        return

    sample_rows = range(3, min(ws.max_row, 80) + 1)
    sample_values = [str(ws.cell(row, 6).value or "") + str(ws.cell(row, 7).value or "") for row in sample_rows]
    cjk_sample_count = sum(1 for value in sample_values if re.search(r"[\u4e00-\u9fff]", value))
    if cjk_sample_count > len(sample_values) * 0.6 and not args.force:
        raise RuntimeError(
            "Target sheet already looks translated. Use --postprocess-only for glossary cleanup, "
            "or --force to translate anyway."
        )
    exact_2024 = load_2024_exact_cn(wb, SOURCE_2024_PATH)

    cache = load_cache(cache_path)
    units = collect_translation_units(ws)
    translate_missing(cache, units, max_chars=args.max_chars)
    save_cache(cache_path, cache)

    stats = {"rows": 0, "reused_2024": 0, "translated": 0, "domains": 0}
    for row in range(3, ws.max_row + 1):
        stats["rows"] += 1
        domain = ws.cell(row, 2).value
        if domain:
            domain_text = str(domain).strip()
            ws.cell(row, 2).value = DOMAIN_TRANSLATIONS.get(
                domain_text, normalize_cn(cache.get(domain_text, domain_text))
            )
            for col in (3, 4):
                original = str(ws.cell(row, col).value or "").strip()
                ws.cell(row, col).value = normalize_cn(cache.get(original, original))
            stats["domains"] += 1

        code = str(ws.cell(row, 5).value or "").strip()
        en_control = str(ws.cell(row, 6).value or "").strip()
        en_desc = str(ws.cell(row, 7).value or "").strip()

        reused = False
        if code in exact_2024:
            match = exact_2024[code]
            if en_control == match["en_control"] and en_desc == match["en_desc"]:
                ws.cell(row, 6).value = match["cn_control"]
                ws.cell(row, 7).value = match["cn_desc"]
                reused = True
                stats["reused_2024"] += 1

        if not reused:
            ws.cell(row, 6).value = normalize_cn(cache.get(en_control, en_control))
            ws.cell(row, 7).value = render_description(en_desc, cache)
            stats["translated"] += 1

        ws.cell(row, 8).value = f"{code} {ws.cell(row, 6).value}" if code else ws.cell(row, 6).value
        ws.cell(row, 9).value = render_question(ws.cell(row, 9).value, cache)

    ws["H1"].value = 2026
    wb.save(workbook_path)
    save_cache(cache_path, cache)
    print(json.dumps(stats, ensure_ascii=False), flush=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default="data/raw-samples/wiki sample.xlsx")
    parser.add_argument(
        "--cache",
        default="data/processed/translation-cache/dsp-2026-google-zh.json",
    )
    parser.add_argument("--max-chars", type=int, default=3200)
    parser.add_argument("--postprocess-only", action="store_true")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()
    apply_translation(args)


if __name__ == "__main__":
    main()
