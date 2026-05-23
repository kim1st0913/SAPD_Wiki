#!/usr/bin/env python3
"""Verify DSP 2026 Chinese translation with a second external translation engine."""

from __future__ import annotations

import argparse
import json
import re
import time
from collections import Counter
from pathlib import Path
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


SHEET_NAME = "DSP策略清单（2026）"
REPORT_PATH = Path("data/processed/translation-cache/dsp-2026-translation-verification.json")

GLOSSARY = {
    "resilience": "韧性",
    "compliance": "合规",
    "governance": "治理",
    "control": "控制",
    "controls": "控制",
    "mechanism": "机制",
    "mechanisms": "机制",
    "Technology Assets, Applications and/or Services": "技术资产、应用和/或服务",
    "TAAS": "TAAS",
    "personally identifiable information": "个人身份信息",
    "personal data": "个人数据",
    "sensitive/regulated data": "敏感/受监管数据",
    "indicators of compromise": "失陷指标",
    "incident response": "事件响应",
    "vulnerability": "漏洞",
    "patch management": "补丁管理",
    "third-party": "第三方",
    "supply chain": "供应链",
    "secure engineering": "安全工程",
    "network segmentation": "网络分段",
    "access control": "访问控制",
}

ZH_TERMS = {
    "韧性",
    "合规",
    "治理",
    "控制",
    "机制",
    "技术资产、应用和/或服务",
    "个人数据",
    "敏感/受监管",
    "失陷指标",
    "事件响应",
    "漏洞",
    "补丁",
    "第三方",
    "供应链",
    "安全工程",
    "网络分段",
    "访问控制",
}

ALLOWED_ENGLISH = {
    "SCF",
    "SCRP",
    "TAAS",
    "TAASD",
    "AI",
    "API",
    "DNS",
    "TLS",
    "HTTP",
    "HTTPS",
    "TCP",
    "UDP",
    "IP",
    "URL",
    "Web",
    "SDLC",
    "CI",
    "CD",
    "IoT",
    "OT",
    "GRC",
    "SCRM",
    "POAM",
    "POA",
    "RPO",
    "RTO",
    "SaaS",
    "PaaS",
    "IaaS",
    "BYOD",
    "MDM",
    "MFA",
    "SMS",
    "SSO",
    "VPN",
    "WAF",
    "XML",
    "SQL",
    "NTP",
    "GPS",
    "RFID",
    "NFC",
    "COTS",
    "DevSecOps",
    "SBOM",
    "VLAN",
    "WiFi",
    "SIEM",
    "SOC",
    "IPv6",
    "IPv4",
    "LLM",
    "OWASP",
    "CVE",
    "CVSS",
    "CIS",
    "NIST",
    "ISO",
    "GDPR",
    "CCPA",
    "HIPAA",
    "PCI",
    "DSS",
    "GOV",
    "AAT",
    "AST",
    "BCD",
    "CAP",
    "CHG",
    "CLD",
    "CPL",
    "CFG",
    "MON",
    "CRY",
    "DCH",
    "END",
    "HRS",
    "IAC",
    "IRO",
    "IAO",
    "MNT",
    "NET",
    "PES",
    "PRI",
    "PRM",
    "RSK",
    "SEA",
    "OPS",
    "SAT",
    "TDA",
    "TPM",
    "THR",
    "VPM",
    "WEB",
    "IOC",
    "IoC",
    "PKI",
    "AAA",
    "SORN",
    "DoS",
    "SOP",
    "FIM",
    "ACL",
    "PIV",
    "CPE",
    "FOCI",
    "DPIA",
    "ABAC",
    "IDS",
    "IPS",
    "DLP",
}


def external_translate(text: str, cache: dict[str, str], engine: str, max_external_chars: int) -> str:
    if len(text) > max_external_chars:
        return "__SKIPPED_LONG_TEXT__"
    cache_key = f"{engine}:{text}"
    if cache_key in cache:
        return cache[cache_key]
    if engine == "mymemory":
        url = "https://api.mymemory.translated.net/get?" + urlencode(
            {"q": text, "langpair": "en|zh-CN"}
        )
    elif engine == "lingva":
        url = "https://lingva.ml/api/v1/en/zh/" + quote(text)
    elif engine == "lingva-lunar":
        url = "https://lingva.lunar.icu/api/v1/en/zh/" + quote(text)
    else:
        raise ValueError(f"Unsupported engine: {engine}")
    request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urlopen(request, timeout=30) as response:
            data = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        return f"__ENGINE_ERROR__:{type(exc).__name__}:{exc}"
    if engine == "mymemory":
        translated = data.get("responseData", {}).get("translatedText", "")
    else:
        translated = data.get("translation", "")
    cache[cache_key] = translated
    time.sleep(0.25)
    return translated


def english_residue(text: str) -> list[str]:
    words = re.findall(r"\b[A-Za-z]{3,}\b", text or "")
    return [word for word in words if word not in ALLOWED_ENGLISH]


def glossary_hits(text: str) -> set[str]:
    lower = (text or "").lower()
    hits = set()
    for source, target in GLOSSARY.items():
        if source.lower() in lower:
            hits.add(target)
    return hits


def zh_term_hits(text: str) -> set[str]:
    return {term for term in ZH_TERMS if term in (text or "")}


def load_records(workbook: Path, source_workbook: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    source_wb = load_workbook(source_workbook, read_only=True, data_only=True)
    source_ws = source_wb["SCF 2026.1"]
    source_by_code = {
        str(row[2]).strip(): {
            "source_control": str(row[1] or "").strip(),
            "source_desc": str(row[3] or "").strip(),
            "source_question": str(row[10] or "").strip(),
        }
        for row in source_ws.iter_rows(min_row=2, values_only=True)
        if row[2]
    }
    records = []
    for row in range(3, ws.max_row + 1):
        code = str(ws.cell(row, 5).value or "").strip()
        source = source_by_code[code]
        records.append(
            {
                "code": code,
                "control_zh": str(ws.cell(row, 6).value or "").strip(),
                "desc_zh": str(ws.cell(row, 7).value or "").strip(),
                "question_zh": str(ws.cell(row, 9).value or "").strip(),
                **source,
            }
        )
    return records


def select_samples(records: list[dict[str, str]], sample_size: int) -> list[dict[str, str]]:
    priority_prefixes = ("AAT", "PRI", "SEA", "TDA", "TPM", "NET", "IAO", "IRO", "AST", "WEB")
    scored = []
    for index, record in enumerate(records):
        code = record["code"]
        source_text = " ".join(
            [record["source_control"], record["source_desc"], record["source_question"]]
        )
        translated = " ".join([record["control_zh"], record["desc_zh"], record["question_zh"]])
        score = 0
        score += len(source_text) // 120
        score += 3 if any(code.startswith(prefix) for prefix in priority_prefixes) else 0
        score += 3 if "." in code else 0
        score += 5 * len(english_residue(translated))
        score += 2 * len(glossary_hits(source_text) - zh_term_hits(translated))
        scored.append((score, index, record))
    selected = [item[2] for item in sorted(scored, reverse=True)[:sample_size]]
    anchor_codes = {"GOV-01", "GOV-01.1", "AST-31", "WEB-14", "IRO-03", "PRI-02", "TDA-19"}
    by_code = {record["code"]: record for record in records}
    for code in anchor_codes:
        if code in by_code and all(record["code"] != code for record in selected):
            selected.append(by_code[code])
    return selected


def verify(
    records: list[dict[str, str]],
    sample_size: int,
    cache_path: Path,
    engine: str,
    max_external_chars: int,
) -> dict:
    cache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.exists() else {}
    samples = select_samples(records, sample_size)
    issues = []
    external_checked = []
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    for index, record in enumerate(samples, start=1):
        source_units = {
            "control": record["source_control"],
            "desc": record["source_desc"],
            "question": record["source_question"],
        }
        external = {}
        for key, value in source_units.items():
            external[key] = external_translate(value, cache, engine, max_external_chars)
            cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        if index % 5 == 0 or index == len(samples):
            print(f"checked={index}/{len(samples)}", flush=True)
        combined_source = " ".join(source_units.values())
        combined_zh = " ".join([record["control_zh"], record["desc_zh"], record["question_zh"]])
        residue = english_residue(combined_zh)
        missing_terms = sorted(glossary_hits(combined_source) - zh_term_hits(combined_zh))
        suspicious = []
        if residue:
            suspicious.append(f"存在英文残留：{', '.join(sorted(set(residue))[:8])}")
        if missing_terms:
            suspicious.append(f"关键术语未命中：{', '.join(missing_terms)}")
        if len(record["desc_zh"]) < max(12, len(record["source_desc"]) * 0.18):
            suspicious.append("描述译文长度异常偏短")
        engine_errors = [value for value in external.values() if value.startswith("__ENGINE_ERROR__")]
        if engine_errors:
            suspicious.append("外部引擎部分字段请求失败")
        external_checked.append(
            {
                "code": record["code"],
                "source_control": record["source_control"],
                "current_control_zh": record["control_zh"],
                "external_control_zh": external["control"],
                "current_desc_zh": record["desc_zh"],
                "external_desc_zh": external["desc"],
                "current_question_zh": record["question_zh"],
                "external_question_zh": external["question"],
                "flags": suspicious,
            }
        )
        if suspicious:
            issues.append(external_checked[-1])
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "engine": engine,
        "sample_size": len(samples),
        "issue_count": len(issues),
        "issues": issues,
        "checked": external_checked,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default="data/raw-samples/wiki sample.xlsx")
    parser.add_argument(
        "--source",
        default="/Users/kim1st/Documents/work@qax/01.战略咨询规划部/技术架构组/安全规划方法论设计研究/架构参考材料/网络安全框架材料/ComplianceForge/SCF 2026/Secure Controls Framework (SCF) - 2026.1.1.xlsx",
    )
    parser.add_argument("--sample-size", type=int, default=80)
    parser.add_argument(
        "--cache",
        default="data/processed/translation-cache/dsp-2026-verification-engine-zh.json",
    )
    parser.add_argument(
        "--engine",
        choices=["mymemory", "lingva", "lingva-lunar"],
        default="lingva",
    )
    parser.add_argument("--max-external-chars", type=int, default=420)
    parser.add_argument("--report", default=str(REPORT_PATH))
    args = parser.parse_args()

    records = load_records(Path(args.workbook), Path(args.source))
    result = verify(
        records,
        args.sample_size,
        Path(args.cache),
        args.engine,
        args.max_external_chars,
    )
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "records": len(records),
                "engine": args.engine,
                "sample_size": result["sample_size"],
                "issue_count": result["issue_count"],
                "report": str(report_path),
            },
            ensure_ascii=False,
        )
    )
    for issue in result["issues"][:20]:
        print(issue["code"], "；".join(issue["flags"]))


if __name__ == "__main__":
    main()
