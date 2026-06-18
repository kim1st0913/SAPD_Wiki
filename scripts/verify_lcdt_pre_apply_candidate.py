#!/usr/bin/env python3
from __future__ import annotations

import copy
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from audit_lcdt_source_update import OUT_DIR, audit, write_json, write_md
from build_lcdt_update_candidate import build_candidate


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "frontend/capability-browser/public/data"
WORKBOOK_PATH = ROOT / "data/raw-samples/wiki sample.xlsx"
BACKUP_DIR = OUT_DIR / "source-excel-backup"
PRIMARY_BACKUP = BACKUP_DIR / "wiki sample.before-lcdt-confirmed-source-updates.20260615T112121Z.xlsx"
PRE_APPLY_DIR = OUT_DIR / "pre-apply-candidate"

LIFECYCLE_SHEET = "LC-DT 数据生命周期"
MAPPING_SHEET = "LC-DT 安全技术服务、模块、策略映射表"
WATCH_TERMS = [
    "数据加密",
    "数据加密和令牌化",
    "数据脱敏",
    "数据脱敏(去标识化)",
    "静态数据脱敏",
    "应用动态数据脱敏",
    "应用页面水印",
    "数据内容水印",
]

EXPECTED_CELL_REASONS = {
    f"{LIFECYCLE_SHEET}!H12": "服务编号按字典从 I-DI&T-AS.AD-03 数据备份 修正为 I-DI&T-AS.DG-03 数据备份。",
    f"{LIFECYCLE_SHEET}!H22": "服务名称按字典修正为 数据内容水印 / 静态数据脱敏，保留网络动态数据脱敏。",
    f"{LIFECYCLE_SHEET}!H26": "服务名称按字典修正为 静态数据脱敏 / 数据内容水印，保留网络动态数据脱敏。",
    f"{LIFECYCLE_SHEET}!I12": "模块短名 数据加密 按字典归一为 数据加密和令牌化。",
    f"{LIFECYCLE_SHEET}!I17": "模块短名 数据加密 / 零信任访问代理长名按字典归一。",
    f"{LIFECYCLE_SHEET}!I22": "模块短名 数据脱敏 按字典归一，并按用户确认补入 数据流转监测和泄漏防护。",
    f"{LIFECYCLE_SHEET}!I26": "模块短名 数据加密 / 数据脱敏 按字典归一。",
    f"{MAPPING_SHEET}!M27": "服务编号按字典从 I-DI&T-AS.AD-03 数据备份 修正为 I-DI&T-AS.DG-03 数据备份。",
    f"{MAPPING_SHEET}!M29": "服务编号按字典从 I-DI&T-AS.AD-03 数据备份 修正为 I-DI&T-AS.DG-03 数据备份。",
    f"{MAPPING_SHEET}!M39": "服务名称按字典修正为 静态数据脱敏，保留网络动态数据脱敏。",
    f"{MAPPING_SHEET}!M40": "服务名称按字典修正为 数据内容水印。",
    f"{MAPPING_SHEET}!M53": "服务名称按字典修正为 静态数据脱敏，保留网络动态数据脱敏。",
    f"{MAPPING_SHEET}!M55": "服务名称按字典修正为 数据内容水印。",
    f"{MAPPING_SHEET}!N14": "模块短名 数据加密 按字典归一为 数据加密和令牌化。",
    f"{MAPPING_SHEET}!N15": "模块短名 数据加密 按字典归一为 数据加密和令牌化。",
    f"{MAPPING_SHEET}!N16": "零信任访问代理长名按字典归一为 零信任访问代理。",
    f"{MAPPING_SHEET}!N26": "模块短名 数据加密 按字典归一为 数据加密和令牌化。",
    f"{MAPPING_SHEET}!N39": "模块短名 数据脱敏 按字典归一为 数据脱敏(去标识化)。",
    f"{MAPPING_SHEET}!N53": "模块短名 数据脱敏 按字典归一为 数据脱敏(去标识化)。",
    f"{MAPPING_SHEET}!N54": "模块短名 数据加密 按字典归一为 数据加密和令牌化。",
}

BACKUP_PLAN_FILES = [
    "lifecycle-workbench.json",
    "maintenance-knowledge.json",
    "maintenance-index.json",
    "maintenance/services.json",
    "maintenance/measures.json",
]

SOURCE_AUDIT_SHEETS = (LIFECYCLE_SHEET, MAPPING_SHEET)


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def norm(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\xa0", " ").replace("\u3000", " ")
    return text.strip()


def split_lines(value: Any) -> list[str]:
    return [line.strip() for line in norm(value).split("\n") if line.strip()]


def sha256_path(path: Path) -> str | None:
    if not path.exists():
        return None
    return hashlib.sha256(path.read_bytes()).hexdigest()


def stable_hash(value: str) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:16]


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def cell_key(sheet: str, coord: str) -> str:
    return f"{sheet}!{coord}"


def merged_ranges(ws) -> dict[str, str]:
    result: dict[str, str] = {}
    for merged_range in ws.merged_cells.ranges:
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                result[ws.cell(row, col).coordinate] = str(merged_range)
    return result


def compare_workbooks(before_path: Path, after_path: Path, label: str) -> dict[str, Any]:
    before = load_workbook(before_path, read_only=False, data_only=False)
    after = load_workbook(after_path, read_only=False, data_only=False)
    changed_cells: list[dict[str, Any]] = []
    changed_merged_ranges: list[dict[str, Any]] = []
    changed_sheets: set[str] = set()
    all_sheets = [sheet for sheet in SOURCE_AUDIT_SHEETS if sheet in before.sheetnames or sheet in after.sheetnames]
    for sheet_name in all_sheets:
        if sheet_name not in before.sheetnames or sheet_name not in after.sheetnames:
            changed_sheets.add(sheet_name)
            changed_merged_ranges.append(
                {
                    "sheet": sheet_name,
                    "before": "missing" if sheet_name not in before.sheetnames else "present",
                    "after": "missing" if sheet_name not in after.sheetnames else "present",
                    "risk": "blocking",
                }
            )
            continue
        ws_before = before[sheet_name]
        ws_after = after[sheet_name]
        before_ranges = {str(rng) for rng in ws_before.merged_cells.ranges}
        after_ranges = {str(rng) for rng in ws_after.merged_cells.ranges}
        if before_ranges != after_ranges:
            changed_sheets.add(sheet_name)
            changed_merged_ranges.append(
                {
                    "sheet": sheet_name,
                    "beforeOnly": sorted(before_ranges - after_ranges),
                    "afterOnly": sorted(after_ranges - before_ranges),
                    "risk": "blocking",
                }
            )
        before_merge_by_cell = merged_ranges(ws_before)
        after_merge_by_cell = merged_ranges(ws_after)
        max_row = max(ws_before.max_row, ws_after.max_row)
        max_col = max(ws_before.max_column, ws_after.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                before_value = ws_before.cell(row, col).value
                after_value = ws_after.cell(row, col).value
                if before_value == after_value:
                    continue
                coord = ws_after.cell(row, col).coordinate
                key = cell_key(sheet_name, coord)
                expected = key in EXPECTED_CELL_REASONS
                changed_sheets.add(sheet_name)
                changed_cells.append(
                    {
                        "comparison": label,
                        "sheet": sheet_name,
                        "cell": coord,
                        "row": row,
                        "column": ws_after.cell(row, col).column_letter,
                        "beforeValue": norm(before_value),
                        "afterValue": norm(after_value),
                        "mergedRangeBefore": before_merge_by_cell.get(coord),
                        "mergedRangeAfter": after_merge_by_cell.get(coord),
                        "expectedChange": expected,
                        "reason": EXPECTED_CELL_REASONS.get(key, "未登记为本轮预期源表修改。"),
                        "risk": "ok" if expected else "blocking",
                    }
                )
    unexpected = [row for row in changed_cells if not row["expectedChange"]]
    blocking = unexpected + [row for row in changed_merged_ranges if row.get("risk") == "blocking"]
    return {
        "comparison": label,
        "beforeWorkbook": str(before_path.relative_to(ROOT)),
        "afterWorkbook": str(after_path.relative_to(ROOT)),
        "changedSheets": sorted(changed_sheets),
        "changedCells": changed_cells,
        "changedMergedRanges": changed_merged_ranges,
        "expectedChanges": [row for row in changed_cells if row["expectedChange"]],
        "unexpectedChanges": unexpected,
        "replacementRiskItems": [row for row in changed_cells if replacement_risk(row["beforeValue"], row["afterValue"]) != "ok"],
        "blockingIssues": blocking,
    }


def replacement_risk(before_value: str, after_value: str) -> str:
    before_lines = split_lines(before_value)
    after_lines = split_lines(after_value)
    for line in after_lines:
        if "网络动态数据脱敏(去标识化)" in line or "静态数据脱敏(去标识化)" in line:
            return "blocking"
        if "数据内容水印(去标识化)" in line or "数据加密和令牌化和令牌化" in line:
            return "blocking"
    if any("应用动态数据脱敏" in line for line in after_lines):
        return "blocking"
    if any("应用页面水印" in line for line in after_lines):
        return "blocking"
    if before_lines and after_lines and before_lines != after_lines:
        return "review"
    return "ok"


def earliest_backup() -> Path:
    candidates = sorted(BACKUP_DIR.glob("wiki sample.before-lcdt-confirmed-source-updates.*.xlsx"))
    return candidates[0] if candidates else PRIMARY_BACKUP


def field_role(sheet: str, column: str) -> str:
    if sheet == LIFECYCLE_SHEET and column == "H":
        return "lifecycle_security_technical_service"
    if sheet == LIFECYCLE_SHEET and column == "I":
        return "lifecycle_module_or_measure"
    if sheet == MAPPING_SHEET and column == "M":
        return "mapping_security_technical_service"
    if sheet == MAPPING_SHEET and column == "N":
        return "mapping_module_or_measure"
    return "other"


def scan_replacement_terms(before_path: Path, after_path: Path) -> dict[str, Any]:
    before = load_workbook(before_path, read_only=False, data_only=False)
    after = load_workbook(after_path, read_only=False, data_only=False)
    rows: list[dict[str, Any]] = []
    blocking: list[dict[str, Any]] = []
    review: list[dict[str, Any]] = []
    for sheet_name in SOURCE_AUDIT_SHEETS:
        if sheet_name not in before.sheetnames or sheet_name not in after.sheetnames:
            continue
        ws_before = before[sheet_name]
        ws_after = after[sheet_name]
        after_merged = merged_ranges(ws_after)
        max_row = max(ws_before.max_row, ws_after.max_row)
        max_col = max(ws_before.max_column, ws_after.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                before_value = norm(ws_before.cell(row, col).value)
                after_value = norm(ws_after.cell(row, col).value)
                combined = "\n".join([before_value, after_value])
                if not any(term in combined for term in WATCH_TERMS):
                    continue
                coord = ws_after.cell(row, col).coordinate
                key = cell_key(sheet_name, coord)
                expected = key in EXPECTED_CELL_REASONS
                role = field_role(sheet_name, ws_after.cell(row, col).column_letter)
                is_lcdt_target_field = role in {
                    "lifecycle_security_technical_service",
                    "lifecycle_module_or_measure",
                    "mapping_security_technical_service",
                    "mapping_module_or_measure",
                }
                risk = "ok"
                reason = "词项未变化或变化属于预期字典归一。"
                before_lines = split_lines(before_value)
                after_lines = split_lines(after_value)
                if before_value != after_value and not expected:
                    risk = "blocking"
                    reason = "该词项所在单元格发生变化，但未登记为本轮预期修改。"
                if any("网络动态数据脱敏(去标识化)" in line for line in after_lines):
                    risk = "blocking"
                    reason = "网络动态数据脱敏被错误附加去标识化后缀。"
                if any("静态数据脱敏(去标识化)" in line for line in after_lines):
                    risk = "blocking"
                    reason = "静态数据脱敏被错误替换为模块名。"
                if is_lcdt_target_field and any(line == "应用动态数据脱敏" or line.endswith(" 应用动态数据脱敏") for line in after_lines):
                    risk = "blocking"
                    reason = "LC-DT 源表仍残留旧服务名 应用动态数据脱敏。"
                if is_lcdt_target_field and any(line == "应用页面水印" or line.endswith(" 应用页面水印") for line in after_lines):
                    risk = "blocking"
                    reason = "LC-DT 源表仍残留旧服务名 应用页面水印。"
                if any("数据加密和令牌化和令牌化" in line for line in after_lines):
                    risk = "blocking"
                    reason = "数据加密发生重复替换。"
                if before_lines != after_lines and risk == "ok" and not expected:
                    risk = "review"
                    reason = "词项发生变化，需要人工复核。"
                item = {
                    "sheet": sheet_name,
                    "cell": coord,
                    "row": row,
                    "column": ws_after.cell(row, col).column_letter,
                    "beforeValue": before_value,
                    "afterValue": after_value,
                    "mergedRange": after_merged.get(coord),
                    "fieldRole": role,
                    "expectedChange": expected,
                    "reason": EXPECTED_CELL_REASONS.get(key, reason),
                    "risk": risk,
                }
                rows.append(item)
                if risk == "blocking":
                    blocking.append(item)
                elif risk == "review":
                    review.append(item)
    return {
        "version": 1,
        "generatedAt": now_iso(),
        "beforeWorkbook": str(before_path.relative_to(ROOT)),
        "afterWorkbook": str(after_path.relative_to(ROOT)),
        "watchTerms": WATCH_TERMS,
        "matchCount": len(rows),
        "riskCounts": dict(Counter(row["risk"] for row in rows)),
        "items": rows,
        "blockingIssues": blocking,
        "needsUserConfirmation": review,
    }


def compact_object(source: dict[str, Any], object_type: str) -> dict[str, Any]:
    title = source.get("title") or source.get("name") or source.get("service", {}).get("title")
    service = source.get("service") if isinstance(source.get("service"), dict) else source
    code = service.get("code") or source.get("code") or ""
    return {
        "id": service.get("id") or source.get("id") or f"{object_type}:{stable_hash(object_type + ':' + (code or title or ''))}",
        "type": object_type,
        "code": code,
        "name": title,
        "title": title,
        "description": service.get("description") if object_type == "security_technical_service" else source.get("description", ""),
        "category": service.get("category") if object_type == "security_technical_service" else source.get("category"),
        "status": source.get("status") or "active",
        "evidenceRefs": [],
    }


def build_capability_indexes(candidate: dict[str, Any]) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    focus_by_code = {obj.get("code"): obj for obj in candidate["objects"].get("capability_focus", {}).values() if obj.get("code")}
    capability_by_code = {obj.get("code"): obj for obj in candidate["objects"].get("capability", {}).values() if obj.get("code")}
    tree_path = DATA_DIR / "capability-tree.json"
    if tree_path.exists():
        tree = read_json(tree_path)
        for category in tree.get("categories", []):
            for domain in category.get("domains", []):
                for capability in domain.get("capabilities", []):
                    cap_code = capability.get("code")
                    if cap_code and cap_code not in capability_by_code:
                        obj = compact_object(capability, "capability")
                        candidate["objects"].setdefault("capability", {})[obj["id"]] = obj
                        capability_by_code[cap_code] = obj
                    for focus in capability.get("focuses", []):
                        focus_code = focus.get("code")
                        if focus_code and focus_code not in focus_by_code:
                            obj = compact_object(focus, "capability_focus")
                            candidate["objects"].setdefault("capability_focus", {})[obj["id"]] = obj
                            focus_by_code[focus_code] = obj
    return capability_by_code, focus_by_code


def service_focus_code(service_code: str) -> str:
    parts = service_code.split("&")
    return parts[-1].strip() if len(parts) > 1 else service_code.strip()


def capability_code_from_focus(focus_code: str) -> str:
    return re.sub(r"-\d+$", "", focus_code)


def relation_id(relation_type: str, source_id: str, target_id: str) -> str:
    return f"relation:{stable_hash(relation_type + '|' + source_id + '|' + target_id)}"


def make_relation(relation_type: str, source: dict[str, Any], target: dict[str, Any], label: str, confidence: str = "explicit") -> dict[str, Any]:
    return {
        "id": relation_id(relation_type, source["id"], target["id"]),
        "type": relation_type,
        "sourceId": source["id"],
        "sourceType": source["type"],
        "targetId": target["id"],
        "targetType": target["type"],
        "label": label,
        "status": "active",
        "confidence": confidence,
        "evidenceRefs": [],
    }


def ensure_target_objects(workbench: dict[str, Any], candidate_update: dict[str, Any]) -> dict[str, dict[str, dict[str, Any]]]:
    objects = workbench["objects"]
    services_by_code = {obj.get("code"): obj for obj in objects.get("security_technical_service", {}).values() if obj.get("code")}
    modules_by_title = {obj.get("title") or obj.get("name"): obj for obj in objects.get("security_technology_module", {}).values() if obj.get("title") or obj.get("name")}
    measures_by_title = {obj.get("title") or obj.get("name"): obj for obj in objects.get("security_technical_measure", {}).values() if obj.get("title") or obj.get("name")}

    maintenance = read_json(DATA_DIR / "maintenance-knowledge.json")
    split_services = read_json(DATA_DIR / "maintenance/services.json")
    split_modules = read_json(DATA_DIR / "maintenance/modules.json")
    split_measures = read_json(DATA_DIR / "maintenance/measures.json")
    for entry in list(maintenance.get("security_technical_services", [])) + list(split_services.get("security_technical_services", [])):
        service = entry.get("service") if isinstance(entry.get("service"), dict) else entry
        code = service.get("code")
        if code and code not in services_by_code:
            obj = compact_object({"service": service}, "security_technical_service")
            objects.setdefault("security_technical_service", {})[obj["id"]] = obj
            services_by_code[code] = obj
    for module in split_modules.get("security_technology_modules", []) + maintenance.get("security_technology_modules", []):
        title = module.get("title") or module.get("name")
        if title and title not in modules_by_title:
            obj = compact_object(module, "security_technology_module")
            objects.setdefault("security_technology_module", {})[obj["id"]] = obj
            modules_by_title[title] = obj
    for measure in split_measures.get("security_technical_measures", []) + maintenance.get("security_technical_measures", []):
        title = measure.get("title") or measure.get("name")
        if title and title not in measures_by_title:
            obj = compact_object(measure, "security_technical_measure")
            objects.setdefault("security_technical_measure", {})[obj["id"]] = obj
            measures_by_title[title] = obj

    missing: list[dict[str, Any]] = []
    for row in candidate_update.get("candidateLifecycleRelations", []):
        if row["targetKind"] == "security_technical_service" and row["targetCode"] not in services_by_code:
            missing.append(row)
        if row["targetKind"] == "security_technology_module" and row["targetTitle"] not in modules_by_title:
            missing.append(row)
        if row["targetKind"] == "security_technical_measure" and row["targetTitle"] not in measures_by_title:
            missing.append(row)
    if missing:
        raise RuntimeError(f"candidate target objects missing: {missing[:5]}")
    return {
        "service": services_by_code,
        "module": modules_by_title,
        "measure": measures_by_title,
    }


def is_data_stage_relation(relation: dict[str, Any], data_stage_ids: set[str]) -> bool:
    return relation.get("sourceType") == "lifecycle_stage" and relation.get("sourceId") in data_stage_ids


def build_candidate_lifecycle_workbench(candidate_update: dict[str, Any]) -> dict[str, Any]:
    current = read_json(DATA_DIR / "lifecycle-workbench.json")
    candidate = copy.deepcopy(current)
    target_indexes = ensure_target_objects(candidate, candidate_update)
    capability_by_code, focus_by_code = build_capability_indexes(candidate)
    stages_by_code = {
        obj.get("code"): obj
        for obj in candidate["objects"].get("lifecycle_stage", {}).values()
        if obj.get("lifecycleType") == "data" and obj.get("code")
    }
    data_stage_ids = {obj["id"] for obj in stages_by_code.values()}
    replaced_relation_types = {"maps_to_service", "implemented_by_module", "uses_measure", "maps_to_focus", "maps_to_capability"}
    kept_relations = [
        relation
        for relation in candidate.get("relations", [])
        if not (is_data_stage_relation(relation, data_stage_ids) and relation.get("type") in replaced_relation_types)
    ]
    relation_by_id = {relation["id"]: relation for relation in kept_relations}

    for row in candidate_update.get("candidateLifecycleRelations", []):
        stage = stages_by_code[row["stageId"]]
        if row["targetKind"] == "security_technical_service":
            target = target_indexes["service"][row["targetCode"]]
            relation = make_relation("maps_to_service", stage, target, "映射安全技术服务")
            relation_by_id[relation["id"]] = relation
            focus_code = service_focus_code(row["targetCode"])
            focus = focus_by_code.get(focus_code)
            if focus:
                focus_relation = make_relation("maps_to_focus", stage, focus, "映射关注点", confidence="derived")
                relation_by_id[focus_relation["id"]] = focus_relation
            capability = capability_by_code.get(capability_code_from_focus(focus_code))
            if capability:
                capability_relation = make_relation("maps_to_capability", stage, capability, "映射能力", confidence="derived")
                relation_by_id[capability_relation["id"]] = capability_relation
        elif row["targetKind"] == "security_technology_module":
            target = target_indexes["module"][row["targetTitle"]]
            relation = make_relation("implemented_by_module", stage, target, "落地安全技术模块")
            relation_by_id[relation["id"]] = relation
        elif row["targetKind"] == "security_technical_measure":
            target = target_indexes["measure"][row["targetTitle"]]
            relation = make_relation("uses_measure", stage, target, "使用安全技术措施")
            relation_by_id[relation["id"]] = relation

    candidate["relations"] = sorted(relation_by_id.values(), key=lambda item: (item.get("sourceType", ""), item.get("sourceId", ""), item.get("type", ""), item.get("targetType", ""), item.get("targetId", "")))
    stats = {key: len(value) for key, value in candidate.get("objects", {}).items() if isinstance(value, dict)}
    stats["objects"] = sum(stats.values())
    stats["relations"] = len(candidate["relations"])
    stats["evidenceRefs"] = len(candidate.get("evidenceRefs", []))
    candidate.setdefault("meta", {}).setdefault("stats", {}).update(stats)
    candidate.setdefault("compatibility", {})["warnings"] = [
        "pre-apply candidate only: generated from current runtime baseline and LC-DT normalized source audit; not applied to formal public/data."
    ]
    return candidate


def direct_lcdt_relation_set(workbench: dict[str, Any]) -> set[tuple[str, str, str, str, str]]:
    objects = workbench.get("objects", {})
    stage_by_id = {
        obj["id"]: obj
        for obj in objects.get("lifecycle_stage", {}).values()
        if obj.get("lifecycleType") == "data"
    }
    result: set[tuple[str, str, str, str, str]] = set()
    for relation in workbench.get("relations", []):
        if relation.get("sourceId") not in stage_by_id:
            continue
        if relation.get("type") not in {"maps_to_service", "implemented_by_module", "uses_measure"}:
            continue
        target = objects.get(relation.get("targetType"), {}).get(relation.get("targetId"), {})
        stage = stage_by_id[relation["sourceId"]]
        result.add(
            (
                stage.get("code") or "",
                stage.get("title") or "",
                relation.get("type") or "",
                relation.get("targetType") or "",
                target.get("code") or target.get("title") or target.get("name") or "",
            )
        )
    return result


def non_lcdt_relation_set(workbench: dict[str, Any]) -> set[tuple[str, str, str, str, str]]:
    objects = workbench.get("objects", {})
    data_stage_ids = {
        obj["id"]
        for obj in objects.get("lifecycle_stage", {}).values()
        if obj.get("lifecycleType") == "data"
    }
    result = set()
    for relation in workbench.get("relations", []):
        if relation.get("sourceId") in data_stage_ids and relation.get("type") in {"maps_to_service", "implemented_by_module", "uses_measure", "maps_to_focus", "maps_to_capability"}:
            continue
        result.add((relation.get("type", ""), relation.get("sourceType", ""), relation.get("sourceId", ""), relation.get("targetType", ""), relation.get("targetId", "")))
    return result


def object_set(workbench: dict[str, Any], object_type: str) -> set[tuple[str, str, str]]:
    return {
        (obj.get("id", ""), obj.get("code", ""), obj.get("title") or obj.get("name") or "")
        for obj in workbench.get("objects", {}).get(object_type, {}).values()
    }


def normalized_diff(current: dict[str, Any], candidate: dict[str, Any]) -> dict[str, Any]:
    current_direct = direct_lcdt_relation_set(current)
    candidate_direct = direct_lcdt_relation_set(candidate)
    added = sorted(candidate_direct - current_direct)
    removed = sorted(current_direct - candidate_direct)
    current_non = non_lcdt_relation_set(current)
    candidate_non = non_lcdt_relation_set(candidate)
    unexpected_non_lcdt = sorted(candidate_non ^ current_non)
    object_changes = []
    for object_type in ["security_technical_service", "security_technology_module", "security_technical_measure", "capability", "capability_focus"]:
        before = object_set(current, object_type)
        after = object_set(candidate, object_type)
        added_objects = sorted(after - before)
        removed_objects = sorted(before - after)
        if added_objects or removed_objects:
            object_changes.append({"objectType": object_type, "added": added_objects, "removed": removed_objects})
    expected_count = len(added) + len(removed) + sum(len(row["added"]) + len(row["removed"]) for row in object_changes)
    unexpected = []
    if unexpected_non_lcdt:
        unexpected.append({"type": "non_lcdt_relation_changed", "items": unexpected_non_lcdt[:50], "count": len(unexpected_non_lcdt)})
    return {
        "version": 1,
        "generatedAt": now_iso(),
        "comparison": "candidate lifecycle-workbench vs current runtime baseline",
        "lifecycleDtDirectRelationAdded": added,
        "lifecycleDtDirectRelationRemoved": removed,
        "objectChanges": object_changes,
        "expectedChangeCount": expected_count,
        "unexpectedChangeCount": len(unexpected),
        "blockingChangeCount": len(unexpected),
        "unexpectedChanges": unexpected,
        "classificationSummary": {
            "expected_change": expected_count,
            "unexpected_change": len(unexpected),
            "blocking_change": len(unexpected),
        },
        "scopeChecks": {
            "affectsLcAp": bool(unexpected_non_lcdt),
            "affectsCapabilityWorkbench": False,
            "affectsEnvironmentWorkbench": False,
            "affectsStandards": False,
            "affectsMaintenance": False,
        },
        "sensitiveTermChanges": [
            item
            for item in added + removed
            if any(term in " ".join(item) for term in ["数据加密", "数据脱敏", "静态数据脱敏", "数据内容水印"])
        ],
    }


def write_candidate_files(candidate_lifecycle: dict[str, Any], diff: dict[str, Any]) -> dict[str, Any]:
    PRE_APPLY_DIR.mkdir(parents=True, exist_ok=True)
    write_json(PRE_APPLY_DIR / "candidate-lifecycle-workbench.json", candidate_lifecycle)
    candidate_names = {
        "maintenance-knowledge.json": "candidate-maintenance-knowledge.json",
        "maintenance-index.json": "candidate-maintenance-index.json",
        "maintenance/services.json": "candidate-maintenance-services.json",
        "maintenance/measures.json": "candidate-maintenance-measures.json",
    }
    for rel, candidate_name in candidate_names.items():
        src = DATA_DIR / rel
        dst = PRE_APPLY_DIR / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        dst.write_bytes(src.read_bytes())
        (PRE_APPLY_DIR / candidate_name).write_bytes(src.read_bytes())
    write_json(PRE_APPLY_DIR / "candidate-normalized-diff.json", diff)
    return {
        "candidateDir": str(PRE_APPLY_DIR.relative_to(ROOT)),
        "files": [
            "candidate-lifecycle-workbench.json",
            "candidate-maintenance-knowledge.json",
            "candidate-maintenance-index.json",
            "candidate-maintenance-services.json",
            "candidate-maintenance-measures.json",
        ],
    }


def backup_plan() -> list[dict[str, Any]]:
    plan = []
    candidate_map = {
        "lifecycle-workbench.json": PRE_APPLY_DIR / "candidate-lifecycle-workbench.json",
        "maintenance-knowledge.json": PRE_APPLY_DIR / "candidate-maintenance-knowledge.json",
        "maintenance-index.json": PRE_APPLY_DIR / "candidate-maintenance-index.json",
        "maintenance/services.json": PRE_APPLY_DIR / "candidate-maintenance-services.json",
        "maintenance/measures.json": PRE_APPLY_DIR / "candidate-maintenance-measures.json",
    }
    for rel in BACKUP_PLAN_FILES:
        current = DATA_DIR / rel
        candidate = candidate_map[rel]
        current_hash = sha256_path(current)
        candidate_hash = sha256_path(candidate)
        will_change = current_hash != candidate_hash
        plan.append(
            {
                "path": f"frontend/capability-browser/public/data/{rel}",
                "currentHash": current_hash,
                "candidateHash": candidate_hash,
                "willModify": will_change,
                "reason": "LC-DT 数据生命周期运行投影候选更新。" if will_change else "候选与当前 runtime baseline 一致，本轮不需要修改。",
                "withinAuthorization": rel == "lifecycle-workbench.json" if will_change else True,
            }
        )
    return plan


def render_source_change_md(payload: dict[str, Any]) -> str:
    lines = ["# LC-DT 源表变更回归审计", ""]
    lines.append(f"- generatedAt: `{payload['generatedAt']}`")
    lines.append(f"- status: `{payload['status']}`")
    for comp in payload["comparisons"]:
        lines.extend(["", f"## {comp['comparison']}", ""])
        lines.append(f"- changedSheets: `{comp['changedSheets']}`")
        lines.append(f"- changedCells: `{len(comp['changedCells'])}`")
        lines.append(f"- unexpectedChanges: `{len(comp['unexpectedChanges'])}`")
        lines.append(f"- blockingIssues: `{len(comp['blockingIssues'])}`")
        for row in comp["changedCells"]:
            lines.append(f"- `{row['sheet']}!{row['cell']}` {row['risk']}: {row['reason']}")
    return "\n".join(lines)


def render_replacement_md(payload: dict[str, Any]) -> str:
    lines = ["# 数据加密 / 数据脱敏专项误替换检查", ""]
    lines.append(f"- generatedAt: `{payload['generatedAt']}`")
    lines.append(f"- matchCount: `{payload['matchCount']}`")
    lines.append(f"- riskCounts: `{payload['riskCounts']}`")
    lines.append(f"- blockingIssues: `{len(payload['blockingIssues'])}`")
    lines.append(f"- needsUserConfirmation: `{len(payload['needsUserConfirmation'])}`")
    lines.extend(["", "## 风险项", ""])
    risks = [item for item in payload["items"] if item["risk"] != "ok"]
    if not risks:
        lines.append("- 无")
    else:
        for item in risks:
            lines.append(f"- `{item['sheet']}!{item['cell']}` {item['risk']}: {item['reason']}")
    return "\n".join(lines)


def render_diff_md(diff: dict[str, Any]) -> str:
    lines = ["# LC-DT pre-apply normalized diff", ""]
    for key, value in diff["classificationSummary"].items():
        lines.append(f"- `{key}`: {value}")
    lines.append(f"- affectsLcAp: `{diff['scopeChecks']['affectsLcAp']}`")
    lines.append(f"- affectsEnvironmentWorkbench: `{diff['scopeChecks']['affectsEnvironmentWorkbench']}`")
    lines.append(f"- affectsStandards: `{diff['scopeChecks']['affectsStandards']}`")
    lines.extend(["", "## LC-DT direct relation added", ""])
    for item in diff["lifecycleDtDirectRelationAdded"][:120]:
        lines.append(f"- `{item}`")
    lines.extend(["", "## LC-DT direct relation removed", ""])
    for item in diff["lifecycleDtDirectRelationRemoved"][:120]:
        lines.append(f"- `{item}`")
    return "\n".join(lines)


def render_backup_plan_md(plan: list[dict[str, Any]]) -> str:
    lines = ["# LC-DT Runtime Package Apply 备份清单", ""]
    for item in plan:
        lines.append(f"- `{item['path']}`")
        lines.append(f"  - willModify: `{item['willModify']}`")
        lines.append(f"  - withinAuthorization: `{item['withinAuthorization']}`")
        lines.append(f"  - currentHash: `{item['currentHash']}`")
        lines.append(f"  - candidateHash: `{item['candidateHash']}`")
        lines.append(f"  - reason: {item['reason']}")
    return "\n".join(lines)


def render_readiness_md(readiness: dict[str, Any]) -> str:
    lines = ["# LC-DT pre-apply readiness", ""]
    lines.append(f"- conclusion: `{readiness['conclusion']}`")
    for key, value in readiness["checks"].items():
        lines.append(f"- `{key}`: {value}")
    if readiness["blockingReasons"]:
        lines.extend(["", "## blockingReasons", ""])
        for reason in readiness["blockingReasons"]:
            lines.append(f"- {reason}")
    if readiness["confirmationReasons"]:
        lines.extend(["", "## confirmationReasons", ""])
        for reason in readiness["confirmationReasons"]:
            lines.append(f"- {reason}")
    return "\n".join(lines)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PRE_APPLY_DIR.mkdir(parents=True, exist_ok=True)
    generated_at = now_iso()

    primary_comparison = compare_workbooks(PRIMARY_BACKUP, WORKBOOK_PATH, "primary_specified_backup")
    full_comparison = compare_workbooks(earliest_backup(), WORKBOOK_PATH, "earliest_available_backup_chain")
    source_payload = {
        "version": 1,
        "generatedAt": generated_at,
        "status": "blocked" if primary_comparison["blockingIssues"] or full_comparison["blockingIssues"] else "pass",
        "changedSheets": sorted(set(primary_comparison["changedSheets"]) | set(full_comparison["changedSheets"])),
        "changedCells": primary_comparison["changedCells"],
        "changedMergedRanges": primary_comparison["changedMergedRanges"],
        "expectedChanges": primary_comparison["expectedChanges"],
        "unexpectedChanges": primary_comparison["unexpectedChanges"],
        "replacementRiskItems": primary_comparison["replacementRiskItems"],
        "blockingIssues": primary_comparison["blockingIssues"],
        "comparisons": [primary_comparison, full_comparison],
    }
    write_json(OUT_DIR / "lcdt-source-excel-change-audit.json", source_payload)
    write_md(OUT_DIR / "lcdt-source-excel-change-audit.md", render_source_change_md(source_payload))

    replacement_payload = scan_replacement_terms(earliest_backup(), WORKBOOK_PATH)
    write_json(OUT_DIR / "lcdt-encryption-desensitization-replacement-audit.json", replacement_payload)
    write_md(OUT_DIR / "lcdt-encryption-desensitization-replacement-audit.md", render_replacement_md(replacement_payload))

    audit_summary = audit()
    update_candidate = build_candidate()
    write_json(OUT_DIR / "lcdt-update-candidate.json", update_candidate)
    current_lifecycle = read_json(DATA_DIR / "lifecycle-workbench.json")
    candidate_lifecycle = build_candidate_lifecycle_workbench(update_candidate)
    diff = normalized_diff(current_lifecycle, candidate_lifecycle)
    write_candidate_files(candidate_lifecycle, diff)
    write_json(PRE_APPLY_DIR / "candidate-normalized-diff.json", diff)
    write_md(PRE_APPLY_DIR / "candidate-normalized-diff.md", render_diff_md(diff))

    apply_scope = {
        "version": 1,
        "generatedAt": generated_at,
        "formalApplyScope": ["frontend/capability-browser/public/data/lifecycle-workbench.json"],
        "excludedScope": [
            "frontend/capability-browser/public/data/maintenance-knowledge.json",
            "frontend/capability-browser/public/data/maintenance-index.json",
            "frontend/capability-browser/public/data/maintenance/services.json",
            "frontend/capability-browser/public/data/maintenance/measures.json",
            "frontend/capability-browser/public/data/capability-workbench.json",
            "frontend/capability-browser/public/data/environment-workbench.json",
            "frontend/capability-browser/public/data/standards-data.json",
            "frontend/capability-browser/public/data/standards-index.json",
            "data/database/sapd_wiki.sqlite3",
            "data/raw-samples/wiki sample.xlsx",
        ],
        "reason": "候选 diff 仅包含 LC-DT 生命周期运行投影关系变更；维护包、能力、环境、标准和 SQLite 不在本轮 apply 范围。",
    }
    write_json(PRE_APPLY_DIR / "candidate-apply-scope.json", apply_scope)
    write_md(
        PRE_APPLY_DIR / "candidate-apply-scope.md",
        "# LC-DT candidate apply scope\n\n"
        + "\n".join(f"- `{item}`" for item in apply_scope["formalApplyScope"])
        + "\n\n## Excluded\n\n"
        + "\n".join(f"- `{item}`" for item in apply_scope["excludedScope"])
        + f"\n\n{apply_scope['reason']}\n",
    )

    plan = backup_plan()
    write_json(PRE_APPLY_DIR / "candidate-backup-plan.json", {"version": 1, "generatedAt": generated_at, "files": plan})
    write_md(PRE_APPLY_DIR / "candidate-backup-plan.md", render_backup_plan_md(plan))

    blocking_reasons = []
    confirmation_reasons = []
    if source_payload["blockingIssues"]:
        blocking_reasons.append("源表变更审计存在 blocking issue。")
    if replacement_payload["blockingIssues"]:
        blocking_reasons.append("数据加密 / 数据脱敏专项检查存在 blocking issue。")
    if audit_summary["status"] != "ready":
        blocking_reasons.append("LC-DT 源表审计未达到 ready。")
    if diff["blockingChangeCount"]:
        blocking_reasons.append("candidate normalized diff 存在 blocking change。")
    if any(item["willModify"] and not item["withinAuthorization"] for item in plan):
        blocking_reasons.append("备份清单中存在越权修改文件。")
    if replacement_payload["needsUserConfirmation"]:
        confirmation_reasons.append("数据加密 / 数据脱敏专项检查存在 review 项。")

    conclusion = "blocked" if blocking_reasons else "needs_user_confirmation" if confirmation_reasons else "ready_for_user_confirmation"
    readiness = {
        "version": 1,
        "generatedAt": generated_at,
        "conclusion": conclusion,
        "checks": {
            "sourceChangeAuditBlockingIssues": len(source_payload["blockingIssues"]),
            "replacementAuditBlockingIssues": len(replacement_payload["blockingIssues"]),
            "replacementAuditReviewItems": len(replacement_payload["needsUserConfirmation"]),
            "lcdtAuditStatus": audit_summary["status"],
            "stagePassed": audit_summary["consistencySummary"].get("stagePassed"),
            "serviceOnlyLifecycle": audit_summary["consistencySummary"].get("serviceOnlyLifecycle"),
            "serviceOnlyMapping": audit_summary["consistencySummary"].get("serviceOnlyMapping"),
            "moduleOnlyLifecycle": audit_summary["consistencySummary"].get("moduleOnlyLifecycle"),
            "moduleOnlyMapping": audit_summary["consistencySummary"].get("moduleOnlyMapping"),
            "serviceReferenceCount": audit_summary["dictionarySummary"].get("serviceReferenceCount"),
            "serviceMatchedCount": audit_summary["dictionarySummary"].get("serviceValidationStatus", {}).get("matched"),
            "newMeasureCandidateCount": audit_summary["dictionarySummary"].get("newMeasureCandidateCount"),
            "candidateExpectedChangeCount": diff["expectedChangeCount"],
            "candidateUnexpectedChangeCount": diff["unexpectedChangeCount"],
            "candidateBlockingChangeCount": diff["blockingChangeCount"],
            "affectsEnvironmentWorkbench": diff["scopeChecks"]["affectsEnvironmentWorkbench"],
            "affectsStandards": diff["scopeChecks"]["affectsStandards"],
            "formalDataReplaced": False,
        },
        "blockingReasons": blocking_reasons,
        "confirmationReasons": confirmation_reasons,
        "candidateDir": str(PRE_APPLY_DIR.relative_to(ROOT)),
    }
    write_json(PRE_APPLY_DIR / "candidate-readiness.json", readiness)
    write_md(PRE_APPLY_DIR / "candidate-readiness.md", render_readiness_md(readiness))

    print(json.dumps(readiness, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
