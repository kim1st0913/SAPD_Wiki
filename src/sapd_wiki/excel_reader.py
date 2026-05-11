from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .paths import display_path, resolve_project_path


CORE_SHEETS = [
    "安全能力目录",
    "安全能力作用域目录",
    "安全能力-安全技术服务",
    "安全技术模块清单",
    "作用域-安全技术服务-安全技术模块映射",
]


@dataclass(frozen=True)
class SheetSummary:
    name: str
    present: bool
    rows: int | None = None
    columns: int | None = None
    header_preview: list[str] | None = None


@dataclass(frozen=True)
class WorkbookSummary:
    path: str
    sheet_count: int
    sheet_names: list[str]
    core_sheets: list[SheetSummary]
    missing_core_sheets: list[str]


def _load_openpyxl():
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required to inspect Excel files. "
            "Install project dependencies or run with the bundled Codex Python runtime."
        ) from exc
    return load_workbook


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _header_preview(ws, max_scan_rows: int = 8, max_values: int = 12) -> list[str]:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows), values_only=True):
        values = [_stringify(value) for value in row]
        non_empty = [value for value in values if value]
        if len(non_empty) >= 2:
            return non_empty[:max_values]
    return []


def inspect_workbook(path: str | Path) -> WorkbookSummary:
    resolved = resolve_project_path(path)
    if not resolved.exists():
        raise FileNotFoundError(f"Excel file not found: {resolved}")

    load_workbook = _load_openpyxl()
    workbook = load_workbook(resolved, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        core_summaries: list[SheetSummary] = []
        for sheet_name in CORE_SHEETS:
            if sheet_name not in workbook.sheetnames:
                core_summaries.append(SheetSummary(name=sheet_name, present=False))
                continue
            ws = workbook[sheet_name]
            core_summaries.append(
                SheetSummary(
                    name=sheet_name,
                    present=True,
                    rows=ws.max_row or 0,
                    columns=ws.max_column or 0,
                    header_preview=_header_preview(ws),
                )
            )
        missing = [sheet.name for sheet in core_summaries if not sheet.present]
        return WorkbookSummary(
            path=display_path(resolved),
            sheet_count=len(sheet_names),
            sheet_names=sheet_names,
            core_sheets=core_summaries,
            missing_core_sheets=missing,
        )
    finally:
        workbook.close()


def workbook_summary_to_dict(summary: WorkbookSummary) -> dict[str, Any]:
    return {
        "path": summary.path,
        "sheet_count": summary.sheet_count,
        "sheet_names": summary.sheet_names,
        "core_sheets": [
            {
                "name": sheet.name,
                "present": sheet.present,
                "rows": sheet.rows,
                "columns": sheet.columns,
                "header_preview": sheet.header_preview or [],
            }
            for sheet in summary.core_sheets
        ],
        "missing_core_sheets": summary.missing_core_sheets,
    }

