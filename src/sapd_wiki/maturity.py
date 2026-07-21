from __future__ import annotations

import hashlib
import html as html_lib
import json
import base64
import math
import re
from io import BytesIO
from copy import deepcopy
from collections import Counter, defaultdict
from datetime import UTC, datetime
from typing import Any


MATURITY_LEVELS = (
    {"id": "L1", "name": "非正式执行", "index": 1},
    {"id": "L2", "name": "计划跟踪", "index": 2},
    {"id": "L3", "name": "充分定义", "index": 3},
    {"id": "L4", "name": "量化控制", "index": 4},
    {"id": "L5", "name": "持续优化", "index": 5},
)

EVIDENCE_LEVELS = (
    {"id": "E0", "name": "无证据"},
    {"id": "E1", "name": "文档证据"},
    {"id": "E2", "name": "配置证据"},
    {"id": "E3", "name": "运行证据"},
    {"id": "E4", "name": "审计证据"},
    {"id": "E5", "name": "持续证据"},
)

PROJECT_STATUS_LABELS = {
    "draft": "草稿",
    "template_configuring": "模板配置中",
    "scoring": "评分中",
    "score_review": "待复核",
    "completed": "评估完成",
    "reported": "已生成报告",
    "archived": "已归档",
}

ASSESSMENT_ITEM_TYPES = {"SERVICE", "FOCUS"}
SERVICE_ROLES = {"ASSESSMENT_POINT", "PLATFORM_EVIDENCE_REFERENCE"}
ALGORITHM_VERSION = "sapd-maturity-v2.2.0"
SCORE_EXCHANGE_SCHEMA = "maturity-score-exchange-v2.2"
TEMPLATE_EXCHANGE_SCHEMA = "maturity-template-exchange-v2.1"
MATURITY_WORKBOOK_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
IMPROVEMENT_ROADMAP_STATUSES = {"待规划", "已确认", "进行中", "已完成", "暂缓"}
MATURITY_INFO_SHEET = "模板基础信息"
MATURITY_ASSESSMENT_SHEET = "评估模板"
MATURITY_WORKBOOK_HEADERS = (
    "安全能力分类",
    "L1 高阶战略能力",
    "L2 安全能力",
    "关注点序号",
    "安全关注点",
    "安全技术服务",
    "评分粒度",
    "适用性",
    "组织与角色",
    "制度与流程",
    "平台与工具",
    "数据与信息",
    "组织与角色说明",
    "制度与流程说明",
    "平台与工具说明",
    "数据与信息说明",
    "目标组织与角色",
    "目标制度与流程",
    "目标平台与工具",
    "目标数据与信息",
    "目标组织与角色说明",
    "目标制度与流程说明",
    "目标平台与工具说明",
    "目标数据与信息说明",
    "其他说明",
)
ELEMENT_KEYS = ("organization", "process", "tool", "data")
ELEMENT_LABELS = {
    "organization": "组织与角色",
    "process": "制度与流程",
    "tool": "平台与工具",
    "data": "数据与信息",
}
CAPABILITY_RADAR_SHORT_LABELS = {
    "T-AS.AD": "体系架构", "T-AS.AM": "资产管理", "T-AS.CM": "配置加固", "T-AS.VM": "漏洞补丁",
    "T-AS.IA": "身份访问", "T-AS.DS": "开发安全", "T-AS.LA": "日志审计", "T-AS.CG": "密码服务",
    "T-AS.DG": "数据治理", "T-PD.PP": "边界防护", "T-PD.AC": "访问控制", "T-PD.TP": "威胁防护",
    "T-PD.DP": "数据防护", "T-AD.SA": "态势感知", "T-AD.IR": "事件响应", "T-AD.SV": "架构评估",
    "T-IN.IO": "情报运营", "T-IN.IP": "情报生产", "T-OF.AT": "进攻反制", "G-SP.SM": "战略管理",
    "M-PM.PL": "安全规划", "M-PM.PR": "安全建设", "M-SA.AM": "安全保障", "M-SA.RM": "风险管理",
    "M-SA.RE": "合规管理", "M-SA.TP": "第三方", "M-SA.OP": "安全运行", "M-SA.CO": "安全协同",
    "M-SE.SE": "监督检查", "M-SE.PE": "绩效考核", "M-PS.HS": "人员管理", "M-PS.CT": "意识培养",
}
RUBRIC_VERSION = "sapd-maturity-generic-rubric-v2.1-2026-07-13"
GENERIC_DIMENSION_RUBRIC = (
    {
        "dimensionCode": "organization",
        "levels": {
            "L1": "没有明确分配责任和职责，安全工作主要依赖临时指派人员，缺乏稳定团队结构和清晰职责定位。",
            "L2": "有人员主动承担部分安全职责，但未正式明确；问题发生时责任边界仍可能不清晰。",
            "L3": "职责和责任已正式分配，各角色定义清楚，流程所有者和责任人明确。",
            "L4": "通过自动化流程促使责任人主动履职并积极参与，形成有效激励和持续参与机制。",
            "L5": "能根据战略变化灵活调整角色责任，责任人自我激励、自驱管理，团队成员高度自主。",
        },
    },
    {
        "dimensionCode": "process",
        "levels": {
            "L1": "流程依靠个人经验，缺乏文档化和标准化；安全要求没有统一明确规定，主要口头传达。",
            "L2": "基于经验形成初步流程或制度，但尚未标准化，执行不严格且覆盖有限。",
            "L3": "已形成标准化运营流程，安全制度正式发布并执行，能够在组织范围内贯彻实施。",
            "L4": "流程健全完整，内嵌内部最佳实践并逐步实现自动化；操作有记录且可重复执行。",
            "L5": "流程、政策和程序高度标准化与自动化，能够随环境变化和威胁发展持续优化改进。",
        },
    },
    {
        "dimensionCode": "tool",
        "levels": {
            "L1": "安全技术工具使用非常有限，覆盖范围窄、功能不完整，主要依靠个别人员手工操作。",
            "L2": "具备基础安全技术措施和工具，但能力有限，尚未覆盖全部关键环节。",
            "L3": "配备完整、专业的安全技术措施，覆盖关键环节和业务流程，并具备必要集成能力。",
            "L4": "各安全技术措施相互融合，实现主要领域的自动化管理和监控，并具备实时响应和调整能力。",
            "L5": "技术措施深度融合，利用人工智能、机器学习等实现高效威胁检测和自适应响应，并与业务高度协同。",
        },
    },
    {
        "dimensionCode": "data",
        "levels": {
            "L1": "安全业务数据没有系统记录和收集，数据分散，无法进行系统分析和利用。",
            "L2": "开始记录和收集部分安全业务数据，但质量和覆盖率有限，难以形成全面、系统的信息支撑。",
            "L3": "全面记录和收集安全业务数据，能够通过技术手段分析处理，形成初步风险事件预警和情报支持。",
            "L4": "安全信息收集系统化，实时支撑安全管理决策和考核，形成闭环管理，并使用分析模型提高决策效率和准确性。",
            "L5": "安全管理决策与考核指标动态调整，能够实时响应安全战略和计划变化，提供前瞻性安全保障。",
        },
    },
)


def _list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _float(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if number == number and abs(number) != float("inf") else default


def _round(value: float | None, digits: int = 2) -> float | None:
    return None if value is None else round(float(value), digits)


def _now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _stable_hash(value: Any, length: int = 16) -> str:
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()[:length]


LEVEL_INDEX = {item["id"]: float(item["index"]) for item in MATURITY_LEVELS}
LEVEL_NAME = {item["id"]: item["name"] for item in MATURITY_LEVELS}


def _rubric_entries_for_item(item_id: str) -> list[dict[str, Any]]:
    return [
        {
            "scoreItemId": item_id,
            "dimensionCode": dimension["dimensionCode"],
            "level": level["id"],
            "levelName": level["name"],
            "criteria": dimension["levels"][level["id"]],
            "sourceType": "GENERIC_FALLBACK",
            "sourceVersion": RUBRIC_VERSION,
        }
        for dimension in GENERIC_DIMENSION_RUBRIC
        for level in MATURITY_LEVELS
    ]


def maturity_level_index(value: Any) -> float | None:
    raw = _text(value).upper()
    if raw in LEVEL_INDEX:
        return LEVEL_INDEX[raw]
    numeric = _float(value, -1.0)
    if 1.0 <= numeric <= 5.0:
        return numeric
    return None


def maturity_level_from_index(value: float | None) -> str:
    if value is None:
        return "Not Scored"
    if value < 1.5:
        return "L1"
    if value < 2.5:
        return "L2"
    if value < 3.5:
        return "L3"
    if value < 4.5:
        return "L4"
    return "L5"


def _weighted_average(rows: list[tuple[float | None, float]]) -> float | None:
    usable = [(value, weight) for value, weight in rows if value is not None and weight > 0]
    if not usable:
        return None
    denominator = sum(weight for _, weight in usable)
    return sum(float(value) * weight for value, weight in usable) / denominator if denominator else None


def _target_achievement_rate(current_index: float | None, target_index: float | None) -> float | None:
    if current_index is None or target_index is None or target_index <= 0:
        return None
    return min(100.0, 100.0 * float(current_index) / float(target_index))


def _object_map(workbench: dict[str, Any], object_type: str) -> dict[str, dict[str, Any]]:
    return {
        _text(key): value
        for key, value in _dict(_dict(workbench.get("objects")).get(object_type)).items()
        if isinstance(value, dict)
    }


def _scope_code_for_service(service: dict[str, Any]) -> str:
    category = _text(service.get("category"))
    if category:
        return category
    code = _text(service.get("code"))
    return code.split("&", 1)[0] if "&" in code else "ALL"


def build_maturity_base_template(capability_workbench: dict[str, Any]) -> dict[str, Any]:
    navigator = _list(_dict(capability_workbench.get("navigator")).get("tree"))
    focus_objects = _object_map(capability_workbench, "capability_focus")
    service_objects = _object_map(capability_workbench, "security_technical_service")
    scope_objects = _object_map(capability_workbench, "scope_type")
    scope_by_code = {_text(item.get("code")): item for item in scope_objects.values()}

    service_ids_by_focus: dict[str, list[str]] = defaultdict(list)
    for relation in _list(capability_workbench.get("relations")):
        if not isinstance(relation, dict) or relation.get("type") != "supports_focus":
            continue
        if relation.get("sourceType") != "security_technical_service" or relation.get("targetType") != "capability_focus":
            continue
        service_id = _text(relation.get("sourceId"))
        focus_id = _text(relation.get("targetId"))
        if service_id and focus_id and service_id not in service_ids_by_focus[focus_id]:
            service_ids_by_focus[focus_id].append(service_id)

    categories: list[dict[str, Any]] = []
    capabilities: list[dict[str, Any]] = []
    focuses: list[dict[str, Any]] = []
    services: dict[str, dict[str, Any]] = {}
    focus_service_mappings: list[dict[str, Any]] = []
    score_items: list[dict[str, Any]] = []

    for top_order, top in enumerate(navigator):
        if not isinstance(top, dict):
            continue
        top_source_id = _text(top.get("id"))
        top_id = f"category:{top_source_id}"
        categories.append(
            {
                "id": top_id,
                "code": _text(top.get("code")) or _text(top.get("name")).rsplit(" ", 1)[-1],
                "name": _text(top.get("name")) or "未命名能力分类",
                "description": "",
                "level": 1,
                "capabilityLevel": "L0",
                "parentId": None,
                "weight": 1,
                "sortOrder": top_order,
                "includedInOverall": True,
                "dictionaryRef": top_source_id,
                "sourceType": "DICTIONARY",
                "sourceSnapshotObjectId": top_source_id,
                "changeAction": "UNCHANGED",
                "originalParentId": None,
                "currentParentId": None,
                "sourceContentHash": _stable_hash(top),
            }
        )
        for domain_order, domain in enumerate(_list(top.get("children"))):
            if not isinstance(domain, dict):
                continue
            domain_source_id = _text(domain.get("id"))
            domain_id = f"domain:{domain_source_id}"
            categories.append(
                {
                    "id": domain_id,
                    "code": _text(domain.get("code")),
                    "name": _text(domain.get("name")) or "未命名能力域",
                    "description": "",
                    "level": 2,
                    "capabilityLevel": "L1",
                    "parentId": top_id,
                    "weight": 1,
                    "sortOrder": domain_order,
                    "includedInOverall": True,
                    "dictionaryRef": domain_source_id,
                    "sourceType": "DICTIONARY",
                    "sourceSnapshotObjectId": domain_source_id,
                    "changeAction": "UNCHANGED",
                    "originalParentId": top_id,
                    "currentParentId": top_id,
                    "sourceContentHash": _stable_hash(domain),
                }
            )
            for capability_order, capability in enumerate(_list(domain.get("children"))):
                if not isinstance(capability, dict):
                    continue
                capability_source_id = _text(capability.get("id"))
                capability_id = f"capability:{capability_source_id}"
                capability_record = {
                    "id": capability_id,
                    "code": _text(capability.get("code")),
                    "name": _text(capability.get("name")) or "未命名安全能力",
                    "description": "",
                    "capabilityLevel": "L2",
                    "categoryId": domain_id,
                    "topCategoryId": top_id,
                    "weight": 1,
                    "sortOrder": capability_order,
                    "included": True,
                    "isCustom": False,
                    "isCritical": False,
                    "businessImportance": 3,
                    "riskUrgency": 3,
                    "targetLevel": "L3",
                    "dictionaryRef": capability_source_id,
                    "sourceType": "DICTIONARY",
                    "sourceSnapshotObjectId": capability_source_id,
                    "changeAction": "UNCHANGED",
                    "originalParentId": domain_id,
                    "currentParentId": domain_id,
                    "sourceContentHash": _stable_hash(capability),
                    "focusIds": [],
                }
                capability_kind = _text(top.get("code")) or _text(top.get("name")).rsplit(" ", 1)[-1]
                for focus_order, focus in enumerate(_list(capability.get("children"))):
                    if not isinstance(focus, dict):
                        continue
                    focus_source_id = _text(focus.get("id"))
                    focus_id = f"focus:{focus_source_id}"
                    focus_object = focus_objects.get(focus_source_id, {})
                    service_source_ids = sorted(
                        service_ids_by_focus.get(focus_source_id, []),
                        key=lambda item_id: (
                            _scope_code_for_service(service_objects.get(item_id, {})),
                            _text(service_objects.get(item_id, {}).get("code")),
                            _text(service_objects.get(item_id, {}).get("name")),
                        ),
                    )
                    item_type = "SERVICE" if capability_kind == "T" and service_source_ids else "FOCUS"
                    focus_record = {
                        "id": focus_id,
                        "code": _text(focus.get("code")),
                        "name": _text(focus.get("name")) or _text(focus_object.get("name")) or "未命名关注点",
                        "description": _text(focus_object.get("description")),
                        "capabilityId": capability_id,
                        "weight": 1,
                        "sortOrder": focus_order,
                        "included": True,
                        "isCustom": False,
                        "isCritical": False,
                        "itemType": item_type,
                        "targetLevel": "L3",
                        "dictionaryRef": focus_source_id,
                        "sourceType": "DICTIONARY",
                        "sourceSnapshotObjectId": focus_source_id,
                        "changeAction": "UNCHANGED",
                        "originalParentId": capability_id,
                        "currentParentId": capability_id,
                        "sourceContentHash": _stable_hash(focus_object or focus),
                        "serviceMappingIds": [],
                        "platformEvidenceServiceIds": [],
                        "scoreItemIds": [],
                    }
                    for service_order, service_source_id in enumerate(service_source_ids):
                        service = service_objects.get(service_source_id, {})
                        service_id = f"service:{service_source_id}"
                        scope_code = _scope_code_for_service(service)
                        scope = scope_by_code.get(scope_code, {})
                        role = "ASSESSMENT_POINT" if capability_kind == "T" else "PLATFORM_EVIDENCE_REFERENCE"
                        services[service_id] = {
                            "id": service_id,
                            "code": _text(service.get("code")),
                            "name": _text(service.get("name")) or "未命名安全技术服务",
                            "scopeCode": scope_code,
                            "scopeName": _text(scope.get("name")) or scope_code,
                            "dictionaryRef": service_source_id,
                            "sourceType": "DICTIONARY",
                            "sourceSnapshotObjectId": service_source_id,
                            "changeAction": "UNCHANGED",
                            "sourceContentHash": _stable_hash(service),
                            "isCustom": False,
                        }
                        mapping_id = f"mapping:{focus_source_id}:{scope_code}:{service_source_id}"
                        focus_service_mappings.append(
                            {
                                "id": mapping_id,
                                "focusId": focus_id,
                                "scopeCode": scope_code,
                                "scopeName": _text(scope.get("name")) or scope_code,
                                "serviceId": service_id,
                                "serviceRole": role,
                                "weight": 1,
                                "sortOrder": service_order,
                                "sourceType": "DICTIONARY",
                                "sourceSnapshotObjectId": service_source_id,
                                "changeAction": "UNCHANGED",
                                "originalParentId": focus_id,
                                "currentParentId": focus_id,
                                "sourceContentHash": _stable_hash((focus_source_id, scope_code, service_source_id)),
                            }
                        )
                        focus_record["serviceMappingIds"].append(mapping_id)
                        if role == "PLATFORM_EVIDENCE_REFERENCE":
                            focus_record["platformEvidenceServiceIds"].append(service_id)
                    if item_type == "SERVICE":
                        for service_order, service_source_id in enumerate(service_source_ids):
                            service = service_objects.get(service_source_id, {})
                            service_id = f"service:{service_source_id}"
                            scope_code = _scope_code_for_service(service)
                            scope = scope_by_code.get(scope_code, {})
                            item_id = f"score:{focus_source_id}:{service_source_id}"
                            score_items.append(
                                {
                                    "id": item_id,
                                    "itemType": "SERVICE",
                                    "capabilityId": capability_id,
                                    "focusId": focus_id,
                                    "serviceId": service_id,
                                    "scopeCode": scope_code,
                                    "scopeName": _text(scope.get("name")) or scope_code,
                                    "weight": 1,
                                    "sortOrder": service_order,
                                    "required": True,
                                    "elementWeights": {key: 0.25 for key in ELEMENT_KEYS},
                                    "rubricEntries": _rubric_entries_for_item(item_id),
                                    "sourceType": "DICTIONARY",
                                    "serviceRole": "ASSESSMENT_POINT",
                                    "sourceMappingId": f"mapping:{focus_source_id}:{scope_code}:{service_source_id}",
                                }
                            )
                            focus_record["scoreItemIds"].append(item_id)
                    else:
                        item_id = f"score:{focus_source_id}:overall"
                        score_items.append(
                            {
                                "id": item_id,
                                "itemType": "FOCUS",
                                "capabilityId": capability_id,
                                "focusId": focus_id,
                                "serviceId": None,
                                "scopeCode": None,
                                "scopeName": None,
                                "weight": 1,
                                "sortOrder": 0,
                                "required": True,
                                "elementWeights": {key: 0.25 for key in ELEMENT_KEYS},
                                "rubricEntries": _rubric_entries_for_item(item_id),
                                "sourceType": "DICTIONARY",
                                "serviceRole": None,
                                "platformEvidenceServiceIds": list(focus_record["platformEvidenceServiceIds"]),
                            }
                        )
                        focus_record["scoreItemIds"].append(item_id)
                    focuses.append(focus_record)
                    capability_record["focusIds"].append(focus_id)
                capabilities.append(capability_record)

    snapshot_basis = {
        "categories": [(item["code"], item["name"], item["parentId"]) for item in categories],
        "capabilities": [(item["code"], item["name"], item["categoryId"]) for item in capabilities],
        "focuses": [(item["code"], item["name"], item["capabilityId"]) for item in focuses],
        "focusServiceMappings": [(item["focusId"], item["scopeCode"], item["serviceId"], item["serviceRole"]) for item in focus_service_mappings],
        "scoreItems": [(item["itemType"], item["focusId"], item["serviceId"], item["scopeCode"]) for item in score_items],
        "rubricVersion": RUBRIC_VERSION,
    }
    snapshot_id = f"maturity-template-{_stable_hash(snapshot_basis)}"
    return {
        "id": "sapd-maturity-base-stable-v2.1",
        "snapshotId": snapshot_id,
        "name": "SAPD标准能力成熟度模板",
        "version": "V2.1",
        "type": "base",
        "status": "validated",
        "readOnly": True,
        "structureMutable": False,
        "weightMutable": False,
        "description": "基于当前稳定能力、关注点、安全技术服务和作用域关系生成的只读评估模板。",
        "rubricVersion": RUBRIC_VERSION,
        "categories": categories,
        "capabilities": capabilities,
        "focuses": focuses,
        "services": list(services.values()),
        "focusServiceMappings": focus_service_mappings,
        "scopes": [
            {
                "id": f"scope:{code}",
                "code": code,
                "name": _text(scope_by_code.get(code, {}).get("name")) or ("全部作用域" if code == "ALL" else code),
                "sourceType": "dictionary",
                "isCustom": False,
            }
            for code in ("ALL", "I-AP", "I-DI", "I-HD", "I-NT", "I-OS", "I-PE", "I-US")
        ],
        "scoreItems": score_items,
        "criticalRules": [],
        "elementWeights": {key: 0.25 for key in ELEMENT_KEYS},
        "stats": {
            "topCategories": sum(1 for item in categories if item["level"] == 1),
            "domains": sum(1 for item in categories if item["level"] == 2),
            "capabilities": len(capabilities),
            "focuses": len(focuses),
            "services": len(services),
            "serviceMappings": len(focus_service_mappings),
            "platformEvidenceReferences": sum(1 for item in focus_service_mappings if item["serviceRole"] == "PLATFORM_EVIDENCE_REFERENCE"),
            "serviceItems": sum(1 for item in score_items if item["itemType"] == "SERVICE"),
            "focusItems": sum(1 for item in score_items if item["itemType"] == "FOCUS"),
            "scoreItems": len(score_items),
        },
    }


def validate_maturity_template(template: dict[str, Any]) -> dict[str, Any]:
    errors: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    categories = [item for item in _list(template.get("categories")) if isinstance(item, dict)]
    capabilities = [item for item in _list(template.get("capabilities")) if isinstance(item, dict)]
    focuses = [item for item in _list(template.get("focuses")) if isinstance(item, dict)]
    services = [item for item in _list(template.get("services")) if isinstance(item, dict)]
    scopes = [item for item in _list(template.get("scopes")) if isinstance(item, dict)]
    focus_service_mappings = [item for item in _list(template.get("focusServiceMappings")) if isinstance(item, dict)]
    score_items = [item for item in _list(template.get("scoreItems")) if isinstance(item, dict)]

    def duplicate_ids(rows: list[dict[str, Any]]) -> set[str]:
        values = [_text(row.get("id")) for row in rows if _text(row.get("id"))]
        return {value for value, count in Counter(values).items() if count > 1}

    for object_type, rows in (
        ("category", categories),
        ("capability", capabilities),
        ("focus", focuses),
        ("service", services),
        ("scope", scopes),
        ("focus_service_mapping", focus_service_mappings),
        ("score_item", score_items),
    ):
        for duplicate_id in sorted(duplicate_ids(rows)):
            errors.append({"code": "duplicate_id", "objectId": duplicate_id, "message": f"{object_type} 存在重复 ID。"})

    category_by_id = {_text(item.get("id")): item for item in categories if _text(item.get("id"))}
    capability_by_id = {_text(item.get("id")): item for item in capabilities if _text(item.get("id"))}
    focus_by_id = {_text(item.get("id")): item for item in focuses if _text(item.get("id"))}
    service_by_id = {_text(item.get("id")): item for item in services if _text(item.get("id"))}
    scope_codes = {_text(item.get("code")) for item in scopes if _text(item.get("code"))}

    if template.get("type") == "base":
        if template.get("readOnly") is not True or template.get("structureMutable") is not False or template.get("weightMutable") is not False:
            errors.append({"code": "fixed_template_not_read_only", "objectId": _text(template.get("id")), "message": "固定知识库模板必须锁定结构和权重。"})

    for category in categories:
        category_id = _text(category.get("id"))
        if not category_id or not _text(category.get("name")):
            errors.append({"code": "category_incomplete", "objectId": category_id, "message": "分类必须有 ID 和名称。"})
        parent_id = _text(category.get("parentId"))
        if parent_id and parent_id not in category_by_id:
            errors.append({"code": "category_parent_missing", "objectId": category_id, "message": "分类引用的上级分类不存在。"})
        capability_level = _text(category.get("capabilityLevel")) or ("L0" if int(_float(category.get("level"), 1)) == 1 else "L1")
        if capability_level not in {"L0", "L1"}:
            errors.append({"code": "category_capability_level_invalid", "objectId": category_id, "message": "模板分类只能表达能力 L0 或能力 L1。"})
        if capability_level == "L0" and parent_id:
            errors.append({"code": "l0_parent_forbidden", "objectId": category_id, "message": "能力 L0 不得存在上级节点。"})
        if capability_level == "L1" and parent_id:
            parent_level = _text(category_by_id.get(parent_id, {}).get("capabilityLevel")) or "L0"
            if parent_level != "L0":
                errors.append({"code": "l1_parent_invalid", "objectId": category_id, "message": "能力 L1 只能归属能力 L0，或在无 L0 时成为顶级。"})

    focus_ids_by_capability: dict[str, list[str]] = defaultdict(list)
    for focus in focuses:
        focus_id = _text(focus.get("id"))
        capability_id = _text(focus.get("capabilityId"))
        if not focus_id or not _text(focus.get("name")):
            errors.append({"code": "focus_incomplete", "objectId": focus_id, "message": "关注点必须有 ID 和名称。"})
        if capability_id not in capability_by_id:
            errors.append({"code": "focus_capability_missing", "objectId": focus_id, "message": "关注点引用的能力不存在。"})
        focus_ids_by_capability[capability_id].append(focus_id)

    for capability in capabilities:
        capability_id = _text(capability.get("id"))
        if capability.get("included") is False:
            continue
        if not capability_id or not _text(capability.get("name")):
            errors.append({"code": "capability_incomplete", "objectId": capability_id, "message": "纳入评估的能力必须有 ID 和名称。"})
        if _text(capability.get("categoryId")) not in category_by_id:
            errors.append({"code": "capability_category_missing", "objectId": capability_id, "message": "纳入评估的能力必须归属有效分类。"})
        if not focus_ids_by_capability.get(capability_id):
            errors.append({"code": "capability_focus_missing", "objectId": capability_id, "message": "纳入评估的能力至少需要一个关注点。"})

    l1_categories = [item for item in categories if (_text(item.get("capabilityLevel")) or ("L0" if int(_float(item.get("level"), 1)) == 1 else "L1")) == "L1"]
    if not l1_categories:
        errors.append({"code": "template_l1_empty", "objectId": _text(template.get("id")), "message": "模板至少需要一个能力 L1。"})
    for category in l1_categories:
        if not any(_text(capability.get("categoryId")) == _text(category.get("id")) and capability.get("included") is not False for capability in capabilities):
            errors.append({"code": "l1_capability_missing", "objectId": _text(category.get("id")), "message": "每个能力 L1 至少需要一个有效能力 L2。"})

    mapping_by_id: dict[str, dict[str, Any]] = {}
    mapping_keys: set[tuple[str, str, str, str]] = set()
    for mapping in focus_service_mappings:
        mapping_id = _text(mapping.get("id"))
        focus_id = _text(mapping.get("focusId"))
        service_id = _text(mapping.get("serviceId"))
        scope_code = _text(mapping.get("scopeCode"))
        service_role = _text(mapping.get("serviceRole"))
        mapping_by_id[mapping_id] = mapping
        if focus_id not in focus_by_id or service_id not in service_by_id or scope_code not in scope_codes:
            errors.append({"code": "focus_service_mapping_reference_invalid", "objectId": mapping_id, "message": "关注点服务映射必须引用有效关注点、作用域和服务。"})
        if service_role not in SERVICE_ROLES:
            errors.append({"code": "service_role_invalid", "objectId": mapping_id, "message": "服务角色必须为 ASSESSMENT_POINT 或 PLATFORM_EVIDENCE_REFERENCE。"})
        mapping_key = (focus_id, scope_code, service_id, service_role)
        if mapping_key in mapping_keys:
            errors.append({"code": "focus_service_mapping_duplicate", "objectId": mapping_id, "message": "关注点、作用域、服务和服务角色组合必须唯一。"})
        mapping_keys.add(mapping_key)

    item_ids_by_focus: dict[str, list[str]] = defaultdict(list)
    item_types_by_focus: dict[str, set[str]] = defaultdict(set)
    item_keys: set[tuple[str, str, str, str]] = set()
    for item in score_items:
        item_id = _text(item.get("id"))
        focus_id = _text(item.get("focusId"))
        item_type = _text(item.get("itemType"))
        if not item_id or focus_id not in focus_by_id:
            errors.append({"code": "score_item_focus_missing", "objectId": item_id, "message": "评估点必须引用有效关注点。"})
        if item_type not in ASSESSMENT_ITEM_TYPES:
            errors.append({"code": "score_item_type_invalid", "objectId": item_id, "message": "评估点类型必须为 SERVICE 或 FOCUS。"})
        if item_type == "SERVICE" and (not _text(item.get("serviceId")) or not _text(item.get("scopeCode"))):
            errors.append({"code": "score_item_service_missing", "objectId": item_id, "message": "服务评估点必须引用安全技术服务和作用域。"})
        if item_type == "SERVICE" and _text(item.get("serviceId")) not in service_by_id:
            errors.append({"code": "score_item_service_reference_invalid", "objectId": item_id, "message": "服务评估点引用的安全技术服务不存在。"})
        if item_type == "SERVICE" and _text(item.get("scopeCode")) not in scope_codes:
            errors.append({"code": "score_item_scope_reference_invalid", "objectId": item_id, "message": "服务评估点引用的作用域不存在。"})
        if item_type == "FOCUS" and _text(item.get("serviceId")):
            errors.append({"code": "focus_item_service_conflict", "objectId": item_id, "message": "关注点评估点不得引用安全技术服务。"})
        source_mapping_id = _text(item.get("sourceMappingId"))
        if item_type == "SERVICE" and source_mapping_id:
            source_mapping = mapping_by_id.get(source_mapping_id, {})
            if _text(source_mapping.get("serviceRole")) != "ASSESSMENT_POINT":
                errors.append({"code": "score_item_mapping_role_invalid", "objectId": item_id, "message": "服务评估点只能来自 ASSESSMENT_POINT 映射。"})
        if _float(item.get("weight"), 1.0) <= 0:
            errors.append({"code": "score_item_weight_invalid", "objectId": item_id, "message": "评估点权重必须大于 0。"})
        element_weights = {**{key: 0.25 for key in ELEMENT_KEYS}, **_dict(item.get("elementWeights"))}
        if any(_float(element_weights.get(key), -1) < 0 for key in ELEMENT_KEYS) or abs(sum(_float(element_weights.get(key), 0) for key in ELEMENT_KEYS) - 1.0) > 0.000001:
            errors.append({"code": "dimension_weight_invalid", "objectId": item_id, "message": "评估点四维权重必须非负且合计为 1。"})
        rubric_entries = [entry for entry in _list(item.get("rubricEntries")) if isinstance(entry, dict)]
        rubric_keys = {
            (_text(entry.get("dimensionCode")), _text(entry.get("level")))
            for entry in rubric_entries
            if _text(entry.get("criteria")) and _text(entry.get("sourceType")) and _text(entry.get("sourceVersion"))
        }
        expected_rubric_keys = {(dimension, level["id"]) for dimension in ELEMENT_KEYS for level in MATURITY_LEVELS}
        if rubric_keys != expected_rubric_keys:
            errors.append({"code": "rubric_missing", "objectId": item_id, "message": "评估点必须具备四个维度各 L1—L5 的完整评分标准。"})
        item_key = (focus_id, item_type, _text(item.get("scopeCode")), _text(item.get("serviceId")))
        if item_key in item_keys:
            errors.append({"code": "score_item_duplicate_mapping", "objectId": item_id, "message": "关注点、作用域、服务和评估点类型组合必须唯一。"})
        item_keys.add(item_key)
        item_ids_by_focus[focus_id].append(item_id)
        item_types_by_focus[focus_id].add(item_type)

    for focus in focuses:
        focus_id = _text(focus.get("id"))
        if focus.get("included") is False:
            continue
        if not item_ids_by_focus.get(focus_id):
            errors.append({"code": "focus_score_item_missing", "objectId": focus_id, "message": "纳入评估的关注点至少需要一个评估点。"})
        if len(item_types_by_focus.get(focus_id, set())) > 1:
            errors.append({"code": "focus_scoring_mode_conflict", "objectId": focus_id, "message": "同一关注点不能同时生成 SERVICE 和 FOCUS 评估点。"})
        assessment_mappings = [item for item in focus_service_mappings if _text(item.get("focusId")) == focus_id and _text(item.get("serviceRole")) == "ASSESSMENT_POINT"]
        item_types = item_types_by_focus.get(focus_id, set())
        if assessment_mappings and item_types != {"SERVICE"}:
            errors.append({"code": "assessment_mapping_requires_service_items", "objectId": focus_id, "message": "存在 ASSESSMENT_POINT 映射时必须只生成 SERVICE 评估点。"})
        if not assessment_mappings and item_types != {"FOCUS"}:
            errors.append({"code": "focus_item_required_without_assessment_mapping", "objectId": focus_id, "message": "不存在 ASSESSMENT_POINT 映射时必须生成 FOCUS 评估点。"})

    if not categories:
        errors.append({"code": "template_category_empty", "objectId": _text(template.get("id")), "message": "模板至少需要一个分类。"})
    if not capabilities:
        errors.append({"code": "template_capability_empty", "objectId": _text(template.get("id")), "message": "模板至少需要一个能力。"})
    if not score_items:
        errors.append({"code": "template_score_item_empty", "objectId": _text(template.get("id")), "message": "模板至少需要一个评估点。"})

    if template.get("type") == "custom" and not any(item.get("isCustom") for item in capabilities):
        warnings.append({"code": "custom_template_without_custom_capability", "objectId": _text(template.get("id")), "message": "当前自定义模板尚未新增模板内能力，仅调整了基础模板。"})

    validation_basis = {
        "categories": [
            (_text(item.get("id")), _text(item.get("name")), _text(item.get("parentId")), _float(item.get("weight"), 1.0))
            for item in categories
        ],
        "capabilities": [
            (_text(item.get("id")), _text(item.get("name")), _text(item.get("categoryId")), item.get("included") is not False)
            for item in capabilities
        ],
        "focuses": [
            (_text(item.get("id")), _text(item.get("name")), _text(item.get("capabilityId")), _text(item.get("itemType")))
            for item in focuses
        ],
        "focusServiceMappings": [
            (_text(item.get("id")), _text(item.get("focusId")), _text(item.get("scopeCode")), _text(item.get("serviceId")), _text(item.get("serviceRole")))
            for item in focus_service_mappings
        ],
        "scoreItems": [
            (_text(item.get("id")), _text(item.get("itemType")), _text(item.get("focusId")), _text(item.get("scopeCode")), _text(item.get("serviceId")), _float(item.get("weight"), 1.0), _dict(item.get("elementWeights")), _text(template.get("rubricVersion")))
            for item in score_items
        ],
    }
    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "snapshotId": f"maturity-template-{_stable_hash(validation_basis)}" if not errors else None,
        "stats": {
            "categories": len(categories),
            "capabilities": sum(1 for item in capabilities if item.get("included") is not False),
            "focuses": sum(1 for item in focuses if item.get("included") is not False),
            "scoreItems": len(score_items),
        },
    }


def _entry_dimensions(entry: dict[str, Any]) -> dict[str, float | None]:
    self_values = _dict(entry.get("elements"))
    review_values = _dict(entry.get("reviewElements"))
    return {
        key: maturity_level_index(review_values.get(key) or self_values.get(key))
        for key in ELEMENT_KEYS
    }


def _entry_target_levels(entry: dict[str, Any]) -> dict[str, Any]:
    if "targetElements" in entry:
        return _dict(entry.get("targetElements"))
    legacy_target = _normalized_level(entry.get("targetLevel"))
    return {key: legacy_target for key in ELEMENT_KEYS} if legacy_target else {}


def _entry_target_dimensions(entry: dict[str, Any]) -> dict[str, float | None]:
    target_values = _entry_target_levels(entry)
    return {key: maturity_level_index(target_values.get(key)) for key in ELEMENT_KEYS}


def _dimension_index(item: dict[str, Any], element_values: dict[str, float | None], template: dict[str, Any]) -> float | None:
    if any(element_values.get(key) is None for key in ELEMENT_KEYS):
        return None
    configured_weights = {
        **{key: 0.25 for key in ELEMENT_KEYS},
        **_dict(template.get("elementWeights")),
        **_dict(item.get("elementWeights")),
    }
    rows = [
        (element_values.get(key), _float(configured_weights.get(key), 0.25))
        for key in ELEMENT_KEYS
    ]
    return _weighted_average(rows)


def _entry_index(item: dict[str, Any], entry: dict[str, Any], template: dict[str, Any]) -> float | None:
    return _dimension_index(item, _entry_dimensions(entry), template)


def _entry_target_index(item: dict[str, Any], entry: dict[str, Any], template: dict[str, Any]) -> float | None:
    return _dimension_index(item, _entry_target_dimensions(entry), template)


def _result_record(
    *,
    object_id: str,
    code: str,
    name: str,
    current_index: float | None,
    target_index: float | None,
    completion_rate: float,
    evidence_coverage: float,
    weight: float = 1.0,
    status: str = "draft",
    **extra: Any,
) -> dict[str, Any]:
    current_level = maturity_level_from_index(current_index)
    target_level = maturity_level_from_index(target_index)
    gap_index = max(0.0, float(target_index) - float(current_index)) if current_index is not None and target_index is not None else None
    target_achievement_rate = _target_achievement_rate(current_index, target_index)
    return {
        "id": object_id,
        "code": code,
        "name": name,
        "currentIndex": _round(current_index),
        "currentLevel": current_level,
        "currentLevelName": LEVEL_NAME.get(current_level, "未评分"),
        "currentPercent": _round(current_index * 20, 1) if current_index is not None else 0,
        "targetIndex": _round(target_index),
        "targetLevel": target_level,
        "targetAchievementRate": _round(target_achievement_rate, 1),
        "gapIndex": _round(gap_index),
        "gapPercent": _round(gap_index * 20, 1) if gap_index is not None else None,
        "completionRate": _round(completion_rate, 1),
        "evidenceCoverage": _round(evidence_coverage, 1),
        "weight": weight,
        "status": status,
        **extra,
    }


def calculate_maturity_assessment(payload: dict[str, Any]) -> dict[str, Any]:
    template = _dict(payload.get("template"))
    project = _dict(payload.get("project"))
    validation = validate_maturity_template(template)
    if not validation["valid"]:
        return {
            "ok": False,
            "dataState": "invalid_template",
            "validation": validation,
            "summary": {},
            "categoryResults": [],
            "capabilityResults": [],
            "focusResults": [],
            "gapItems": [],
        }

    categories = [item for item in _list(template.get("categories")) if isinstance(item, dict)]
    capabilities = [item for item in _list(template.get("capabilities")) if isinstance(item, dict) and item.get("included") is not False]
    focuses = [item for item in _list(template.get("focuses")) if isinstance(item, dict) and item.get("included") is not False]
    score_items = [item for item in _list(template.get("scoreItems")) if isinstance(item, dict)]
    services = {_text(item.get("id")): item for item in _list(template.get("services")) if isinstance(item, dict)}
    entries = {
        _text(item.get("scoreItemId")): item
        for item in _list(payload.get("scoreEntries"))
        if isinstance(item, dict) and _text(item.get("scoreItemId"))
    }
    category_by_id = {_text(item.get("id")): item for item in categories}
    capability_by_id = {_text(item.get("id")): item for item in capabilities}
    focus_by_id = {_text(item.get("id")): item for item in focuses}
    included_capability_ids = set(capability_by_id)
    included_focus_ids = {item_id for item_id, item in focus_by_id.items() if _text(item.get("capabilityId")) in included_capability_ids}
    item_results: list[dict[str, Any]] = []
    evidence_counts = Counter({item["id"]: 0 for item in EVIDENCE_LEVELS})
    score_status_counts = Counter()
    four_element_rows: list[dict[str, Any]] = []
    enforce_target_floor = _text(project.get("status")) not in {"completed", "reported", "archived"}

    for item in score_items:
        focus_id = _text(item.get("focusId"))
        if focus_id not in included_focus_ids:
            continue
        entry = entries.get(_text(item.get("id")), {})
        is_applicable = entry.get("isApplicable") is not False
        status = "not_applicable" if not is_applicable else "not_scored"
        current_index = None if not is_applicable else _entry_index(item, entry, template)
        dimension_results = _entry_dimensions(entry) if is_applicable else {key: None for key in ELEMENT_KEYS}
        target_dimensions = _entry_target_dimensions(entry) if is_applicable else {key: None for key in ELEMENT_KEYS}
        target_index = _entry_target_index(item, entry, template) if is_applicable else None
        current_level = maturity_level_from_index(current_index)
        minimum_target_index = current_index if current_index is not None else None
        target_dimension_conflicts = [
            {
                "dimension": key,
                "dimensionLabel": ELEMENT_LABELS[key],
                "currentLevel": maturity_level_from_index(dimension_results.get(key)),
                "targetLevel": maturity_level_from_index(target_dimensions.get(key)),
                "minimumTargetLevel": maturity_level_from_index(dimension_results.get(key)),
            }
            for key in ELEMENT_KEYS
            if dimension_results.get(key) is not None
            and target_dimensions.get(key) is not None
            and float(target_dimensions[key]) < float(dimension_results[key])
        ]
        overall_target_below_current = bool(
            target_index is not None
            and minimum_target_index is not None
            and target_index < minimum_target_index
        )
        target_below_current = bool(
            enforce_target_floor
            and is_applicable
            and (target_dimension_conflicts or overall_target_below_current)
        )
        target_dimension_notes = _dict(entry.get("targetDimensionNotes")) if "targetDimensionNotes" in entry else {
            key: _text(entry.get("targetReason")) for key in ELEMENT_KEYS
        }
        target_reason = _text(entry.get("targetReason")) or next(
            (_text(target_dimension_notes.get(key)) for key in ELEMENT_KEYS if _text(target_dimension_notes.get(key))),
            "",
        )
        na_reason = _text(entry.get("naReason"))
        is_complete = True if not is_applicable else current_index is not None and target_index is not None and not target_below_current
        if is_applicable and current_index is not None:
            has_review = all(maturity_level_index(_dict(entry.get("reviewElements")).get(key)) is not None for key in ELEMENT_KEYS)
            status = "invalid_target" if target_below_current else "confirmed" if has_review or entry.get("status") == "confirmed" else "scored" if is_complete else "incomplete"
        elif is_applicable:
            status = "incomplete"
        evidence_level = _text(entry.get("evidenceLevel") or "E0").upper()
        if evidence_level not in {item["id"] for item in EVIDENCE_LEVELS}:
            evidence_level = "E0"
        if current_index is not None:
            evidence_counts[evidence_level] += 1
        score_status_counts[status] += 1
        item_type = _text(item.get("itemType"))
        service = services.get(_text(item.get("serviceId")), {})
        result = {
            "id": _text(item.get("id")),
            "itemType": item_type,
            "capabilityId": _text(item.get("capabilityId")),
            "focusId": focus_id,
            "serviceId": _text(item.get("serviceId")) or None,
            "serviceCode": _text(service.get("code")),
            "serviceName": _text(service.get("name")),
            "scopeCode": _text(item.get("scopeCode")),
            "scopeName": _text(item.get("scopeName")),
            "currentIndex": current_index,
            "currentLevel": current_level,
            "minimumTargetLevel": current_level if current_index is not None else None,
            "minimumTargetDimensions": {
                key: maturity_level_from_index(dimension_results.get(key))
                for key in ELEMENT_KEYS
            },
            "currentPercent": _round(current_index * 20, 1) if current_index is not None else 0,
            "dimensionResults": {key: _round(dimension_results.get(key)) for key in ELEMENT_KEYS},
            "targetDimensionResults": {key: _round(target_dimensions.get(key)) for key in ELEMENT_KEYS},
            "targetIndex": target_index,
            "targetLevel": maturity_level_from_index(target_index),
            "targetAchievementRate": _round(_target_achievement_rate(current_index, target_index), 1),
            "targetReason": target_reason,
            "targetDimensionNotes": {key: _text(target_dimension_notes.get(key)) for key in ELEMENT_KEYS},
            "targetConfirmed": bool(target_index is not None),
            "weight": max(_float(item.get("weight"), 1.0), 0.0001),
            "isApplicable": is_applicable,
            "evidenceLevel": evidence_level,
            "hasEvidence": evidence_level != "E0" or bool(_text(entry.get("evidenceSummary"))),
            "status": status,
            "isComplete": is_complete,
            "targetBelowCurrent": target_below_current,
            "targetDimensionConflicts": target_dimension_conflicts if enforce_target_floor else [],
            "naReason": na_reason,
            "note": _text(entry.get("note")),
        }
        item_results.append(result)
        if current_index is not None:
            four_element_rows.append(
                {
                    "focusId": focus_id,
                    **dimension_results,
                    **{f"target_{key}": target_dimensions.get(key) for key in ELEMENT_KEYS},
                }
            )

    item_results_by_focus: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for result in item_results:
        item_results_by_focus[result["focusId"]].append(result)

    focus_results: list[dict[str, Any]] = []
    for focus in focuses:
        focus_id = _text(focus.get("id"))
        if focus_id not in included_focus_ids:
            continue
        rows = item_results_by_focus.get(focus_id, [])
        applicable = [row for row in rows if row["isApplicable"]]
        scored = [row for row in applicable if row["currentIndex"] is not None]
        current_index = _weighted_average([(row["currentIndex"], row["weight"]) for row in scored])
        target_index = _weighted_average([(row["targetIndex"], row["weight"]) for row in applicable if row["targetIndex"] is not None])
        dimension_results = {
            key: _weighted_average([(row["dimensionResults"].get(key), row["weight"]) for row in scored])
            for key in ELEMENT_KEYS
        }
        target_dimension_results = {
            key: _weighted_average([(row["targetDimensionResults"].get(key), row["weight"]) for row in applicable if row["targetDimensionResults"].get(key) is not None])
            for key in ELEMENT_KEYS
        }
        completed = [row for row in applicable if row["isComplete"]]
        completion_rate = 100.0 * len(completed) / len(applicable) if applicable else 0.0
        evidence_coverage = 100.0 * sum(1 for row in scored if row["hasEvidence"]) / len(scored) if scored else 0.0
        result_status = "not_applicable" if rows and not applicable else "ready" if len(completed) == len(applicable) and applicable else "partial"
        focus_results.append(
            _result_record(
                object_id=focus_id,
                code=_text(focus.get("code")),
                name=_text(focus.get("name")),
                current_index=current_index,
                target_index=target_index,
                completion_rate=completion_rate,
                evidence_coverage=evidence_coverage,
                weight=max(_float(focus.get("weight"), 1.0), 0.0001),
                status=result_status,
                capabilityId=_text(focus.get("capabilityId")),
                itemType=_text(focus.get("itemType")),
                dimensionResults={key: _round(value) for key, value in dimension_results.items()},
                targetDimensionResults={key: _round(value) for key, value in target_dimension_results.items()},
                scoreItemCount=len(rows),
                applicableItemCount=len(applicable),
                scoredItemCount=len(scored),
                completedItemCount=len(completed),
                notApplicableItemCount=len(rows) - len(applicable),
                evidenceItemCount=sum(1 for row in scored if row["hasEvidence"]),
            )
        )

    focus_results_by_capability: dict[str, list[dict[str, Any]]] = defaultdict(list)
    focus_result_by_id = {item["id"]: item for item in focus_results}
    for result in focus_results:
        focus_results_by_capability[result["capabilityId"]].append(result)

    capability_results: list[dict[str, Any]] = []
    for capability in capabilities:
        capability_id = _text(capability.get("id"))
        rows = focus_results_by_capability.get(capability_id, [])
        applicable_rows = [row for row in rows if row["status"] != "not_applicable"]
        scored = [row for row in rows if row["currentIndex"] is not None]
        target_rows = [row for row in rows if row["targetIndex"] is not None and row["status"] != "not_applicable"]
        current_index = _weighted_average([(row["currentIndex"], row["weight"]) for row in scored])
        target_index = _weighted_average([(row["targetIndex"], row["weight"]) for row in target_rows])
        applicable_item_count = sum(int(row.get("applicableItemCount", 0)) for row in applicable_rows)
        completed_item_count = sum(int(row.get("completedItemCount", 0)) for row in applicable_rows)
        not_applicable_item_count = sum(int(row.get("notApplicableItemCount", 0)) for row in rows)
        completion_rate = 100.0 * completed_item_count / applicable_item_count if applicable_item_count else 0.0
        evidence_coverage = sum(row["evidenceCoverage"] for row in scored) / len(scored) if scored else 0.0
        dimension_results = {
            key: _weighted_average([(row.get("dimensionResults", {}).get(key), row["weight"]) for row in scored])
            for key in ELEMENT_KEYS
        }
        target_dimension_results = {
            key: _weighted_average([(row.get("targetDimensionResults", {}).get(key), row["weight"]) for row in target_rows])
            for key in ELEMENT_KEYS
        }
        applied_rules: list[dict[str, Any]] = []
        for rule in _list(template.get("criticalRules")):
            if not isinstance(rule, dict) or _text(rule.get("capabilityId")) != capability_id:
                continue
            trigger = focus_result_by_id.get(_text(rule.get("triggerFocusId")))
            threshold = maturity_level_index(rule.get("thresholdLevel"))
            cap_index = maturity_level_index(rule.get("capLevel"))
            if trigger and trigger["currentIndex"] is not None and threshold is not None and cap_index is not None and trigger["currentIndex"] < threshold:
                before = current_index
                current_index = min(current_index, cap_index) if current_index is not None else current_index
                applied_rules.append(
                    {
                        "id": _text(rule.get("id")),
                        "name": _text(rule.get("name")) or "关键项上限",
                        "triggerFocusId": trigger["id"],
                        "capLevel": maturity_level_from_index(cap_index),
                        "beforeIndex": _round(before),
                        "afterIndex": _round(current_index),
                    }
                )
        capability_results.append(
            _result_record(
                object_id=capability_id,
                code=_text(capability.get("code")),
                name=_text(capability.get("name")),
                current_index=current_index,
                target_index=target_index,
                completion_rate=completion_rate,
                evidence_coverage=evidence_coverage,
                weight=max(_float(capability.get("weight"), 1.0), 0.0001),
                status="not_applicable" if rows and not applicable_rows else "ready" if completion_rate >= 100 and applicable_item_count else "partial",
                categoryId=_text(capability.get("categoryId")),
                topCategoryId=_text(capability.get("topCategoryId")),
                focusCount=len(rows),
                applicableFocusCount=len(applicable_rows),
                applicableItemCount=applicable_item_count,
                completedItemCount=completed_item_count,
                notApplicableItemCount=not_applicable_item_count,
                dimensionResults={key: _round(value) for key, value in dimension_results.items()},
                targetDimensionResults={key: _round(value) for key, value in target_dimension_results.items()},
                appliedCriticalRules=applied_rules,
                businessImportance=max(0.0, min(5.0, _float(capability.get("businessImportance"), 3.0))),
                riskUrgency=max(0.0, min(5.0, _float(capability.get("riskUrgency"), 3.0))),
            )
        )

    capability_results_by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for result in capability_results:
        capability_results_by_category[result["categoryId"]].append(result)

    def aggregate_category(category: dict[str, Any], rows: list[dict[str, Any]], *, child_count: int = 0) -> dict[str, Any]:
        scored = [row for row in rows if row["currentIndex"] is not None]
        applicable_rows = [row for row in rows if row["status"] != "not_applicable"]
        targeted = [row for row in applicable_rows if row["targetIndex"] is not None]
        current_index = _weighted_average([(row["currentIndex"], row["weight"]) for row in scored])
        target_index = _weighted_average([(row["targetIndex"], row["weight"]) for row in targeted])
        applicable_item_count = sum(int(row.get("applicableItemCount", 0)) for row in applicable_rows)
        completed_item_count = sum(int(row.get("completedItemCount", 0)) for row in applicable_rows)
        not_applicable_item_count = sum(int(row.get("notApplicableItemCount", 0)) for row in rows)
        completion_rate = 100.0 * completed_item_count / applicable_item_count if applicable_item_count else 0.0
        evidence_coverage = sum(row["evidenceCoverage"] for row in scored) / len(scored) if scored else 0.0
        dimension_results = {
            key: _weighted_average([(row.get("dimensionResults", {}).get(key), row["weight"]) for row in scored])
            for key in ELEMENT_KEYS
        }
        target_dimension_results = {
            key: _weighted_average([(row.get("targetDimensionResults", {}).get(key), row["weight"]) for row in targeted])
            for key in ELEMENT_KEYS
        }
        return _result_record(
            object_id=_text(category.get("id")),
            code=_text(category.get("code")),
            name=_text(category.get("name")),
            current_index=current_index,
            target_index=target_index,
            completion_rate=completion_rate,
            evidence_coverage=evidence_coverage,
            weight=max(_float(category.get("weight"), 1.0), 0.0001),
            status="not_applicable" if rows and not applicable_rows else "ready" if completion_rate >= 100 and applicable_item_count else "partial",
            parentId=_text(category.get("parentId")) or None,
            level=int(_float(category.get("level"), 1)),
            capabilityCount=sum(int(row.get("capabilityCount", 1)) for row in rows),
            childCategoryCount=child_count,
            applicableItemCount=applicable_item_count,
            completedItemCount=completed_item_count,
            notApplicableItemCount=not_applicable_item_count,
            dimensionResults={key: _round(value) for key, value in dimension_results.items()},
            targetDimensionResults={key: _round(value) for key, value in target_dimension_results.items()},
        )

    def category_capability_level(category: dict[str, Any]) -> str:
        return _text(category.get("capabilityLevel")) or ("L0" if int(_float(category.get("level"), 1)) == 1 else "L1")

    leaf_category_results: list[dict[str, Any]] = []
    for category in categories:
        category_id = _text(category.get("id"))
        if category_capability_level(category) != "L1":
            continue
        leaf_category_results.append(aggregate_category(category, capability_results_by_category.get(category_id, [])))

    leaf_by_parent: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for result in leaf_category_results:
        leaf_by_parent[_text(result.get("parentId"))].append(result)

    top_category_results: list[dict[str, Any]] = []
    l0_ids = {_text(category.get("id")) for category in categories if category_capability_level(category) == "L0"}
    for category in categories:
        if category_capability_level(category) != "L0":
            continue
        category_id = _text(category.get("id"))
        child_rows = leaf_by_parent.get(category_id, [])
        direct_rows = capability_results_by_category.get(category_id, [])
        aggregate_rows = child_rows + direct_rows
        top_category_results.append(aggregate_category(category, aggregate_rows, child_count=len(child_rows)))
    top_category_results.extend(
        result
        for result in leaf_category_results
        if not _text(result.get("parentId")) or _text(result.get("parentId")) not in l0_ids
    )

    top_for_overall = [
        result
        for result in top_category_results
        if category_by_id.get(result["id"], {}).get("includedInOverall") is not False
    ]
    scored_top = [row for row in top_for_overall if row["currentIndex"] is not None]
    target_top = [row for row in top_for_overall if row["targetIndex"] is not None]
    overall_index = _weighted_average([(row["currentIndex"], row["weight"]) for row in scored_top])
    overall_target = _weighted_average([(row["targetIndex"], row["weight"]) for row in target_top])
    overall_dimensions = {
        key: _weighted_average([(row.get("dimensionResults", {}).get(key), row["weight"]) for row in scored_top])
        for key in ELEMENT_KEYS
    }
    overall_target_dimensions = {
        key: _weighted_average([(row.get("targetDimensionResults", {}).get(key), row["weight"]) for row in target_top])
        for key in ELEMENT_KEYS
    }
    applicable_item_count = sum(1 for row in item_results if row["isApplicable"])
    scored_item_count = sum(1 for row in item_results if row["isApplicable"] and row["isComplete"])
    evidence_item_count = sum(1 for row in item_results if row["currentIndex"] is not None and row["hasEvidence"])
    completion_rate = 100.0 * scored_item_count / applicable_item_count if applicable_item_count else 0.0
    evidence_coverage = 100.0 * evidence_item_count / scored_item_count if scored_item_count else 0.0
    invalid_na_reason_count = sum(1 for row in item_results if not row["isApplicable"] and not row["naReason"])
    target_below_current_count = score_status_counts["invalid_target"]
    statistics_ready = bool(
        applicable_item_count
        and scored_item_count == applicable_item_count
        and target_below_current_count == 0
    )
    result_status = "reviewed" if project.get("status") in {"completed", "reported", "archived"} else "draft"
    summary = _result_record(
        object_id=_text(project.get("id")) or "assessment-project",
        code="",
        name=_text(project.get("name")) or "成熟度评估项目",
        current_index=overall_index,
        target_index=overall_target,
        completion_rate=completion_rate,
        evidence_coverage=evidence_coverage,
        status=result_status,
        scoreItemCount=len(item_results),
        capabilityCount=len(capability_results),
        applicableCapabilityCount=sum(1 for row in capability_results if row["status"] != "not_applicable"),
        completedCapabilityCount=sum(1 for row in capability_results if row["status"] == "ready"),
        focusCount=len(focus_results),
        applicableFocusCount=sum(1 for row in focus_results if row["status"] != "not_applicable"),
        completedFocusCount=sum(1 for row in focus_results if row["status"] == "ready"),
        applicableItemCount=applicable_item_count,
        scoredItemCount=scored_item_count,
        evidenceItemCount=evidence_item_count,
        notApplicableCount=score_status_counts["not_applicable"],
        notScoredCount=score_status_counts["not_scored"] + score_status_counts["incomplete"] + target_below_current_count,
        missingTargetCount=sum(1 for row in item_results if row["isApplicable"] and row["targetIndex"] is None),
        missingTargetReasonCount=sum(1 for row in item_results if row["isApplicable"] and not row["targetReason"]),
        invalidNaReasonCount=invalid_na_reason_count,
        targetBelowCurrentCount=target_below_current_count,
        statisticsReady=statistics_ready,
        resultAvailability="ready" if statistics_ready else "incomplete",
        confirmedCount=score_status_counts["confirmed"],
        templateSnapshotId=_text(template.get("snapshotId")),
        knowledgeSnapshotId=_text(project.get("knowledgeSnapshotId") or template.get("knowledgeSnapshotId") or template.get("snapshotId")),
        algorithmVersion=ALGORITHM_VERSION,
        assessmentObjectType="ENTERPRISE_ORGANIZATION",
        customerProfileSnapshot=_dict(project.get("customerContextSnapshot")) or {
            "organization": _text(project.get("organization")),
            "industry": _text(project.get("industry")),
            "companySize": _text(project.get("companySize")),
            "customerCharacteristics": _text(project.get("customerCharacteristics")),
            "constraints": _text(project.get("constraints")),
        },
        dimensionResults={key: _round(value) for key, value in overall_dimensions.items()},
        targetDimensionResults={key: _round(value) for key, value in overall_target_dimensions.items()},
    )

    gap_items: list[dict[str, Any]] = []
    for capability in capability_results:
        if capability["gapIndex"] is None or capability["gapIndex"] <= 0:
            continue
        gap_norm = capability["gapIndex"] / 4.0
        business_norm = capability["businessImportance"] / 5.0
        risk_norm = capability["riskUrgency"] / 5.0
        priority_score = 100.0 * (0.50 * gap_norm + 0.25 * business_norm + 0.25 * risk_norm)
        priority = "高" if priority_score >= 70 else "中" if priority_score >= 40 else "低"
        category = category_by_id.get(capability["categoryId"], {})
        related_focus_ids = [item["id"] for item in focus_results_by_capability.get(capability["id"], []) if item["gapIndex"] and item["gapIndex"] > 0]
        related_services = []
        for item in item_results:
            if item["focusId"] not in related_focus_ids or not item["serviceId"]:
                continue
            service_label = " ".join(part for part in (item["serviceCode"], item["serviceName"]) if part)
            if service_label and service_label not in related_services:
                related_services.append(service_label)
        gap_items.append(
            {
                "id": f"gap:{capability['id']}",
                "categoryId": capability["categoryId"],
                "categoryName": _text(category.get("name")),
                "capabilityId": capability["id"],
                "capabilityCode": capability["code"],
                "capabilityName": capability["name"],
                "currentIndex": capability["currentIndex"],
                "currentLevel": capability["currentLevel"],
                "targetIndex": capability["targetIndex"],
                "targetLevel": capability["targetLevel"],
                "gapIndex": capability["gapIndex"],
                "gapPercent": capability["gapPercent"],
                "priorityScore": _round(priority_score, 1),
                "priority": priority,
                "evidenceCoverage": capability["evidenceCoverage"],
                "relatedServices": related_services[:6],
                "recommendationStatus": "candidate",
                "recommendations": [
                    {"type": "组织与角色", "text": "明确责任部门、岗位职责、流程所有者和协同关系。"},
                    {"type": "制度与流程", "text": "建立或完善制度、流程、审批、复核和例外管理。"},
                    {"type": "平台与工具", "text": "结合关注点、作用域和已关联安全技术服务确认建设范围。"},
                    {"type": "数据与信息", "text": "补充日志、工单、指标、报告、审计和持续监测证据。"},
                ],
            }
        )
    gap_items.sort(key=lambda item: (-float(item["priorityScore"]), -float(item["gapIndex"]), item["capabilityCode"]))

    maturity_distribution = Counter({item["id"]: 0 for item in MATURITY_LEVELS})
    for capability in capability_results:
        if capability["currentLevel"] in LEVEL_INDEX:
            maturity_distribution[capability["currentLevel"]] += 1

    service_distribution = Counter({item["id"]: 0 for item in MATURITY_LEVELS})
    service_distribution.update({"N/A": 0, "Not Scored": 0})
    for item in item_results:
        if item["itemType"] != "SERVICE":
            continue
        if not item["isApplicable"]:
            service_distribution["N/A"] += 1
        elif item["currentLevel"] in LEVEL_INDEX:
            service_distribution[item["currentLevel"]] += 1
        else:
            service_distribution["Not Scored"] += 1

    four_element_summary = []
    for key in ELEMENT_KEYS:
        average = _weighted_average([(row.get(key), 1.0) for row in four_element_rows])
        target_average = _weighted_average([(row.get(f"target_{key}"), 1.0) for row in four_element_rows])
        four_element_summary.append(
            {
                "id": key,
                "name": ELEMENT_LABELS[key],
                "currentIndex": _round(average),
                "currentLevel": maturity_level_from_index(average),
                "currentPercent": _round(average * 20, 1) if average is not None else 0,
                "targetIndex": _round(target_average),
                "targetLevel": maturity_level_from_index(target_average),
                "targetPercent": _round(target_average * 20, 1) if target_average is not None else 0,
                "gapIndex": _round(max(0.0, target_average - average)) if average is not None and target_average is not None else None,
            }
        )

    input_basis = {
        "projectId": project.get("id"),
        "templateSnapshotId": template.get("snapshotId"),
        "knowledgeSnapshotId": summary.get("knowledgeSnapshotId"),
        "algorithmVersion": ALGORITHM_VERSION,
        "entries": _list(payload.get("scoreEntries")),
    }
    response = {
        "ok": True,
        "dataState": "ready",
        "mode": "controlled_demo",
        "calculatedAt": _now(),
        "validation": validation,
        "summary": summary,
        "categoryResults": top_category_results,
        "subCategoryResults": leaf_category_results,
        "capabilityResults": capability_results,
        "focusResults": focus_results,
        "scoreItemResults": item_results,
        "gapItems": gap_items,
        "maturityDistribution": [
            {"level": item["id"], "name": item["name"], "count": maturity_distribution[item["id"]]}
            for item in MATURITY_LEVELS
        ],
        "serviceDistribution": [
            {"level": level, "count": service_distribution[level]}
            for level in ("L1", "L2", "L3", "L4", "L5", "N/A", "Not Scored")
        ],
        "evidenceDistribution": [
            {"level": item["id"], "name": item["name"], "count": evidence_counts[item["id"]]}
            for item in EVIDENCE_LEVELS
        ],
        "fourElementResults": four_element_summary,
    }
    response["calculationRun"] = {
        "id": f"maturity-calc-{_stable_hash(input_basis, 20)}",
        "algorithmVersion": ALGORITHM_VERSION,
        "templateSnapshotId": _text(template.get("snapshotId")),
        "knowledgeSnapshotId": _text(summary.get("knowledgeSnapshotId")),
        "inputHash": _stable_hash(input_basis, 32),
        "resultHash": _stable_hash({key: response[key] for key in ("summary", "categoryResults", "capabilityResults", "gapItems")}, 32),
        "status": "formal" if summary.get("completionRate") == 100 and project.get("status") in {"completed", "reported", "archived"} else "draft",
    }
    return response


def _demo_score_entries(
    template: dict[str, Any],
    *,
    variant: int,
    complete: bool,
    reviewed: bool,
    target_conflict_count: int | None = None,
) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    remaining_target_conflicts = max(0, int(target_conflict_count or 0))
    for index, item in enumerate(_list(template.get("scoreItems"))):
        item_id = _text(item.get("id"))
        digest = int(hashlib.sha1(f"{variant}:{item_id}".encode("utf-8")).hexdigest()[:8], 16)
        is_applicable = digest % 37 != 0
        is_unscored = not complete and digest % 11 == 0
        level_number = 2 + digest % 3
        evidence_number = 1 + digest % 4
        elements = {
            key: None if is_unscored or not is_applicable else f"L{max(1, min(5, level_number + ((digest >> offset) % 3) - 1))}"
            for offset, key in enumerate(ELEMENT_KEYS)
        }
        target_level = "L4" if digest % 7 else "L3"
        if is_applicable and not is_unscored:
            current_level_indices = [
                int(maturity_level_index(value) or 1)
                for value in elements.values()
            ]
            current_max_index = max(current_level_indices, default=1)
            target_level_index = max(int(maturity_level_index(target_level) or 1), current_max_index)
            target_level = f"L{target_level_index}"
            if target_conflict_count is not None and remaining_target_conflicts and current_max_index > 1:
                target_level = f"L{current_max_index - 1}"
                remaining_target_conflicts -= 1
        entry = {
            "scoreItemId": item_id,
            "isApplicable": is_applicable,
            "elements": elements,
            "reviewElements": elements if reviewed and is_applicable and not is_unscored else {},
            "targetLevel": target_level,
            "targetReason": "结合业务重要性、风险与实施可行性建议目标等级。" if is_applicable and not is_unscored else "",
            "targetConfirmed": bool(is_applicable and not is_unscored),
            "evidenceLevel": "E0" if is_unscored or not is_applicable else f"E{evidence_number}",
            "evidenceSummary": "" if is_unscored or not is_applicable else "当前评估证据说明，用于记录工作流判断依据。",
            "note": "" if index % 5 else "当前项目仍为本地草稿，未经复核不代表正式客户结论。",
            "naReason": "当前评估范围不适用。" if not is_applicable else "",
            "status": "not_applicable" if not is_applicable else "incomplete" if is_unscored else "confirmed" if reviewed else "scored",
        }
        entries.append(entry)
    if target_conflict_count is not None and remaining_target_conflicts:
        raise ValueError(f"受控 demo 无法生成约定的 {target_conflict_count} 个目标等级冲突。")
    return entries


def _project_detail(
    template: dict[str, Any],
    *,
    project_id: str,
    name: str,
    organization: str,
    status: str,
    variant: int,
    complete: bool,
    reviewed: bool,
    industry: str,
    company_size: str,
    target_conflict_count: int | None = None,
    controlled_demo_revision: str | None = None,
) -> dict[str, Any]:
    template = deepcopy(template)
    project = {
        "id": project_id,
        "name": name,
        "organization": organization,
        "assessmentObjectType": "ENTERPRISE_ORGANIZATION",
        "industry": industry,
        "companySize": company_size,
        "customerCharacteristics": "当前企业组织的评估背景与目标范围。",
        "constraints": "当前项目约束不改变已记录的事实评分。",
        "plannedStartDate": "2026-07-01",
        "plannedEndDate": "2026-07-31",
        "owner": "项目负责人",
        "assessors": ["评估人员"],
        "status": status,
        "statusLabel": PROJECT_STATUS_LABELS.get(status, status),
        "templateId": template["id"],
        "templateName": template["name"],
        "templateType": template["type"],
        "templateSnapshotId": template["snapshotId"],
        "knowledgeSnapshotId": template["snapshotId"],
        "algorithmVersion": ALGORITHM_VERSION,
        "customerContextSnapshot": {
            "organization": organization,
            "industry": industry,
            "companySize": company_size,
            "customerCharacteristics": "当前企业组织的评估背景与目标范围。",
            "constraints": "当前项目约束不改变已记录的事实评分。",
        },
        "updatedAt": f"2026-07-{10 - variant:02d} {9 + variant:02d}:20",
        "mode": "controlled_demo",
        "readOnly": status in {"completed", "reported", "archived"},
    }
    if controlled_demo_revision:
        project["controlledDemoRevision"] = controlled_demo_revision
    elif target_conflict_count is not None:
        project["controlledDemoRevision"] = f"target-conflicts-{target_conflict_count}-20260716"
    entries = _demo_score_entries(
        template,
        variant=variant,
        complete=complete,
        reviewed=reviewed,
        target_conflict_count=target_conflict_count,
    )
    result = calculate_maturity_assessment({"project": project, "template": template, "scoreEntries": entries})
    return {"project": project, "template": template, "scoreEntries": entries, "result": result}


def build_maturity_workspace(
    capability_workbench: dict[str, Any],
    *,
    project_profile: str = "development",
) -> dict[str, Any]:
    template = build_maturity_base_template(capability_workbench)
    project_details = {
        "demo-project-001": _project_detail(
            template,
            project_id="demo-project-001",
            name="某集团网络安全成熟度评估",
            organization="某集团",
            status="scoring",
            variant=1,
            complete=False,
            reviewed=False,
            industry="综合集团",
            company_size="大型企业",
            target_conflict_count=3,
            controlled_demo_revision="two-test-cases-in-progress-20260720",
        ),
        "demo-project-002": _project_detail(
            template,
            project_id="demo-project-002",
            name="某金融企业成熟度评估",
            organization="某金融企业",
            status="completed",
            variant=2,
            complete=True,
            reviewed=True,
            industry="金融",
            company_size="大型企业",
            controlled_demo_revision="two-test-cases-completed-20260720",
        ),
        "demo-project-003": _project_detail(
            template,
            project_id="demo-project-003",
            name="某科技企业成熟度评估",
            organization="某科技企业",
            status="reported",
            variant=3,
            complete=True,
            reviewed=True,
            industry="科技",
            company_size="中型企业",
        ),
        "demo-project-004": _project_detail(
            template,
            project_id="demo-project-004",
            name="某制造企业成熟度评估",
            organization="某制造企业",
            status="score_review",
            variant=4,
            complete=True,
            reviewed=False,
            industry="制造业",
            company_size="大型企业",
        ),
        "demo-project-005": _project_detail(
            template,
            project_id="demo-project-005",
            name="某零售企业成熟度评估",
            organization="某零售企业",
            status="template_configuring",
            variant=5,
            complete=False,
            reviewed=False,
            industry="零售",
            company_size="中型企业",
        ),
    }
    normalized_profile = "delivery" if str(project_profile).strip().lower() == "delivery" else "development"
    if normalized_profile == "delivery":
        project_details = {project_id: project_details[project_id] for project_id in ("demo-project-001", "demo-project-002")}
    projects = []
    for detail in project_details.values():
        project = detail["project"]
        summary = _dict(detail["result"].get("summary"))
        projects.append(
            {
                **project,
                "currentIndex": summary.get("currentIndex"),
                "currentLevel": summary.get("currentLevel"),
                "currentPercent": summary.get("currentPercent"),
                "completionRate": summary.get("completionRate"),
                "evidenceCoverage": summary.get("evidenceCoverage"),
                "notScoredCount": summary.get("notScoredCount"),
            }
        )
    return {
        "dataState": "ready",
        "mode": "controlled_demo",
        "projectProfile": normalized_profile,
        "persistence": "browser_local_only",
        "notice": "项目草稿保存在当前浏览器，不写正式 SQLite、正式 JSON 或用户库。",
        "levels": list(MATURITY_LEVELS),
        "evidenceLevels": list(EVIDENCE_LEVELS),
        "projectStatuses": [{"id": key, "name": value} for key, value in PROJECT_STATUS_LABELS.items()],
        "template": template,
        "projects": projects,
        "projectDetails": project_details,
        "dictionarySnapshot": {
            "id": template["snapshotId"],
            "label": "当前稳定能力字典快照",
            "scopePolicy": "dictionary_mappings_only",
            "scopeCodes": ["ALL", "I-AP", "I-DI", "I-HD", "I-NT", "I-OS", "I-PE", "I-US"],
            "stats": template["stats"],
        },
    }


def _template_structure_basis(template: dict[str, Any]) -> dict[str, Any]:
    return {
        "templateId": _text(template.get("id")),
        "templateVersion": _text(template.get("version")),
        "templateSnapshotId": _text(template.get("snapshotId")),
        "categories": [
            (_text(item.get("id")), _text(item.get("capabilityLevel")), _text(item.get("parentId")), _text(item.get("name")))
            for item in _list(template.get("categories"))
            if isinstance(item, dict)
        ],
        "capabilities": [
            (_text(item.get("id")), _text(item.get("categoryId")), _text(item.get("name")), item.get("included") is not False)
            for item in _list(template.get("capabilities"))
            if isinstance(item, dict)
        ],
        "focuses": [
            (_text(item.get("id")), _text(item.get("capabilityId")), _text(item.get("name")), item.get("included") is not False)
            for item in _list(template.get("focuses"))
            if isinstance(item, dict)
        ],
        "focusServiceMappings": [
            (_text(item.get("id")), _text(item.get("focusId")), _text(item.get("scopeCode")), _text(item.get("serviceId")), _text(item.get("serviceRole")))
            for item in _list(template.get("focusServiceMappings"))
            if isinstance(item, dict)
        ],
        "scoreItems": [
            (_text(item.get("id")), _text(item.get("itemType")), _text(item.get("focusId")), _text(item.get("scopeCode")), _text(item.get("serviceId")), _float(item.get("weight"), 1.0), _dict(item.get("elementWeights")))
            for item in _list(template.get("scoreItems"))
            if isinstance(item, dict)
        ],
    }


def _template_structure_hash(template: dict[str, Any]) -> str:
    return _stable_hash(_template_structure_basis(template), 32)


def _normalized_level(value: Any) -> str | None:
    raw = _text(value).upper().replace("等级", "").strip()
    if raw in LEVEL_INDEX:
        return raw
    if raw in {"1", "2", "3", "4", "5"}:
        return f"L{raw}"
    return None


def _business_cell(value: Any) -> str:
    return re.sub(r"[ \t\r\f\v]+", " ", _text(value)).strip()


def _business_service_names(value: Any) -> list[str]:
    names = [_business_cell(item) for item in re.split(r"[\n；;]+", _text(value))]
    return [name for name in names if name]


def _template_business_rows(template: dict[str, Any], score_entries: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    categories = {_text(item.get("id")): item for item in _list(template.get("categories")) if isinstance(item, dict)}
    capabilities = {_text(item.get("id")): item for item in _list(template.get("capabilities")) if isinstance(item, dict)}
    focuses = {_text(item.get("id")): item for item in _list(template.get("focuses")) if isinstance(item, dict)}
    services = {_text(item.get("id")): item for item in _list(template.get("services")) if isinstance(item, dict)}
    mappings_by_focus: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for mapping in _list(template.get("focusServiceMappings")):
        if isinstance(mapping, dict):
            mappings_by_focus[_text(mapping.get("focusId"))].append(mapping)
    entries = {
        _text(item.get("scoreItemId")): item
        for item in _list(score_entries)
        if isinstance(item, dict) and _text(item.get("scoreItemId"))
    }
    rows: list[dict[str, Any]] = []
    for item in _list(template.get("scoreItems")):
        if not isinstance(item, dict):
            continue
        focus = focuses.get(_text(item.get("focusId")), {})
        capability = capabilities.get(_text(focus.get("capabilityId") or item.get("capabilityId")), {})
        if not focus or not capability or focus.get("included") is False or capability.get("included") is False:
            continue
        l1 = categories.get(_text(capability.get("categoryId")), {})
        l0 = categories.get(_text(capability.get("topCategoryId") or l1.get("parentId")), {})
        item_type = _text(item.get("itemType")) or "FOCUS"
        if item_type == "SERVICE":
            service_names = [_text(services.get(_text(item.get("serviceId")), {}).get("name"))]
        else:
            reference_ids = [
                _text(mapping.get("serviceId"))
                for mapping in sorted(mappings_by_focus.get(_text(focus.get("id")), []), key=lambda row: (_float(row.get("sortOrder")), _text(row.get("id"))))
                if _text(mapping.get("serviceRole")) == "PLATFORM_EVIDENCE_REFERENCE"
            ]
            if not reference_ids:
                reference_ids = [_text(value) for value in _list(focus.get("platformEvidenceServiceIds"))]
            service_names = [_text(services.get(service_id, {}).get("name")) for service_id in reference_ids]
        service_names = [name for name in service_names if name]
        entry = entries.get(_text(item.get("id")), {})
        applicable = entry.get("isApplicable") is not False
        elements = _dict(entry.get("elements"))
        dimension_notes = _dict(entry.get("dimensionNotes"))
        target_elements = _entry_target_levels(entry)
        target_dimension_notes = _dict(entry.get("targetDimensionNotes")) if "targetDimensionNotes" in entry else {
            key: _text(entry.get("targetReason")) for key in ELEMENT_KEYS
        }
        note = _text(entry.get("naReason")) if not applicable else _text(entry.get("note") or entry.get("evidenceSummary"))
        rows.append(
            {
                "itemId": _text(item.get("id")),
                "l0": _text(l0.get("name")) or "未分组能力分类",
                "l1": _text(l1.get("name")) or "未分组高阶能力",
                "l2": _text(capability.get("name")) or "未命名安全能力",
                "focusNo": _text(focus.get("code")) or str(int(_float(focus.get("sortOrder"), 0)) + 1),
                "focus": _text(focus.get("name")) or "未命名关注点",
                "services": "\n".join(service_names),
                "granularity": "安全技术服务" if item_type == "SERVICE" else "关注点",
                "applicability": "不适用" if score_entries is not None and not applicable else "适用" if score_entries is not None else "",
                "organization": _normalized_level(elements.get("organization")) or "" if score_entries is not None else "",
                "process": _normalized_level(elements.get("process")) or "" if score_entries is not None else "",
                "tool": _normalized_level(elements.get("tool")) or "" if score_entries is not None else "",
                "data": _normalized_level(elements.get("data")) or "" if score_entries is not None else "",
                "currentNotes": {key: _text(dimension_notes.get(key)) if score_entries is not None else "" for key in ELEMENT_KEYS},
                "targets": {key: _normalized_level(target_elements.get(key)) or "" if score_entries is not None else "" for key in ELEMENT_KEYS},
                "targetNotes": {key: _text(target_dimension_notes.get(key)) if score_entries is not None else "" for key in ELEMENT_KEYS},
                "note": note if score_entries is not None else "",
                "sort": (
                    _float(l0.get("sortOrder")),
                    _float(l1.get("sortOrder")),
                    _float(capability.get("sortOrder")),
                    _float(focus.get("sortOrder")),
                    _float(item.get("sortOrder")),
                    _text(item.get("id")),
                ),
            }
        )
    return sorted(rows, key=lambda row: row["sort"])


def _workbook_row_values(row: dict[str, Any]) -> list[str]:
    return [
        row["l0"], row["l1"], row["l2"], row["focusNo"], row["focus"], row["services"], row["granularity"],
        row["applicability"], row["organization"], row["process"], row["tool"], row["data"],
        *[row["currentNotes"].get(key, "") for key in ELEMENT_KEYS],
        *[row["targets"].get(key, "") for key in ELEMENT_KEYS],
        *[row["targetNotes"].get(key, "") for key in ELEMENT_KEYS],
        row["note"],
    ]


def _business_row_key(row: dict[str, Any]) -> tuple[str, ...]:
    return (
        _business_cell(row.get("l0")),
        _business_cell(row.get("l1")),
        _business_cell(row.get("l2")),
        _business_cell(row.get("focusNo")),
        _business_cell(row.get("focus")),
        "\n".join(sorted(_business_service_names(row.get("services")))),
        _business_cell(row.get("granularity")),
    )


def _merge_business_hierarchy(sheet: Any, rows: list[dict[str, Any]], start_row: int = 3) -> None:
    specifications = (
        (1, lambda row: (row["l0"],)),
        (2, lambda row: (row["l0"], row["l1"])),
        (3, lambda row: (row["l0"], row["l1"], row["l2"])),
        (4, lambda row: (row["l0"], row["l1"], row["l2"], row["focusNo"], row["focus"])),
        (5, lambda row: (row["l0"], row["l1"], row["l2"], row["focusNo"], row["focus"])),
    )
    for column, key_of in specifications:
        group_start = 0
        while group_start < len(rows):
            key = key_of(rows[group_start])
            group_end = group_start + 1
            while group_end < len(rows) and key_of(rows[group_end]) == key:
                group_end += 1
            if group_end - group_start > 1 and _text(sheet.cell(start_row + group_start, column).value):
                sheet.merge_cells(start_row=start_row + group_start, start_column=column, end_row=start_row + group_end - 1, end_column=column)
            group_start = group_end


def _build_maturity_workbook(template: dict[str, Any], *, purpose: str, score_entries: list[dict[str, Any]] | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    info = workbook.active
    info.title = MATURITY_INFO_SHEET
    assessment = workbook.create_sheet(MATURITY_ASSESSMENT_SHEET)
    workbook.calculation.fullCalcOnLoad = True

    line = Side(style="thin", color="CBD5E1")
    border = Border(left=line, right=line, top=line, bottom=line)
    title_fill = PatternFill("solid", fgColor="DDE7F1")
    header_fill = PatternFill("solid", fgColor="EAF0F6")
    label_fill = PatternFill("solid", fgColor="F4F7FA")
    input_fill = PatternFill("solid", fgColor="FFF9E8")
    title_font = Font(name="微软雅黑", size=15, bold=True, color="23384D")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="334155")
    body_font = Font(name="微软雅黑", size=10, color="3F4E5E")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    info.merge_cells("A1:D1")
    info["A1"] = "成熟度评估模板基础信息"
    info["A1"].font = title_font
    info["A1"].fill = title_fill
    info["A1"].alignment = left
    info["A1"].border = border
    is_template_workbook = purpose == "模板配置"
    filling_note = (
        "评分标题和评分列必须保留，但适用性、当前与目标四维评分及说明的单元格必须全部为空；"
        "将模板类型切换为“自定义模板”后，请填写新的模板名称和模板说明。"
        if is_template_workbook
        else "评分标题和评分列不得删除；请分别填写当前与目标四维评分，说明为可选。不适用项中已有的评分会在导入时忽略。"
    )
    info_rows = (
        ("模板名称", _text(template.get("name")) or "未命名成熟度模板"),
        ("模板类型", "标准模板" if template.get("type") == "base" else "自定义模板"),
        ("模板版本", _text(template.get("version")) or "V2.1"),
        ("模板说明", _text(template.get("description")) or "用于成熟度评估的业务模板。"),
        ("评分维度", "组织与角色、制度与流程、平台与工具、数据与信息"),
        ("等级范围", "L1—L5"),
        ("填写说明", filling_note),
    )
    for row_index, (label, value) in enumerate(info_rows, start=2):
        info.cell(row_index, 1, label)
        info.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=4)
        info.cell(row_index, 2, value)
        for column in range(1, 5):
            cell = info.cell(row_index, column)
            cell.border = border
            cell.font = header_font if column == 1 else body_font
            cell.fill = label_fill if column == 1 else PatternFill("solid", fgColor="FFFFFF")
            cell.alignment = left
        info.row_dimensions[row_index].height = 46 if label in {"评分维度", "填写说明"} else 32
    if is_template_workbook:
        template_type_validation = DataValidation(type="list", formula1='"标准模板,自定义模板"', allow_blank=False)
        template_type_validation.error = "模板类型只能选择“标准模板”或“自定义模板”。"
        template_type_validation.errorTitle = "模板类型无效"
        template_type_validation.prompt = "导入自定义模板前，请切换为“自定义模板”，并填写新的模板名称和模板说明。"
        template_type_validation.promptTitle = "模板导入要求"
        template_type_validation.showInputMessage = True
        template_type_validation.showErrorMessage = True
        info.add_data_validation(template_type_validation)
        template_type_validation.add(info["B3"])
        info["B3"].comment = Comment("可选择“标准模板”或“自定义模板”。自定义模板导入时必须填写新的模板名称和模板说明。", "SAPD Wiki")
        for row_index in (2, 3, 5):
            for column in range(2, 5):
                info.cell(row_index, column).fill = input_fill
    info.row_dimensions[1].height = 34
    info.column_dimensions["A"].width = 18
    for column in ("B", "C", "D"):
        info.column_dimensions[column].width = 24
    info.freeze_panes = "A2"
    info.sheet_view.showGridLines = False
    info.print_area = f"A1:D{len(info_rows) + 1}"

    assessment.merge_cells("A1:A2")
    assessment.merge_cells("B1:B2")
    assessment.merge_cells("C1:C2")
    assessment.merge_cells("D1:E1")
    for column in range(6, 9):
        assessment.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
    assessment.merge_cells(start_row=1, start_column=9, end_row=1, end_column=16)
    assessment.merge_cells(start_row=1, start_column=17, end_row=1, end_column=24)
    assessment.merge_cells(start_row=1, start_column=25, end_row=2, end_column=25)
    first_row = ("安全能力分类", "L1 高阶战略能力", "L2 安全能力", "安全关注点", "", "安全技术服务", "评分粒度", "适用性", "当前状态", "", "", "", "", "", "", "", "目标状态", "", "", "", "", "", "", "", "其他说明")
    second_row = ("", "", "", "序号", "关注点", "", "", "", "组织与角色", "制度与流程", "平台与工具", "数据与信息", "组织说明", "流程说明", "工具说明", "数据说明", "组织与角色", "制度与流程", "平台与工具", "数据与信息", "组织说明", "流程说明", "工具说明", "数据说明", "")
    for column, value in enumerate(first_row, start=1):
        if value:
            assessment.cell(1, column, value)
    for column, value in enumerate(second_row, start=1):
        if value:
            assessment.cell(2, column, value)
    for row_index in (1, 2):
        for column in range(1, 26):
            cell = assessment.cell(row_index, column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
    assessment.row_dimensions[1].height = 30
    assessment.row_dimensions[2].height = 26

    business_rows = _template_business_rows(template, score_entries)
    for row_index, row in enumerate(business_rows, start=3):
        for column, value in enumerate(_workbook_row_values(row), start=1):
            cell = assessment.cell(row_index, column, value)
            cell.font = body_font
            cell.border = border
            cell.alignment = center if column in {4, 7, 8, 9, 10, 11, 12, 17, 18, 19, 20} else left
            if column >= 8:
                cell.fill = input_fill if score_entries is None else PatternFill("solid", fgColor="FFFDF5")
        assessment.row_dimensions[row_index].height = 38
    _merge_business_hierarchy(assessment, business_rows)
    for merged_range in assessment.merged_cells.ranges:
        for row in assessment.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row, min_col=merged_range.min_col, max_col=merged_range.max_col):
            for cell in row:
                cell.border = border
                cell.alignment = center

    widths = (22, 24, 26, 12, 30, 34, 16, 12, 15, 15, 15, 15, 26, 26, 26, 26, 15, 15, 15, 15, 26, 26, 26, 26, 28)
    for index, width in enumerate(widths, start=1):
        assessment.column_dimensions[get_column_letter(index)].width = width
    last_row = max(3, len(business_rows) + 2)
    applicability_validation = DataValidation(type="list", formula1='"适用,不适用"', allow_blank=True)
    level_validation = DataValidation(type="list", formula1='"L1,L2,L3,L4,L5"', allow_blank=True)
    assessment.add_data_validation(applicability_validation)
    applicability_validation.add(f"H3:H{last_row}")
    assessment.add_data_validation(level_validation)
    level_validation.add(f"I3:L{last_row}")
    level_validation.add(f"Q3:T{last_row}")
    assessment.freeze_panes = "H3"
    assessment.sheet_view.showGridLines = False
    assessment.auto_filter.ref = f"A2:Y{last_row}"
    assessment.print_title_rows = "1:2"
    assessment.print_options.horizontalCentered = True
    assessment.page_setup.orientation = "landscape"
    assessment.page_setup.fitToWidth = 1
    assessment.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _load_maturity_workbook(exchange: dict[str, Any]) -> tuple[Any | None, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    encoded = _text(exchange.get("workbookBase64"))
    if not encoded:
        return None, [{"row": 0, "code": "xlsx_missing", "message": "请选择 XLSX 文件。"}]
    try:
        raw = base64.b64decode(encoded, validate=True)
    except (ValueError, TypeError):
        return None, [{"row": 0, "code": "xlsx_base64_invalid", "message": "评分文件内容无法读取。"}]
    if len(raw) > 20 * 1024 * 1024:
        return None, [{"row": 0, "code": "xlsx_too_large", "message": "XLSX 文件不能超过 20 MB。"}]
    try:
        workbook = load_workbook(BytesIO(raw), data_only=False, read_only=False)
    except Exception:
        return None, [{"row": 0, "code": "xlsx_invalid", "message": "文件不是有效的 XLSX 工作簿。"}]
    required = {MATURITY_INFO_SHEET, MATURITY_ASSESSMENT_SHEET}
    if set(workbook.sheetnames) != required or len(workbook.sheetnames) != 2:
        return None, [{"row": 0, "code": "sheet_contract_invalid", "message": "工作簿必须且只能包含“模板基础信息”和“评估模板”两张工作表。"}]
    return workbook, []


def _workbook_metadata(workbook: Any) -> dict[str, str]:
    sheet = workbook[MATURITY_INFO_SHEET]
    return {
        _business_cell(sheet.cell(row, 1).value): _business_cell(sheet.cell(row, 2).value)
        for row in range(1, min(sheet.max_row, 30) + 1)
        if _business_cell(sheet.cell(row, 1).value) and _business_cell(sheet.cell(row, 2).value)
    }


def _merged_value_map(sheet: Any) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        value = sheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row, column)] = value
    return values


def _read_assessment_rows(workbook: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sheet = workbook[MATURITY_ASSESSMENT_SHEET]
    modern_targets = _business_cell(sheet.cell(1, 9).value) == "当前状态"
    header_checks = {
        (1, 1): "安全能力分类",
        (1, 2): "L1 高阶战略能力",
        (1, 3): "L2 安全能力",
        (1, 4): "安全关注点",
        (2, 4): "序号",
        (2, 5): "关注点",
        (1, 6): "安全技术服务",
        (1, 7): "评分粒度",
        (1, 8): "适用性",
    }
    header_checks.update(
        {
            (1, 9): "当前状态",
            (1, 17): "目标状态",
            (2, 9): "组织与角色",
            (2, 10): "制度与流程",
            (2, 11): "平台与工具",
            (2, 12): "数据与信息",
            (2, 17): "组织与角色",
            (2, 18): "制度与流程",
            (2, 19): "平台与工具",
            (2, 20): "数据与信息",
        }
        if modern_targets
        else {
            (1, 9): "组织与角色",
            (1, 10): "制度与流程",
            (1, 11): "平台与工具",
            (1, 12): "数据与信息",
            (1, 13): "目标等级",
        }
    )
    for (row, column), expected in header_checks.items():
        if _business_cell(sheet.cell(row, column).value) != expected:
            return [], [{"row": row, "code": "header_invalid", "message": f"评估模板第 {column} 列表头应为“{expected}”。"}]
    merged_values = _merged_value_map(sheet)
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for row_number in range(3, sheet.max_row + 1):
        column_count = 25 if modern_targets else 14
        values = [merged_values.get((row_number, column), sheet.cell(row_number, column).value) for column in range(1, column_count + 1)]
        if not any(_text(value) for value in values):
            continue
        row = {
            "row": row_number,
            "l0": _business_cell(values[0]),
            "l1": _business_cell(values[1]),
            "l2": _business_cell(values[2]),
            "focusNo": _business_cell(values[3]) or str(row_number - 2),
            "focus": _business_cell(values[4]),
            "services": _text(values[5]).strip(),
            "granularity": _business_cell(values[6]) or ("安全技术服务" if _business_cell(values[5]) else "关注点"),
            "applicability": _business_cell(values[7]),
            "organization": values[8],
            "process": values[9],
            "tool": values[10],
            "data": values[11],
            "currentNotes": {
                key: _text(values[12 + index]).strip() if modern_targets else ""
                for index, key in enumerate(ELEMENT_KEYS)
            },
            "targets": {
                key: values[16 + index] if modern_targets else values[12]
                for index, key in enumerate(ELEMENT_KEYS)
            },
            "targetNotes": {
                key: _text(values[20 + index]).strip() if modern_targets else _text(values[13]).strip()
                for index, key in enumerate(ELEMENT_KEYS)
            },
            "note": _text(values[24]).strip() if modern_targets else _text(values[13]).strip(),
            "workbookSchema": "v2.2" if modern_targets else "v2.1",
        }
        missing = [label for label, value in (("安全能力分类", row["l0"]), ("L1 高阶战略能力", row["l1"]), ("L2 安全能力", row["l2"]), ("安全关注点", row["focus"])) if not value]
        if missing:
            errors.append({"row": row_number, "code": "business_path_incomplete", "message": f"缺少业务结构字段：{'、'.join(missing)}。"})
            continue
        if row["granularity"] not in {"关注点", "安全技术服务"}:
            errors.append({"row": row_number, "code": "granularity_invalid", "message": "评分粒度只能填写“关注点”或“安全技术服务”。"})
            continue
        if row["granularity"] == "安全技术服务" and not _business_service_names(row["services"]):
            errors.append({"row": row_number, "code": "service_missing", "message": "按安全技术服务评分时必须填写安全技术服务。"})
            continue
        rows.append(row)
    if not rows and not errors:
        errors.append({"row": 0, "code": "assessment_rows_empty", "message": "评估模板中没有可导入的业务评估行。"})
    return rows, errors


def _custom_template_from_business_rows(metadata: dict[str, str], rows: list[dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    categories: list[dict[str, Any]] = []
    capabilities: list[dict[str, Any]] = []
    focuses: list[dict[str, Any]] = []
    services: list[dict[str, Any]] = []
    mappings: list[dict[str, Any]] = []
    score_items: list[dict[str, Any]] = []
    category_by_key: dict[tuple[str, ...], dict[str, Any]] = {}
    capability_by_key: dict[tuple[str, ...], dict[str, Any]] = {}
    focus_by_key: dict[tuple[str, ...], dict[str, Any]] = {}
    service_by_name: dict[str, dict[str, Any]] = {}
    focus_modes: dict[tuple[str, ...], str] = {}
    row_errors: list[dict[str, Any]] = []
    seen_items: set[tuple[str, ...]] = set()

    for row in rows:
        l0_key = (row["l0"],)
        if l0_key not in category_by_key:
            category = {
                "id": f"custom:l0:{_stable_hash(l0_key, 16)}", "code": f"L0-{len([item for item in categories if item['capabilityLevel'] == 'L0']) + 1:02d}",
                "name": row["l0"], "level": 1, "capabilityLevel": "L0", "parentId": None, "weight": 1, "sortOrder": len(categories), "includedInOverall": True, "sourceType": "WORKBOOK_IMPORT",
            }
            categories.append(category)
            category_by_key[l0_key] = category
        l0 = category_by_key[l0_key]
        l1_key = (row["l0"], row["l1"])
        if l1_key not in category_by_key:
            category = {
                "id": f"custom:l1:{_stable_hash(l1_key, 16)}", "code": f"L1-{len([item for item in categories if item['capabilityLevel'] == 'L1']) + 1:02d}",
                "name": row["l1"], "level": 2, "capabilityLevel": "L1", "parentId": l0["id"], "weight": 1, "sortOrder": len(categories), "includedInOverall": True, "sourceType": "WORKBOOK_IMPORT",
            }
            categories.append(category)
            category_by_key[l1_key] = category
        l1 = category_by_key[l1_key]
        l2_key = (row["l0"], row["l1"], row["l2"])
        if l2_key not in capability_by_key:
            capability = {
                "id": f"custom:l2:{_stable_hash(l2_key, 16)}", "code": f"L2-{len(capabilities) + 1:03d}", "name": row["l2"], "description": "",
                "capabilityLevel": "L2", "categoryId": l1["id"], "topCategoryId": l0["id"], "weight": 1, "sortOrder": len(capabilities), "included": True,
                "isCustom": True, "isCritical": False, "businessImportance": 3, "riskUrgency": 3, "targetLevel": "L3", "sourceType": "WORKBOOK_IMPORT", "focusIds": [],
            }
            capabilities.append(capability)
            capability_by_key[l2_key] = capability
        capability = capability_by_key[l2_key]
        focus_key = (*l2_key, row["focusNo"], row["focus"])
        previous_mode = focus_modes.get(focus_key)
        if previous_mode and previous_mode != row["granularity"]:
            row_errors.append({"row": row["row"], "code": "focus_granularity_conflict", "message": "同一关注点不能同时按关注点和安全技术服务评分。"})
            continue
        focus_modes[focus_key] = row["granularity"]
        if focus_key not in focus_by_key:
            focus_id = f"custom:focus:{_stable_hash(focus_key, 16)}"
            focus = {
                "id": focus_id, "code": row["focusNo"], "name": row["focus"], "description": "", "capabilityId": capability["id"], "weight": 1,
                "sortOrder": len(focuses), "included": True, "isCustom": True, "isCritical": False, "itemType": "SERVICE" if row["granularity"] == "安全技术服务" else "FOCUS",
                "targetLevel": "L3", "sourceType": "WORKBOOK_IMPORT", "serviceMappingIds": [], "platformEvidenceServiceIds": [], "scoreItemIds": [],
            }
            focuses.append(focus)
            focus_by_key[focus_key] = focus
            capability["focusIds"].append(focus_id)
        focus = focus_by_key[focus_key]
        service_names = _business_service_names(row["services"])
        for service_name in service_names:
            if service_name not in service_by_name:
                service = {
                    "id": f"custom:service:{_stable_hash(service_name, 16)}", "code": f"SVC-{len(services) + 1:03d}", "name": service_name,
                    "scopeCode": "ALL", "scopeName": "全部作用域", "sourceType": "WORKBOOK_IMPORT", "isCustom": True,
                }
                services.append(service)
                service_by_name[service_name] = service
            service = service_by_name[service_name]
            item_key = (*focus_key, row["granularity"], service_name)
            if item_key in seen_items:
                if row["granularity"] == "安全技术服务":
                    row_errors.append({"row": row["row"], "code": "assessment_point_duplicate", "message": "同一关注点下的安全技术服务评估点重复。"})
                continue
            seen_items.add(item_key)
            role = "ASSESSMENT_POINT" if row["granularity"] == "安全技术服务" else "PLATFORM_EVIDENCE_REFERENCE"
            mapping_id = f"custom:mapping:{_stable_hash((focus['id'], service['id'], role), 16)}"
            mapping = {"id": mapping_id, "focusId": focus["id"], "scopeCode": "ALL", "scopeName": "全部作用域", "serviceId": service["id"], "serviceRole": role, "weight": 1, "sortOrder": len(mappings), "sourceType": "WORKBOOK_IMPORT"}
            mappings.append(mapping)
            focus["serviceMappingIds"].append(mapping_id)
            if role == "PLATFORM_EVIDENCE_REFERENCE":
                focus["platformEvidenceServiceIds"].append(service["id"])
            else:
                item_id = f"custom:score:{_stable_hash((focus['id'], service['id']), 16)}"
                score_items.append({"id": item_id, "itemType": "SERVICE", "capabilityId": capability["id"], "focusId": focus["id"], "serviceId": service["id"], "scopeCode": "ALL", "scopeName": "全部作用域", "weight": 1, "sortOrder": len(score_items), "required": True, "elementWeights": {key: 0.25 for key in ELEMENT_KEYS}, "rubricEntries": _rubric_entries_for_item(item_id), "sourceType": "WORKBOOK_IMPORT", "serviceRole": "ASSESSMENT_POINT", "sourceMappingId": mapping_id})
                focus["scoreItemIds"].append(item_id)

    for focus_key, focus in focus_by_key.items():
        if focus_modes.get(focus_key) != "关注点":
            continue
        item_id = f"custom:score:{_stable_hash((focus['id'], 'overall'), 16)}"
        capability = capability_by_key[focus_key[:3]]
        score_items.append({"id": item_id, "itemType": "FOCUS", "capabilityId": capability["id"], "focusId": focus["id"], "serviceId": None, "scopeCode": None, "scopeName": None, "weight": 1, "sortOrder": len(score_items), "required": True, "elementWeights": {key: 0.25 for key in ELEMENT_KEYS}, "rubricEntries": _rubric_entries_for_item(item_id), "sourceType": "WORKBOOK_IMPORT", "serviceRole": None, "platformEvidenceServiceIds": list(focus["platformEvidenceServiceIds"])})
        focus["scoreItemIds"].append(item_id)

    template_name = metadata.get("模板名称") or "导入的自定义成熟度模板"
    structure_basis = [(row["l0"], row["l1"], row["l2"], row["focusNo"], row["focus"], tuple(_business_service_names(row["services"])), row["granularity"]) for row in rows]
    template = {
        "id": f"custom-template-import-{_stable_hash(structure_basis, 16)}", "snapshotId": f"custom-template-snapshot-{_stable_hash(structure_basis, 20)}",
        "name": template_name, "version": metadata.get("模板版本") or "V2.1", "type": "custom", "status": "validated", "readOnly": False,
        "structureMutable": True, "weightMutable": True, "description": metadata.get("模板说明") or "由业务 XLSX 导入的自定义成熟度模板。", "rubricVersion": RUBRIC_VERSION,
        "categories": categories, "capabilities": capabilities, "focuses": focuses, "services": services, "focusServiceMappings": mappings,
        "scopes": [{"id": "scope:ALL", "code": "ALL", "name": "全部作用域", "sourceType": "WORKBOOK_IMPORT", "isCustom": True}],
        "scoreItems": score_items, "criticalRules": [], "elementWeights": {key: 0.25 for key in ELEMENT_KEYS},
        "stats": {"topCategories": sum(1 for item in categories if item["capabilityLevel"] == "L0"), "domains": sum(1 for item in categories if item["capabilityLevel"] == "L1"), "capabilities": len(capabilities), "focuses": len(focuses), "services": len(services), "serviceMappings": len(mappings), "scoreItems": len(score_items)},
    }
    return template, row_errors


def _legacy_export_maturity_score_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    project = _dict(payload.get("project"))
    template = _dict(payload.get("template"))
    validation = validate_maturity_template(template)
    if not validation["valid"]:
        return {"ok": False, "dataState": "invalid_template", "validation": validation}

    entry_by_id = {
        _text(item.get("scoreItemId")): item
        for item in _list(payload.get("scoreEntries"))
        if isinstance(item, dict) and _text(item.get("scoreItemId"))
    }
    capability_by_id = {_text(item.get("id")): item for item in _list(template.get("capabilities")) if isinstance(item, dict)}
    focus_by_id = {_text(item.get("id")): item for item in _list(template.get("focuses")) if isinstance(item, dict)}
    service_by_id = {_text(item.get("id")): item for item in _list(template.get("services")) if isinstance(item, dict)}
    assessment_items = []
    score_input = []
    for item in _list(template.get("scoreItems")):
        if not isinstance(item, dict):
            continue
        item_id = _text(item.get("id"))
        focus = focus_by_id.get(_text(item.get("focusId")), {})
        capability = capability_by_id.get(_text(focus.get("capabilityId")), {})
        service = service_by_id.get(_text(item.get("serviceId")), {})
        assessment_items.append(
            {
                "itemInstanceId": item_id,
                "templateItemId": item_id,
                "capabilityL2Code": _text(capability.get("code")),
                "capabilityL2Name": _text(capability.get("name")),
                "focusCode": _text(focus.get("code")),
                "focusName": _text(focus.get("name")),
                "itemType": _text(item.get("itemType")),
                "scopeCode": _text(item.get("scopeCode")) or None,
                "serviceCode": _text(service.get("code")) or None,
                "serviceName": _text(service.get("name")) or None,
                "sortOrder": item.get("sortOrder"),
            }
        )
        entry = entry_by_id.get(item_id, {})
        score_input.append(
            {
                "itemInstanceId": item_id,
                "isApplicable": entry.get("isApplicable") is not False,
                "organizationLevel": _dict(entry.get("elements")).get("organization"),
                "processLevel": _dict(entry.get("elements")).get("process"),
                "toolLevel": _dict(entry.get("elements")).get("tool"),
                "dataLevel": _dict(entry.get("elements")).get("data"),
                "targetLevel": entry.get("targetLevel"),
                "targetReason": _text(entry.get("targetReason")),
                "note": _text(entry.get("note")),
                "evidenceLevel": _text(entry.get("evidenceLevel") or "E0"),
                "evidenceSummary": _text(entry.get("evidenceSummary")),
                "naReason": _text(entry.get("naReason")),
            }
        )

    structure_hash = _template_structure_hash(template)
    assessment_items_hash = _stable_hash(assessment_items, 32)
    package = {
        "schemaVersion": SCORE_EXCHANGE_SCHEMA,
        "fileInfo": {
            "projectId": _text(project.get("id")),
            "templateId": _text(template.get("id")),
            "templateVersion": _text(template.get("version")),
            "templateSnapshotId": _text(template.get("snapshotId")),
            "knowledgeSnapshotId": _text(project.get("knowledgeSnapshotId") or template.get("knowledgeSnapshotId") or template.get("snapshotId")),
            "structureHash": structure_hash,
            "assessmentItemsHash": assessment_items_hash,
            "exportedAt": _now(),
        },
        "assessmentItems": assessment_items,
        "scoreInput": score_input,
        "rubricReference": {
            "version": _text(template.get("rubricVersion") or RUBRIC_VERSION),
            "dimensions": [
                {
                    "dimensionCode": dimension["dimensionCode"],
                    "dimensionName": ELEMENT_LABELS[dimension["dimensionCode"]],
                    "levels": [
                        {
                            "level": level["id"],
                            "levelName": level["name"],
                            "criteria": dimension["levels"][level["id"]],
                        }
                        for level in MATURITY_LEVELS
                    ],
                }
                for dimension in GENERIC_DIMENSION_RUBRIC
            ],
        },
    }
    batch_id = f"maturity-export-{_stable_hash(package, 20)}"
    return {
        "ok": True,
        "dataState": "ready",
        "batch": {"id": batch_id, "direction": "EXPORT", "exchangeType": "SCORE_DATA", "status": "success", "rowCount": len(score_input)},
        "fileName": f"{_text(project.get('id')) or 'maturity-project'}-score-exchange-v2.1.json",
        "package": package,
    }


def _legacy_import_maturity_score_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    project = _dict(payload.get("project"))
    template = _dict(payload.get("template"))
    exchange = _dict(payload.get("exchange"))
    existing_entries = {
        _text(item.get("scoreItemId")): deepcopy(item)
        for item in _list(payload.get("scoreEntries"))
        if isinstance(item, dict) and _text(item.get("scoreItemId"))
    }
    row_errors: list[dict[str, Any]] = []
    batch_id = f"maturity-import-{_stable_hash({'project': project.get('id'), 'exchange': exchange}, 20)}"

    if _text(exchange.get("schemaVersion")) != SCORE_EXCHANGE_SCHEMA:
        return {"ok": False, "dataState": "invalid_file", "batch": {"id": batch_id, "status": "failed", "successCount": 0, "failureCount": 1}, "rowErrors": [{"row": 0, "code": "schema_version_invalid", "message": "评分文件版本不受支持。"}]}
    file_info = _dict(exchange.get("fileInfo"))
    if _text(file_info.get("projectId")) != _text(project.get("id")):
        row_errors.append({"row": 0, "code": "project_mismatch", "message": "评分文件不属于当前项目。"})
    if _text(file_info.get("templateSnapshotId")) != _text(template.get("snapshotId")):
        row_errors.append({"row": 0, "code": "template_snapshot_mismatch", "message": "评分文件模板版本与当前项目不一致。"})
    if _text(file_info.get("structureHash")) != _template_structure_hash(template):
        row_errors.append({"row": 0, "code": "structure_hash_mismatch", "message": "评分文件结构与当前模板不一致。"})
    if _text(file_info.get("assessmentItemsHash")) != _stable_hash(_list(exchange.get("assessmentItems")), 32):
        row_errors.append({"row": 0, "code": "assessment_items_changed", "message": "只读评估项结构已被修改。"})
    if row_errors:
        return {"ok": False, "dataState": "invalid_structure", "batch": {"id": batch_id, "status": "failed", "successCount": 0, "failureCount": len(row_errors)}, "rowErrors": row_errors}

    valid_item_by_id = {
        _text(item.get("id")): item
        for item in _list(template.get("scoreItems"))
        if isinstance(item, dict) and _text(item.get("id"))
    }
    valid_item_ids = set(valid_item_by_id)
    imported_entries: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(_list(exchange.get("scoreInput")), start=2):
        if not isinstance(row, dict):
            row_errors.append({"row": row_number, "code": "row_invalid", "message": "评分行格式无效。"})
            continue
        item_id = _text(row.get("itemInstanceId"))
        if not item_id or item_id not in valid_item_ids:
            row_errors.append({"row": row_number, "itemInstanceId": item_id, "code": "item_not_found", "message": "评估点 ID 不存在。"})
            continue
        if item_id in seen:
            row_errors.append({"row": row_number, "itemInstanceId": item_id, "code": "item_duplicate", "message": "评分文件包含重复评估点。"})
            continue
        seen.add(item_id)
        is_applicable = row.get("isApplicable") is not False
        elements = {
            "organization": row.get("organizationLevel"),
            "process": row.get("processLevel"),
            "tool": row.get("toolLevel"),
            "data": row.get("dataLevel"),
        }
        if is_applicable and any(maturity_level_index(value) is None for value in elements.values()):
            row_errors.append({"row": row_number, "itemInstanceId": item_id, "code": "dimension_level_invalid", "message": "适用项必须填写四个有效成熟度等级。"})
            continue
        if is_applicable and maturity_level_index(row.get("targetLevel")) is None:
            row_errors.append({"row": row_number, "itemInstanceId": item_id, "code": "target_incomplete", "message": "适用项必须填写目标等级；评估说明为可选。"})
            continue
        target_index = maturity_level_index(row.get("targetLevel")) if is_applicable else None
        current_dimension_indices = {
            key: maturity_level_index(value)
            for key, value in elements.items()
        }
        minimum_target_index = max(
            (value for value in current_dimension_indices.values() if value is not None),
            default=None,
        ) if is_applicable else None
        if is_applicable and target_index is not None and minimum_target_index is not None and target_index < minimum_target_index:
            minimum_target_level = maturity_level_from_index(minimum_target_index)
            row_errors.append({"row": row_number, "itemInstanceId": item_id, "code": "target_below_current", "message": f"目标等级不能低于四维当前状态中的最高等级 {minimum_target_level}。"})
            continue
        entry = existing_entries.get(item_id, {"scoreItemId": item_id, "reviewElements": {}})
        entry.update(
            {
                "scoreItemId": item_id,
                "isApplicable": is_applicable,
                "elements": elements if is_applicable else {},
                "targetLevel": row.get("targetLevel") if is_applicable else None,
                "targetReason": _text(row.get("targetReason")) if is_applicable else "",
                "targetConfirmed": bool(is_applicable),
                "note": _text(row.get("note")),
                "evidenceLevel": _text(row.get("evidenceLevel") or "E0"),
                "evidenceSummary": _text(row.get("evidenceSummary")),
                "naReason": _text(row.get("naReason")) if not is_applicable else "",
                "status": "scored" if is_applicable else "not_applicable",
            }
        )
        imported_entries.append(entry)

    status = "success" if not row_errors else "partial_success" if imported_entries else "failed"
    merged_entries = {**existing_entries, **{item["scoreItemId"]: item for item in imported_entries}}
    return {
        "ok": bool(imported_entries) or not row_errors,
        "dataState": "ready" if status != "failed" else "invalid_rows",
        "batch": {
            "id": batch_id,
            "direction": "IMPORT",
            "exchangeType": "SCORE_DATA",
            "status": status,
            "rowCount": len(_list(exchange.get("scoreInput"))),
            "successCount": len(imported_entries),
            "failureCount": len(row_errors),
            "structureHash": _template_structure_hash(template),
        },
        "rowErrors": row_errors,
        "scoreEntries": list(merged_entries.values()),
    }


def _legacy_export_maturity_template_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    template = _dict(payload.get("template") or payload)
    validation = validate_maturity_template(template)
    if template.get("type") not in {"base", "custom"} or not validation["valid"]:
        return {"ok": False, "dataState": "invalid_template", "validation": validation, "message": "只有校验通过的默认或自定义模板可以导出结构文件。"}
    package = {
        "schemaVersion": TEMPLATE_EXCHANGE_SCHEMA,
        "fileInfo": {"templateId": _text(template.get("id")), "templateSnapshotId": _text(template.get("snapshotId")), "templateType": _text(template.get("type")), "structureHash": _template_structure_hash(template), "exportedAt": _now()},
        "template": template,
    }
    return {"ok": True, "dataState": "ready", "batch": {"id": f"maturity-template-export-{_stable_hash(package, 20)}", "status": "success", "direction": "EXPORT", "exchangeType": "TEMPLATE_STRUCTURE", "sourceTemplateType": template.get("type")}, "fileName": f"{_text(template.get('id')) or 'maturity-template'}-structure-v2.1.json", "package": package}


def _legacy_import_maturity_template_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    exchange = _dict(payload.get("exchange") or payload)
    batch_id = f"maturity-template-import-{_stable_hash(exchange, 20)}"
    if _text(exchange.get("schemaVersion")) != TEMPLATE_EXCHANGE_SCHEMA:
        return {"ok": False, "dataState": "invalid_file", "batch": {"id": batch_id, "status": "failed"}, "rowErrors": [{"row": 0, "code": "schema_version_invalid", "message": "模板结构文件版本不受支持。"}]}
    template = _dict(exchange.get("template"))
    file_info = _dict(exchange.get("fileInfo"))
    source_template_type = _text(template.get("type"))
    if source_template_type not in {"base", "custom"} or _text(file_info.get("structureHash")) != _template_structure_hash(template):
        return {"ok": False, "dataState": "invalid_structure", "batch": {"id": batch_id, "status": "failed"}, "rowErrors": [{"row": 0, "code": "structure_hash_mismatch", "message": "默认或自定义模板结构哈希不一致。"}]}
    validation = validate_maturity_template(template)
    imported_template = deepcopy(template)
    if validation["valid"]:
        source_template_id = _text(template.get("id"))
        source_snapshot_id = _text(template.get("snapshotId"))
        imported_template.update(
            {
                "id": f"custom-template-import-{_stable_hash({'batch': batch_id, 'source': source_template_id}, 16)}",
                "snapshotId": f"custom-template-snapshot-{_stable_hash({'batch': batch_id, 'structure': file_info.get('structureHash')}, 20)}",
                "name": f"{_text(template.get('name')) or '成熟度模板'}（导入副本）",
                "type": "custom",
                "status": "validated",
                "readOnly": False,
                "structureMutable": True,
                "weightMutable": True,
                "sourceTemplateId": source_template_id,
                "sourceTemplateSnapshotId": source_snapshot_id,
                "importSourceTemplateType": source_template_type,
            }
        )
    return {
        "ok": validation["valid"],
        "dataState": "ready" if validation["valid"] else "invalid_template",
        "batch": {"id": batch_id, "status": "success" if validation["valid"] else "failed", "direction": "IMPORT", "exchangeType": "TEMPLATE_STRUCTURE", "sourceTemplateType": source_template_type, "successCount": 1 if validation["valid"] else 0, "failureCount": len(validation["errors"])},
        "rowErrors": validation["errors"],
        "template": imported_template if validation["valid"] else None,
        "sourceTemplateType": source_template_type,
        "validation": validation,
    }


def export_maturity_score_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    project = _dict(payload.get("project"))
    template = _dict(payload.get("template"))
    validation = validate_maturity_template(template)
    if not validation["valid"]:
        return {"ok": False, "dataState": "invalid_template", "validation": validation}
    workbook_bytes = _build_maturity_workbook(template, purpose="项目评分", score_entries=_list(payload.get("scoreEntries")))
    structure_hash = _template_structure_hash(template)
    workbook_base64 = base64.b64encode(workbook_bytes).decode("ascii")
    package = {
        "schemaVersion": SCORE_EXCHANGE_SCHEMA,
        "fileInfo": {
            "templateVersion": _text(template.get("version")),
            "structureHash": structure_hash,
            "exportedAt": _now(),
            "businessWorkbook": True,
            "containsCustomerInformation": False,
        },
        "workbookBase64": workbook_base64,
        "mimeType": MATURITY_WORKBOOK_MIME,
    }
    batch_id = f"maturity-score-export-{_stable_hash((project.get('id'), structure_hash, len(workbook_bytes)), 20)}"
    return {
        "ok": True,
        "dataState": "ready",
        "batch": {"id": batch_id, "direction": "EXPORT", "exchangeType": "SCORE_DATA", "status": "success", "rowCount": len(_template_business_rows(template, _list(payload.get("scoreEntries"))))},
        "fileName": f"{_text(template.get('name')) or '成熟度评估'}-评分表.xlsx",
        "mimeType": MATURITY_WORKBOOK_MIME,
        "package": package,
    }


def import_maturity_score_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    project = _dict(payload.get("project"))
    template = _dict(payload.get("template"))
    exchange = _dict(payload.get("exchange") or payload)
    existing_entries = {
        _text(item.get("scoreItemId")): deepcopy(item)
        for item in _list(payload.get("scoreEntries"))
        if isinstance(item, dict) and _text(item.get("scoreItemId"))
    }
    batch_id = f"maturity-score-import-{_stable_hash((_text(project.get('id')), _text(exchange.get('fileName')), len(_text(exchange.get('workbookBase64')))), 20)}"
    workbook, file_errors = _load_maturity_workbook(exchange)
    if file_errors:
        return {"ok": False, "dataState": "invalid_file", "batch": {"id": batch_id, "direction": "IMPORT", "exchangeType": "SCORE_DATA", "status": "failed", "successCount": 0, "failureCount": len(file_errors)}, "rowErrors": file_errors, "scoreEntries": list(existing_entries.values())}
    metadata = _workbook_metadata(workbook)
    rows, row_errors = _read_assessment_rows(workbook)
    if metadata.get("模板名称") and metadata.get("模板名称") != _text(template.get("name")):
        row_errors.insert(0, {"row": 0, "code": "template_name_mismatch", "message": "评分文件的模板名称与当前项目不一致。"})
    expected_rows = _template_business_rows(template)
    expected_by_key = {_business_row_key(row): row for row in expected_rows}
    received_by_key: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        key = _business_row_key(row)
        if key in received_by_key:
            row_errors.append({"row": row["row"], "code": "assessment_row_duplicate", "message": "评分文件包含重复的业务评估行。"})
        else:
            received_by_key[key] = row
    missing_keys = set(expected_by_key) - set(received_by_key)
    unexpected_keys = set(received_by_key) - set(expected_by_key)
    if missing_keys:
        row_errors.append({"row": 0, "code": "assessment_rows_missing", "message": f"评分文件缺少 {len(missing_keys)} 个当前模板评估点，不能用于裁剪标准模板。"})
    if unexpected_keys:
        row_errors.append({"row": 0, "code": "assessment_rows_changed", "message": f"评分文件包含 {len(unexpected_keys)} 个当前模板之外的业务评估点。"})
    structural_codes = {"header_invalid", "business_path_incomplete", "granularity_invalid", "service_missing", "assessment_rows_empty", "template_name_mismatch", "assessment_row_duplicate", "assessment_rows_missing", "assessment_rows_changed"}
    if any(error.get("code") in structural_codes for error in row_errors):
        return {"ok": False, "dataState": "invalid_structure", "batch": {"id": batch_id, "direction": "IMPORT", "exchangeType": "SCORE_DATA", "status": "failed", "rowCount": len(rows), "successCount": 0, "failureCount": len(row_errors), "structureHash": _template_structure_hash(template)}, "rowErrors": row_errors, "scoreEntries": list(existing_entries.values())}

    valid_items = {_text(item.get("id")): item for item in _list(template.get("scoreItems")) if isinstance(item, dict)}
    imported_entries: list[dict[str, Any]] = []
    for key, expected in expected_by_key.items():
        row = received_by_key[key]
        raw_applicability = _business_cell(row.get("applicability")).upper()
        if raw_applicability in {"", "适用", "是", "Y", "YES", "TRUE"}:
            is_applicable = True
        elif raw_applicability in {"不适用", "否", "N", "NO", "FALSE"}:
            is_applicable = False
        else:
            row_errors.append({"row": row["row"], "itemInstanceId": expected["itemId"], "code": "applicability_invalid", "message": "适用性只能填写“适用”或“不适用”。"})
            continue
        normalized_elements = {
            "organization": _normalized_level(row.get("organization")),
            "process": _normalized_level(row.get("process")),
            "tool": _normalized_level(row.get("tool")),
            "data": _normalized_level(row.get("data")),
        }
        target_elements = {
            key_name: _normalized_level(_dict(row.get("targets")).get(key_name))
            for key_name in ELEMENT_KEYS
        }
        if not is_applicable:
            normalized_elements = {}
            target_elements = {}
        else:
            invalid_dimensions = [ELEMENT_LABELS[key_name] for key_name, value in normalized_elements.items() if value is None]
            if invalid_dimensions:
                row_errors.append({"row": row["row"], "itemInstanceId": expected["itemId"], "code": "dimension_level_invalid", "message": f"适用项必须填写有效的 L1—L5 四维评分：{'、'.join(invalid_dimensions)}。"})
                continue
            invalid_targets = [ELEMENT_LABELS[key_name] for key_name, value in target_elements.items() if value is None]
            if invalid_targets:
                row_errors.append({"row": row["row"], "itemInstanceId": expected["itemId"], "code": "target_incomplete", "message": f"适用项必须填写有效的 L1—L5 四维目标：{'、'.join(invalid_targets)}。"})
                continue
            target_conflict_dimensions = [
                f"{ELEMENT_LABELS[key_name]}（当前 {normalized_elements[key_name]}，目标 {target_elements[key_name]}）"
                for key_name in ELEMENT_KEYS
                if maturity_level_index(target_elements[key_name]) < maturity_level_index(normalized_elements[key_name])
            ]
            if target_conflict_dimensions:
                row_errors.append({"row": row["row"], "itemInstanceId": expected["itemId"], "code": "target_below_current", "message": f"目标状态不能低于同维度当前状态：{'、'.join(target_conflict_dimensions)}。"})
                continue
        target_notes = {key_name: _text(_dict(row.get("targetNotes")).get(key_name)) for key_name in ELEMENT_KEYS}
        current_notes = {key_name: _text(_dict(row.get("currentNotes")).get(key_name)) for key_name in ELEMENT_KEYS}
        target_index = _entry_target_index(valid_items.get(expected["itemId"], {}), {"targetElements": target_elements}, template) if is_applicable else None
        derived_target_level = maturity_level_from_index(target_index)
        legacy_target_reason = next((_text(target_notes.get(key_name)) for key_name in ELEMENT_KEYS if _text(target_notes.get(key_name))), "")
        entry = existing_entries.get(expected["itemId"], {"scoreItemId": expected["itemId"], "reviewElements": {}, "dimensionNotes": {}, "targetElements": {}, "targetDimensionNotes": {}})
        entry.update(
            {
                "scoreItemId": expected["itemId"],
                "isApplicable": is_applicable,
                "elements": normalized_elements if is_applicable else {},
                "dimensionNotes": current_notes if is_applicable else {},
                "targetElements": target_elements if is_applicable else {},
                "targetDimensionNotes": target_notes if is_applicable else {},
                "targetLevel": derived_target_level if is_applicable else None,
                "targetReason": legacy_target_reason if is_applicable else "",
                "targetConfirmed": bool(is_applicable and target_index is not None),
                "naReason": _text(row.get("note")) if not is_applicable else "",
                "status": "scored" if is_applicable else "not_applicable",
                "lastUpdateScope": "XLSX_IMPORT",
                "lastUpdatedAt": _now(),
            }
        )
        imported_entries.append(entry)

    status = "success" if not row_errors else "partial_success" if imported_entries else "failed"
    merged_entries = {**existing_entries, **{item["scoreItemId"]: item for item in imported_entries}}
    return {
        "ok": bool(imported_entries) or not row_errors,
        "dataState": "ready" if status != "failed" else "invalid_rows",
        "batch": {"id": batch_id, "direction": "IMPORT", "exchangeType": "SCORE_DATA", "status": status, "rowCount": len(rows), "successCount": len(imported_entries), "failureCount": len(row_errors), "structureHash": _template_structure_hash(template)},
        "rowErrors": row_errors,
        "scoreEntries": list(merged_entries.values()),
    }


def export_maturity_template_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    template = _dict(payload.get("template") or payload)
    validation = validate_maturity_template(template)
    if template.get("type") not in {"base", "custom"} or not validation["valid"]:
        return {"ok": False, "dataState": "invalid_template", "validation": validation, "message": "只有校验通过的标准或自定义模板可以导出。"}
    workbook_bytes = _build_maturity_workbook(template, purpose="模板配置", score_entries=None)
    workbook_base64 = base64.b64encode(workbook_bytes).decode("ascii")
    package = {
        "schemaVersion": TEMPLATE_EXCHANGE_SCHEMA,
        "fileInfo": {"templateType": _text(template.get("type")), "structureHash": _template_structure_hash(template), "exportedAt": _now(), "businessWorkbook": True, "containsCustomerInformation": False, "containsScores": False},
        "workbookBase64": workbook_base64,
        "mimeType": MATURITY_WORKBOOK_MIME,
    }
    return {"ok": True, "dataState": "ready", "batch": {"id": f"maturity-template-export-{_stable_hash((_template_structure_hash(template), len(workbook_bytes)), 20)}", "status": "success", "direction": "EXPORT", "exchangeType": "TEMPLATE_STRUCTURE", "sourceTemplateType": template.get("type"), "rowCount": len(_template_business_rows(template))}, "fileName": f"{_text(template.get('name')) or '成熟度模板'}-业务模板.xlsx", "mimeType": MATURITY_WORKBOOK_MIME, "package": package}


def import_maturity_template_exchange(payload: dict[str, Any]) -> dict[str, Any]:
    exchange = _dict(payload.get("exchange") or payload)
    batch_id = f"maturity-template-import-{_stable_hash((_text(exchange.get('fileName')), len(_text(exchange.get('workbookBase64')))), 20)}"
    workbook, file_errors = _load_maturity_workbook(exchange)
    if file_errors:
        return {"ok": False, "dataState": "invalid_file", "batch": {"id": batch_id, "direction": "IMPORT", "exchangeType": "TEMPLATE_STRUCTURE", "status": "failed", "successCount": 0, "failureCount": len(file_errors)}, "rowErrors": file_errors}
    metadata = _workbook_metadata(workbook)
    rows, row_errors = _read_assessment_rows(workbook)
    if metadata.get("模板类型") != "自定义模板":
        row_errors.insert(0, {"row": 0, "code": "standard_template_import_forbidden", "message": "只允许导入自定义模板；请在“模板基础信息”中把模板类型下拉切换为“自定义模板”，再填写新的模板名称和模板说明。"})
    elif not metadata.get("模板名称") or not metadata.get("模板说明"):
        row_errors.insert(0, {"row": 0, "code": "custom_template_metadata_required", "message": "自定义模板必须填写模板名称和模板说明。"})
    elif metadata.get("模板名称") == "SAPD标准能力成熟度模板" or metadata.get("模板说明") == "基于当前稳定能力、关注点、安全技术服务和作用域关系生成的只读评估模板。":
        row_errors.insert(0, {"row": 0, "code": "custom_template_metadata_unchanged", "message": "切换为自定义模板后，必须重新填写模板名称和模板说明，不能沿用标准模板信息。"})
    for row in rows:
        populated_fields = [label for label, key in (("适用性", "applicability"), ("组织与角色", "organization"), ("制度与流程", "process"), ("平台与工具", "tool"), ("数据与信息", "data"), ("其他说明", "note")) if _text(row.get(key))]
        populated_fields.extend(f"当前{ELEMENT_LABELS[key]}说明" for key in ELEMENT_KEYS if _text(_dict(row.get("currentNotes")).get(key)))
        populated_fields.extend(f"目标{ELEMENT_LABELS[key]}" for key in ELEMENT_KEYS if _text(_dict(row.get("targets")).get(key)))
        populated_fields.extend(f"目标{ELEMENT_LABELS[key]}说明" for key in ELEMENT_KEYS if _text(_dict(row.get("targetNotes")).get(key)))
        if populated_fields:
            row_errors.append({"row": row["row"], "code": "template_contains_scores", "message": f"评分标题和评分列必须保留，但自定义模板中的评分数据单元格必须为空；检测到：{'、'.join(populated_fields)}。"})
    if row_errors:
        return {"ok": False, "dataState": "invalid_template_file", "batch": {"id": batch_id, "direction": "IMPORT", "exchangeType": "TEMPLATE_STRUCTURE", "status": "failed", "successCount": 0, "failureCount": len(row_errors)}, "rowErrors": row_errors, "sourceTemplateType": "custom"}
    template, build_errors = _custom_template_from_business_rows(metadata, rows)
    validation = validate_maturity_template(template)
    row_errors = [*build_errors, *validation["errors"]]
    ok = not row_errors and validation["valid"]
    if ok and validation.get("snapshotId"):
        template["snapshotId"] = validation["snapshotId"]
    return {
        "ok": ok,
        "dataState": "ready" if ok else "invalid_template",
        "batch": {"id": batch_id, "status": "success" if ok else "failed", "direction": "IMPORT", "exchangeType": "TEMPLATE_STRUCTURE", "sourceTemplateType": "custom", "successCount": 1 if ok else 0, "failureCount": len(row_errors), "rowCount": len(rows)},
        "rowErrors": row_errors,
        "template": template if ok else None,
        "sourceTemplateType": "custom",
        "validation": validation,
    }


def _legacy_create_maturity_report_snapshot(payload: dict[str, Any]) -> dict[str, Any]:
    project = _dict(payload.get("project"))
    template = _dict(payload.get("template"))
    score_entries = _list(payload.get("scoreEntries"))
    narrative = _dict(payload.get("narrative"))
    result = calculate_maturity_assessment({"project": project, "template": template, "scoreEntries": score_entries})
    if not result.get("ok"):
        return {"ok": False, "dataState": result.get("dataState", "invalid"), "validation": result.get("validation", {})}

    summary = _dict(result.get("summary"))
    top_gaps = _list(result.get("gapItems"))[:10]
    submitted_roadmap = {
        _text(item.get("capabilityId")): item
        for item in _list(payload.get("improvementRoadmap"))
        if isinstance(item, dict) and _text(item.get("capabilityId"))
    }
    improvement_roadmap: list[dict[str, Any]] = []
    for rank, gap in enumerate(top_gaps, start=1):
        capability_id = _text(gap.get("capabilityId"))
        manual = _dict(submitted_roadmap.get(capability_id))
        default_action = " ".join(
            _text(item.get("text"))
            for item in _list(gap.get("recommendations"))[:2]
            if _text(item.get("text"))
        )
        status = _text(manual.get("status"))
        improvement_roadmap.append(
            {
                "rank": rank,
                "capabilityId": capability_id,
                "capabilityCode": _text(gap.get("capabilityCode")),
                "capabilityName": _text(gap.get("capabilityName")),
                "priority": _text(gap.get("priority")),
                "priorityScore": gap.get("priorityScore"),
                "currentLevel": _text(gap.get("currentLevel")),
                "targetLevel": _text(gap.get("targetLevel")),
                "gapIndex": gap.get("gapIndex"),
                "action": _text(manual.get("action")) if "action" in manual else default_action,
                "owner": _text(manual.get("owner")),
                "resources": _text(manual.get("resources")),
                "dependencies": _text(manual.get("dependencies")),
                "status": status if status in IMPROVEMENT_ROADMAP_STATUSES else "待规划",
            }
        )
    snapshot_basis = {
        "projectId": project.get("id"),
        "templateSnapshotId": template.get("snapshotId"),
        "summary": summary,
        "entries": score_entries,
        "narrative": narrative,
        "improvementRoadmap": improvement_roadmap,
    }
    snapshot_id = f"maturity-report-{_stable_hash(snapshot_basis, 20)}"
    generated_at = _now()
    is_formal = (
        project.get("status") in {"completed", "reported", "archived"}
        and summary.get("statisticsReady") is True
        and summary.get("completionRate") == 100
    )
    report_status = "snapshot" if is_formal else "draft_preview"
    conclusion_heading = "总体结论" if is_formal else "当前试算结论"
    report_notice = "正式报告快照；全部适用评估点已完成，不适用与无证据作为信息口径保留。" if is_formal else f"草稿报告预览；仍有 {summary.get('notScoredCount')} 条适用项待评分，当前等级仅基于已评分项，不是正式评估结论。"
    project_name = _text(project.get("name")) or "成熟度评估项目"
    organization = _text(project.get("organization")) or "未填写"
    capability_results = _list(result.get("capabilityResults"))
    dimension_labels = {"organization": "组织与角色", "process": "制度与流程", "tool": "平台与工具", "data": "数据与信息"}
    dimension_results = _dict(summary.get("dimensionResults"))
    no_evidence_count = next((int(_float(item.get("count"), 0.0)) for item in _list(result.get("evidenceDistribution")) if _text(item.get("level")) == "E0"), 0)

    def narrative_value(key: str, placeholder: str) -> str:
        return _text(narrative.get(key)) or f"[待填写：{placeholder}]"

    executive_summary = narrative_value("executiveSummary", "管理层摘要")
    key_findings = narrative_value("keyFindings", "关键发现")
    management_recommendations = narrative_value("managementRecommendations", "管理建议")
    next_steps = narrative_value("nextSteps", "下一步计划与复评安排")

    category_lines = [
        f"| {item.get('code') or '-'} | {item.get('name')} | {item.get('currentIndex') or '-'} | {item.get('currentLevel')} | {item.get('targetLevel')} | {item.get('gapIndex') if item.get('gapIndex') is not None else '-'} |"
        for item in _list(result.get("categoryResults"))
    ]
    gap_lines = [
        f"| {item.get('priority')} | {item.get('capabilityCode') or '-'} | {item.get('capabilityName')} | {item.get('currentLevel')} | {item.get('targetLevel')} | {item.get('priorityScore')} |"
        for item in top_gaps
    ]
    def markdown_cell(value: Any, fallback: str = "-") -> str:
        cleaned = _text(value).replace("|", "\\|").replace("\r\n", "<br />").replace("\n", "<br />")
        return cleaned or fallback

    roadmap_lines = [
        "| {rank} | {priority} | {code} {name} | {current} → {target} | {gap} | {action} | {owner} | {resources} | {dependencies} | {status} |".format(
            rank=item.get("rank"),
            priority=markdown_cell(item.get("priority")),
            code=markdown_cell(item.get("capabilityCode")),
            name=markdown_cell(item.get("capabilityName")),
            current=markdown_cell(item.get("currentLevel")),
            target=markdown_cell(item.get("targetLevel")),
            gap=item.get("gapIndex") if item.get("gapIndex") is not None else "-",
            action=markdown_cell(item.get("action"), "待明确"),
            owner=markdown_cell(item.get("owner"), "待指定"),
            resources=markdown_cell(item.get("resources"), "待评估"),
            dependencies=markdown_cell(item.get("dependencies"), "无"),
            status=markdown_cell(item.get("status"), "待规划"),
        )
        for item in improvement_roadmap
    ]
    capability_lines = [
        f"| {item.get('code') or '-'} | {item.get('name')} | {item.get('currentIndex') if item.get('currentIndex') is not None else '-'} | {item.get('currentLevel')} | {item.get('targetIndex') if item.get('targetIndex') is not None else '-'} | {item.get('targetLevel')} | {item.get('gapIndex') if item.get('gapIndex') is not None else '-'} | {item.get('targetAchievementRate') if item.get('targetAchievementRate') is not None else '-'}% | {item.get('evidenceCoverage')}% |"
        for item in capability_results
    ]
    dimension_lines = [
        f"| {label} | {dimension_results.get(key) if dimension_results.get(key) is not None else '-'} |"
        for key, label in dimension_labels.items()
    ]
    markdown = "\n".join(
        [
            f"# {project_name}",
            "",
            f"> 报告状态：{report_notice}",
            "",
            "## 项目信息",
            "",
            f"- 客户或组织：{organization}",
            "- 评估对象：企业组织",
            f"- 所属行业：{_text(project.get('industry')) or '未填写'}",
            f"- 企业规模：{_text(project.get('companySize')) or '未填写'}",
            f"- 模板：{_text(template.get('name'))} { _text(template.get('version'))}",
            f"- 报告快照：{snapshot_id}",
            "",
            "## 管理层摘要",
            "",
            executive_summary,
            "",
            f"## {conclusion_heading}",
            "",
            f"- 成熟度指数：{summary.get('currentIndex')}",
            f"- 成熟度等级：{summary.get('currentLevel')}",
            f"- 百分制得分：{summary.get('currentPercent')}",
            f"- 目标等级：{summary.get('targetLevel')}",
            f"- 目标达成率：{summary.get('targetAchievementRate') if summary.get('targetAchievementRate') is not None else '-'}%",
            f"- 评分完成度：{summary.get('completionRate')}%",
            f"- 证据覆盖率：{summary.get('evidenceCoverage')}%",
            "",
            "## 四维成熟度",
            "",
            "| 维度 | 当前指数 |",
            "|---|---:|",
            *dimension_lines,
            "",
            "## 能力类别评分",
            "",
            "| 编码 | 分类 | 当前指数 | 当前等级 | 目标等级 | 差距 |",
            "|---|---|---:|:---:|:---:|---:|",
            *category_lines,
            "",
            "## L2 安全能力结果",
            "",
            "| 编码 | L2 安全能力 | 当前指数 | 当前等级 | 目标指数 | 目标等级 | 差距 | 目标达成率 | 证据覆盖 |",
            "|---|---|---:|:---:|---:|:---:|---:|---:|---:|",
            *capability_lines,
            "",
            "## 高优先级差距",
            "",
            "| 优先级 | 能力编码 | 能力 | 当前 | 目标 | 优先级分数 |",
            "|:---:|---|---|:---:|:---:|---:|",
            *(gap_lines or ["| - | - | 当前没有已计算差距 | - | - | - |"]),
            "",
            "## 关键发现",
            "",
            key_findings,
            "",
            "## 管理建议",
            "",
            management_recommendations,
            "",
            "## 改进路线图",
            "",
            "| 排名 | 优先级 | L2 能力 | 当前 / 目标 | 差距 | 改进行动 | 负责人 | 资源投入 | 依赖事项 | 状态 |",
            "|---:|:---:|---|:---:|---:|---|---|---|---|:---:|",
            *(roadmap_lines or ["| - | - | 当前没有可执行的改进差距 | - | - | - | - | - | - | - |"]),
            "",
            "## 下一步计划",
            "",
            next_steps,
            "",
            "## 证据与限制",
            "",
            f"- 待评分项：{summary.get('notScoredCount')}。",
            f"- N/A 项：{summary.get('notApplicableCount')}。",
            f"- 无证据项：{no_evidence_count}；证据材料为可选补充，不阻塞完成评估。",
            "- 客户主结果最细展示到 L2 安全能力；关注点、服务、四维原始评分和证据保留在内部明细及附录。",
            "- 改进建议为结构化候选，需要评估人员确认后再进入正式报告。",
            "- 当前本地草稿不写入正式数据库、正式数据包或用户库。",
        ]
    )

    gap_rows_html = "".join(
        f"<tr><td>{html_lib.escape(_text(item.get('priority')))}</td><td>{html_lib.escape(_text(item.get('capabilityCode')))}</td><td>{html_lib.escape(_text(item.get('capabilityName')))}</td><td>{html_lib.escape(_text(item.get('currentLevel')))}</td><td>{html_lib.escape(_text(item.get('targetLevel')))}</td><td>{item.get('priorityScore')}</td></tr>"
        for item in top_gaps
    )
    roadmap_rows_html = "".join(
        "<tr><td>{rank}</td><td>{priority}</td><td><strong>{code}</strong><br />{name}</td><td>{current} → {target}</td><td>{gap}</td><td>{action}</td><td>{owner}</td><td>{resources}</td><td>{dependencies}</td><td>{status}</td></tr>".format(
            rank=item.get("rank"),
            priority=html_lib.escape(_text(item.get("priority"))),
            code=html_lib.escape(_text(item.get("capabilityCode"))),
            name=html_lib.escape(_text(item.get("capabilityName"))),
            current=html_lib.escape(_text(item.get("currentLevel"))),
            target=html_lib.escape(_text(item.get("targetLevel"))),
            gap=item.get("gapIndex") if item.get("gapIndex") is not None else "-",
            action=html_lib.escape(_text(item.get("action")) or "待明确"),
            owner=html_lib.escape(_text(item.get("owner")) or "待指定"),
            resources=html_lib.escape(_text(item.get("resources")) or "待评估"),
            dependencies=html_lib.escape(_text(item.get("dependencies")) or "无"),
            status=html_lib.escape(_text(item.get("status")) or "待规划"),
        )
        for item in improvement_roadmap
    )
    capability_rows_html = "".join(
        f"<tr><td>{html_lib.escape(_text(item.get('code')))}</td><td>{html_lib.escape(_text(item.get('name')))}</td><td>{item.get('currentIndex') if item.get('currentIndex') is not None else '-'}</td><td>{html_lib.escape(_text(item.get('currentLevel')))}</td><td>{item.get('targetIndex') if item.get('targetIndex') is not None else '-'}</td><td>{html_lib.escape(_text(item.get('targetLevel')))}</td><td>{item.get('gapIndex') if item.get('gapIndex') is not None else '-'}</td><td>{item.get('targetAchievementRate') if item.get('targetAchievementRate') is not None else '-'}%</td><td>{item.get('evidenceCoverage')}%</td></tr>"
        for item in capability_results
    )
    dimension_bars_html = "".join(
        f"<div class='bar-row'><span>{html_lib.escape(label)}</span><i><b style='width:{max(0.0, min(100.0, _float(dimension_results.get(key), 0.0) * 20.0)):.1f}%'></b></i><strong>{dimension_results.get(key) if dimension_results.get(key) is not None else '-'}</strong></div>"
        for key, label in dimension_labels.items()
    )
    tgm_bars_html = "".join(
        f"<div class='bar-row is-{html_lib.escape(_text(item.get('code')).lower())}'><span>{html_lib.escape(_text(item.get('code')))} {html_lib.escape(_text(item.get('name')))}</span><i><b style='width:{max(0.0, min(100.0, _float(item.get('currentIndex'), 0.0) * 20.0)):.1f}%'></b><em style='left:{max(0.0, min(100.0, _float(item.get('targetIndex'), 0.0) * 20.0)):.1f}%'></em></i><strong>{item.get('currentIndex') if item.get('currentIndex') is not None else '-'} / {item.get('targetIndex') if item.get('targetIndex') is not None else '-'}</strong></div>"
        for item in _list(result.get("categoryResults"))[:3]
    )

    def editable_html(value: str) -> str:
        return html_lib.escape(value).replace("\n", "<br />")

    html_report = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{html_lib.escape(project_name)}</title>
  <style>
    :root{{--ink:#24384b;--muted:#667b8e;--line:#d8e2ea;--blue:#2479be;--gold:#a87934;--paper:#fff}}
    *{{box-sizing:border-box}} body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;margin:0;color:var(--ink);background:#eef3f7;line-height:1.55}}
    main{{max-width:1120px;margin:0 auto;background:var(--paper);box-shadow:0 12px 36px rgba(31,52,72,.12)}}
    .cover{{min-height:340px;display:grid;align-content:end;padding:56px;background:linear-gradient(145deg,#eaf4fb 0%,#fff 58%,#f7f0e5 100%);border-bottom:1px solid var(--line)}}
    .eyebrow{{color:var(--blue);font-size:12px;font-weight:750;letter-spacing:.08em}} h1{{max-width:760px;font-size:36px;line-height:1.2;margin:12px 0}} h2{{font-size:21px;margin:0 0 16px}} h3{{font-size:15px;margin:0 0 10px}}
    .meta{{color:var(--muted);font-size:13px}} .content{{padding:42px 56px 64px}} .notice{{border-left:3px solid var(--blue);background:#f1f7fb;padding:12px 14px;color:#4b6479}}
    .summary{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));margin:28px 0;border:1px solid var(--line)}} .metric{{padding:18px;border-right:1px solid var(--line)}} .metric:last-child{{border:0}} .metric span{{display:block;color:var(--muted);font-size:11px}} .metric strong{{display:block;margin-top:6px;font-size:26px;letter-spacing:-.02em}}
    .section{{margin-top:38px;break-inside:avoid}} .section-head{{display:flex;align-items:end;justify-content:space-between;gap:16px;border-bottom:1px solid var(--line);margin-bottom:16px}} .section-head p{{margin:0 0 10px;color:var(--muted);font-size:11px}}
    .editable{{min-height:104px;padding:16px;border:1px dashed #aebfcd;background:#fbfcfd;white-space:normal}} .editable:focus{{outline:2px solid rgba(36,121,190,.24)}}
    .profile-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:28px}} .bar-list{{display:grid;gap:11px}} .bar-row{{display:grid;grid-template-columns:120px minmax(120px,1fr) 70px;align-items:center;gap:10px;font-size:12px}} .bar-row i{{position:relative;height:8px;background:#e6edf2}} .bar-row i b{{display:block;height:100%;background:var(--blue)}} .bar-row i em{{position:absolute;top:-3px;width:2px;height:14px;background:var(--gold)}} .bar-row strong{{text-align:right;font-variant-numeric:tabular-nums}}
    table{{width:100%;border-collapse:collapse;font-size:11px}} th,td{{padding:9px 10px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}} th{{background:#eef3f7;color:#516a7e}} tbody tr:nth-child(even){{background:#fafcfd}}
    .roadmap-section{{break-inside:auto}} .roadmap-table{{table-layout:fixed;font-size:9px}} .roadmap-table th,.roadmap-table td{{padding:7px 6px;overflow-wrap:anywhere}} .roadmap-table tr{{break-inside:avoid}} .roadmap-table th:nth-child(1){{width:4%}} .roadmap-table th:nth-child(2){{width:6%}} .roadmap-table th:nth-child(3){{width:14%}} .roadmap-table th:nth-child(4){{width:8%}} .roadmap-table th:nth-child(5){{width:5%}} .roadmap-table th:nth-child(6){{width:23%}} .roadmap-table th:nth-child(7){{width:9%}} .roadmap-table th:nth-child(8){{width:10%}} .roadmap-table th:nth-child(9){{width:13%}} .roadmap-table th:nth-child(10){{width:8%}}
    .limits{{color:var(--muted);font-size:11px}}
    @media(max-width:760px){{.cover{{padding:32px;min-height:280px}}.content{{padding:28px 24px 44px}}.summary{{grid-template-columns:repeat(2,minmax(0,1fr))}}.metric:nth-child(2){{border-right:0}}.profile-grid{{grid-template-columns:1fr}}.bar-row{{grid-template-columns:100px minmax(100px,1fr) 62px}}}}
    @media print{{body{{background:#fff}}main{{box-shadow:none}}.editable{{border-color:#cbd5de}}.section{{break-inside:avoid}}}}
  </style>
</head>
<body><main><header class="cover"><span class="eyebrow">SAPD 成熟度评估报告</span><h1>{html_lib.escape(project_name)}</h1><p class="meta">{html_lib.escape(organization)} · {html_lib.escape(_text(template.get('name')))} · {html_lib.escape(generated_at[:10])}</p></header><div class="content">
  <p class="notice">{html_lib.escape(report_notice)}</p>
  <section class="section"><div class="section-head"><h2>管理层摘要</h2><p>可直接在导出的 HTML 中继续编辑</p></div><div class="editable" contenteditable="true">{editable_html(executive_summary)}</div></section>
  <section class="summary"><div class="metric"><span>当前成熟度</span><strong>{html_lib.escape(_text(summary.get('currentLevel')))}</strong></div><div class="metric"><span>成熟度指数 / 5.00</span><strong>{summary.get('currentIndex')}</strong></div><div class="metric"><span>目标成熟度</span><strong>{html_lib.escape(_text(summary.get('targetLevel')))}</strong></div><div class="metric"><span>目标达成率</span><strong>{summary.get('targetAchievementRate') if summary.get('targetAchievementRate') is not None else '-'}%</strong></div></section>
  <section class="section"><div class="section-head"><h2>成熟度轮廓</h2><p>实色为当前指数；金色标记为目标指数</p></div><div class="profile-grid"><div><h3>四维成熟度</h3><div class="bar-list">{dimension_bars_html}</div></div><div><h3>能力类别评分</h3><div class="bar-list">{tgm_bars_html}</div></div></div></section>
  <section class="section"><div class="section-head"><h2>主要差距</h2><p>优先级与分数沿用后端结果</p></div><table><thead><tr><th>优先级</th><th>能力编码</th><th>能力</th><th>当前</th><th>目标</th><th>优先级分数</th></tr></thead><tbody>{gap_rows_html or '<tr><td colspan="6">当前没有已计算差距。</td></tr>'}</tbody></table></section>
  <section class="section"><div class="section-head"><h2>关键发现</h2></div><div class="editable" contenteditable="true">{editable_html(key_findings)}</div></section>
  <section class="section"><div class="section-head"><h2>管理建议</h2></div><div class="editable" contenteditable="true">{editable_html(management_recommendations)}</div></section>
  <section class="section roadmap-section"><div class="section-head"><h2>改进路线图</h2><p>由当前改进优先 Top 10 转化，行动信息由评估人员维护</p></div><table class="roadmap-table"><thead><tr><th>排名</th><th>优先级</th><th>L2 能力</th><th>当前 / 目标</th><th>差距</th><th>改进行动</th><th>负责人</th><th>资源投入</th><th>依赖事项</th><th>状态</th></tr></thead><tbody>{roadmap_rows_html or '<tr><td colspan="10">当前没有可执行的改进差距。</td></tr>'}</tbody></table></section>
  <section class="section"><div class="section-head"><h2>下一步计划</h2></div><div class="editable" contenteditable="true">{editable_html(next_steps)}</div></section>
  <section class="section"><div class="section-head"><h2>L2 安全能力结果</h2><p>{len(capability_results)} 项能力</p></div><table><thead><tr><th>编码</th><th>L2 安全能力</th><th>当前指数</th><th>当前等级</th><th>目标指数</th><th>目标等级</th><th>差距</th><th>达成率</th><th>证据覆盖</th></tr></thead><tbody>{capability_rows_html}</tbody></table></section>
  <section class="section limits"><div class="section-head"><h2>口径与限制</h2></div><p>适用项完成度 {summary.get('completionRate')}%；不适用项 {summary.get('notApplicableCount')}；无证据项 {no_evidence_count}。不适用和无证据为信息口径，不阻塞完成评估。</p><p>客户主结果最细到 L2；关注点、服务、四维原始评分和证据保留在内部明细。快照编号：{html_lib.escape(snapshot_id)}</p></section>
</div></main></body></html>"""

    return {
        "ok": True,
        "dataState": "ready",
        "id": snapshot_id,
        "status": report_status,
        "generatedAt": generated_at,
        "formal": is_formal,
        "summary": summary,
        "improvementRoadmap": improvement_roadmap,
        "markdown": markdown,
        "html": html_report,
        "fileNames": {
            "markdown": f"{snapshot_id}.md",
            "html": f"{snapshot_id}.html",
        },
    }


def create_maturity_report_snapshot(payload: dict[str, Any]) -> dict[str, Any]:
    """Create one immutable, self-contained assessment report from the backend result.

    ``reportModel`` is the versioned source consumed by both renderers.  The
    renderer layer only formats values already present in the calculation
    result (plus project facts and assessor-authored narrative); it does not
    calculate maturity scores or alter the assessment grain.
    """

    project = _dict(payload.get("project"))
    template = _dict(payload.get("template"))
    score_entries = _list(payload.get("scoreEntries"))
    narrative = _dict(payload.get("narrative"))
    result = calculate_maturity_assessment(
        {"project": project, "template": template, "scoreEntries": score_entries}
    )
    if not result.get("ok"):
        return {
            "ok": False,
            "dataState": result.get("dataState", "invalid"),
            "validation": result.get("validation", {}),
        }

    summary = deepcopy(_dict(result.get("summary")))
    capability_results = [deepcopy(item) for item in _list(result.get("capabilityResults")) if isinstance(item, dict)]
    category_results = [deepcopy(item) for item in _list(result.get("categoryResults")) if isinstance(item, dict)]
    subcategory_results = [deepcopy(item) for item in _list(result.get("subCategoryResults")) if isinstance(item, dict)]
    template_category_order = {
        _text(item.get("id")): index
        for index, item in enumerate(
            sorted(
                [item for item in _list(template.get("categories")) if isinstance(item, dict)],
                key=lambda item: (
                    _float(item.get("sortOrder"), float("inf")) if item.get("sortOrder") not in (None, "") else float("inf"),
                    _text(item.get("code") or item.get("name")),
                ),
            )
        )
    }
    category_results.sort(
        key=lambda item: (
            template_category_order.get(_text(item.get("id")), len(template_category_order)),
            _text(item.get("code") or item.get("name")),
        )
    )
    gap_items = [deepcopy(item) for item in _list(result.get("gapItems")) if isinstance(item, dict)]
    score_item_results = [deepcopy(item) for item in _list(result.get("scoreItemResults")) if isinstance(item, dict)]
    calculation_run = deepcopy(_dict(result.get("calculationRun")))
    project_name = _text(project.get("name")) or "成熟度评估项目"
    organization = _text(project.get("organization")) or "未填写"
    generated_at = _now()
    is_formal = (
        project.get("status") in {"completed", "reported", "archived"}
        and summary.get("statisticsReady") is True
        and summary.get("completionRate") == 100
    )
    if not is_formal:
        return {
            "ok": False,
            "dataState": "assessment_incomplete",
            "error": "assessment_must_be_completed_before_report_generation",
            "message": "请先完成全部适用评估点并正式完成评估，再生成评估报告。",
            "validation": {
                "errors": [
                    {
                        "code": "assessment_incomplete",
                        "message": "评估尚未正式完成，不能生成或持久化评估报告。",
                    }
                ]
            },
            "summary": summary,
        }

    def narrative_value(key: str, label: str) -> str:
        return _text(narrative.get(key)) or f"[待填写：{label}]"

    narrative_fields = [
        {
            "key": "executiveSummary",
            "label": "评估概况",
            "legacyLabel": "管理层摘要",
            "value": narrative_value("executiveSummary", "评估概况"),
            "editable": True,
        },
        {
            "key": "keyFindings",
            "label": "关键发现",
            "value": narrative_value("keyFindings", "关键发现"),
            "editable": True,
        },
        {
            "key": "managementRecommendations",
            "label": "提升建议",
            "legacyLabel": "管理建议",
            "value": narrative_value("managementRecommendations", "提升建议"),
            "editable": True,
        },
        {
            "key": "nextSteps",
            "label": "下一步计划",
            "value": narrative_value("nextSteps", "下一步计划"),
            "editable": True,
        },
    ]

    top_gaps = gap_items[:10]
    submitted_roadmap = {
        _text(item.get("capabilityId")): item
        for item in _list(payload.get("improvementRoadmap"))
        if isinstance(item, dict) and _text(item.get("capabilityId"))
    }
    improvement_roadmap: list[dict[str, Any]] = []
    for rank, gap in enumerate(top_gaps, start=1):
        capability_id = _text(gap.get("capabilityId"))
        manual = _dict(submitted_roadmap.get(capability_id))
        default_action = " ".join(
            _text(item.get("text"))
            for item in _list(gap.get("recommendations"))[:2]
            if _text(item.get("text"))
        )
        status = _text(manual.get("status"))
        improvement_roadmap.append(
            {
                "rank": rank,
                "capabilityId": capability_id,
                "capabilityCode": _text(gap.get("capabilityCode")),
                "capabilityName": _text(gap.get("capabilityName")),
                "priority": _text(gap.get("priority")),
                "priorityScore": gap.get("priorityScore"),
                "currentIndex": gap.get("currentIndex"),
                "currentLevel": _text(gap.get("currentLevel")),
                "targetIndex": gap.get("targetIndex"),
                "targetLevel": _text(gap.get("targetLevel")),
                "gapIndex": gap.get("gapIndex"),
                "action": _text(manual.get("action")) if "action" in manual else default_action,
                "owner": _text(manual.get("owner")),
                "resources": _text(manual.get("resources")),
                "dependencies": _text(manual.get("dependencies")),
                "status": status if status in IMPROVEMENT_ROADMAP_STATUSES else "待规划",
            }
        )

    category_by_id = {_text(item.get("id")): item for item in category_results}
    priority_by_capability = {
        _text(item.get("capabilityId")): item for item in gap_items if _text(item.get("capabilityId"))
    }

    def priority_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
        counts = {"高": 0, "中": 0, "低": 0}
        for row in rows:
            priority = _text(row.get("priority"))
            if priority in counts:
                counts[priority] += 1
        return counts

    capability_groups: list[dict[str, Any]] = []
    known_capability_ids: set[str] = set()
    for category in category_results:
        category_id = _text(category.get("id"))
        rows = [row for row in capability_results if _text(row.get("topCategoryId")) == category_id]
        if not rows:
            continue
        known_capability_ids.update(_text(row.get("id")) for row in rows)
        capability_groups.append(
            {
                "id": category_id,
                "code": _text(category.get("code")) or "—",
                "name": _text(category.get("name")) or "未命名能力分类",
                "currentIndex": category.get("currentIndex"),
                "targetIndex": category.get("targetIndex"),
                "l2Count": len(rows),
                "belowTargetCount": sum(1 for row in rows if _float(row.get("gapIndex"), 0.0) > 0),
                "priorityCounts": priority_counts(
                    [priority_by_capability[_text(row.get("id"))] for row in rows if _text(row.get("id")) in priority_by_capability]
                ),
                "capabilities": rows,
            }
        )
    unmatched = [row for row in capability_results if _text(row.get("id")) not in known_capability_ids]
    if unmatched:
        capability_groups.append(
            {
                "id": "ungrouped",
                "code": "—",
                "name": "未分组能力",
                "currentIndex": None,
                "targetIndex": None,
                "l2Count": len(unmatched),
                "belowTargetCount": sum(1 for row in unmatched if _float(row.get("gapIndex"), 0.0) > 0),
                "priorityCounts": priority_counts(
                    [priority_by_capability[_text(row.get("id"))] for row in unmatched if _text(row.get("id")) in priority_by_capability]
                ),
                "capabilities": unmatched,
            }
        )

    l1_statistics: list[dict[str, Any]] = []
    for row in subcategory_results:
        l2_rows = [item for item in capability_results if _text(item.get("categoryId")) == _text(row.get("id"))]
        group = next((item for item in capability_groups if item["id"] == _text(row.get("parentId"))), None)
        l1_statistics.append(
            {
                **row,
                "groupCode": _text(_dict(group).get("code")),
                "l2Count": len(l2_rows),
                "priorityCounts": priority_counts(
                    [priority_by_capability[_text(item.get("id"))] for item in l2_rows if _text(item.get("id")) in priority_by_capability]
                ),
            }
        )

    leading_overall = sorted(
        [row for row in capability_results if row.get("currentIndex") is not None],
        key=lambda row: (-_float(row.get("currentIndex")), -_float(row.get("targetAchievementRate")), _text(row.get("code"))),
    )[:10]
    improvement_overall = gap_items[:10]
    dimension_rankings: list[dict[str, Any]] = []
    for key in ELEMENT_KEYS:
        usable = [row for row in capability_results if _dict(row.get("dimensionResults")).get(key) is not None]
        leading = sorted(usable, key=lambda row: (-_float(_dict(row.get("dimensionResults")).get(key)), _text(row.get("code"))))[:10]
        improvement = sorted(
            [row for row in usable if _text(row.get("id")) in priority_by_capability],
            key=lambda row: (
                _float(_dict(row.get("dimensionResults")).get(key)),
                -_float(_dict(priority_by_capability.get(_text(row.get("id")))).get("priorityScore")),
                _text(row.get("code")),
            ),
        )[:10]
        dimension_rankings.append(
            {
                "dimension": key,
                "label": ELEMENT_LABELS[key],
                "leading": leading,
                "improvement": improvement,
            }
        )

    maturity_distribution = deepcopy(_list(result.get("maturityDistribution")))
    evidence_distribution = deepcopy(_list(result.get("evidenceDistribution")))
    service_distribution = deepcopy(_list(result.get("serviceDistribution")))
    evidence_total = sum(int(_float(item.get("count"))) for item in evidence_distribution if isinstance(item, dict))
    evidence_missing = next(
        (int(_float(item.get("count"))) for item in evidence_distribution if _text(item.get("level")) == "E0"),
        0,
    )
    dimension_values = [
        {"id": key, "label": ELEMENT_LABELS[key], "value": _dict(summary.get("dimensionResults")).get(key)}
        for key in ELEMENT_KEYS
    ]
    comparable_dimensions = [item for item in dimension_values if item["value"] is not None]
    strongest = max(comparable_dimensions, key=lambda item: _float(item["value"])) if comparable_dimensions else None
    weakest = min(comparable_dimensions, key=lambda item: _float(item["value"])) if comparable_dimensions else None
    evaluation = {
        "l2TotalCount": len(capability_results),
        "l2ScoredCount": sum(1 for row in capability_results if row.get("currentIndex") is not None),
        "l2BelowTargetCount": sum(1 for row in capability_results if _float(row.get("gapIndex"), 0.0) > 0),
        "l2ReachedTargetCount": sum(
            1 for row in capability_results if row.get("gapIndex") is not None and _float(row.get("gapIndex")) <= 0
        ),
        "leadingL1": sorted(
            [row for row in subcategory_results if row.get("currentIndex") is not None],
            key=lambda row: (-_float(row.get("currentIndex")), _text(row.get("code"))),
        )[:3],
        "improvementL1": sorted(
            [row for row in subcategory_results if _float(row.get("gapIndex"), 0.0) > 0],
            key=lambda row: (-_float(row.get("gapIndex")), _text(row.get("code"))),
        )[:3],
        "strongestDimension": strongest,
        "weakestDimension": weakest,
        "dimensionSpread": _round(_float(strongest["value"]) - _float(weakest["value"])) if strongest and weakest else None,
        "evidenceTotalCount": evidence_total,
        "evidenceFilledCount": max(0, evidence_total - evidence_missing),
        "evidenceMissingCount": evidence_missing,
    }

    focus_by_id = {_text(item.get("id")): item for item in _list(template.get("focuses")) if isinstance(item, dict)}
    score_entry_by_id = {
        _text(item.get("scoreItemId")): item for item in score_entries if isinstance(item, dict) and _text(item.get("scoreItemId"))
    }
    score_details: list[dict[str, Any]] = []
    for row in score_item_results:
        focus = _dict(focus_by_id.get(_text(row.get("focusId"))))
        entry = _dict(score_entry_by_id.get(_text(row.get("id"))))
        score_details.append(
            {
                **row,
                "focusCode": _text(focus.get("code")),
                "focusName": _text(focus.get("name")),
                "evidenceSummary": _text(entry.get("evidenceSummary")),
                "reviewElements": deepcopy(_dict(entry.get("reviewElements"))),
                "dimensionNotes": deepcopy(_dict(entry.get("dimensionNotes"))),
                "targetElements": deepcopy(_entry_target_levels(entry)),
                "targetDimensionNotes": deepcopy(
                    _dict(entry.get("targetDimensionNotes")) if "targetDimensionNotes" in entry else {key: _text(entry.get("targetReason")) for key in ELEMENT_KEYS}
                ),
            }
        )

    l2_rows: list[dict[str, Any]] = []
    for row in capability_results:
        priority = _dict(priority_by_capability.get(_text(row.get("id"))))
        top_category = _dict(category_by_id.get(_text(row.get("topCategoryId"))))
        l2_rows.append(
            {
                **row,
                "groupCode": _text(top_category.get("code")),
                "groupName": _text(top_category.get("name")),
                "priority": _text(priority.get("priority")),
                "priorityScore": priority.get("priorityScore"),
            }
        )

    capability_radar = {
        "id": "l2-capability-by-top-category",
        "type": "radar",
        "title": "全 L2 能力分组雷达",
        "scale": {"minimum": 1, "maximum": 5},
        "groups": [
            {"id": group["id"], "code": group["code"], "name": group["name"], "count": group["l2Count"]}
            for group in capability_groups
        ],
        "axes": [
            {
                "id": _text(row.get("id")),
                "code": _text(row.get("code")),
                "label": _text(row.get("name")),
                "displayLabel": CAPABILITY_RADAR_SHORT_LABELS.get(_text(row.get("code")), _text(row.get("code")) or _text(row.get("name"))),
                "groupCode": group["code"],
                "current": row.get("currentIndex"),
                "target": row.get("targetIndex"),
            }
            for group in capability_groups
            for row in group["capabilities"]
        ],
    }
    dimension_radar = {
        "id": "overall-four-dimension",
        "type": "radar",
        "title": "总体四维成熟度雷达",
        "scale": {"minimum": 1, "maximum": 5},
        "groups": [],
        "axes": [
            {
                "id": item["id"],
                "code": item["label"],
                "label": item["label"],
                "current": item["value"],
                "target": _dict(summary.get("targetDimensionResults")).get(item["id"]),
            }
            for item in dimension_values
        ],
        "targetNote": "目标状态来自各评估点四个目标维度的聚合结果。",
    }

    report_model: dict[str, Any] = {
        "schemaVersion": "sapd-maturity-report-model-v3",
        "rendererVersion": "sapd-maturity-report-renderer-v3",
        "resultVersion": {
            "algorithmVersion": _text(calculation_run.get("algorithmVersion") or summary.get("algorithmVersion")),
            "calculationRunId": _text(calculation_run.get("id")),
            "resultHash": _text(calculation_run.get("resultHash")),
            "templateSnapshotId": _text(summary.get("templateSnapshotId") or template.get("snapshotId")),
            "knowledgeSnapshotId": _text(summary.get("knowledgeSnapshotId")),
        },
        "report": {
            "title": "评估报告",
            "projectName": project_name,
            "organization": organization,
            "status": "snapshot" if is_formal else "draft_preview",
            "formal": is_formal,
        },
        "project": {
            "id": _text(project.get("id")),
            "name": project_name,
            "organization": organization,
            "industry": _text(project.get("industry")) or "未填写",
            "companySize": _text(project.get("companySize")) or "未填写",
            "assessmentObjectType": _text(summary.get("assessmentObjectType")) or "ENTERPRISE_ORGANIZATION",
            "templateName": _text(template.get("name")),
            "templateVersion": _text(template.get("version")),
        },
        "narrativeFields": narrative_fields,
        "sections": [
            {"id": "overall", "title": "总体结果（关键结论）", "renderer": "executive_overview", "data": {"summary": summary, "categories": category_results}},
            {"id": "narratives", "title": "管理层研判", "renderer": "editable_narratives", "data": narrative_fields},
            {"id": "radars", "title": "成熟度雷达", "renderer": "radar_suite", "data": {"capabilityRadar": capability_radar, "dimensionRadar": dimension_radar}},
            {"id": "hierarchy_statistics", "title": "T / G / M 与 L1 分层统计", "renderer": "hierarchy_statistics", "data": {"groups": capability_groups, "l1": l1_statistics, "priorityCounts": priority_counts(gap_items)}},
            {"id": "evaluation", "title": "结果评价与分布", "renderer": "evaluation_distributions", "data": {"evaluation": evaluation, "maturityDistribution": maturity_distribution, "evidenceDistribution": evidence_distribution, "serviceDistribution": service_distribution}},
            {"id": "capability_results", "title": "完整 L2 能力结果", "renderer": "l2_result_table", "data": l2_rows},
            {"id": "overall_rankings", "title": "总体 L2 能力 Top 10", "renderer": "overall_rankings", "data": {"leading": leading_overall, "improvement": improvement_overall}},
            {"id": "dimension_rankings", "title": "四维 L2 能力 Top 10", "renderer": "dimension_rankings", "data": dimension_rankings},
            {"id": "improvement_roadmap", "title": "改进路线图", "renderer": "improvement_roadmap", "data": improvement_roadmap},
            {"id": "score_appendix", "title": "完整评分明细附录", "renderer": "score_appendix", "data": score_details},
            {"id": "traceability", "title": "评估口径说明", "renderer": "traceability", "data": {"summary": summary}},
        ],
        # Keep the complete calculation output so future backend statistics can be
        # registered as another section without changing or reconstructing inputs.
        "resultSnapshot": deepcopy(result),
    }
    snapshot_basis = deepcopy(report_model)
    snapshot_basis["improvementRoadmap"] = improvement_roadmap
    report_model_hash = _stable_hash(snapshot_basis, 32)
    snapshot_id = f"maturity-report-{report_model_hash[:20]}"
    report_model["modelHash"] = report_model_hash
    report_model["snapshotId"] = snapshot_id
    report_model["generatedAt"] = generated_at

    def md_cell(value: Any, fallback: str = "-") -> str:
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False, sort_keys=True)
        cleaned = _text(value).replace("|", "\\|").replace("\r\n", "<br />").replace("\n", "<br />")
        return cleaned or fallback

    def md_table(headers: list[str], rows: list[list[Any]]) -> str:
        lines = ["| " + " | ".join(headers) + " |", "|" + "|".join("---" for _ in headers) + "|"]
        lines.extend("| " + " | ".join(md_cell(value) for value in row) + " |" for row in rows)
        if not rows:
            lines.append("| " + " | ".join(["暂无数据", *(["-"] * (len(headers) - 1))]) + " |")
        return "\n".join(lines)

    def value_or_dash(value: Any) -> Any:
        return "-" if value is None or value == "" else value

    def radar_data_markdown(chart: dict[str, Any]) -> str:
        return md_table(
            ["分组", "编码 / 维度", "名称", "当前", "目标"],
            [
                [axis.get("groupCode") or "总体", axis.get("code"), axis.get("label"), value_or_dash(axis.get("current")), value_or_dash(axis.get("target"))]
                for axis in _list(chart.get("axes"))
            ],
        )

    def render_markdown_section(section: dict[str, Any]) -> str:
        section_id = _text(section.get("id"))
        data = section.get("data")
        title = _text(section.get("title"))
        lines = [f"## {title}", ""]
        if section_id == "overall":
            section_summary = _dict(_dict(data).get("summary"))
            total = int(_float(section_summary.get("applicableItemCount"))) + int(_float(section_summary.get("notApplicableCount")))
            lines.extend(
                [
                    md_table(
                        ["当前成熟度", "目标成熟度", "适用性", "评估进度", "完成率", "证据覆盖"],
                        [[f"{section_summary.get('currentLevel')} / {value_or_dash(section_summary.get('currentIndex'))}", f"{section_summary.get('targetLevel')} / {value_or_dash(section_summary.get('targetIndex'))}", f"{section_summary.get('applicableItemCount')} / {total}", f"{section_summary.get('scoredItemCount')} / {section_summary.get('applicableItemCount')}", f"{section_summary.get('completionRate')}%", f"{section_summary.get('evidenceCoverage')}%"]],
                    ),
                    "",
                    md_table(
                        ["类别", "名称", "当前", "目标", "差距"],
                        [[item.get("code"), item.get("name"), value_or_dash(item.get("currentIndex")), value_or_dash(item.get("targetIndex")), value_or_dash(item.get("gapIndex"))] for item in _list(_dict(data).get("categories"))],
                    ),
                ]
            )
        elif section_id == "narratives":
            for field in _list(data):
                lines.extend([f"### {_text(field.get('label'))}", "", _text(field.get("value")), ""])
        elif section_id == "radars":
            for chart_key in ("capabilityRadar", "dimensionRadar"):
                chart = _dict(_dict(data).get(chart_key))
                lines.extend([f"### {_text(chart.get('title'))}", "", radar_data_markdown(chart), ""])
            lines.append(_text(_dict(_dict(data).get("dimensionRadar")).get("targetNote")))
        elif section_id == "hierarchy_statistics":
            lines.extend(
                [
                    md_table(
                        ["类别", "当前 / 目标", "L1", "L2", "低于目标", "高 / 中 / 低优先级"],
                        [[group.get("code"), f"{value_or_dash(group.get('currentIndex'))} / {value_or_dash(group.get('targetIndex'))}", sum(1 for item in _list(_dict(data).get("l1")) if _text(item.get("groupCode")) == _text(group.get("code"))), group.get("l2Count"), group.get("belowTargetCount"), f"{_dict(group.get('priorityCounts')).get('高', 0)} / {_dict(group.get('priorityCounts')).get('中', 0)} / {_dict(group.get('priorityCounts')).get('低', 0)}"] for group in _list(_dict(data).get("groups"))],
                    ),
                    "",
                    md_table(
                        ["类别", "L1 能力域", "当前 / 目标", "L2 数量", "高 / 中 / 低优先级"],
                        [[item.get("groupCode"), f"{item.get('code')} {item.get('name')}", f"{value_or_dash(item.get('currentIndex'))} / {value_or_dash(item.get('targetIndex'))}", item.get("l2Count"), f"{_dict(item.get('priorityCounts')).get('高', 0)} / {_dict(item.get('priorityCounts')).get('中', 0)} / {_dict(item.get('priorityCounts')).get('低', 0)}"] for item in _list(_dict(data).get("l1"))],
                    ),
                ]
            )
        elif section_id == "evaluation":
            evaluation_data = _dict(_dict(data).get("evaluation"))
            lines.extend(
                [
                    md_table(
                        ["L2 已评分 / 总数", "达到目标", "低于目标", "证据 E1+ / 总数", "四维极差"],
                        [[f"{evaluation_data.get('l2ScoredCount')} / {evaluation_data.get('l2TotalCount')}", evaluation_data.get("l2ReachedTargetCount"), evaluation_data.get("l2BelowTargetCount"), f"{evaluation_data.get('evidenceFilledCount')} / {evaluation_data.get('evidenceTotalCount')}", value_or_dash(evaluation_data.get("dimensionSpread"))]],
                    ),
                    "",
                    "### L1 结果评价依据",
                    "",
                    md_table(
                        ["评价类型", "L1 能力域", "当前", "目标", "差距"],
                        [[label, f"{item.get('code')} {item.get('name')}", item.get("currentIndex"), item.get("targetIndex"), item.get("gapIndex")] for key, label in (("leadingL1", "优势能力域"), ("improvementL1", "重点加强")) for item in _list(evaluation_data.get(key))],
                    ),
                    "",
                ]
            )
            for key, label in (("maturityDistribution", "成熟度分布"), ("evidenceDistribution", "证据分布"), ("serviceDistribution", "服务评估点分布")):
                lines.extend([f"### {label}", "", md_table(["等级", "名称", "数量"], [[item.get("level"), item.get("name"), item.get("count")] for item in _list(_dict(data).get(key))]), ""])
        elif section_id == "capability_results":
            lines.append(md_table(["类别", "L2 能力", "当前", "目标", "差距", "优先级", "达成率", "组织", "流程", "工具", "数据", "证据"], [[item.get("groupCode"), f"{item.get('code')} {item.get('name')}", f"{item.get('currentLevel')} / {value_or_dash(item.get('currentIndex'))}", f"{item.get('targetLevel')} / {value_or_dash(item.get('targetIndex'))}", value_or_dash(item.get("gapIndex")), value_or_dash(item.get("priority")), f"{value_or_dash(item.get('targetAchievementRate'))}%", _dict(item.get("dimensionResults")).get("organization"), _dict(item.get("dimensionResults")).get("process"), _dict(item.get("dimensionResults")).get("tool"), _dict(item.get("dimensionResults")).get("data"), f"{item.get('evidenceCoverage')}%"] for item in _list(data)]))
        elif section_id == "overall_rankings":
            for key, label in (("leading", "成熟度领先 Top 10"), ("improvement", "改进优先 Top 10")):
                rows = _list(_dict(data).get(key))
                lines.extend([f"### {label}", "", md_table(["排名", "能力", "当前", "目标", "达成率 / 优先级分数"], [[index, f"{item.get('code') or item.get('capabilityCode')} {item.get('name') or item.get('capabilityName')}", f"{item.get('currentLevel')} / {value_or_dash(item.get('currentIndex'))}", f"{item.get('targetLevel')} / {value_or_dash(item.get('targetIndex'))}", item.get("targetAchievementRate") if key == "leading" else f"{item.get('priority')} / {item.get('priorityScore')}"] for index, item in enumerate(rows, 1)]), ""])
        elif section_id == "dimension_rankings":
            for ranking in _list(data):
                lines.extend([f"### {ranking.get('label')}", "", md_table(["类型", "排名", "能力", "当前维度得分", "优先级"], [[label, index, f"{item.get('code')} {item.get('name')}", _dict(item.get("dimensionResults")).get(ranking.get("dimension")), _dict(priority_by_capability.get(_text(item.get("id")))).get("priority") or "-"] for key, label in (("leading", "领先"), ("improvement", "改进")) for index, item in enumerate(_list(ranking.get(key)), 1)]), ""])
        elif section_id == "improvement_roadmap":
            lines.append(md_table(["排名", "优先级", "L2 能力", "当前 / 目标", "差距", "改进行动", "负责人", "资源投入", "依赖事项", "状态"], [[item.get("rank"), item.get("priority"), f"{item.get('capabilityCode')} {item.get('capabilityName')}", f"{item.get('currentLevel')} → {item.get('targetLevel')}", item.get("gapIndex"), item.get("action") or "待明确", item.get("owner") or "待指定", item.get("resources") or "待评估", item.get("dependencies") or "无", item.get("status")] for item in _list(data)]))
        elif section_id == "score_appendix":
            lines.append(md_table(["关注点 / 评估点", "类型 / 作用域", "组织", "流程", "工具", "数据", "当前", "目标", "达成率", "评估说明", "不适用说明", "证据等级", "证据摘要", "状态"], [[f"{item.get('focusCode')} {item.get('focusName')} / {item.get('serviceCode')} {item.get('serviceName')}", f"{item.get('itemType')} / {item.get('scopeCode')}", _dict(item.get("dimensionResults")).get("organization"), _dict(item.get("dimensionResults")).get("process"), _dict(item.get("dimensionResults")).get("tool"), _dict(item.get("dimensionResults")).get("data"), f"{item.get('currentLevel')} / {value_or_dash(item.get('currentIndex'))}", f"{item.get('targetLevel')} / {value_or_dash(item.get('targetIndex'))}", f"{value_or_dash(item.get('targetAchievementRate'))}%", item.get("targetReason") or item.get("note"), item.get("naReason"), item.get("evidenceLevel"), item.get("evidenceSummary"), item.get("status")] for item in _list(data)]))
        elif section_id == "traceability":
            trace = _dict(data)
            trace_summary = _dict(trace.get("summary"))
            lines.extend(
                [
                    f"- 评估对象：{_text(project.get('assessmentObjectType')) or '企业组织'}",
                    f"- 评估模板：{_text(template.get('name'))} {_text(template.get('version'))}",
                    f"- 不适用项：{trace_summary.get('notApplicableCount')}；无证据项作为信息口径保留，不阻塞评估完成。",
                    "- 报告中的图表、排名与明细均来自本次评估结果。",
                ]
            )
        return "\n".join(lines).rstrip()

    report_notice = "正式评估报告；全部适用评估点已完成，评估结果已锁定。" if is_formal else f"草稿报告预览；仍有 {summary.get('notScoredCount')} 条适用项待完成，当前结果不是正式结论。"
    markdown = "\n\n".join(
        [
            f"# {project_name}｜评估报告",
            f"> {report_notice}",
            f"客户或组织：{organization}  ",
            *(render_markdown_section(section) for section in report_model["sections"]),
        ]
    )

    def h(value: Any) -> str:
        return html_lib.escape(_text(value))

    def html_table(headers: list[str], rows: list[list[Any]], class_name: str = "") -> str:
        head = "".join(f"<th>{h(item)}</th>" for item in headers)
        body = "".join("<tr>" + "".join(f"<td>{h(value_or_dash(value))}</td>" for value in row) + "</tr>" for row in rows)
        if not body:
            body = f"<tr><td colspan='{len(headers)}'>暂无数据</td></tr>"
        return f"<div class='table-wrap'><table class='{h(class_name)}'><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"

    def radar_svg(chart: dict[str, Any], *, compact: bool = False) -> str:
        axes = [item for item in _list(chart.get("axes")) if isinstance(item, dict)]
        if len(axes) < 3:
            return "<div class='empty'>有效轴不足，无法形成雷达图。</div>"
        width, height = (560, 440) if compact else (920, 560)
        center_x, center_y = width / 2, height / 2 + (8 if compact else 18)
        radius = 150 if compact else 205
        angles = [-math.pi / 2 + 2 * math.pi * index / len(axes) for index in range(len(axes))]

        def point(angle: float, value: float, extra: float = 0.0) -> tuple[float, float]:
            distance = radius * max(0.0, min(5.0, value)) / 5.0 + extra
            return center_x + math.cos(angle) * distance, center_y + math.sin(angle) * distance

        grid = []
        for level in range(1, 6):
            points = " ".join(f"{point(angle, level)[0]:.1f},{point(angle, level)[1]:.1f}" for angle in angles)
            grid.append(f"<polygon points='{points}' fill='none' stroke='{'#aebdca' if level == 5 else '#dfe7ee'}' stroke-width='1'/>")
        sectors = []
        if not compact:
            group_colors = {"T": "#2f78c4", "G": "#7467b8", "M": "#3d8969", "—": "#738394"}
            offset = 0
            for group in _list(chart.get("groups")):
                count = int(_float(group.get("count")))
                if count <= 0:
                    continue
                start_angle = -math.pi / 2 + 2 * math.pi * (offset - 0.5) / len(axes)
                end_angle = -math.pi / 2 + 2 * math.pi * (offset + count - 0.5) / len(axes)
                start = point(start_angle, 5, 11)
                end = point(end_angle, 5, 11)
                large = 1 if count / len(axes) > 0.5 else 0
                color = group_colors.get(_text(group.get("code")), "#738394")
                sectors.append(f"<path d='M {center_x:.1f} {center_y:.1f} L {start[0]:.1f} {start[1]:.1f} A {radius + 11:.1f} {radius + 11:.1f} 0 {large} 1 {end[0]:.1f} {end[1]:.1f} Z' fill='{color}' opacity='.08'/>")
                offset += count
        spokes = []
        labels = []
        for index, axis in enumerate(axes):
            edge = point(angles[index], 5)
            label_point = point(angles[index], 5, 31 if compact else 37 + (index % 2) * 11)
            anchor = "start" if math.cos(angles[index]) > 0.22 else "end" if math.cos(angles[index]) < -0.22 else "middle"
            label = _text(axis.get("displayLabel")) or _text(axis.get("code")) or _text(axis.get("label"))
            if not compact and len(label) > 12:
                label = label[:12]
            spokes.append(f"<line x1='{center_x:.1f}' y1='{center_y:.1f}' x2='{edge[0]:.1f}' y2='{edge[1]:.1f}' stroke='#d7e0e8'/>")
            labels.append(f"<text x='{label_point[0]:.1f}' y='{label_point[1]:.1f}' text-anchor='{anchor}' dominant-baseline='middle'>{h(label)}</text>")

        def series(key: str, color: str, fill: str, dashed: bool = False) -> str:
            values = [None if axis.get(key) is None else _float(axis.get(key)) for axis in axes]
            dash = " stroke-dasharray='7 6'" if dashed else ""
            if all(value is not None for value in values):
                points = " ".join(f"{point(angles[index], value)[0]:.1f},{point(angles[index], value)[1]:.1f}" for index, value in enumerate(values))
                shape = f"<polygon points='{points}' fill='{fill}' stroke='{color}' stroke-width='2.4'{dash}/>"
            else:
                edges = []
                for index, value in enumerate(values):
                    next_index = (index + 1) % len(values)
                    next_value = values[next_index]
                    if value is None or next_value is None:
                        continue
                    start = point(angles[index], value)
                    end = point(angles[next_index], next_value)
                    edges.append(f"<line x1='{start[0]:.1f}' y1='{start[1]:.1f}' x2='{end[0]:.1f}' y2='{end[1]:.1f}' stroke='{color}' stroke-width='2.4'{dash}/>")
                shape = "".join(edges)
            circles = "" if dashed else "".join(f"<circle cx='{point(angles[index], value)[0]:.1f}' cy='{point(angles[index], value)[1]:.1f}' r='3' fill='white' stroke='{color}' stroke-width='2'/>" for index, value in enumerate(values) if value is not None)
            return f"<g class='radar-series radar-series-{h(key)}'>{shape}{circles}</g>"

        incomplete = any(axis.get("current") is None or axis.get("target") is None for axis in axes)
        incomplete_note = "<text class='chart-note' x='20' y='28'>缺失轴保留为空，其余已评估数据正常绘制。</text>" if incomplete else ""
        return f"<svg class='radar-svg' viewBox='0 0 {width} {height}' role='img' aria-label='{h(chart.get('title'))}'>{''.join(sectors)}{''.join(grid)}{''.join(spokes)}{series('target', '#9a6d2f', 'none', True)}{series('current', '#1676c5', 'rgba(22,118,197,.12)')}{''.join(labels)}{incomplete_note}</svg>"

    def html_section(section: dict[str, Any]) -> str:
        section_id = _text(section.get("id"))
        data = section.get("data")
        body = ""

        def section_heading(number: str, title: Any, eyebrow: str = "") -> str:
            eyebrow_html = f"<p>{h(eyebrow)}</p>" if eyebrow else ""
            return f"<header class='section-heading'><span>{h(number)}</span><div><h2>{h(title)}</h2>{eyebrow_html}</div></header>"

        if section_id == "overall":
            section_summary = _dict(_dict(data).get("summary"))
            total = int(_float(section_summary.get("applicableItemCount"))) + int(_float(section_summary.get("notApplicableCount")))
            current_index = section_summary.get("currentIndex")
            target_index = section_summary.get("targetIndex")
            maturity_gap = _round(_float(target_index) - _float(current_index)) if current_index is not None and target_index is not None else None
            evidence_total = int(_float(evaluation.get("evidenceTotalCount")))
            evidence_filled = int(_float(evaluation.get("evidenceFilledCount")))
            evidence_missing = int(_float(evaluation.get("evidenceMissingCount")))
            current_level = _text(section_summary.get("currentLevel"))
            target_level = _text(section_summary.get("targetLevel"))
            current_level_name = _text(section_summary.get("currentLevelName")) or LEVEL_NAME.get(current_level, "未评分")
            target_level_name = LEVEL_NAME.get(target_level, "待确认")
            completed_count = int(_float(section_summary.get("scoredItemCount")))
            applicable_count = int(_float(section_summary.get("applicableItemCount")))
            completion_rate = _round(_float(section_summary.get("completionRate")), 1) or 0
            l2_total = int(_float(evaluation.get("l2TotalCount")))
            l2_below_target = int(_float(evaluation.get("l2BelowTargetCount")))
            current_position = max(0.0, min(100.0, (_float(current_index) - 1.0) / 4.0 * 100.0)) if current_index is not None else 0.0
            target_position = max(0.0, min(100.0, (_float(target_index) - 1.0) / 4.0 * 100.0)) if target_index is not None else 0.0
            progress_split = max(0.0, min(100.0, current_position / target_position * 100.0)) if target_position > 0 else 0.0
            target_readiness = _round(_float(current_index) / _float(target_index) * 100.0, 1) if current_index is not None and _float(target_index) > 0 else None
            target_readiness_width = max(0.0, min(100.0, _float(target_readiness)))
            if current_index is not None and target_index is not None and _float(current_index) >= _float(target_index):
                executive_headline = f"核心判断：整体达到 {current_level}，当前结果已达到既定目标"
                executive_statement = f"本次评估已完成 {completed_count} 个适用评估点，整体成熟度处于 {current_level} {current_level_name} 阶段，已达到 {target_level} {target_level_name} 目标。"
            elif current_index is not None and target_index is not None:
                executive_headline = f"核心判断：整体达到 {current_level}，{target_level_name}仍是下一阶段主线"
                executive_statement = f"本次评估已完成 {completed_count} 个适用评估点，整体成熟度处于 {current_level} {current_level_name} 阶段，向 {target_level} {target_level_name} 目标仍存在 {value_or_dash(maturity_gap)} 的指数差距。"
            else:
                executive_headline = "核心判断：当前成熟度尚未形成完整结论"
                executive_statement = f"本次评估已完成 {completed_count} / {applicable_count} 个适用评估点，待完整评分与目标确认后形成正式管理层结论。"
            decision_items = sorted(
                (item for item in improvement_roadmap if item.get("gapIndex") is not None),
                key=lambda item: _float(item.get("gapIndex")),
                reverse=True,
            )[:3]
            decision_rows = "".join(
                f"<li><span class='decision-rank'>{index:02d}</span><div><strong>{h(item.get('capabilityCode'))} {h(item.get('capabilityName'))}</strong><p><span>当前 {h(item.get('currentLevel'))}</span><i>→</i><span>目标 {h(item.get('targetLevel'))}</span><b>差距 {h(value_or_dash(item.get('gapIndex')))}</b></p></div></li>"
                for index, item in enumerate(decision_items, 1)
            ) or "<li><div><strong>暂无需重点研究的成熟度差距</strong></div></li>"
            body = f"""
<div class='executive-brief'>
  <header class='executive-brief__lead'>
    <div class='executive-brief__title'>
      <p class='executive-eyebrow'><span>EXECUTIVE BRIEF</span> 管理层摘要</p>
      <h2>{h(executive_headline)}</h2>
      <p>{h(executive_statement)}</p>
    </div>
    <div class='executive-completion' aria-label='评估完成度 {h(value_or_dash(completion_rate))}%'>
      <span>评估完成度</span><strong>{h(value_or_dash(completion_rate))}%</strong><small>{completed_count} / {applicable_count}</small>
    </div>
  </header>
  <div class='executive-brief__grid'>
    <section class='maturity-path' aria-labelledby='maturity-path-title'>
      <header class='maturity-path__header'><div><span>MATURITY TRAJECTORY</span><h3 id='maturity-path-title'>成熟度跃迁路径</h3></div><p>当前结果与人工确认目标</p></header>
      <div class='maturity-scoreline'>
        <div class='maturity-score is-current'><span>当前成熟度</span><strong>{h(value_or_dash(current_index))}</strong><small><b>{h(current_level)}</b>{h(current_level_name)}</small></div>
        <div class='maturity-gap-card'><span>成熟度差距</span><strong>{h(value_or_dash(maturity_gap))}</strong><small>{'已达到目标' if maturity_gap is not None and maturity_gap <= 0 else '仍需提升'}</small></div>
        <div class='maturity-score is-target'><span>目标成熟度</span><strong>{h(value_or_dash(target_index))}</strong><small><b>{h(target_level)}</b>{h(target_level_name)}</small></div>
      </div>
      <div class='maturity-track' role='img' aria-label='五级成熟度路径，当前 {h(value_or_dash(current_index))}，目标 {h(value_or_dash(target_index))}'>
        <div class='maturity-track__labels' aria-hidden='true'><span>L1</span><span>L2</span><span>L3</span><span>L4</span><span>L5</span></div>
        <div class='maturity-track__rail' aria-hidden='true'>
          <i class='maturity-track__progress' style='width:{target_position:.1f}%;background:linear-gradient(90deg,var(--blue) 0 {progress_split:.1f}%,var(--copper) {progress_split:.1f}% 100%)'></i>
          <i class='maturity-track__marker is-current' style='left:{current_position:.1f}%'><b>当前</b></i>
          <i class='maturity-track__marker is-target' style='left:{target_position:.1f}%'><b>目标</b></i>
        </div>
      </div>
      <div class='target-readiness'><div><span>目标达成度</span><small>当前成熟度 ÷ 目标成熟度</small></div><div class='target-readiness__bar' aria-hidden='true'><i style='width:{target_readiness_width:.1f}%'></i></div><strong>{h(value_or_dash(target_readiness))}%</strong></div>
    </section>
    <aside class='decision-focus'><header><div><span>TOP 3</span><h3>本次需重点研究</h3></div><p>按成熟度差距排序</p></header><ol>{decision_rows}</ol></aside>
  </div>
  <div class='executive-facts' aria-label='评估关键事实'>
    <article class='is-attention'><span>低于目标</span><strong>{l2_below_target}<small>/ {l2_total}</small></strong><p>L2 能力需提升</p></article>
    <article><span>纳入范围</span><strong>{applicable_count}<small>/ {total}</small></strong><p>适用评估点</p></article>
    <article><span>完成进度</span><strong>{completed_count}<small>/ {applicable_count}</small></strong><p>完成率 {h(value_or_dash(completion_rate))}%</p></article>
    <article><span>证据覆盖</span><strong>{evidence_filled}<small>/ {evidence_total}</small></strong><p>E0 无证据 {evidence_missing} 项</p></article>
  </div>
</div>"""
        elif section_id == "narratives":
            body = "<div class='narrative-grid'>" + "".join(f"<article><header><span>{index + 3:02d}</span><h2>{h(field.get('label'))}</h2></header><div class='editable' contenteditable='true' data-report-field='{h(field.get('key'))}'>{h(field.get('value')).replace(chr(10), '<br />')}</div></article>" for index, field in enumerate(_list(data))) + "</div>"
        elif section_id == "radars":
            capability_chart = _dict(_dict(data).get("capabilityRadar"))
            dimension_chart = _dict(_dict(data).get("dimensionRadar"))
            legend = "<div class='legend'><span class='current'><i></i>当前成熟度</span><span class='target'><i></i>目标成熟度</span>" + "".join(f"<span class='group is-{h(group.get('code')).lower()}'><i></i>{h(group.get('code'))} {h(group.get('name'))} · {group.get('count')} L2</span>" for group in _list(capability_chart.get("groups"))) + "</div>"
            dimension_rows = [[axis.get("label"), axis.get("current"), axis.get("target"), _round(_float(axis.get("target")) - _float(axis.get("current"))) if axis.get("current") is not None and axis.get("target") is not None else None] for axis in _list(dimension_chart.get("axes"))]
            priority_rows = [[item.get("rank"), f"{item.get('capabilityCode')} {item.get('capabilityName')}", item.get("currentIndex"), item.get("targetIndex"), item.get("gapIndex")] for item in improvement_roadmap[:5]]
            l2_summary_rows = [["L2 评估点总数", evaluation.get("l2TotalCount")], ["L2 已评分", f"{evaluation.get('l2ScoredCount')} / {evaluation.get('l2TotalCount')}"], ["达标（当前 ≥ 目标）", evaluation.get("l2ReachedTargetCount")], ["低于目标", evaluation.get("l2BelowTargetCount")], ["未评估", int(_float(evaluation.get("l2TotalCount"))) - int(_float(evaluation.get("l2ScoredCount")))]]
            category_rows = [[f"{item.get('code')} {item.get('name')}", item.get("currentIndex"), item.get("targetIndex"), item.get("gapIndex")] for item in category_results]
            body = f"""<header class='scorecard-heading'><h2>L2 能力成熟度雷达 <small>32 维</small></h2><p>关键比较与缺口按本次正式评估结果呈现</p></header><div class='scorecard-main'><figure class='capability-radar'><figcaption>{legend}</figcaption>{radar_svg(capability_chart)}<p class='figure-note'>未评分能力保留为空，不按 0 计入；其余已评估轴继续绘制。</p></figure><aside class='priority-panel'><h3>优先级缺口 <small>按差距降序</small></h3>{html_table(['排名', '能力编号与名称', '当前', '目标', '差距'], priority_rows, 'priority-table')}<h3>L2 评分概况</h3>{html_table(['项目', '数值'], l2_summary_rows, 'l2-summary-table')}</aside></div><div class='scorecard-lower'><section><h3>维度表现与证据覆盖</h3>{html_table(['能力类别', '当前', '目标', '差距'], category_rows, 'category-profile-table')}</section><figure class='dimension-radar-card'><figcaption><h3>{h(dimension_chart.get('title'))}</h3><div class='legend'><span class='current'><i></i>当前</span><span class='target'><i></i>目标</span></div></figcaption>{radar_svg(dimension_chart, compact=True)}</figure><section class='evidence-profile'><h3>证据覆盖与完成度</h3><dl><div><dt>适用评估点</dt><dd>{summary.get('applicableItemCount')} / {int(_float(summary.get('applicableItemCount'))) + int(_float(summary.get('notApplicableCount')))}</dd></div><div><dt>完成度</dt><dd>{summary.get('scoredItemCount')} / {summary.get('applicableItemCount')}</dd></div><div><dt>证据 E1+</dt><dd>{evaluation.get('evidenceFilledCount')} / {evaluation.get('evidenceTotalCount')}</dd></div><div><dt>证据 E0</dt><dd>{evaluation.get('evidenceMissingCount')}</dd></div></dl></section></div>"""
        elif section_id == "hierarchy_statistics":
            body = html_table(["类别", "当前 / 目标", "L1", "L2", "低于目标", "高 / 中 / 低"], [[group.get("code"), f"{value_or_dash(group.get('currentIndex'))} / {value_or_dash(group.get('targetIndex'))}", sum(1 for item in _list(_dict(data).get("l1")) if _text(item.get("groupCode")) == _text(group.get("code"))), group.get("l2Count"), group.get("belowTargetCount"), f"{_dict(group.get('priorityCounts')).get('高', 0)} / {_dict(group.get('priorityCounts')).get('中', 0)} / {_dict(group.get('priorityCounts')).get('低', 0)}"] for group in _list(_dict(data).get("groups"))])
            body = section_heading("A1", "分类与 L1 分层统计", "T / G / M 分类、能力域与优先级分布") + body
            body += html_table(["类别", "L1 能力域", "当前 / 目标", "L2 数量", "高 / 中 / 低"], [[item.get("groupCode"), f"{item.get('code')} {item.get('name')}", f"{value_or_dash(item.get('currentIndex'))} / {value_or_dash(item.get('targetIndex'))}", item.get("l2Count"), f"{_dict(item.get('priorityCounts')).get('高', 0)} / {_dict(item.get('priorityCounts')).get('中', 0)} / {_dict(item.get('priorityCounts')).get('低', 0)}"] for item in _list(_dict(data).get("l1"))])
        elif section_id == "evaluation":
            evaluation_data = _dict(_dict(data).get("evaluation"))
            strongest_row = _dict(evaluation_data.get("strongestDimension"))
            weakest_row = _dict(evaluation_data.get("weakestDimension"))
            body = section_heading("A2", "结果评价与分布", "成熟度、证据与服务评估点统计")
            body += f"<div class='evaluation'><p><strong>{evaluation_data.get('l2ScoredCount')} / {evaluation_data.get('l2TotalCount')}</strong> 项 L2 已评分，<strong>{evaluation_data.get('l2ReachedTargetCount')}</strong> 项达到目标，<strong>{evaluation_data.get('l2BelowTargetCount')}</strong> 项低于目标。</p><p>四维最高为 <strong>{h(strongest_row.get('label'))} {h(strongest_row.get('value'))}</strong>，最低为 <strong>{h(weakest_row.get('label'))} {h(weakest_row.get('value'))}</strong>，极差 {h(value_or_dash(evaluation_data.get('dimensionSpread')))}。</p><p>证据 E1 及以上 <strong>{evaluation_data.get('evidenceFilledCount')} / {evaluation_data.get('evidenceTotalCount')}</strong>，E0 无证据 {evaluation_data.get('evidenceMissingCount')} 项。</p></div>"
            body += html_table(["评价类型", "L1 能力域", "当前", "目标", "差距"], [[label, f"{item.get('code')} {item.get('name')}", item.get("currentIndex"), item.get("targetIndex"), item.get("gapIndex")] for key, label in (("leadingL1", "优势能力域"), ("improvementL1", "重点加强")) for item in _list(evaluation_data.get(key))])
            body += "<div class='distribution-grid'>" + "".join(f"<article><h3>{label}</h3>{html_table(['等级', '名称', '数量'], [[item.get('level'), item.get('name'), item.get('count')] for item in _list(_dict(data).get(key))])}</article>" for key, label in (("maturityDistribution", "成熟度分布"), ("evidenceDistribution", "证据分布"), ("serviceDistribution", "服务评估点分布"))) + "</div>"
        elif section_id == "capability_results":
            body = section_heading("A4", "完整 L2 能力结果", "包含能力成熟度、四维评分、目标差距与证据覆盖")
            body += html_table(["类别", "L2 能力", "当前", "目标", "差距", "优先级", "达成率", "组织", "流程", "工具", "数据", "证据"], [[item.get("groupCode"), f"{item.get('code')} {item.get('name')}", f"{item.get('currentLevel')} / {value_or_dash(item.get('currentIndex'))}", f"{item.get('targetLevel')} / {value_or_dash(item.get('targetIndex'))}", item.get("gapIndex"), item.get("priority"), f"{value_or_dash(item.get('targetAchievementRate'))}%", _dict(item.get("dimensionResults")).get("organization"), _dict(item.get("dimensionResults")).get("process"), _dict(item.get("dimensionResults")).get("tool"), _dict(item.get("dimensionResults")).get("data"), f"{item.get('evidenceCoverage')}%"] for item in _list(data)], "dense")
        elif section_id == "overall_rankings":
            body = section_heading("A5", "总体 L2 能力排名", "成熟度领先与改进优先 Top 10") + "<div class='ranking-grid'>"
            for key, label in (("leading", "成熟度领先 Top 10"), ("improvement", "改进优先 Top 10")):
                body += f"<article><h3>{label}</h3>" + html_table(["排名", "能力", "当前", "目标", "达成率 / 优先级"], [[index, f"{item.get('code') or item.get('capabilityCode')} {item.get('name') or item.get('capabilityName')}", f"{item.get('currentLevel')} / {value_or_dash(item.get('currentIndex'))}", f"{item.get('targetLevel')} / {value_or_dash(item.get('targetIndex'))}", f"{value_or_dash(item.get('targetAchievementRate'))}%" if key == "leading" else f"{item.get('priority')} / {item.get('priorityScore')}"] for index, item in enumerate(_list(_dict(data).get(key)), 1)]) + "</article>"
            body += "</div>"
        elif section_id == "dimension_rankings":
            body = section_heading("A3", "四维能力统计", "组织与角色、制度与流程、平台与工具、数据与信息四维排名") + "<div class='dimension-rankings'>"
            for ranking in _list(data):
                rows = [[label, index, f"{item.get('code')} {item.get('name')}", _dict(item.get("dimensionResults")).get(ranking.get("dimension")), _dict(priority_by_capability.get(_text(item.get("id")))).get("priority")] for key, label in (("leading", "领先"), ("improvement", "改进")) for index, item in enumerate(_list(ranking.get(key)), 1)]
                body += f"<article><h3>{h(ranking.get('label'))}</h3>{html_table(['类型', '排名', '能力', '当前维度得分', '优先级'], rows)}</article>"
            body += "</div>"
        elif section_id == "improvement_roadmap":
            body = section_heading("07", "风险与改进登记册", "优先能力差距、建议行动、责任安排与推进状态")
            body += html_table(["排名", "优先级", "L2 能力", "当前 / 目标", "差距", "改进行动", "负责人", "资源", "依赖", "状态"], [[item.get("rank"), f"{item.get('priority')} / {item.get('priorityScore')}", f"{item.get('capabilityCode')} {item.get('capabilityName')}", f"{item.get('currentLevel')} / {item.get('targetLevel')}", item.get("gapIndex"), item.get("action") or "待明确", item.get("owner") or "待指定", item.get("resources") or "待评估", item.get("dependencies") or "无", item.get("status")] for item in _list(data)], "dense")
        elif section_id == "score_appendix":
            body = section_heading("A6", "完整评分明细附录", "关注点、评估点、四维评分、证据与状态")
            body += html_table(["关注点 / 评估点", "类型 / 作用域", "组织", "流程", "工具", "数据", "当前", "目标", "达成率", "评估说明", "不适用说明", "证据等级", "证据摘要", "状态"], [[f"{item.get('focusCode')} {item.get('focusName')} / {item.get('serviceCode')} {item.get('serviceName')}", f"{item.get('itemType')} / {item.get('scopeCode')}", _dict(item.get("dimensionResults")).get("organization"), _dict(item.get("dimensionResults")).get("process"), _dict(item.get("dimensionResults")).get("tool"), _dict(item.get("dimensionResults")).get("data"), f"{item.get('currentLevel')} / {value_or_dash(item.get('currentIndex'))}", f"{item.get('targetLevel')} / {value_or_dash(item.get('targetIndex'))}", f"{value_or_dash(item.get('targetAchievementRate'))}%", item.get("targetReason") or item.get("note"), item.get("naReason"), item.get("evidenceLevel"), item.get("evidenceSummary"), item.get("status")] for item in _list(data)], "dense appendix")
        elif section_id == "traceability":
            trace_summary = _dict(_dict(data).get("summary"))
            body = section_heading("A7", "评估口径说明", "仅保留对管理决策有意义的范围与口径")
            body += f"<div class='method-note'><p><strong>评估对象：</strong>{h(project.get('assessmentObjectType') or '企业组织')}</p><p><strong>评估模板：</strong>{h(template.get('name'))} {h(template.get('version'))}</p><p><strong>适用性：</strong>不适用项 {trace_summary.get('notApplicableCount')}；无证据项作为信息口径保留，不阻塞评估完成。</p><p><strong>数据说明：</strong>报告中的图表、排名与明细均来自本次评估结果。</p></div>"
        return f"<section class='report-section section-{h(section_id)}' data-report-section='{h(section_id)}'>{body}</section>"

    sections_by_id = {_text(section.get("id")): section for section in report_model["sections"]}

    def render_sections(*section_ids: str) -> str:
        return "".join(html_section(sections_by_id[section_id]) for section_id in section_ids if section_id in sections_by_id)

    report_meta = f"{h(organization)} · {h(template.get('name'))} · 评估日期 {h(generated_at[:10])}"
    report_footer = f"SAPD 标准能力成熟度模板 · {h(organization)} · {h(generated_at[:10])}"
    executive_report_css = r"""
/* Executive report refinement: presentation hierarchy, edit states, responsive and print fidelity. */
html{scroll-behavior:smooth}body{background:#e6ebf0;color:#183148;-webkit-font-smoothing:antialiased;text-rendering:optimizeLegibility}::selection{background:#cfe2f2;color:#082c4f}
.report-page{border-top:4px solid var(--navy);border-radius:2px;box-shadow:0 18px 48px rgba(8,44,79,.11)}
.brief-masthead{position:relative;isolation:isolate;padding:29px 38px 25px;background:var(--navy);border-bottom:4px solid var(--copper);color:#fff;overflow:hidden}
.brief-masthead::after{content:"";position:absolute;z-index:-1;right:-64px;bottom:-132px;width:360px;height:250px;border:54px solid rgba(255,255,255,.035);border-radius:50%;pointer-events:none}
.brief-masthead .kicker{color:#9fc6e3;letter-spacing:.17em}.brief-masthead h1{color:#fff;font-size:31px;letter-spacing:.035em}.brief-masthead .meta{color:#d4e2ec}.report-badge{border-color:rgba(255,255,255,.45);background:rgba(255,255,255,.07);color:#fff;letter-spacing:.06em}
.page-one .page-content{padding-top:20px}.transition-panel{min-height:142px;border-top:1px solid var(--line);background:#fff}.transition-level strong{font-size:62px}.transition-level b{letter-spacing:.01em}.transition-gap{background:var(--warm)}
.decision-agenda{padding:0 16px 2px;border-top:0;border-left:4px solid var(--navy);background:#f5f8fa}.decision-agenda h2{padding-top:11px}.decision-agenda li>span{color:var(--copper)}.decision-agenda li:last-child{border-bottom:0}
.executive-conclusion{margin:14px 0;padding:12px 15px;border:0;border-left:4px solid var(--copper);background:var(--warm);font-weight:650;color:#243e53}
.metric-strip{background:#fbfcfd}.metric-strip article{position:relative;padding-top:12px;padding-bottom:12px}.metric-strip strong{font-size:23px}.metric-strip article:nth-child(2) strong,.metric-strip article:nth-child(4) strong{color:var(--sage)}
.scorecard-heading{margin-top:26px}.scorecard-heading h2{letter-spacing:.015em}.scorecard-main{gap:26px}.scorecard-main .capability-radar{padding:8px 26px 0 4px}.priority-panel{padding:13px 15px;background:#f7f9fa;border-top:3px solid var(--navy)}.priority-panel .table-wrap{margin-bottom:16px}.priority-table tbody tr:first-child{background:#fff3df;font-weight:750}.priority-table tbody tr:first-child td:first-child{color:var(--danger)}
.scorecard-lower{background:#fbfcfd;padding:15px 14px 0}.scorecard-lower h3{letter-spacing:.02em}.evidence-profile dd{font-size:13px}
.page-intro{position:relative;padding:4px 0 13px}.page-intro::before{content:"";position:absolute;left:0;bottom:-2px;width:74px;height:4px;background:var(--copper)}.page-intro h2{letter-spacing:.025em}.page-intro>p{max-width:420px;text-align:right}
.narrative-grid article{position:relative;padding:12px 14px 0;border:1px solid var(--line);border-top:3px solid var(--navy);background:#fff}.narrative-grid article header{margin-bottom:3px}.editable{min-height:100px;margin:0 -3px 10px;padding:10px 11px;border:1px solid transparent;border-radius:2px;transition:border-color .16s ease,background .16s ease,box-shadow .16s ease}.editable:hover{border-color:#cad8e3;background:#f8fafb}.editable:focus{outline:0;border-color:var(--blue);background:#fff;box-shadow:0 0 0 3px rgba(23,107,181,.12)}
.section-improvement_roadmap .table-wrap{border-top:3px solid var(--navy)}.section-improvement_roadmap tbody tr:nth-child(-n+3){background:#fff7e9}.section-improvement_roadmap tbody tr:nth-child(-n+3) td:first-child{box-shadow:inset 3px 0 var(--copper);font-weight:800;color:var(--navy)}
.page-three .page-intro{margin-bottom:19px}.page-three .section-heading{margin-top:23px}.page-three .section-heading>span{color:var(--copper)}
.table-wrap{scrollbar-color:#aebdca #edf2f5;scrollbar-width:thin}thead th{position:sticky;top:0;z-index:1;border-top:1px solid var(--line);background:#eaf0f4;color:#28475f}tbody tr{transition:background-color .12s ease}tbody tr:hover{background:#f1f6f9}td.is-placeholder{color:#8f551d;background:#fff7e8;font-weight:750}td.is-empty-placeholder{color:#8a99a5;font-style:italic}
[contenteditable="true"]{caret-color:var(--blue)}[contenteditable="true"]:focus-visible{outline:0}

/* Management-first executive brief. */
.executive-brief{padding:0 0 17px;border-bottom:1px solid var(--line)}
.executive-brief__lead{display:grid;grid-template-columns:minmax(0,1fr) 172px;gap:28px;align-items:end;padding:0 0 16px;border-bottom:1px solid var(--line)}
.executive-eyebrow{display:flex;align-items:center;gap:9px;margin:0 0 7px;color:var(--muted);font-size:10px;font-weight:720;letter-spacing:.07em}.executive-eyebrow span{color:var(--copper);font-weight:820;letter-spacing:.16em}
.executive-brief__title h2{font-size:25px;letter-spacing:.015em}.executive-brief__title>p:last-child{max-width:790px;margin:7px 0 0;color:#425a6d;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:14px;font-weight:600;line-height:1.65}
.executive-completion{display:grid;grid-template-columns:1fr auto;gap:2px 12px;align-items:center;padding:13px 14px;background:var(--navy);border-bottom:3px solid var(--sage);color:#fff}.executive-completion span{white-space:nowrap;font-size:10px;color:#bcd0df}.executive-completion strong{grid-row:1 / 3;grid-column:2;font-size:28px;line-height:1}.executive-completion small{white-space:nowrap;font-size:9px;color:#8fb4ce}
.executive-brief__grid{display:grid;grid-template-columns:minmax(0,1.45fr) minmax(330px,.72fr);gap:18px;margin-top:18px}
.maturity-path{padding:17px 19px 15px;border:1px solid var(--line);border-top:4px solid var(--blue);background:#f8fafb}
.maturity-path__header,.decision-focus>header{display:flex;align-items:flex-end;justify-content:space-between;gap:14px}.maturity-path__header span,.decision-focus>header span{color:var(--blue);font-size:8.5px;font-weight:820;letter-spacing:.15em}.maturity-path__header h3,.decision-focus>header h3{margin:1px 0 0;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:16px}.maturity-path__header p,.decision-focus>header p{margin:0;color:var(--muted);font-size:9px}
.maturity-scoreline{display:grid;grid-template-columns:minmax(0,1fr) 112px minmax(0,1fr);gap:10px;margin-top:14px}.maturity-score{position:relative;display:grid;grid-template-columns:1fr auto;gap:2px 10px;align-items:end;padding:11px 13px;background:#fff;border:1px solid var(--line)}.maturity-score::before{content:"";position:absolute;inset:-1px auto -1px -1px;width:4px;background:var(--blue)}.maturity-score.is-target::before{background:var(--copper)}.maturity-score>span{font-size:9.5px;color:var(--muted)}.maturity-score>strong{grid-row:1 / 3;grid-column:2;font-size:34px;line-height:.95;color:var(--blue)}.maturity-score.is-target>strong{color:var(--copper)}.maturity-score small{font-size:10px;color:#52697b}.maturity-score small b{margin-right:4px;padding:2px 5px;background:#e8f2fa;color:var(--blue)}.maturity-score.is-target small b{background:#f7eadc;color:var(--copper)}
.maturity-gap-card{display:grid;place-content:center;text-align:center;background:#f3eee7;border:1px solid #e4d6c5}.maturity-gap-card span,.maturity-gap-card small{font-size:8.5px;color:#7a6b5d}.maturity-gap-card strong{font-size:25px;line-height:1.15;color:var(--navy)}
.maturity-track{margin:20px 2px 0}.maturity-track__labels{display:grid;grid-template-columns:repeat(5,1fr);color:#7e909e;font-size:8px}.maturity-track__labels span{text-align:center}.maturity-track__labels span:first-child{text-align:left}.maturity-track__labels span:last-child{text-align:right}.maturity-track__rail{position:relative;height:5px;margin-top:7px;background:#dce5eb}.maturity-track__rail::before{content:"";position:absolute;inset:-4px 0;background:repeating-linear-gradient(to right,transparent 0,transparent calc(25% - 1px),#a9bac7 calc(25% - 1px),#a9bac7 25%);opacity:.6}.maturity-track__progress{position:absolute;z-index:1;inset:0 auto 0 0;width:67%;background:linear-gradient(90deg,var(--blue) 0 68%,var(--copper) 68% 100%)}.maturity-track__marker{position:absolute;z-index:2;top:50%;width:13px;height:13px;background:#fff;border:3px solid var(--blue);border-radius:50%;transform:translate(-50%,-50%)}.maturity-track__marker b{position:absolute;top:13px;left:50%;transform:translateX(-50%);color:var(--blue);font-size:8px;white-space:nowrap}.maturity-track__marker.is-current{left:45.5%}.maturity-track__marker.is-target{left:67%;border-color:var(--copper)}.maturity-track__marker.is-target b{color:var(--copper)}
.target-readiness{display:grid;grid-template-columns:auto minmax(80px,1fr) auto;gap:12px;align-items:center;margin-top:27px;padding-top:10px;border-top:1px solid var(--line)}.target-readiness span,.target-readiness small{display:block}.target-readiness span{font-size:9.5px;font-weight:750;color:var(--navy)}.target-readiness small{font-size:8px;color:var(--muted)}.target-readiness__bar{height:7px;background:#e2e9ee}.target-readiness__bar i{display:block;width:76.6%;height:100%;background:var(--sage)}.target-readiness>strong{font-size:18px;color:var(--sage)}
.decision-focus{padding:17px 18px 13px;background:var(--navy);border-top:4px solid var(--copper);color:#fff}.decision-focus>header{padding-bottom:11px;border-bottom:1px solid rgba(255,255,255,.18)}.decision-focus>header span{color:#dda267}.decision-focus>header h3{color:#fff}.decision-focus>header p{color:#9db7ca}.decision-focus ol{list-style:none;margin:0;padding:0}.decision-focus li{display:grid;grid-template-columns:38px 1fr;gap:10px;padding:13px 0;border-bottom:1px solid rgba(255,255,255,.14)}.decision-focus li:last-child{border-bottom:0}.decision-rank{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:25px;line-height:1;color:#dda267}.decision-focus li strong{display:block;color:#fff;font-size:11px}.decision-focus li p{display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin:5px 0 0;color:#a9c0d1;font-size:8.5px}.decision-focus li p i{font-style:normal;color:#6f91aa}.decision-focus li p b{margin-left:auto;padding:3px 6px;background:rgba(218,150,82,.18);color:#f0bd88;font-weight:750}
.executive-facts{display:grid;grid-template-columns:repeat(4,1fr);margin-top:14px;border:1px solid var(--line);background:#fff}.executive-facts article{display:grid;grid-template-columns:1fr auto;gap:2px 10px;align-items:center;padding:10px 14px;border-right:1px solid var(--line)}.executive-facts article:last-child{border-right:0}.executive-facts article.is-attention{background:#fbf3e8;box-shadow:inset 4px 0 var(--copper)}.executive-facts span{color:var(--muted);font-size:9px}.executive-facts strong{grid-row:1 / 3;grid-column:2;color:var(--navy);font-size:24px;line-height:1}.executive-facts strong small{margin-left:3px;color:#7d8f9e;font-size:12px;font-weight:650}.executive-facts p{margin:0;color:#6c8090;font-size:8.5px}.executive-facts .is-attention strong{color:var(--copper)}

@media(max-width:1100px) and (min-width:821px){.executive-brief__grid{grid-template-columns:minmax(0,1.25fr) minmax(300px,.75fr)}.maturity-scoreline{grid-template-columns:minmax(0,1fr) 92px minmax(0,1fr)}.maturity-score>strong{font-size:30px}}
@media(max-width:820px){.executive-brief__lead{grid-template-columns:1fr auto;gap:16px;align-items:start}.executive-brief__title h2{font-size:22px}.executive-brief__title>p:last-child{font-size:13px}.executive-completion{min-width:116px}.executive-brief__grid{grid-template-columns:1fr}.executive-facts{grid-template-columns:repeat(2,1fr)}.executive-facts article:nth-child(2){border-right:0}.executive-facts article:nth-child(-n+2){border-bottom:1px solid var(--line)}}
@media(max-width:520px){.executive-brief__lead{grid-template-columns:1fr}.executive-completion{grid-template-columns:1fr auto;width:100%}.maturity-path{padding:15px 14px}.maturity-scoreline{grid-template-columns:1fr 74px 1fr;gap:7px}.maturity-score{grid-template-columns:1fr;padding:10px}.maturity-score>strong{grid-row:auto;grid-column:auto;font-size:29px}.maturity-score small{font-size:8.5px}.maturity-gap-card strong{font-size:21px}.target-readiness{grid-template-columns:1fr auto}.target-readiness__bar{grid-column:1 / 3;grid-row:2}.decision-focus{padding:15px}.executive-facts{grid-template-columns:1fr}.executive-facts article{border-right:0;border-bottom:1px solid var(--line)}.executive-facts article:last-child{border-bottom:0}}

@media(max-width:1100px) and (min-width:821px){.report-page{width:calc(100vw - 24px)}.page-content{padding-left:30px;padding-right:30px}.brief-masthead{padding-left:30px;padding-right:30px}.executive-grid{grid-template-columns:minmax(0,1fr) 292px;gap:24px}.transition-panel{grid-template-columns:126px 34px 154px minmax(108px,1fr)}.scorecard-main{grid-template-columns:minmax(0,1.45fr) minmax(315px,.8fr)}}
@media(max-width:820px){body{background:#fff}.report-page{box-shadow:none;border-top-width:3px}.brief-masthead::after{display:none}.brief-masthead h1{font-size:27px}.report-badge{align-self:flex-start}.decision-agenda{margin-top:10px}.executive-conclusion{font-size:13px}.metric-strip article{border-right:0;border-bottom:1px solid var(--line)}.metric-strip article:last-child{border-bottom:0}.scorecard-main .capability-radar{max-width:100%;overflow-x:auto;padding-right:0}.scorecard-main .capability-radar .radar-svg{min-width:720px;height:auto}.priority-panel{margin-top:12px}.scorecard-lower{padding:15px}.page-intro{align-items:flex-start;gap:12px}.page-intro>p{text-align:left}.narrative-grid article{min-height:0}.table-wrap{width:100%;max-width:100%;overflow-x:auto}table.appendix{min-width:1120px}}
@media(max-width:520px){body{font-size:12.5px}.brief-masthead,.page-content{padding-left:18px;padding-right:18px}.brief-masthead{padding-top:24px}.brief-masthead h1{font-size:23px}.transition-level strong{font-size:52px}.transition-level b{font-size:17px}.decision-agenda{padding-left:12px;padding-right:12px}.report-footer{margin-left:18px;margin-right:18px}.page-intro{display:block}.page-intro>p{margin-top:7px}}

@media print{*{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important}html,body{background:#fff}.report-page{border-top:0;border-radius:0;box-shadow:none}.brief-masthead{background:#082c4f!important;border-bottom-color:#a06125!important}.brief-masthead::after{display:block}.page-one .page-content{padding-top:6mm}.priority-panel,.scorecard-lower,.decision-agenda{background-color:#f7f9fa!important}thead th{position:static}.editable,.editable:hover,.editable:focus{border-color:transparent;background:transparent;box-shadow:none}.section-improvement_roadmap tbody tr:nth-child(-n+3){background:#fff7e9!important}tbody tr:hover{background:inherit}td.is-placeholder{background:#fff7e8!important}}

/* Print compaction keeps the two leadership pages intact on A4 landscape. */
@media print{
  .brief-masthead{padding:5mm 10mm 4mm}.brief-masthead h1{font-size:20pt}.brief-masthead .kicker{margin-bottom:1mm;font-size:6.5pt}.brief-masthead .meta{margin-top:1mm;font-size:7.5pt}.report-badge{padding:1.5mm 2.5mm;font-size:6.5pt}
  .page-one,.page-two{height:210mm;min-height:210mm;overflow:hidden}.page-one .page-content{padding:4mm 10mm 10mm}.executive-grid{grid-template-columns:minmax(0,1fr) 82mm;gap:5mm}.transition-panel{min-height:24mm;grid-template-columns:30mm 8mm 36mm minmax(25mm,1fr)}.transition-level{column-gap:1.5mm}.transition-level strong{font-size:32pt}.transition-level span{font-size:6.5pt}.transition-level b{font-size:10.5pt}.transition-arrow{font-size:18pt}.transition-gap span{font-size:6pt}.transition-gap strong{font-size:16pt}
  .decision-agenda{padding:0 3mm 1mm}.decision-agenda h2{padding:1.2mm 0;font-size:9.5pt}.decision-agenda h2 small{font-size:5.5pt}.decision-agenda li{grid-template-columns:7mm 1fr;gap:1.5mm;padding:1mm 0}.decision-agenda li>span{font-size:13pt}.decision-agenda strong{font-size:6.3pt}.decision-agenda p{margin-top:.4mm;font-size:5.6pt}.executive-conclusion{margin:2mm 0;padding:2mm 3mm;font-size:7.6pt}.metric-strip article{padding:1.5mm 4mm}.metric-strip span,.metric-strip small{font-size:5.8pt}.metric-strip strong{margin:.4mm 0;font-size:12pt}
  .scorecard-heading{margin-top:2.5mm;padding-bottom:1mm}.scorecard-heading h2{font-size:10.5pt}.scorecard-heading h2 small,.scorecard-heading p{font-size:5.6pt}.scorecard-main{grid-template-columns:minmax(0,1.55fr) minmax(76mm,.75fr);gap:3mm;padding-top:1mm}.scorecard-main .capability-radar{padding:0 3mm 0 0}.scorecard-main .capability-radar .radar-svg{height:48mm}.legend{gap:1.4mm;font-size:4.7pt}.legend i{width:3mm}.figure-note{font-size:4.8pt}.priority-panel{padding:1mm 1.5mm}.priority-panel h3{margin:0;padding-bottom:.5mm;font-size:6.3pt}.priority-panel .table-wrap{margin:.4mm 0 .8mm}.priority-table,.l2-summary-table{font-size:4.6pt}.priority-table th,.priority-table td,.l2-summary-table th,.l2-summary-table td{padding:.35mm .65mm;line-height:1.18}
  .scorecard-lower{gap:2mm;margin-top:1mm;padding:1mm 1.5mm 0}.scorecard-lower h3{margin-bottom:.5mm;font-size:6.1pt}.scorecard-lower>section+figure,.scorecard-lower>figure+section{padding-left:2mm}.category-profile-table{font-size:4.8pt}.category-profile-table th,.category-profile-table td{padding:.45mm .7mm;line-height:1.2}.dimension-radar-card .radar-svg{height:22mm}.evidence-profile dl>div{padding:.65mm 0}.evidence-profile dt,.evidence-profile dd{font-size:5.3pt}.report-footer{bottom:2mm;font-size:5.5pt}
  .page-two .page-content{padding:6mm 10mm 11mm}.page-two .page-intro{margin-bottom:4mm;padding-bottom:2.5mm}.page-two .page-intro span{font-size:6.5pt}.page-two .page-intro h2{font-size:16pt}.page-two .page-intro p{font-size:7pt}.page-two .section-narratives{margin-top:0}.page-two .narrative-grid{gap:4mm 5mm}.page-two .narrative-grid article{padding:2.5mm 3mm 0}.page-two .narrative-grid header{margin-bottom:1mm}.page-two .narrative-grid header span{font-size:13pt}.page-two .narrative-grid h2{font-size:11pt}.page-two .editable{min-height:25mm;margin:0;padding:2mm 1mm;font-size:8pt;line-height:1.45}.page-two .section-improvement_roadmap{margin-top:4mm}.page-two .section-heading{grid-template-columns:9mm 1fr;gap:2mm;margin-bottom:2mm;padding-bottom:1.5mm}.page-two .section-heading>span{font-size:14pt}.page-two .section-heading h2{font-size:12pt}.page-two .section-heading p{font-size:6.3pt}.page-two .section-improvement_roadmap .table-wrap{margin:0}.page-two .section-improvement_roadmap table{font-size:6.1pt}.page-two .section-improvement_roadmap th,.page-two .section-improvement_roadmap td{padding:.85mm 1mm;line-height:1.3}
}

@media print{
  .executive-brief{padding-bottom:1mm}.executive-brief__lead{grid-template-columns:minmax(0,1fr) 34mm;gap:4mm;padding-bottom:1.5mm}.executive-eyebrow{gap:1.5mm;margin-bottom:1mm;font-size:4.6pt}.executive-brief__title h2{font-size:10.5pt}.executive-brief__title>p:last-child{max-width:none;margin-top:1mm;font-size:6.2pt;line-height:1.4}.executive-completion{gap:.5mm 1.5mm;padding:1.5mm 2mm;border-bottom-width:1mm}.executive-completion span{font-size:4.7pt}.executive-completion strong{font-size:13pt}.executive-completion small{font-size:4.5pt}
  .executive-brief__grid{grid-template-columns:minmax(0,1.45fr) 78mm;gap:3mm;margin-top:2mm}.maturity-path{padding:1.7mm 2mm 1.5mm;border-top-width:1mm}.maturity-path__header,.decision-focus>header{gap:2mm}.maturity-path__header span,.decision-focus>header span{font-size:4.2pt}.maturity-path__header h3,.decision-focus>header h3{font-size:7pt}.maturity-path__header p,.decision-focus>header p{font-size:4.3pt}.maturity-scoreline{grid-template-columns:minmax(0,1fr) 22mm minmax(0,1fr);gap:1mm;margin-top:1.5mm}.maturity-score{gap:.5mm 1mm;padding:1.2mm 1.5mm}.maturity-score::before{width:1mm}.maturity-score>span{font-size:4.5pt}.maturity-score>strong{font-size:15pt}.maturity-score small{font-size:4.8pt}.maturity-score small b{margin-right:.7mm;padding:.3mm .8mm}.maturity-gap-card span,.maturity-gap-card small{font-size:4.1pt}.maturity-gap-card strong{font-size:12pt}
  .maturity-track{margin:2.3mm .5mm 0}.maturity-track__labels{font-size:3.8pt}.maturity-track__rail{height:1mm;margin-top:1mm}.maturity-track__marker{width:2.8mm;height:2.8mm;border-width:.6mm}.maturity-track__marker b{top:2.8mm;font-size:3.6pt}.target-readiness{grid-template-columns:auto minmax(20mm,1fr) auto;gap:2mm;margin-top:4.2mm;padding-top:1.5mm}.target-readiness span{font-size:4.5pt}.target-readiness small{font-size:3.7pt}.target-readiness__bar{height:1.3mm}.target-readiness>strong{font-size:8pt}
  .decision-focus{padding:1.7mm 2mm 1.2mm;border-top-width:1mm}.decision-focus>header{padding-bottom:1.2mm}.decision-focus li{grid-template-columns:7mm 1fr;gap:1mm;padding:1.55mm 0}.decision-rank{font-size:12pt}.decision-focus li strong{font-size:5.4pt}.decision-focus li p{gap:1mm;margin-top:.6mm;font-size:4.1pt}.decision-focus li p b{padding:.4mm .8mm}
  .executive-facts{margin-top:1.5mm}.executive-facts article{gap:.5mm 1mm;padding:1mm 1.5mm}.executive-facts span{font-size:4.2pt}.executive-facts strong{font-size:10pt}.executive-facts strong small{font-size:5.5pt}.executive-facts p{font-size:4pt}.scorecard-main .capability-radar .radar-svg{height:43mm}
}
"""
    executive_report_script = r"""<script>
(() => {
  const placeholders = new Set(["待指定", "待评估", "待规划", "待模拟", "待确认", "待补充", "待完善"]);
  const emptyPlaceholders = new Set(["-", "—"]);
  document.querySelectorAll("td").forEach((cell) => {
    const value = cell.textContent.trim();
    if (placeholders.has(value)) {
      cell.classList.add("is-placeholder");
      cell.dataset.placeholder = "true";
    } else if (emptyPlaceholders.has(value)) {
      cell.classList.add("is-empty-placeholder");
      cell.dataset.placeholder = "empty";
    }
  });
  document.querySelectorAll('[contenteditable="true"]').forEach((field) => {
    field.setAttribute("role", "textbox");
    field.setAttribute("aria-multiline", "true");
    if (!field.hasAttribute("tabindex")) field.setAttribute("tabindex", "0");
  });
})();
</script>"""
    html_report = f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/><meta name="color-scheme" content="light"/><meta name="description" content="SAPD 企业网络安全能力成熟度正式评估报告"/><title>{h(project_name)}｜评估报告</title>
<style>
:root{{--navy:#082c4f;--navy-2:#123d66;--ink:#152b3f;--muted:#64788b;--line:#cbd6df;--line-strong:#93a8ba;--blue:#176bb5;--copper:#a06125;--sage:#3f7f68;--violet:#6e63ac;--paper:#fff;--wash:#f4f6f7;--warm:#f8f4ed;--danger:#9d3e34}}
*{{box-sizing:border-box}}html{{background:#e8edf1}}body{{margin:0;color:var(--ink);font:13px/1.56 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;font-variant-numeric:tabular-nums}}main{{padding:22px 0 48px}}.report-page{{width:min(1180px,calc(100vw - 32px));min-height:820px;margin:0 auto 22px;background:var(--paper);border:1px solid #d7dfe6;position:relative;overflow:hidden}}.report-page.is-appendix{{overflow:visible}}.page-content{{padding:24px 38px 48px}}
.report-masthead{{display:grid;grid-template-columns:1fr auto;gap:28px;padding:24px 38px;background:var(--navy);color:#fff}}.report-masthead .kicker{{margin:0 0 6px;color:#b8cee0;font-size:10px;font-weight:760;letter-spacing:.16em}}h1{{margin:0;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:28px;line-height:1.22;letter-spacing:.02em}}.report-masthead .meta{{margin:6px 0 0;color:#d8e5ee}}.report-meta{{display:grid;grid-template-columns:auto auto;gap:5px 14px;align-content:center;font-size:11px;color:#d8e5ee}}.report-meta dt{{color:#94b5ce}}.report-meta dd{{margin:0;text-align:right}}.report-status{{display:flex;justify-content:space-between;gap:18px;padding:7px 38px;border-bottom:1px solid var(--line);color:var(--muted);font-size:11px}}.report-status strong{{color:var(--sage)}}
.report-section{{margin-top:22px}}.report-section:first-child{{margin-top:0}}.section-heading{{display:grid;grid-template-columns:42px 1fr;gap:10px;align-items:start;margin-bottom:13px;border-bottom:2px solid var(--navy);padding-bottom:7px}}.section-heading>span{{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:22px;font-weight:800;color:var(--navy)}}h2{{margin:0;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:20px;line-height:1.3;color:var(--navy)}}.section-heading p{{margin:3px 0 0;color:var(--muted);font-size:11px}}h3{{margin:0 0 10px;font-size:13px;color:var(--navy)}}
.maturity-comparison{{display:grid;grid-template-columns:minmax(0,1fr) 150px minmax(0,1fr);align-items:stretch;border-bottom:1px solid var(--line);margin-bottom:12px}}.maturity-state{{padding:13px 22px;color:#fff}}.maturity-state.is-current{{background:var(--navy-2)}}.maturity-state.is-target{{background:var(--copper)}}.maturity-state span,.maturity-state b,.maturity-state p{{display:block}}.maturity-state span{{font-size:11px;letter-spacing:.08em}}.maturity-state strong{{display:inline-block;margin:3px 10px 0 0;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:43px;line-height:1}}.maturity-state b{{display:inline-block;font-size:14px}}.maturity-state p{{margin:6px 0 0;color:#e3edf4;font-size:10px}}.maturity-gap{{display:grid;place-content:center;text-align:center;background:var(--warm);border-left:1px solid var(--line);border-right:1px solid var(--line)}}.maturity-gap span,.maturity-gap small{{color:var(--muted);font-size:10px}}.maturity-gap strong{{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:28px;color:var(--ink)}}
.metric-strip{{display:grid;grid-template-columns:repeat(4,1fr);border-top:1px solid var(--line);border-bottom:1px solid var(--line)}}.metric-strip article{{padding:10px 16px;border-right:1px solid var(--line)}}.metric-strip article:last-child{{border-right:0}}.metric-strip span,.metric-strip small{{display:block;color:var(--muted);font-size:10px}}.metric-strip strong{{display:block;margin:2px 0;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:21px;color:var(--navy)}}
.radar-grid{{display:grid;grid-template-columns:minmax(0,1.62fr) minmax(310px,.88fr);gap:22px}}figure{{margin:0}}.capability-radar{{padding-right:22px;border-right:1px solid var(--line)}}figcaption{{display:flex;align-items:flex-start;justify-content:space-between;gap:12px}}.dimension-analysis figure{{border-bottom:1px solid var(--line);padding-bottom:5px;margin-bottom:7px}}.radar-svg{{display:block;width:100%;height:auto}}.capability-radar .radar-svg{{height:285px}}.dimension-analysis .radar-svg{{height:195px}}.radar-svg text{{font-size:9px;fill:#405a70;font-weight:650}}.radar-svg .chart-note{{font-size:10px;fill:var(--copper)}}.legend{{display:flex;gap:10px;flex-wrap:wrap;color:var(--muted);font-size:9px}}.legend span{{display:flex;align-items:center;gap:4px}}.legend i{{width:20px;height:2px;background:var(--blue)}}.legend .target i{{height:0;background:none;border-top:2px dashed var(--copper)}}.legend .group i{{width:7px;height:7px;border-radius:50%;background:#738394}}.legend .is-t i{{background:#2f78c4}}.legend .is-g i{{background:var(--violet)}}.legend .is-m i{{background:var(--sage)}}.figure-note{{margin:2px 0 0;color:var(--muted);font-size:9.5px}}
.narrative-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:24px 30px}}.narrative-grid article{{border-top:2px solid var(--navy);padding-top:10px}}.narrative-grid header{{display:flex;gap:10px;align-items:baseline;margin-bottom:8px}}.narrative-grid header span{{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:20px;font-weight:800;color:var(--navy)}}.editable{{min-height:112px;padding:8px 0;color:#334b5f;white-space:normal}}.editable:hover{{background:#f7f9fa}}.editable:focus{{outline:2px solid #176bb533;outline-offset:4px;background:#fff}}
.table-wrap{{overflow:auto;margin:10px 0 18px}}table{{width:100%;border-collapse:collapse;font-size:10.5px}}th,td{{text-align:left;vertical-align:top;padding:7px 8px;border-bottom:1px solid var(--line)}}th{{color:#405a70;background:#eef2f4;white-space:nowrap;font-weight:760}}tbody tr:nth-child(even){{background:#fafbfb}}table.dense{{font-size:9px}}table.dense th,table.dense td{{padding:5px;overflow-wrap:anywhere}}table.appendix{{min-width:1120px}}.category-table td:first-child{{font-weight:700}}.dimension-table{{font-size:9.5px}}
.distribution-grid,.ranking-grid,.dimension-rankings{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:18px}}.distribution-grid{{grid-template-columns:repeat(3,minmax(0,1fr))}}.evaluation{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;padding:14px 16px;background:var(--wash);border-left:3px solid var(--blue)}}.evaluation p{{margin:0}}.method-note{{display:grid;grid-template-columns:repeat(2,1fr);gap:8px 24px;color:var(--muted)}}.method-note p{{margin:0;padding:8px 0;border-bottom:1px solid var(--line)}}.page-intro{{display:flex;justify-content:space-between;align-items:end;border-bottom:2px solid var(--navy);padding-bottom:12px;margin-bottom:24px}}.page-intro span{{color:var(--copper);font-size:10px;font-weight:800;letter-spacing:.12em}}.page-intro h2{{font-size:24px}}.page-intro p{{margin:0;color:var(--muted);font-size:11px}}
.brief-masthead{{display:flex;justify-content:space-between;gap:28px;padding:26px 38px 18px;border-bottom:1px solid var(--line)}}.brief-masthead .kicker{{margin:0 0 4px;color:var(--blue);font-size:10px;font-weight:800;letter-spacing:.14em}}.brief-masthead h1{{color:var(--navy);font-size:30px}}.brief-masthead .meta{{margin:5px 0 0;color:var(--muted)}}.report-badge{{align-self:flex-start;padding:6px 10px;border:1px solid var(--line-strong);color:var(--navy);font-size:10px;font-weight:750}}.executive-grid{{display:grid;grid-template-columns:minmax(0,1fr) 320px;gap:34px}}.transition-panel{{display:grid;grid-template-columns:138px 42px 168px minmax(118px,1fr);align-items:center;min-height:138px;border-bottom:1px solid var(--line)}}.transition-level{{display:grid;grid-template-columns:auto 1fr;align-items:end;column-gap:9px}}.transition-level strong{{grid-row:1 / 3;font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:66px;line-height:.9}}.transition-level span{{font-size:11px;color:var(--muted)}}.transition-level b{{font-size:20px}}.transition-level.is-current strong,.transition-level.is-current b{{color:var(--blue)}}.transition-level.is-target strong,.transition-level.is-target b{{color:var(--copper)}}.transition-arrow{{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:32px;color:#9aa7b2;text-align:center}}.transition-gap{{align-self:stretch;display:grid;place-content:center;border-left:1px solid var(--line);text-align:center}}.transition-gap span{{font-size:10px;color:var(--muted)}}.transition-gap strong{{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:26px;color:var(--navy)}}.decision-agenda{{border-top:2px solid var(--navy)}}.decision-agenda h2{{padding:8px 0;border-bottom:1px solid var(--line);font-size:16px}}.decision-agenda h2 small{{margin-left:6px;color:var(--muted);font:10px/1.4 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC",sans-serif}}.decision-agenda ol{{list-style:none;margin:0;padding:0}}.decision-agenda li{{display:grid;grid-template-columns:34px 1fr;gap:8px;padding:8px 0;border-bottom:1px solid var(--line)}}.decision-agenda li>span{{font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:22px;color:var(--navy)}}.decision-agenda strong{{font-size:10.5px}}.decision-agenda p{{margin:2px 0 0;color:var(--muted);font-size:9.5px}}.executive-conclusion{{margin:14px 0;padding:11px 0;border-top:1px solid var(--line);border-bottom:1px solid var(--line);font-family:"Songti SC","STSong","Noto Serif SC",serif;font-size:14px;color:#2b4255}}
.scorecard-heading{{display:flex;align-items:end;justify-content:space-between;margin-top:24px;padding-bottom:8px;border-bottom:2px solid var(--navy)}}.scorecard-heading h2{{font-size:20px}}.scorecard-heading h2 small{{font:11px/1.4 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC",sans-serif;color:var(--muted)}}.scorecard-heading p{{margin:0;color:var(--muted);font-size:10px}}.scorecard-main{{display:grid;grid-template-columns:minmax(0,1.52fr) minmax(340px,.82fr);gap:24px;padding-top:12px}}.scorecard-main .capability-radar{{padding-right:24px;border-right:1px solid var(--line)}}.scorecard-main .capability-radar .radar-svg{{height:430px}}.priority-panel h3{{display:flex;justify-content:space-between;padding-bottom:7px;border-bottom:1px solid var(--line)}}.priority-panel h3 small{{color:var(--muted);font-size:9px;font-weight:500}}.priority-table{{font-size:9px}}.priority-table td:nth-child(2){{max-width:150px}}.l2-summary-table{{font-size:9.5px}}.scorecard-lower{{display:grid;grid-template-columns:1.15fr .85fr .85fr;gap:20px;margin-top:10px;padding-top:14px;border-top:2px solid var(--navy)}}.scorecard-lower>section+figure,.scorecard-lower>figure+section{{padding-left:20px;border-left:1px solid var(--line)}}.dimension-radar-card .radar-svg{{height:170px}}.dimension-radar-card figcaption{{align-items:center}}.evidence-profile dl{{margin:0}}.evidence-profile dl>div{{display:flex;justify-content:space-between;padding:8px 0;border-bottom:1px solid var(--line)}}.evidence-profile dt{{color:var(--muted)}}.evidence-profile dd{{margin:0;font-weight:760;color:var(--navy)}}
.report-footer{{position:absolute;left:38px;right:38px;bottom:16px;display:flex;justify-content:space-between;padding-top:8px;border-top:1px solid var(--line);color:var(--muted);font-size:9px}}.is-appendix .report-footer{{position:static;margin:28px 38px 16px}}code{{overflow-wrap:anywhere}}
@media(max-width:820px){{main{{padding-top:0}}.report-page{{width:100%;min-height:0;margin:0;border:0;border-bottom:10px solid #e8edf1}}.report-masthead{{grid-template-columns:1fr;padding:24px}}.report-meta{{display:none}}.report-status,.page-content{{padding-left:22px;padding-right:22px}}.brief-masthead{{padding:22px;flex-direction:column}}.executive-grid,.scorecard-main,.scorecard-lower,.transition-panel,.metric-strip,.narrative-grid,.distribution-grid,.ranking-grid,.dimension-rankings,.evaluation,.method-note{{grid-template-columns:1fr}}.transition-panel{{gap:8px}}.transition-arrow{{transform:rotate(90deg)}}.transition-gap{{padding:10px;border-left:0;border-top:1px solid var(--line)}}.scorecard-main .capability-radar,.scorecard-lower>section+figure,.scorecard-lower>figure+section{{border-left:0;border-right:0;padding-left:0;padding-right:0}}.capability-radar{{border-left:0;border-right:0;padding-left:0;padding-right:0}}.report-footer{{position:static;margin:24px 22px 16px}}}}
@page{{size:A4 landscape;margin:0}}@media print{{html{{background:#fff}}body{{font-size:10pt}}main{{padding:0}}.report-page{{width:297mm;min-height:210mm;margin:0;border:0;break-after:page;page-break-after:always;overflow:visible}}.report-page:last-child{{break-after:auto;page-break-after:auto}}.page-content{{padding:8mm 10mm 14mm}}.report-masthead{{padding:8mm 10mm}}.report-status{{padding:2.5mm 10mm}}.report-section{{break-inside:auto}}.section-heading,.maturity-comparison,.metric-strip,.radar-grid,.narrative-grid article,figure,table{{break-inside:avoid}}.table-wrap{{overflow:visible}}table.appendix{{min-width:0;font-size:6.6pt}}table.dense{{font-size:6.8pt}}.editable:hover{{background:transparent}}.report-footer{{left:10mm;right:10mm;bottom:4mm}}}}
{executive_report_css}
</style></head><body><main data-report-model='sapd-maturity-report-model-v2'>
<article class="report-page page-one"><header class="brief-masthead"><div><p class="kicker">SAPD WIKI · MATURITY ASSESSMENT REPORT</p><h1>{h(project_name)}｜评估报告</h1><p class="meta">{report_meta}</p></div><span class="report-badge">{'正式评估报告' if is_formal else '草稿预览'}</span></header><div class="page-content">{render_sections('overall', 'radars')}</div><footer class="report-footer"><span>{report_footer}</span><span>第 1 页</span></footer></article>
<article class="report-page page-two"><div class="page-content"><header class="page-intro"><div><span>MANAGEMENT JUDGEMENT</span><h2>管理层研判与改进安排</h2></div><p>人工填写内容与正式评估结果共同构成报告正文</p></header>{render_sections('narratives', 'improvement_roadmap')}</div><footer class="report-footer"><span>{report_footer}</span><span>第 2 页</span></footer></article>
<article class="report-page page-three is-appendix"><div class="page-content"><header class="page-intro"><div><span>APPENDIX · PAGE THREE</span><h2>维度数据与完整评估附录</h2></div><p>各维度统计、能力结果、排名与评分明细</p></header>{render_sections('hierarchy_statistics', 'evaluation', 'dimension_rankings', 'capability_results', 'overall_rankings', 'score_appendix', 'traceability')}</div><footer class="report-footer"><span>{report_footer}</span><span>第 3 页 · 附录</span></footer></article>
</main>{executive_report_script}</body></html>"""

    return {
        "ok": True,
        "dataState": "ready",
        "id": snapshot_id,
        "status": "snapshot" if is_formal else "draft_preview",
        "generatedAt": generated_at,
        "formal": is_formal,
        "summary": summary,
        "improvementRoadmap": improvement_roadmap,
        "reportModel": report_model,
        "markdown": markdown,
        "html": html_report,
        "fileNames": {
            "markdown": f"{snapshot_id}.md",
            "html": f"{snapshot_id}.html",
        },
    }
