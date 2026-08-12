from __future__ import annotations

import json
import sqlite3
import unittest

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from sapd_wiki.candidates import ObjectCandidate
from sapd_wiki.loader import _existing_relation_for_upsert
from sapd_wiki.parsers import parse_data_lifecycle_mapping_sheet, parse_scene_sheet
from sapd_wiki.staging import _match_item


def _connection() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    return conn


class EnvironmentSegmentMatchingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.conn = _connection()
        self.conn.execute(
            """
            CREATE TABLE knowledge_items (
              id TEXT PRIMARY KEY,
              type TEXT NOT NULL,
              code TEXT,
              title TEXT NOT NULL,
              metadata_json TEXT
            )
            """
        )

    def tearDown(self) -> None:
        self.conn.close()

    def _insert_segment(self, item_id: str, environment: str) -> None:
        object_key = f"environment_segment::::网络::{environment}"
        self.conn.execute(
            """
            INSERT INTO knowledge_items (id, type, title, metadata_json)
            VALUES (?, 'environment_segment', '网络', ?)
            """,
            (item_id, json.dumps({"object_key": object_key}, ensure_ascii=False)),
        )

    def test_matches_exact_environment_context_independent_of_insert_order(self) -> None:
        self._insert_segment("segment-b", "云数据中心")
        self._insert_segment("segment-a", "园区网")

        matched = _match_item(
            self.conn,
            ObjectCandidate(type="environment_segment", title="网络", qualifier="园区网"),
        )

        self.assertEqual(matched, "segment-a")

    def test_does_not_fall_back_to_same_title_in_other_environment(self) -> None:
        self._insert_segment("segment-a", "园区网")

        matched = _match_item(
            self.conn,
            ObjectCandidate(type="environment_segment", title="网络", qualifier="云数据中心"),
        )

        self.assertIsNone(matched)

    def test_requires_environment_qualifier(self) -> None:
        with self.assertRaisesRegex(ValueError, "qualifier"):
            _match_item(
                self.conn,
                ObjectCandidate(type="environment_segment", title="网络"),
            )

    def test_rejects_duplicate_exact_context_identity(self) -> None:
        self._insert_segment("segment-a", "园区网")
        self._insert_segment("segment-b", "园区网")

        with self.assertRaisesRegex(ValueError, "上下文身份重复"):
            _match_item(
                self.conn,
                ObjectCandidate(type="environment_segment", title="网络", qualifier="园区网"),
            )

    def test_other_item_types_keep_existing_title_matching(self) -> None:
        self.conn.execute(
            """
            INSERT INTO knowledge_items (id, type, title, metadata_json)
            VALUES ('object-a', 'information_object', '网络设备', '{}')
            """
        )

        matched = _match_item(
            self.conn,
            ObjectCandidate(type="information_object", title="网络设备"),
        )

        self.assertEqual(matched, "object-a")


class EnvironmentSegmentTypeIdentityTests(unittest.TestCase):
    def test_code_is_stable_identity_when_title_changes(self) -> None:
        before = ObjectCandidate(
            type="environment_segment_type",
            code="ES-001",
            title="互联网边界",
        )
        after = ObjectCandidate(
            type="environment_segment_type",
            code="ES-001",
            title="互联网连接边界",
        )

        self.assertEqual(before.key, "environment_segment_type::ES-001")
        self.assertEqual(after.key, before.key)


class InstanceOfCardinalityTests(unittest.TestCase):
    def setUp(self) -> None:
        self.conn = _connection()
        self.conn.execute(
            """
            CREATE TABLE knowledge_relations (
              id TEXT PRIMARY KEY,
              source_item_id TEXT NOT NULL,
              target_item_id TEXT NOT NULL,
              relation_type TEXT NOT NULL
            )
            """
        )

    def tearDown(self) -> None:
        self.conn.close()

    def test_reuses_same_instance_of_target(self) -> None:
        self.conn.execute(
            """
            INSERT INTO knowledge_relations
              (id, source_item_id, target_item_id, relation_type)
            VALUES ('relation-a', 'segment-a', 'type-a', 'instance_of')
            """
        )

        relation_id = _existing_relation_for_upsert(
            self.conn,
            source_item_id="segment-a",
            target_item_id="type-a",
            relation_type="instance_of",
        )

        self.assertEqual(relation_id, "relation-a")

    def test_rejects_instance_of_target_change(self) -> None:
        self.conn.execute(
            """
            INSERT INTO knowledge_relations
              (id, source_item_id, target_item_id, relation_type)
            VALUES ('relation-a', 'segment-a', 'type-a', 'instance_of')
            """
        )

        with self.assertRaisesRegex(ValueError, "目标变更"):
            _existing_relation_for_upsert(
                self.conn,
                source_item_id="segment-a",
                target_item_id="type-b",
                relation_type="instance_of",
            )

    def test_rejects_existing_multiple_instance_of_relations(self) -> None:
        self.conn.executemany(
            """
            INSERT INTO knowledge_relations
              (id, source_item_id, target_item_id, relation_type)
            VALUES (?, 'segment-a', ?, 'instance_of')
            """,
            [("relation-a", "type-a"), ("relation-b", "type-b")],
        )

        with self.assertRaisesRegex(ValueError, "多重关系"):
            _existing_relation_for_upsert(
                self.conn,
                source_item_id="segment-a",
                target_item_id="type-a",
                relation_type="instance_of",
            )


class SceneParserParentResetTests(unittest.TestCase):
    def test_environment_change_does_not_inherit_previous_segment(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "作用域-安全技术服务-安全技术模块映射"
        sheet.append(["", "", "", "", "", "", "", ""])
        sheet.append(["", "", "", "", "", "", "", ""])
        sheet.append(["", "园区网", "网络", "办公终端", "", "", "", ""])
        sheet.append(["", "远程办公", "", "移动办公终端", "", "", "", ""])

        result = parse_scene_sheet(workbook)

        forbidden_key = "environment_segment::::网络::远程办公"
        self.assertNotIn(forbidden_key, {item.key for item in result.objects})
        self.assertIn(
            (
                "information_object::::移动办公终端",
                "belongs_to",
                "information_environment::::远程办公",
            ),
            {
                (relation.source_key, relation.relation_type, relation.target_key)
                for relation in result.relations
            },
        )

    @staticmethod
    def _measure_fill() -> PatternFill:
        return PatternFill(
            fill_type="solid",
            fgColor=Color(
                type="theme",
                theme=6,
                tint=0.5999938962981048,
            ),
        )

    def test_measure_relations_use_local_merged_anchor_and_both_endpoint_sources(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "作用域-安全技术服务-安全技术模块映射"
        sheet.append(["", "", "", "", "", "", "", ""])
        sheet.append(["", "", "", "", "", "", "", ""])
        sheet.append(["", "园区网", "网络", "认证场景", "I-AP 应用系统", "I-AP&T-AS.IA-02 应用身份认证", "应用系统自身认证模块", ""])
        sheet.append(["", "", "", "", "", "I-US&T-AS.IA-02 用户认证", "", ""])
        sheet.append(["", "", "", "另一个场景", "I-DI 数据", "I-DI&T-AS.IA-03 数据资源授权", "", ""])
        sheet.merge_cells("D3:D4")
        sheet.merge_cells("E3:E4")
        sheet.merge_cells("G3:G4")
        sheet["G3"].fill = self._measure_fill()

        result = parse_scene_sheet(workbook)
        relations = [
            relation
            for relation in result.relations
            if relation.relation_type == "uses_measure"
        ]

        self.assertEqual(
            {(relation.source_key, relation.target_key) for relation in relations},
            {
                (
                    "security_technical_service::I-AP&T-AS.IA-02",
                    "security_technical_measure::::应用系统自身认证模块",
                ),
                (
                    "security_technical_service::I-US&T-AS.IA-02",
                    "security_technical_measure::::应用系统自身认证模块",
                ),
            },
        )
        self.assertFalse(
            any("I-DI&T-AS.IA-03" in relation.source_key for relation in relations)
        )
        for relation in relations:
            self.assertEqual(
                {source.column for source in relation.sources},
                {"安全技术服务", "安全技术模块/措施"},
            )
            self.assertTrue(any(source.cell == "G3" for source in relation.sources))

    def test_multivalue_lcdt_row_does_not_create_service_measure_cartesian_product(self) -> None:
        workbook = Workbook()
        lifecycle = workbook.active
        lifecycle.title = "LC-DT 数据生命周期"
        for _ in range(3):
            lifecycle.append([])
        lifecycle.append([1, "数据使用"])
        mapping = workbook.create_sheet("LC-DT 安全技术服务、模块、策略映射表")
        for _ in range(5):
            mapping.append([])
        mapping.cell(6, 2, "认证")
        mapping.cell(
            6,
            13,
            "I-US&T-AS.IA-02 用户认证\nI-DI&T-AS.IA-03 数据资源授权",
        )
        mapping.cell(6, 14, "应用系统自身认证模块")

        result = parse_data_lifecycle_mapping_sheet(workbook)

        self.assertFalse(
            any(relation.relation_type == "uses_measure" for relation in result.relations)
        )


if __name__ == "__main__":
    unittest.main()
